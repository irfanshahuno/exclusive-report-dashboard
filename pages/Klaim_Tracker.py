import streamlit as st
import pandas as pd
import boto3
import io
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Klaim Financial Tracker", layout="wide", page_icon="💊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.main { background: #f0f4f8; }
.block-container { padding-top: 1rem; }

.kpi-card {
    background: white; border-radius: 12px; padding: 18px 22px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07); margin-bottom: 10px;
    border-left: 5px solid #1F3864;
}
.kpi-card.loss  { border-left-color: #e74c3c; }
.kpi-card.gain  { border-left-color: #27ae60; }
.kpi-card.warn  { border-left-color: #f39c12; }
.kpi-card.blue  { border-left-color: #2980b9; }
.kpi-label { font-size: 11px; color: #888; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
.kpi-value { font-size: 24px; font-weight: 700; color: #1F3864; margin: 4px 0 0 0; }
.kpi-value.red    { color: #e74c3c; }
.kpi-value.green  { color: #27ae60; }
.kpi-value.orange { color: #f39c12; }
.kpi-sub { font-size: 11px; color: #aaa; margin-top: 2px; }

.section-hdr {
    font-size: 15px; font-weight: 700; color: #1F3864;
    border-bottom: 2px solid #1F3864; padding-bottom: 5px; margin: 20px 0 12px 0;
}
.pill-paid     { background:#e2f4ea; color:#27ae60; padding:2px 10px; border-radius:20px; font-size:11px; font-weight:600; }
.pill-rejected { background:#fde8e8; color:#e74c3c; padding:2px 10px; border-radius:20px; font-size:11px; font-weight:600; }
.pill-pending  { background:#fff4e0; color:#f39c12; padding:2px 10px; border-radius:20px; font-size:11px; font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ── S3 helpers ───────────────────────────────────────────────────────────
BUCKET = "emc-rcm-storage-2026"
KLAIM_PREFIX   = "klaim-tracker/klaim-files/"
BILLING_PREFIX = "klaim-tracker/billing-files/"

@st.cache_resource
def get_s3():
    return boto3.client(
        "s3",
        aws_access_key_id=st.secrets["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=st.secrets["AWS_SECRET_ACCESS_KEY"],
        region_name=st.secrets.get("AWS_REGION", "us-east-1"),
    )

def list_s3_files(prefix):
    s3 = get_s3()
    resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=prefix)
    files = []
    for obj in resp.get("Contents", []):
        key = obj["Key"]
        if key != prefix:
            files.append(key)
    return files

def read_s3_file(key):
    s3 = get_s3()
    obj = s3.get_object(Bucket=BUCKET, Key=key)
    data = obj["Body"].read()
    if key.endswith(".csv"):
        return pd.read_csv(io.BytesIO(data))
    else:
        return pd.read_excel(io.BytesIO(data))

def upload_s3(file_bytes, key):
    s3 = get_s3()
    s3.put_object(Bucket=BUCKET, Key=key, Body=file_bytes)

def clean_cols(df):
    df.columns = df.columns.str.strip()
    return df

# ── Financial calculations ───────────────────────────────────────────────
def calc_klaim_metrics(kdf):
    """Given a Klaim dataframe, return per-claim financial summary."""
    kdf = clean_cols(kdf)
    kdf["Claim net"]                  = pd.to_numeric(kdf.get("Claim net", 0), errors="coerce").fillna(0)
    kdf["Paid by insurance"]          = pd.to_numeric(kdf.get("Paid by insurance", 0), errors="coerce").fillna(0)
    kdf["Denied by insurance"]        = pd.to_numeric(kdf.get("Denied by insurance", 0), errors="coerce").fillna(0)
    kdf["Pending insurance response"] = pd.to_numeric(kdf.get("Pending insurance response", 0), errors="coerce").fillna(0)
    kdf["Pending reconciliation"]     = pd.to_numeric(kdf.get("Pending reconciliation", 0), errors="coerce").fillna(0)
    return kdf

def merge_billing_klaim(bdf, kdf):
    """Merge billing detail with klaim file on Claim ID = UniqueID."""
    bdf = clean_cols(bdf)
    kdf = clean_cols(kdf)
    bdf["SubInsShare"] = pd.to_numeric(bdf.get("SubInsShare", 0), errors="coerce").fillna(0)

    # Aggregate billing to claim level
    billing_grp = bdf.groupby("UniqueID").agg(
        Insurance    =("Insurance", "first"),
        Doctor       =("DocName", "first"),
        Department   =("DepName", "first"),
        SubDate      =("SubDate", "first"),
        Month        =("Month", "first"),
        Year         =("Year", "first"),
        Billing_Net  =("SubInsShare", "sum"),
        Activities   =("UniqueID", "count"),
    ).reset_index()

    # Aggregate klaim to claim level
    kdf = calc_klaim_metrics(kdf)
    klaim_grp = kdf.groupby("Claim ID").agg(
        Payer            =("Payer", "first"),
        Deal_Reference   =("Deal reference", "first"),
        Deal_Date        =("Deal date", "first"),
        Klaim_Net        =("Claim net", "sum"),
        Paid_Insurance   =("Paid by insurance", "sum"),
        Denied_Insurance =("Denied by insurance", "sum"),
        Pending_Response =("Pending insurance response", "sum"),
        Pending_Recon    =("Pending reconciliation", "sum"),
        Status           =("Status", "first"),
    ).reset_index()

    merged = pd.merge(klaim_grp, billing_grp,
                      left_on="Claim ID", right_on="UniqueID", how="left")

    # Financial metrics
    merged["Discount_Loss"]     = merged["Klaim_Net"] - merged["Paid_Insurance"]
    merged["Rejection_Charge"]  = merged["Denied_Insurance"]
    merged["Net_Position"]      = merged["Paid_Insurance"] - merged["Rejection_Charge"]
    return merged

# ── KPI card helper ──────────────────────────────────────────────────────
def kpi(label, value, style="", sub="", fmt="aed"):
    if fmt == "aed":
        v_str = f"AED {value:,.0f}"
    elif fmt == "pct":
        v_str = f"{value:.1f}%"
    else:
        v_str = f"{value:,.0f}"
    color = {"loss":"red","gain":"green","warn":"orange"}.get(style,"")
    st.markdown(f"""
    <div class="kpi-card {style}">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value {color}">{v_str}</div>
        {"<div class='kpi-sub'>"+sub+"</div>" if sub else ""}
    </div>""", unsafe_allow_html=True)

# ── Header ───────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:linear-gradient(135deg,#1F3864,#2c5f8c);padding:22px 30px;
     border-radius:14px;margin-bottom:20px;">
  <h1 style="color:white;margin:0;font-size:24px;font-weight:700;">💊 Klaim Financial Tracker</h1>
  <p style="color:#a8c4e0;margin:4px 0 0 0;font-size:13px;">
     Track discount losses, rejection charges &amp; net position across all RPA runs
  </p>
</div>
""", unsafe_allow_html=True)

# ── Tabs ─────────────────────────────────────────────────────────────────
tab_overview, tab_rpa, tab_claims, tab_upload = st.tabs([
    "📊 Overall Exposure", "📋 RPA Level", "🔍 Claim Level", "⬆️ Upload Files"
])

# ════════════════════════════════════════════════════════════════════════
# UPLOAD TAB
# ════════════════════════════════════════════════════════════════════════
with tab_upload:
    st.markdown('<div class="section-hdr">Upload New Files to S3</div>', unsafe_allow_html=True)

    col_k, col_b = st.columns(2)

    with col_k:
        st.markdown("#### 📄 Klaim File (CSV or Excel)")
        facility_k = st.selectbox("Facility", ["EXCELLENT", "PHARMACY", "EASYHEALTH"], key="fac_k")
        rpa_ref_k  = st.text_input("RPA Reference (e.g. 2603/008173/EXCELLENT/0145)", key="rpa_ref_k")
        klaim_file = st.file_uploader("Upload Klaim File", type=["csv","xlsx"], key="kf")
        if klaim_file and rpa_ref_k:
            if st.button("Upload Klaim File to S3", key="btn_k"):
                safe_ref = rpa_ref_k.replace("/", "_")
                key = f"{KLAIM_PREFIX}{facility_k}/{safe_ref}_{klaim_file.name}"
                upload_s3(klaim_file.read(), key)
                st.success(f"✅ Uploaded: {key}")

    with col_b:
        st.markdown("#### 📊 Billing File (Excel)")
        facility_b = st.selectbox("Facility", ["EXCELLENT", "PHARMACY", "EASYHEALTH"], key="fac_b")
        rpa_ref_b  = st.text_input("RPA Reference (matching Klaim file)", key="rpa_ref_b")
        billing_file = st.file_uploader("Upload Billing File", type=["xlsx","csv"], key="bf")
        if billing_file and rpa_ref_b:
            if st.button("Upload Billing File to S3", key="btn_b"):
                safe_ref = rpa_ref_b.replace("/", "_")
                key = f"{BILLING_PREFIX}{facility_b}/{safe_ref}_{billing_file.name}"
                upload_s3(billing_file.read(), key)
                st.success(f"✅ Uploaded: {key}")

    st.info("💡 Files are stored in S3 and automatically available in all tabs once uploaded.")

# ── Load all data from S3 ────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_all_data():
    """Load and combine all Klaim + Billing files from S3."""
    try:
        klaim_files   = list_s3_files(KLAIM_PREFIX)
        billing_files = list_s3_files(BILLING_PREFIX)
    except Exception as e:
        return None, None, str(e)

    all_klaim   = []
    all_billing = []

    for key in klaim_files:
        try:
            df = read_s3_file(key)
            df = clean_cols(df)
            # Extract facility and RPA ref from path
            parts = key.replace(KLAIM_PREFIX,"").split("/")
            df["_facility"] = parts[0] if len(parts) > 0 else "UNKNOWN"
            df["_s3_key"]   = key
            all_klaim.append(df)
        except:
            pass

    for key in billing_files:
        try:
            df = read_s3_file(key)
            df = clean_cols(df)
            parts = key.replace(BILLING_PREFIX,"").split("/")
            df["_facility"] = parts[0] if len(parts) > 0 else "UNKNOWN"
            df["_s3_key"]   = key
            all_billing.append(df)
        except:
            pass

    klaim_all   = pd.concat(all_klaim,   ignore_index=True) if all_klaim   else pd.DataFrame()
    billing_all = pd.concat(all_billing, ignore_index=True) if all_billing else pd.DataFrame()
    return klaim_all, billing_all, None

klaim_all, billing_all, load_err = load_all_data()

# ── Fallback: use uploaded sample files for demo ─────────────────────────
if klaim_all is None or klaim_all.empty:
    st.warning("⚠️ No data loaded from S3. Showing demo mode — upload files in the Upload tab.")
    # Demo: use session state uploaded files if available
    klaim_all   = pd.DataFrame()
    billing_all = pd.DataFrame()

# ── Helper: merge all data ───────────────────────────────────────────────
def get_merged_all(kdf, bdf):
    if kdf.empty:
        return pd.DataFrame()
    kdf = calc_klaim_metrics(kdf)
    # Aggregate klaim
    klaim_grp = kdf.groupby(["Claim ID","_facility"]).agg(
        Payer            =("Payer", "first"),
        Deal_Reference   =("Deal reference", "first"),
        Deal_Date        =("Deal date", "first"),
        Klaim_Net        =("Claim net", "sum"),
        Paid_Insurance   =("Paid by insurance", "sum"),
        Denied_Insurance =("Denied by insurance", "sum"),
        Pending_Response =("Pending insurance response", "sum"),
        Status           =("Status", "first"),
    ).reset_index()

    if not bdf.empty and "UniqueID" in bdf.columns:
        bdf["SubInsShare"] = pd.to_numeric(bdf.get("SubInsShare",0), errors="coerce").fillna(0)
        billing_grp = bdf.groupby("UniqueID").agg(
            Insurance  =("Insurance","first"),
            Doctor     =("DocName","first"),
            Department =("DepName","first"),
            Month      =("Month","first"),
            Year       =("Year","first"),
        ).reset_index()
        merged = pd.merge(klaim_grp, billing_grp, left_on="Claim ID", right_on="UniqueID", how="left")
    else:
        merged = klaim_grp.copy()
        merged["Insurance"] = merged.get("Payer", "")

    merged["Discount_Loss"]    = (merged["Klaim_Net"] - merged["Paid_Insurance"]).clip(lower=0)
    merged["Rejection_Charge"] = merged["Denied_Insurance"]
    merged["Net_Position"]     = merged["Paid_Insurance"] - merged["Rejection_Charge"]
    merged["Deal_Date"]        = pd.to_datetime(merged["Deal_Date"], errors="coerce")
    merged["Month_Year"]       = merged["Deal_Date"].dt.to_period("M").astype(str)
    return merged

merged_all = get_merged_all(klaim_all, billing_all)

# ════════════════════════════════════════════════════════════════════════
# OVERALL EXPOSURE TAB
# ════════════════════════════════════════════════════════════════════════
with tab_overview:
    if merged_all.empty:
        st.info("📂 No data yet. Please upload Klaim and Billing files in the **Upload Files** tab.")
    else:
        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        facilities = ["All"] + sorted(merged_all["_facility"].dropna().unique().tolist())
        sel_fac    = col_f1.selectbox("Facility", facilities, key="ov_fac")
        months     = ["All"] + sorted(merged_all["Month_Year"].dropna().unique().tolist(), reverse=True)
        sel_month  = col_f2.selectbox("Month", months, key="ov_month")
        insurers   = ["All"] + sorted(merged_all["Payer"].dropna().unique().tolist())
        sel_ins    = col_f3.selectbox("Insurer (Klaim)", insurers, key="ov_ins")

        df_ov = merged_all.copy()
        if sel_fac   != "All": df_ov = df_ov[df_ov["_facility"]     == sel_fac]
        if sel_month != "All": df_ov = df_ov[df_ov["Month_Year"]    == sel_month]
        if sel_ins   != "All": df_ov = df_ov[df_ov["Payer"]         == sel_ins]

        st.markdown('<div class="section-hdr">Overall KPIs</div>', unsafe_allow_html=True)
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1: kpi("Total Claims Sold",       len(df_ov),                       fmt="num")
        with c2: kpi("Total Value Sold",         df_ov["Klaim_Net"].sum(),          style="blue")
        with c3: kpi("Total Discount Loss",      df_ov["Discount_Loss"].sum(),      style="loss", sub="Lost by selling early")
        with c4: kpi("Total Rejection Charges",  df_ov["Rejection_Charge"].sum(),   style="warn", sub="Clawed back by Klaim")
        with c5:
            net = df_ov["Net_Position"].sum()
            kpi("Net Position", net, style="gain" if net >= 0 else "loss")

        # Breakdown by facility
        st.markdown('<div class="section-hdr">Breakdown by Facility</div>', unsafe_allow_html=True)
        fac_grp = df_ov.groupby("_facility").agg(
            Claims   =("Claim ID","count"),
            Value    =("Klaim_Net","sum"),
            Discount =("Discount_Loss","sum"),
            Rejected =("Rejection_Charge","sum"),
            Net      =("Net_Position","sum"),
        ).reset_index().rename(columns={"_facility":"Facility"})
        fac_grp["Discount %"] = (fac_grp["Discount"] / fac_grp["Value"] * 100).round(2)
        st.dataframe(
            fac_grp.style
                .format({"Value":"AED {:,.0f}","Discount":"AED {:,.0f}","Rejected":"AED {:,.0f}",
                         "Net":"AED {:,.0f}","Discount %":"{:.1f}%"})
                .background_gradient(subset=["Discount %"], cmap="Reds"),
            use_container_width=True, hide_index=True
        )

        # Breakdown by Insurer
        st.markdown('<div class="section-hdr">Breakdown by Insurer</div>', unsafe_allow_html=True)
        ins_grp = df_ov.groupby("Payer").agg(
            Claims   =("Claim ID","count"),
            Value    =("Klaim_Net","sum"),
            Discount =("Discount_Loss","sum"),
            Rejected =("Rejection_Charge","sum"),
            Net      =("Net_Position","sum"),
        ).reset_index().rename(columns={"Payer":"Insurer"}).sort_values("Value", ascending=False)
        ins_grp["Discount %"] = (ins_grp["Discount"] / ins_grp["Value"] * 100).round(2)
        st.dataframe(
            ins_grp.style
                .format({"Value":"AED {:,.0f}","Discount":"AED {:,.0f}","Rejected":"AED {:,.0f}",
                         "Net":"AED {:,.0f}","Discount %":"{:.1f}%"})
                .background_gradient(subset=["Discount %"], cmap="Reds"),
            use_container_width=True, hide_index=True
        )

        # Monthly trend
        if "Month_Year" in df_ov.columns:
            st.markdown('<div class="section-hdr">Monthly Trend</div>', unsafe_allow_html=True)
            trend = df_ov.groupby("Month_Year").agg(
                Value    =("Klaim_Net","sum"),
                Discount =("Discount_Loss","sum"),
                Rejected =("Rejection_Charge","sum"),
            ).reset_index().sort_values("Month_Year")
            st.bar_chart(trend.set_index("Month_Year")[["Value","Discount","Rejected"]])

# ════════════════════════════════════════════════════════════════════════
# RPA LEVEL TAB
# ════════════════════════════════════════════════════════════════════════
with tab_rpa:
    if merged_all.empty:
        st.info("📂 No data yet. Please upload files in the **Upload Files** tab.")
    else:
        st.markdown('<div class="section-hdr">Select Facility & RPA Run</div>', unsafe_allow_html=True)
        col_r1, col_r2 = st.columns(2)
        facs_rpa  = sorted(merged_all["_facility"].dropna().unique().tolist())
        sel_fac_r = col_r1.selectbox("Facility", facs_rpa, key="rpa_fac")
        df_rpa    = merged_all[merged_all["_facility"] == sel_fac_r]

        rpa_list  = sorted(df_rpa["Deal_Reference"].dropna().unique().tolist(), reverse=True)
        sel_rpa   = col_r2.selectbox("RPA Reference", rpa_list, key="rpa_sel")

        df_sel = df_rpa[df_rpa["Deal_Reference"] == sel_rpa]

        st.markdown(f'<div class="section-hdr">RPA: {sel_rpa}</div>', unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi("Claims in this RPA",   len(df_sel),                     fmt="num")
        with c2: kpi("Total Value",           df_sel["Klaim_Net"].sum())
        with c3: kpi("Discount Loss",         df_sel["Discount_Loss"].sum(),   style="loss")
        with c4: kpi("Rejection Charges",     df_sel["Rejection_Charge"].sum(),style="warn")

        # Status breakdown
        st.markdown('<div class="section-hdr">Status Breakdown</div>', unsafe_allow_html=True)
        stat_grp = df_sel.groupby("Status").agg(
            Claims=("Claim ID","count"),
            Value =("Klaim_Net","sum"),
        ).reset_index()
        st.dataframe(stat_grp.style.format({"Value":"AED {:,.0f}"}),
                     use_container_width=True, hide_index=True)

        # Insurer breakdown for this RPA
        st.markdown('<div class="section-hdr">By Insurer</div>', unsafe_allow_html=True)
        ins_rpa = df_sel.groupby("Payer").agg(
            Claims   =("Claim ID","count"),
            Value    =("Klaim_Net","sum"),
            Discount =("Discount_Loss","sum"),
            Rejected =("Rejection_Charge","sum"),
        ).reset_index()
        ins_rpa["Discount %"] = (ins_rpa["Discount"] / ins_rpa["Value"] * 100).round(2)
        st.dataframe(
            ins_rpa.style.format({"Value":"AED {:,.0f}","Discount":"AED {:,.0f}",
                                  "Rejected":"AED {:,.0f}","Discount %":"{:.1f}%"}),
            use_container_width=True, hide_index=True
        )

# ════════════════════════════════════════════════════════════════════════
# CLAIM LEVEL TAB
# ════════════════════════════════════════════════════════════════════════
with tab_claims:
    if merged_all.empty:
        st.info("📂 No data yet. Please upload files in the **Upload Files** tab.")
    else:
        st.markdown('<div class="section-hdr">Claim-Level Detail</div>', unsafe_allow_html=True)

        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        facs_cl   = ["All"] + sorted(merged_all["_facility"].dropna().unique().tolist())
        sel_fac_c = col_c1.selectbox("Facility", facs_cl, key="cl_fac")
        rpas_cl   = ["All"] + sorted(merged_all["Deal_Reference"].dropna().unique().tolist(), reverse=True)
        sel_rpa_c = col_c2.selectbox("RPA", rpas_cl, key="cl_rpa")
        stats_cl  = ["All"] + sorted(merged_all["Status"].dropna().unique().tolist())
        sel_sta_c = col_c3.selectbox("Status", stats_cl, key="cl_sta")
        payers_cl = ["All"] + sorted(merged_all["Payer"].dropna().unique().tolist())
        sel_pay_c = col_c4.selectbox("Payer", payers_cl, key="cl_pay")

        df_cl = merged_all.copy()
        if sel_fac_c != "All": df_cl = df_cl[df_cl["_facility"]     == sel_fac_c]
        if sel_rpa_c != "All": df_cl = df_cl[df_cl["Deal_Reference"] == sel_rpa_c]
        if sel_sta_c != "All": df_cl = df_cl[df_cl["Status"]         == sel_sta_c]
        if sel_pay_c != "All": df_cl = df_cl[df_cl["Payer"]          == sel_pay_c]

        display_cols = {
            "Claim ID"       : "Claim ID",
            "_facility"      : "Facility",
            "Deal_Reference" : "RPA Reference",
            "Deal_Date"      : "Deal Date",
            "Payer"          : "Insurer",
            "Klaim_Net"      : "Claim Net (AED)",
            "Paid_Insurance" : "Paid (AED)",
            "Denied_Insurance":"Rejected (AED)",
            "Pending_Response":"Pending (AED)",
            "Discount_Loss"  : "Discount Loss (AED)",
            "Rejection_Charge":"Rejection Charge (AED)",
            "Net_Position"   : "Net Position (AED)",
            "Status"         : "Status",
        }
        # Add billing cols if available
        if "Insurance" in df_cl.columns:
            display_cols["Insurance"] = "Insurance (Billing)"
        if "Doctor" in df_cl.columns:
            display_cols["Doctor"] = "Doctor"

        df_show = df_cl[[c for c in display_cols.keys() if c in df_cl.columns]].copy()
        df_show = df_show.rename(columns=display_cols)

        st.markdown(f"**{len(df_show):,} claims** | Total Value: **AED {df_cl['Klaim_Net'].sum():,.0f}** | "
                    f"Discount Lost: **AED {df_cl['Discount_Loss'].sum():,.0f}** | "
                    f"Rejected: **AED {df_cl['Rejection_Charge'].sum():,.0f}**")

        def color_status(val):
            if val == "Paid":     return "background-color:#e2f4ea;color:#27ae60;font-weight:600"
            if val == "Rejected": return "background-color:#fde8e8;color:#e74c3c;font-weight:600"
            if val == "Pending":  return "background-color:#fff4e0;color:#f39c12;font-weight:600"
            return ""

        aed_cols = [c for c in df_show.columns if "AED" in c]
        fmt_dict = {c: "AED {:,.0f}" for c in aed_cols}

        st.dataframe(
            df_show.style
                .applymap(color_status, subset=["Status"] if "Status" in df_show.columns else [])
                .format(fmt_dict),
            use_container_width=True, hide_index=True, height=500
        )

        # Excel download
        if st.button("📥 Download as Excel", key="dl_claims"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Claim Detail"
            NAVY = "1F3864"
            headers = df_show.columns.tolist()
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font    = Font(bold=True, color="FFFFFF", name="Arial")
                c.fill    = PatternFill("solid", start_color=NAVY)
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = max(len(h)+4, 14)

            for ri, row in enumerate(df_show.itertuples(index=False), 2):
                bg = "F2F2F2" if ri % 2 == 0 else "FFFFFF"
                for ci, v in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=v)
                    cell.font      = Font(name="Arial", size=10)
                    cell.fill      = PatternFill("solid", start_color=bg)
                    cell.alignment = Alignment(horizontal="center")

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button("⬇️ Download Excel", buf,
                               file_name=f"klaim_claims_{datetime.now().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
