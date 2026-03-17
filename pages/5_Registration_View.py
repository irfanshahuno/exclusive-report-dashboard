#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Streamlit Page: Registration Summary (View Only)

Purpose
- Management should ONLY view results (no upload).
- Loads the latest saved summary from S3 created by:
    pages/4_Registration_Summary.py  (Process & Save to S3)

Important
- This viewer MUST read the SAME S3 folder structure as the uploader page:
    registration/<center>/<YYYY-MM-DD>/summary.pkl
    registration/<center>/history.csv

So we intentionally IGNORE any `year=` query param for storage paths, unless you
also change the uploader to save year-wise.
"""

import io
import os
import re
import pickle
from datetime import datetime, date
from typing import Dict, Optional, List, Tuple

import pandas as pd
import streamlit as st


# ==========================
# Email helpers (SMTP)
# ==========================
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib


def _dfs_to_html(dfs: dict, title: str, picked_label: str) -> str:
    """HTML email body — exact same colors as Excel:
    - Dark navy (#0D1B2A) headers, orange (#FF6600) grand total, light-blue (#F0F4FF) alt rows
    - KPI cards matching Streamlit UI (white card, navy label, bold number)
    - Doctor x Insurance: doctor name shown once via HTML rowspan
    """
    def _num(x, default=0):
        try:
            return float(pd.to_numeric(x, errors="coerce"))
        except Exception:
            return default

    def _kpi_value(metric, default=0):
        kpi = dfs.get("KPI")
        if isinstance(kpi, pd.DataFrame) and not kpi.empty and {"Metric", "Value"}.issubset(kpi.columns):
            try:
                return kpi.set_index("Metric")["Value"].get(metric, default)
            except Exception:
                return default
        return default

    def _safe_df(df):
        if not isinstance(df, pd.DataFrame) or df.empty:
            return pd.DataFrame()
        out = df.copy()
        bad = [c for c in out.columns if str(c).strip() == "" or str(c).strip().lower().startswith("unnamed")]
        out = out.drop(columns=bad, errors="ignore")
        # Drop misleading % cols
        for drop_c in ["Radiology_%", "Procedure_%"]:
            if drop_c in out.columns:
                out = out.drop(columns=[drop_c])
        # Compute per-visit ratios
        visit_col = next((c for c in out.columns if str(c).strip().lower() in
                          ["total_visit", "total visit", "visits", "visit"]), None)
        if visit_col is not None:
            denom = pd.to_numeric(out[visit_col], errors="coerce").replace(0, pd.NA)
            if "Procedure" in out.columns and "Procedure_Per_Visit" not in out.columns:
                out["Procedure_Per_Visit"] = (pd.to_numeric(out["Procedure"], errors="coerce") / denom).round(2).fillna(0)
            if "Radiology" in out.columns and "Radiology_Per_Visit" not in out.columns:
                out["Radiology_Per_Visit"] = (pd.to_numeric(out["Radiology"], errors="coerce") / denom).round(2).fillna(0)
        # Rename to match Excel display names
        _EMAIL_RENAME = {
            "Total_Amount_Service":   "Total Service",
            "Total_Amount_Insuance":  "Total Insurance",
            "Total_Amount_Insurance": "Total Insurance",
            "Avg_Amount_Service":     "Avg Service",
            "Avg_Amount_Insuance":    "Avg Insurance",
            "Avg_Amount_Insurance":   "Avg Insurance",
            "Avg.Amount":             "Avg Insurance",
            "Lab_%":                  "Lab %",
            "Procedure_Per_Visit":    "Procedure %",
            "Radiology_Per_Visit":    "Radiology %",
            "Total_Visit":            "Visits",
            "Department":             "Dept",
        }
        out = out.rename(columns={k: v for k, v in _EMAIL_RENAME.items() if k in out.columns})
        # Remove any duplicate columns (keep first occurrence)
        out = out.loc[:, ~out.columns.duplicated(keep="first")]
        # Reorder to match Excel
        _ORDER = [
            "Dept", "Doctor", "Insurance",
            "Consultation", "Lab", "Radiology", "Procedure",
            "Visits",
            "Total Service", "Total Insurance",
            "Avg Service", "Avg Insurance",
            "Lab %", "Procedure %", "Radiology %",
        ]
        ordered = [c for c in _ORDER if c in out.columns]
        remaining = [c for c in out.columns if c not in ordered]
        return out[ordered + remaining]

    def _round1_df(df):
        out = df.copy()
        _int_cols = {"Consultation","Lab","Radiology","Procedure","Visits","Total_Visit","Total Visit"}
        _pct_cols = {"Lab_%","Lab %","Procedure_Per_Visit","Procedure %","Radiology_Per_Visit","Radiology %",
                     "Avg Service","Avg Insurance","Avg Svc","Avg Ins",
                     "Avg_Amount_Service","Avg_Amount_Insuance","Avg_Amount_Insurance"}
        for c in out.columns:
            if not pd.api.types.is_numeric_dtype(out[c]):
                continue
            if c in _int_cols:
                out[c] = pd.to_numeric(out[c], errors="coerce").round(0).fillna(0).astype(int)
            elif c in _pct_cols:
                out[c] = pd.to_numeric(out[c], errors="coerce").round(2)
            else:
                series = pd.to_numeric(out[c], errors="coerce").round(1)
                if series.dropna().apply(lambda x: x == int(x)).all():
                    out[c] = series.fillna(0).astype(int)
                else:
                    out[c] = series
        return out

    # ── exact Excel colors ────────────────────────────────────────────────────
    C_NAVY      = "#0D1B2A"
    C_NAVY_SECT = "#1E3A5F"
    C_ORANGE    = "#FF6600"
    C_ALT       = "#F0F4FF"
    C_WHITE     = "#FFFFFF"
    C_BORDER    = "#CCCCCC"
    C_DARK      = "#0f172a"
    C_MUTED     = "#64748b"

    def _td(val, align="right", extra=""):
        v = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else val
        return (f"<td style='padding:7px 10px;border:1px solid {C_BORDER};"
                f"text-align:{align};{extra}'>{v}</td>")

    def _th(label):
        return (f"<th style='padding:8px 10px;border:1px solid {C_BORDER};"
                f"text-align:center;color:{C_WHITE};background:{C_NAVY};font-size:12px;"
                f"font-weight:700;white-space:nowrap;'>{label}</th>")

    def _section_banner(text, n_cols=99):
        return (f"<tr><td colspan='{n_cols}' style='background:{C_NAVY_SECT};color:{C_WHITE};"
                f"font-size:14px;font-weight:900;padding:10px 14px;border:none;'>"
                f"{text}</td></tr>")

    def _render_plain(df):
        df = _round1_df(_safe_df(df))
        if df.empty:
            return ""
        cols = list(df.columns)
        h = (f"<table style='width:100%;border-collapse:collapse;font-size:12px;"
             f"margin-bottom:2px;'>")
        h += f"<tr>{''.join(_th(c) for c in cols)}</tr>"
        for ri, row in enumerate(df.itertuples(index=False, name=None)):
            first = str(row[0] or "").strip().upper()
            is_tot = first in ("GRAND TOTAL", "TOTAL")
            bg = C_ORANGE if is_tot else (C_ALT if ri % 2 == 0 else C_WHITE)
            fx = f"color:{C_WHITE};font-weight:900;" if is_tot else ""
            tds = [_td(v, align="left" if ci == 0 else "right",
                       extra=f"background:{bg};{fx}")
                   for ci, v in enumerate(row)]
            h += "<tr>" + "".join(tds) + "</tr>"
        return h + "</table>"

    def _render_dx(df):
        df = _round1_df(_safe_df(df))
        if df.empty:
            return ""
        cols = list(df.columns)
        doc_ci = cols.index("Doctor") if "Doctor" in cols else None
        ins_cols = [c for c in cols if c != "Doctor"]
        n_ins = len(ins_cols)

        h = (f"<table style='width:100%;border-collapse:collapse;font-size:12px;margin-bottom:2px;'>")
        # Header: Doctor col + insurance cols
        h += (f"<tr>{_th('Doctor')}"
              + "".join(_th(c) for c in ins_cols)
              + "</tr>")

        rows_list = list(df.itertuples(index=False, name=None))
        ri = 0
        alt = 0

        while ri < len(rows_list):
            row = rows_list[ri]
            first = str(row[0] or "").strip().upper()
            is_tot = first in ("GRAND TOTAL", "TOTAL")

            if is_tot or doc_ci is None:
                bg = C_ORANGE if is_tot else (C_ALT if alt % 2 == 0 else C_WHITE)
                fx = f"color:{C_WHITE};font-weight:900;" if is_tot else ""
                # Grand total: spans Doctor col + all ins_cols
                ins_vals = [row[cols.index(c)] for c in ins_cols]
                h += (f"<tr>"
                      f"<td style='padding:7px 10px;border:1px solid {C_BORDER};"
                      f"background:{bg};{fx};font-weight:900;text-align:left;'>GRAND TOTAL</td>"
                      + "".join(_td(v, "right", f"background:{bg};{fx}") for v in ins_vals)
                      + "</tr>")
                alt += 1; ri += 1; continue

            # Gather group
            cur = str(row[doc_ci] or "").strip().upper()
            grp = []
            j = ri
            while j < len(rows_list):
                rd = rows_list[j]
                fv = str(rd[0] or "").strip().upper()
                if fv in ("GRAND TOTAL", "TOTAL"): break
                if str(rd[doc_ci] or "").strip().upper() != cur: break
                grp.append(rd); j += 1

            doc_display = str(row[doc_ci] or "").strip()
            n_grp = len(grp)
            # +1 for the TOTAL row
            total_rows = n_grp + 1

            group_totals = {c: 0.0 for c in ins_cols if c != "Insurance"}

            for g_idx, g_row in enumerate(grp):
                bg = C_ALT if alt % 2 == 0 else C_WHITE
                ins_vals = [g_row[cols.index(c)] for c in ins_cols]

                # Top border style: thick navy separator for first row of each group (except first)
                top_border = (f"border-top:2px solid {C_NAVY_SECT};" if (g_idx == 0 and not (ri == 0 and alt == 0)) else "")

                if g_idx == 0:
                    # Doctor cell: rowspan covers data rows + TOTAL row
                    doc_cell = (
                        f"<td rowspan='{total_rows}' style='padding:8px 10px;"
                        f"border:1px solid {C_BORDER};border-top:2px solid {C_NAVY_SECT};"
                        f"text-align:center;font-weight:900;vertical-align:middle;"
                        f"background:#D6EAF8;white-space:nowrap;font-size:12px;'>"
                        f"{doc_display}</td>"
                    )
                    h += ("<tr>" + doc_cell
                          + "".join(_td(v, "left" if ci == 0 else "right",
                                        f"background:{bg};{top_border}")
                                    for ci, v in enumerate(ins_vals)) + "</tr>")
                else:
                    h += "<tr>" + "".join(_td(v, "left" if ci == 0 else "right", f"background:{bg};") for ci, v in enumerate(ins_vals)) + "</tr>"

                for c in ins_cols:
                    if c != "Insurance":
                        try:
                            group_totals[c] += float(pd.to_numeric(g_row[cols.index(c)], errors="coerce") or 0)
                        except Exception:
                            pass
                alt += 1

            # Doctor TOTAL row (medium blue) — doctor cell already covered by rowspan
            BLUE = "#2E6DA4"
            tot_tds = []
            for ci, c in enumerate(ins_cols):
                v = "TOTAL" if c == "Insurance" else round(group_totals.get(c, 0), 1)
                tot_tds.append(
                    f"<td style='padding:7px 10px;border:1px solid {C_BORDER};"
                    f"border-bottom:2px solid {C_NAVY_SECT};"
                    f"text-align:{'left' if ci==0 else 'right'};"
                    f"background:{BLUE};color:{C_WHITE};font-weight:700;'>{v}</td>"
                )
            h += "<tr>" + "".join(tot_tds) + "</tr>"
            ri = j

        return h + "</table>"

    # ── KPI values ────────────────────────────────────────────────────────────
    total_visits     = int(_num(_kpi_value("Total Visits", 0)) or 0)
    new_patients     = int(_num(_kpi_value("New Patients", 0)) or 0)
    established      = int(_num(_kpi_value("Established Patients", 0)) or 0)
    follow_up        = int(_num(_kpi_value("Follow Up", 0)) or 0)
    unclassified     = int(_num(_kpi_value("Unclassified Visits", 0)) or 0)
    pending_patients = int(_num(_kpi_value("Pending Patients", 0)) or 0)

    df_doc = _safe_df(dfs.get("Income | Doctor Wise Revenue"))
    df_ins = _safe_df(dfs.get("Income | Insurance Wise Revenue"))
    df_dx  = _safe_df(dfs.get("Income | Doctor x Insurance Revenue"))

    def _kpi_card(label, val):
        return (
            f"<td style='padding:6px;'>"
            f"<div style='background:linear-gradient(145deg,#ffffff 0%,#f8faff 100%);"
            f"border:1.5px solid #dde8f5;border-radius:16px;padding:16px 20px;"
            f"box-shadow:0 4px 16px rgba(10,38,71,0.08),inset 0 1px 0 rgba(255,255,255,0.95);'>"
            f"<div style='color:#8A9BB5;font-size:11px;font-weight:700;"
            f"text-transform:uppercase;letter-spacing:0.7px;margin-bottom:10px;'>{label}</div>"
            f"<div style='color:#0D1B2E;font-size:28px;font-weight:900;"
            f"letter-spacing:-0.5px;line-height:1.05;'>{val}</div>"
            f"</div></td>"
        )

    parts = [f"""<!DOCTYPE html><html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#EDF2FB;font-family:Segoe UI,Inter,Arial,sans-serif;">
<div style="max-width:960px;margin:20px auto;background:#EDF2FB;border-radius:18px;
     box-shadow:0 12px 40px rgba(10,38,71,0.13);overflow:hidden;">

  <!-- ── Header ── -->
  <div style="background:linear-gradient(135deg,#0D1B2A 0%,#1a3353 100%);padding:20px 26px;">
    <div style="color:#ffffff;font-size:19px;font-weight:900;letter-spacing:-0.03em;">
      📌 EMC Income Analysis Report
    </div>
    <div style="color:#94a3b8;font-size:12px;margin-top:5px;">
      {picked_label} &nbsp;·&nbsp; Generated: {pd.Timestamp.now().strftime('%d %b %Y %H:%M')}
    </div>
  </div>

  <div style="padding:20px 24px;">

    <!-- ── KPI Cards row 1 ── -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:6px;">
      <tr>
        {_kpi_card("Total Visits", total_visits)}
        {_kpi_card("New Patients", new_patients)}
        {_kpi_card("Established Patients", established)}
      </tr>
    </table>
    <!-- ── KPI Cards row 2 ── -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:22px;">
      <tr>
        {_kpi_card("Follow Up", follow_up)}
        {_kpi_card("Pending Patients", pending_patients)}
        <td style="padding:6px;"></td>
      </tr>
    </table>

    <!-- ── Income Analysis banner ── -->
    <div style="background:linear-gradient(135deg,#1E3A5F 0%,#2d5282 100%);
                border-radius:10px;padding:11px 16px;margin-bottom:16px;">
      <span style="color:#ffffff;font-size:14px;font-weight:900;">
        📊 Income Analysis (Doctor Revenue)
      </span>
    </div>
"""]

    if not df_doc.empty:
        parts.append(
            f"<table style='width:100%;border-collapse:collapse;margin-bottom:2px;'>"
            f"{_section_banner('Doctor Wise Revenue')}</table>"
            f"{_render_plain(df_doc)}"
            f"<div style='height:20px;'></div>"
        )

    if not df_ins.empty:
        parts.append(
            f"<table style='width:100%;border-collapse:collapse;margin-bottom:2px;'>"
            f"{_section_banner('Insurance Wise Revenue')}</table>"
            f"{_render_plain(df_ins)}"
            f"<div style='height:20px;'></div>"
        )

    if not df_dx.empty:
        parts.append(
            f"<table style='width:100%;border-collapse:collapse;margin-bottom:2px;'>"
            f"{_section_banner('Doctor x Insurance Revenue')}</table>"
            f"{_render_dx(df_dx)}"
            f"<div style='height:20px;'></div>"
        )

    parts.append(f"""
    <div style="color:{C_MUTED};font-size:11px;border-top:1px solid #e2e8f0;padding-top:10px;margin-top:8px;">
      This is an automated report generated by the EMC dashboard.
    </div>
  </div>
</div>
</body></html>""")

    return "".join(parts)

def _send_email_smtp(
    subject: str,
    html_body: str,
    attachment_bytes: bytes = None,
    attachment_filename: str = None,
) -> None:
    """Send an HTML email via SMTP. Optionally attach a file (e.g. Excel)."""
    host = st.secrets.get("SMTP_HOST", "")
    port = int(st.secrets.get("SMTP_PORT", 465))
    user = st.secrets.get("SMTP_USER", "")
    pwd  = st.secrets.get("SMTP_PASS", "")

    to_addr = st.secrets.get("EMAIL_TO", "")
    cc_addr = st.secrets.get("EMAIL_CC", "")

    if not (host and user and pwd and to_addr):
        raise ValueError("Missing SMTP secrets (SMTP_HOST/SMTP_PORT/SMTP_USER/SMTP_PASS/EMAIL_TO).")

    # Use 'mixed' when we have an attachment, 'alternative' otherwise
    msg = MIMEMultipart("mixed" if attachment_bytes else "alternative")
    msg["Subject"] = subject
    msg["From"] = user
    msg["To"] = to_addr
    if cc_addr:
        msg["Cc"] = cc_addr

    # Wrap HTML in an 'alternative' sub-part so email clients render it correctly
    alt_part = MIMEMultipart("alternative")
    alt_part.attach(MIMEText(html_body, "html"))
    msg.attach(alt_part)

    # Attach Excel file if provided
    if attachment_bytes and attachment_filename:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(attachment_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=attachment_filename)
        msg.attach(part)

    recipients = [x.strip() for x in (to_addr.split(",") + (cc_addr.split(",") if cc_addr else [])) if x.strip()]

    with smtplib.SMTP_SSL(host, port) as s:
        s.login(user, pwd)
        s.sendmail(user, recipients, msg.as_string())


def _build_income_excel(dfs: dict, period_label: str) -> bytes:
    """Build ONE combined Excel sheet: Income Analysis (Doctor Revenue).

    Layout (single sheet 'Income Analysis'):
        Section 1 — Doctor Wise Revenue
        [blank row]
        Section 2 — Insurance Wise Revenue
        [blank row]
        Section 3 — Doctor x Insurance Revenue  (doctor name shown ONCE, insurances listed under it)

    Formatting:
        - Numbers rounded to 1 decimal place
        - Dark navy headers, orange Grand Total rows, alternating light-blue rows
        - Doctor name merged/shown only once in Doctor x Insurance section
    """
    import io as _io
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="0D1B2A")   # dark navy
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TOTAL_FILL  = PatternFill("solid", fgColor="FF6600")   # orange
    TOTAL_FONT  = Font(color="FFFFFF", bold=True, size=11)
    ALT_FILL    = PatternFill("solid", fgColor="F0F4FF")   # light blue
    THIN        = Side(border_style="thin", color="CCCCCC")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    DOC_FONT    = Font(bold=True, size=11)                 # bold for doctor name in merged block

    # Columns that should be integers (no decimals)
    INT_COLS = {
        "Consultation", "Lab", "Radiology", "Procedure",
        "Visits", "Total_Visit", "Total Visit",
    }
    # Columns that need 2 decimal places
    PCT_COLS = {
        "Lab_%", "Lab %",
        "Procedure_Per_Visit", "Procedure %",
        "Radiology_Per_Visit", "Radiology %",
        "Avg Service", "Avg Insurance",
        "Avg Svc", "Avg Ins",
        "Avg.Amount", "Avg_Amount",
        "Avg_Amount_Service", "Avg_Amount_Insuance", "Avg_Amount_Insurance",
    }

    def _smart_round(df: pd.DataFrame) -> pd.DataFrame:
        """Integers for count/amount cols, 2dp for % and avg cols, 1dp for everything else."""
        out = df.copy()
        for c in out.columns:
            if not pd.api.types.is_numeric_dtype(out[c]):
                continue
            if c in INT_COLS:
                out[c] = pd.to_numeric(out[c], errors="coerce").round(0).astype("Int64")
            elif c in PCT_COLS:
                out[c] = pd.to_numeric(out[c], errors="coerce").round(2)
            else:
                # amounts (Total Svc, Total Ins etc) — 1dp but convert whole numbers to int
                series = pd.to_numeric(out[c], errors="coerce").round(1)
                # if all values are whole numbers, store as int
                if series.dropna().apply(lambda x: x == int(x) if pd.notna(x) else True).all():
                    out[c] = series.round(0).astype("Int64")
                else:
                    out[c] = series
        return out

    # Clean display names for Excel — full readable, consistent
    COL_RENAME = {
        "Total_Amount_Service":    "Total Service",
        "Total_Amount_Insuance":   "Total Insurance",
        "Total_Amount_Insurance":  "Total Insurance",
        "Avg_Amount_Service":      "Avg Service",
        "Avg_Amount_Insuance":     "Avg Insurance",
        "Avg_Amount_Insurance":    "Avg Insurance",
        "Avg.Amount":              "Avg Insurance",
        "Avg_Amount":              "Avg Insurance",
        "Lab_%":                   "Lab %",
        "Procedure_Per_Visit":     "Procedure %",
        "Radiology_Per_Visit":     "Radiology %",
        "Total_Visit":             "Visits",
        "Total_Amount":            "Total Insurance",
        "Department":              "Dept",
    }

    # Desired column order — Avg cols BEFORE % cols (matches Excel image)
    PREFERRED_ORDER = [
        "Dept", "Doctor", "Insurance",
        "Consultation", "Lab", "Radiology", "Procedure",
        "Visits",
        "Total Service", "Total Insurance",
        "Avg Service", "Avg Insurance",
        "Lab %", "Procedure %", "Radiology %",
    ]

    def _clean_df(df) -> pd.DataFrame:
        if not isinstance(df, pd.DataFrame) or df.empty:
            return pd.DataFrame()
        out = df.copy()
        # Drop unnamed/blank cols
        bad = [c for c in out.columns if str(c).strip() == "" or str(c).strip().lower().startswith("unnamed")]
        out = out.drop(columns=bad, errors="ignore")
        # Drop misleading % cols (Radiology_% and Procedure_% = % of service amount, not useful)
        for drop_c in ["Radiology_%", "Procedure_%"]:
            if drop_c in out.columns:
                out = out.drop(columns=[drop_c])
        # Compute per-visit ratios (2 decimal places)
        visit_col = next((c for c in out.columns if str(c).strip().lower() in
                          ["total_visit", "total visit", "visits", "visit"]), None)
        if visit_col is not None:
            denom = pd.to_numeric(out[visit_col], errors="coerce").replace(0, pd.NA)
            if "Procedure" in out.columns:
                out["Procedure_Per_Visit"] = (pd.to_numeric(out["Procedure"], errors="coerce") / denom).round(2).fillna(0)
            if "Radiology" in out.columns:
                out["Radiology_Per_Visit"] = (pd.to_numeric(out["Radiology"], errors="coerce") / denom).round(2).fillna(0)
        # Apply smart rounding (integers for counts, 2dp for %, 1dp for amounts)
        out = _smart_round(out)
        # Rename to clean display names
        out = out.rename(columns={k: v for k, v in COL_RENAME.items() if k in out.columns})
        # Reorder columns
        ordered = [c for c in PREFERRED_ORDER if c in out.columns]
        remaining = [c for c in out.columns if c not in ordered]
        out = out[ordered + remaining]
        return out

    df_doc = _clean_df(dfs.get("Income | Doctor Wise Revenue"))
    df_ins = _clean_df(dfs.get("Income | Insurance Wise Revenue"))
    df_dx  = _clean_df(dfs.get("Income | Doctor x Insurance Revenue"))

    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Income Analysis", index=False)
        ws = writer.sheets["Income Analysis"]

        current_row = 1  # 1-based

        # ── TITLE ROW ──────────────────────────────────────────────────────────
        title_text = f"EMC - INCOME ANALYSIS REPORT - {period_label.upper()}"
        title_cell = ws.cell(row=current_row, column=1, value=title_text)
        title_cell.font = Font(bold=True, size=14, color="0D1B2A")
        title_cell.fill = PatternFill("solid", fgColor="D6EAF8")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # We'll merge across all columns after we know max_col — store row for later
        title_row = current_row
        ws.row_dimensions[current_row].height = 28
        current_row += 1
        # Blank row after title
        ws.row_dimensions[current_row].height = 6
        current_row += 1

        def _write_section_header(ws, row, text, n_cols):
            """Write a section title spanning all columns."""
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(bold=True, size=13, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
            if n_cols > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
            ws.row_dimensions[row].height = 20
            return row + 1

        def _write_df_section(ws, start_row, df: pd.DataFrame, section_title: str) -> int:
            """Write a plain table (Doctor Wise / Insurance Wise) and return next free row."""
            if df.empty:
                return start_row

            cols = list(df.columns)
            n_cols = len(cols)

            start_row = _write_section_header(ws, start_row, section_title, n_cols)

            # Detect which columns are numeric (for right-alignment)
            num_flags = [pd.api.types.is_numeric_dtype(df[c]) for c in cols]

            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=start_row, column=ci, value=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDER
            ws.row_dimensions[start_row].height = 30
            start_row += 1

            for ri, row_data in enumerate(df.itertuples(index=False, name=None)):
                first_val = str(row_data[0] or "").strip().upper()
                is_total = first_val in ("GRAND TOTAL", "TOTAL")
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=start_row, column=ci, value=val if val is not None else "")
                    cell.border = BORDER
                    # Left-align text columns, right-align numeric columns
                    is_num = num_flags[ci - 1]
                    cell.alignment = Alignment(
                        horizontal="right" if is_num else "left",
                        vertical="center"
                    )
                    if is_total:
                        cell.fill = TOTAL_FILL
                        cell.font = TOTAL_FONT
                    elif ri % 2 == 0:
                        cell.fill = ALT_FILL
                start_row += 1

            start_row += 1
            return start_row

        def _write_dx_section(ws, start_row, df: pd.DataFrame) -> int:
            """Doctor x Insurance: doctor name as side-merged cell (light blue),
            insurance rows beside it, per-doctor TOTAL row, thick separator between groups,
            Grand Total at end. All columns sized to fit screen without scrolling.
            """
            if df.empty:
                return start_row

            cols = list(df.columns)
            n_cols = len(cols)

            start_row = _write_section_header(ws, start_row, "Doctor x Insurance Revenue", n_cols)

            doc_col_idx = None
            if "Doctor" in cols:
                doc_col_idx = cols.index("Doctor") + 1

            ins_cols = [c for c in cols if c != "Doctor"]
            n_ins = len(ins_cols)

            # Header row
            header_cols = ["Doctor"] + ins_cols
            for ci, col in enumerate(header_cols, 1):
                cell = ws.cell(row=start_row, column=ci, value=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDER
            ws.row_dimensions[start_row].height = 30
            start_row += 1

            rows_list = list(df.itertuples(index=False, name=None))
            alt = 0
            i = 0
            is_first_group = True

            DOC_FILL  = PatternFill("solid", fgColor="D6EAF8")
            DTOT_FILL = PatternFill("solid", fgColor="2E6DA4")
            DTOT_FONT = Font(color="FFFFFF", bold=True, size=10)

            # Thick separator border (top of doctor group)
            SEP_SIDE  = Side(border_style="medium", color="1E3A5F")
            THIN      = Side(border_style="thin", color="CCCCCC")

            while i < len(rows_list):
                row_data = rows_list[i]
                first_val = str(row_data[0] or "").strip().upper()
                is_total = first_val in ("GRAND TOTAL", "TOTAL")

                if is_total:
                    all_vals = ["GRAND TOTAL"] + [row_data[cols.index(c)] for c in ins_cols]
                    for ci, val in enumerate(all_vals, 1):
                        cell = ws.cell(row=start_row, column=ci, value=val if val is not None else "")
                        cell.fill = TOTAL_FILL
                        cell.font = TOTAL_FONT
                        cell.border = BORDER
                        cell.alignment = Alignment(horizontal="right" if ci > 1 else "left", vertical="center")
                    start_row += 1; i += 1; continue

                if doc_col_idx is None:
                    ins_vals = list(row_data)
                    for ci, val in enumerate(ins_vals, 1):
                        cell = ws.cell(row=start_row, column=ci, value=val if val is not None else "")
                        cell.border = BORDER
                        cell.alignment = Alignment(horizontal="right" if ci > 1 else "left", vertical="center")
                        if alt % 2 == 0: cell.fill = ALT_FILL
                    alt += 1; start_row += 1; i += 1; continue

                # Gather group
                cur = str(row_data[doc_col_idx - 1] or "").strip().upper()
                group = []
                j = i
                while j < len(rows_list):
                    rd = rows_list[j]
                    fv = str(rd[0] or "").strip().upper()
                    if fv in ("GRAND TOTAL", "TOTAL"): break
                    if str(rd[doc_col_idx - 1] or "").strip().upper() != cur: break
                    group.append(rd); j += 1

                doc_display = str(row_data[doc_col_idx - 1] or "").strip()
                n_grp = len(group)
                total_rows = n_grp + 1
                group_start = start_row
                group_totals = {c: 0.0 for c in ins_cols}

                # Write insurance data rows
                for g_idx, g_row in enumerate(group):
                    # Top border: thick for first row of each group (except very first)
                    top_side = SEP_SIDE if (g_idx == 0 and not is_first_group) else THIN

                    if g_idx == 0:
                        cell = ws.cell(row=start_row, column=1, value=doc_display)
                        cell.fill = DOC_FILL
                        cell.font = Font(bold=True, size=11)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = Border(left=SEP_SIDE if not is_first_group else THIN,
                                             right=THIN, top=top_side, bottom=THIN)
                    else:
                        cell = ws.cell(row=start_row, column=1, value="")
                        cell.fill = DOC_FILL
                        cell.border = Border(left=SEP_SIDE if not is_first_group else THIN,
                                             right=THIN, top=THIN, bottom=THIN)

                    for ci, c in enumerate(ins_cols, 2):
                        val = g_row[cols.index(c)]
                        cell = ws.cell(row=start_row, column=ci, value=val if val is not None else "")
                        # Left-align text columns, right-align numeric
                        col_is_num = pd.api.types.is_numeric_dtype(df.dtypes.get(c, object))
                        cell.alignment = Alignment(horizontal="right" if col_is_num else "left", vertical="center")
                        cell.border = Border(left=THIN, right=THIN, top=top_side if g_idx == 0 else THIN, bottom=THIN)
                        if alt % 2 == 0: cell.fill = ALT_FILL
                        try:
                            group_totals[c] += float(pd.to_numeric(val, errors="coerce") or 0)
                        except Exception:
                            pass

                    alt += 1
                    start_row += 1

                # Doctor TOTAL row
                cell = ws.cell(row=start_row, column=1, value="")
                cell.fill = DOC_FILL
                cell.border = BORDER
                for ci, c in enumerate(ins_cols, 2):
                    v = "TOTAL" if c == "Insurance" else round(group_totals.get(c, 0), 1)
                    cell = ws.cell(row=start_row, column=ci, value=v)
                    cell.fill = DTOT_FILL
                    cell.font = DTOT_FONT
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="right" if ci > 2 else "left", vertical="center")
                start_row += 1

                # Merge doctor cell vertically across all rows in group
                if total_rows > 1:
                    ws.merge_cells(start_row=group_start, start_column=1,
                                   end_row=group_start + total_rows - 1, end_column=1)
                    merged = ws.cell(row=group_start, column=1)
                    merged.fill = DOC_FILL
                    merged.font = Font(bold=True, size=11)
                    merged.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    merged.border = Border(
                        left=SEP_SIDE if not is_first_group else THIN,
                        right=THIN,
                        top=SEP_SIDE if not is_first_group else THIN,
                        bottom=THIN
                    )

                is_first_group = False
                i = j

            start_row += 1
            return start_row

        # --- Write 3 sections ---
        current_row = _write_df_section(ws, current_row, df_doc, "Doctor Wise Revenue")
        current_row = _write_df_section(ws, current_row, df_ins, "Insurance Wise Revenue")
        current_row = _write_dx_section(ws, current_row, df_dx)

        # --- Merge title row across all columns now that we know max_col ---
        max_col = ws.max_column
        if max_col > 1:
            ws.merge_cells(start_row=title_row, start_column=1,
                           end_row=title_row, end_column=max_col)
            # Re-apply style on merged cell
            tc = ws.cell(row=title_row, column=1)
            tc.font = Font(bold=True, size=14, color="0D1B2A")
            tc.fill = PatternFill("solid", fgColor="D6EAF8")
            tc.alignment = Alignment(horizontal="center", vertical="center")

        # --- Smart column widths based on header name ---
        max_row_used = ws.max_row

        # Minimum widths by header name (ensures text fits without wrap)
        MIN_WIDTHS = {
            "Total Service":   16,
            "Total Insurance": 16,
            "Avg Service":     14,
            "Avg Insurance":   14,
            "Lab %":           10,
            "Procedure %":     13,
            "Radiology %":     13,
            "Consultation":    14,
            "Visits":          9,
            "Doctor":          22,
            "Insurance":       28,
            "Dept":            18,
        }

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            header_val = str(ws.cell(row=3, column=col_idx).value or "")  # row 3 = first section header row
            # Also check row 1+2 area for any header
            for r in range(1, min(max_row_used + 1, 8)):
                v = ws.cell(row=r, column=col_idx).value
                if v and str(v).strip() and str(v).strip() != header_val:
                    if str(v).strip() in MIN_WIDTHS:
                        header_val = str(v).strip()
                        break

            # Measure max content width in this column
            max_content = max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(1, max_row_used + 1)),
                default=6,
            )
            # Base width from content
            base = max_content + 2
            # Apply minimum from header name map
            min_w = MIN_WIDTHS.get(header_val, 0)
            final_w = max(base, min_w, 9)
            # Cap very wide text cols
            if final_w > 30:
                final_w = 30
            ws.column_dimensions[col_letter].width = round(final_w, 1)

        ws.freeze_panes = "A4"  # freeze below title + blank row

    return buf.getvalue()


# Employer normalization map (from Employer names.csv 'check' column)
EMPLOYER_CANON_MAP = {
    "A D C CONTRACTING": "A.D.C Energy And Contracting",
    "A D C ENERGY SYSTEMS LLC": "A.D.C Energy And Contracting",
    "A G FACILITIES SOLUTIONS FOR BUILDING": "AG FACILITIES SOLUTIONS",
    "A.D.C ENERGY SYSTEMS (L.L.C)": "A.D.C Energy And Contracting",
    "ABDULLA MOHSEN HADI HUSAIN AL HAMED": "ABDULLA MOHSEN HADI HUSAIN AL HAMED",
    "ABUDHABI BERKELEY SERVICES LLC": "ABUDHABI BERKELEY",
    "ADC CONTARCTING": "A.D.C Energy And Contracting",
    "ADEEB ELECTRICAL AND ELECTRONICS SERVICES": "ADEEB ELECTRICAL AND ELECTRONICS SERVICES",
    "AG FACILITIES SOLUTIONS BUILDINGS MAINTENANCE": "AG FACILITIES SOLUTIONS",
    "AG FACILITIES SOLUTIONS L.L.C": "AG FACILITIES SOLUTIONS",
    "AG FACILITIES SOLUTIONS LLC": "AG FACILITIES SOLUTIONS",
    "AGILITY ENGINEERING AND CONTRACTING COMPANY": "AGILITY",
    "AGILITY ENGINEERING AND CONTRACTING COMPANY LLC": "AGILITY",
    "AL BAYADER IRRIGATION AND CONTRACTING LLC": "AL BAYADER IRRIGATION AND CONTRACTING LLC",
    "AL GEEMI & PARTNERS CONTRACTING COMPANY LLC.": "AL GEEMI",
    "AL GEEMI CONTRACTING COMPANY LLC": "AL GEEMI",
    "AL NASIYA": "AL NASIYA",
    "AL RAWAI CONTRACTING GENERAL MAINATENANCE": "AL RAWAI CONTRACTING",
    "AL RAWAI CONTRACTING GENERAL MAINTENANCE ESTABLISHMENT": "AL RAWAI CONTRACTING",
    "AL RAYUM CONT GEN TRASPORT EST": "ALRYUM CONTRACTING",
    "AL RYUM CONTRACTING GENERAL TRANSPORT L L C": "ALRYUM CONTRACTING",
    "AL SAGR NATIONAL INSURANCE CO. (PSC)-3": "ALRYUM CONTRACTING",
    "AL SAIF GRAPHICS L.L.C": "AL SAIF GRAPHICS L.L.C",
    "AL SHOUMOKH MANPOWER": "AL SHOUMOKH",
    "AL SHOUMOKH MANPOWER RECRUITMENT": "AL SHOUMOKH",
    "ALBA TEC": "ALBA TEC",
    "ALBA TEC EMPLOYMENT - SOLE PROPRIETORSHIP": "ALBA TEC EMPLOYMENT",
    "ALBA TEC EMPLOYMENT SERVICES": "ALBA TEC EMPLOYMENT",
    "ALBA TEC MODERN CONSTURCTION AND DEVELOPMENT": "ALBA TEC",
    "ALBARRAK ELECTRICAL CONTRACTING COMPANY": "ALBA TEC EMPLOYMENT",
    "ALBATEC CONST AND DEVELOPMENT": "ALBA TEC",
    "ALGEEMI PARTNERS CONT CO LLC": "AL GEEMI",
    "ALKALINE ELECTROMECHANICAL LLC.": "ALKALINE ELECTROMECHANICAL LLC.",
    "ALRYUM CONT GEN TRANSPORT": "ALRYUM CONTRACTING",
    "ALRYUM CONTRACTING GENERAL": "ALRYUM CONTRACTING",
    "ALRYUM CONTRACTING GENERAL TRANSPORT": "ALRYUM CONTRACTING",
    "ALRYUM CONTRACTING GENERAL TRANSPORT LLC": "ALRYUM CONTRACTING",
    "ALRYUM CONTRRACTING": "ALRYUM CONTRACTING",
    "ALWATHBA CEMENT INDUSTRIES -SOLE PROPRIETORSHIP": "ALWATHBA CEMENT INDUSTRIES -SOLE PROPRIETORSHIP",
    "ARABIAN COMPANY LLC": "GULF",
    "ARABIAN GULF STEEL INDUSTRIES L L C": "GULF",
    "ARABIAN GULF STEEL INDUSTRIES LLC": "GULF",
    "ARCO ELECTRO MECHANICAL L.L.C": "ARCO",
    "ARCO ELECTRO MECHANICAL LLC": "ARCO",
    "ARCO ELECTROMECHANICAL L.L.C": "ARCO",
    "ARCO ELECTROMECHANICAL L.L.C-DUBAI BRANCH": "ARCO",
    "ARCO ELECTROMECHANICAL LLC": "ARCO",
    "ARCO GENERAL CONTRACTING": "ARCO GENERAL",
    "ARCO GENERAL CONTRACTING LLC": "ARCO GENERAL",
    "ARCO INTL CONTRACTING COMPANY": "ARCO INTL",
    "BGC INTERNATIONAL GENERAL CONTRACTING- LLC": "BGC INTERNATIONAL GENERAL CONTRACTING- LLC",
    "BHATTI GENTS SALOON": "BHATTI GENTS SALOON",
    "CLEANPRO FACILITIES MANAGEMENT SERVICES": "CLEANPRO FACILITIES MANAGEMENT SERVICES",
    "CLIFTON GENERAL CONTRACTION L.L.C": "DOLPHIN OILFIELD EQUIPMENT",
    "CYLINGAS COMPANY LLC": "CYLINGAS COMPANY LLC",
    "DELMON AUTOCLAVED AERATED CONCRETE": "DELMON AUTOCLAVED AERATED CONCRETE",
    "DOLPHIN OILFIELD EQUIPMENT SERVICES CO- LLC": "DOLPHIN OILFIELD EQUIPMENT",
    "DOLPHIN OILFIELD EQUIPMENT SERVICES COMPANY LLC": "DOLPHIN OILFIELD EQUIPMENT",
    "E F S FACILITEIS MANAGEMENT SERVICES L L C": "E F S FACILITEIS",
    "E F S INVESTMENT L L C": "E F S FACILITEIS",
    "EDAN GARDREN": "EDAN GARDREN",
    "EFS FACILITIES MANAGEMENT SERVICES LLC": "E F S FACILITEIS",
    "EFS FACILITIES MANGEMNT SERVICES LLC": "E F S FACILITEIS",
    "EMARAT ALOULA INDUSTRIES SOLE PROPRITORSHIP": "EMARAT ALOULA INDUSTRIES SOLE PROPRITORSHIP",
    "EMIRATES ELECTRICAL AND INSTRUMENTATION": "EMIRATES ELECTRICAL",
    "EMIRATES GATEWAY SECURITY SERVICES LLC": "EMIRATES GATEWAY",
    "EMIRATES LINK CONTRACTING LLC": "EMIRATES ELECTRICAL",
    "EXCEED PRECAST": "EXEED INDUSTRIES",
    "EXCEED PRECAST OWNED BY EXCEED INDUSTRIES LLC": "EXCEED PRECAST",
    "EXCEED PRECAST-SOLE PROPRIETORSHIP LLC": "EXCEED PRECAST",
    "EXCELLENT MEDICAL CENTER": "EXCELLENT MEDICAL CENTER",
    "EXCELLENT MEDICAL CENTER LLC": "EXCELLENT MEDICAL CENTER",
    "EXCELLENT MEDICAL CENTERQ": "EXCELLENT MEDICAL CENTER",
    "EXEED PRECAST": "EXCEED PRECAST",
    "EXEED PRECAST - SOLE PROPRIETORSHIP L L C": "EXCEED PRECAST",
    "EXEED PRECAST - SOLE PROPRIETORSHIP LLC": "EXCEED PRECAST",
    "EXEED PRECAST LLC": "EXCEED PRECAST",
    "EXEED PRECAST OWNED BY EXEED INDUSTIRES LLC": "EXCEED PRECAST",
    "EXEED PRECAST OWNED BY EXEED INDUSTRIES LLC": "EXCEED PRECAST",
    "EXEED PRECAST OWNED BY EXZEED INDUSTRIES LLC": "EXEED INDUSTRIES",
    "EXEED PRECAST-SOLE PROPRIETORSHIP L L C": "EXCEED PRECAST",
    "FALCON ZINC METAL INDUSTRIES LLC": "FALCON ZINC METAL INDUSTRIES LLC",
    "FALCON ZINC STEEL WORKS L.L.C": "AG FACILITIES SOLUTIONS",
    "FATEMA ALI WIDOW MOHAMED K AL MANSOORI": "FATEMA ALI WIDOW MOHAMED K AL MANSOORI",
    "FIBREX L L C BR 1": "FIBREX L L C",
    "FIBREX L L C BRANCH 1": "FIBREX L L C",
    "FIBREX LLC": "FIBREX L L C",
    "FOCUS SECURITY SERVICES": "FOCUS SECURITY SERVICES",
    "FURSAN SECURITY SERVICES L.L.C": "FURSAN SECURITY SERVICES L.L.C",
    "G4S SECURE SOLUTIONS L.L.C.": "G4S SECURE SOLUTIONS L.L.C.",
    "GIFT ACTION TRADING": "GIFT ACTION TRADING",
    "GISCO": "GISCO",
    "GREAT MART GENERAL TRADING LLC.": "GREAT MART GENERAL TRADING LLC.",
    "GULF  CONTRACTING AND LANDSCAPING LLC": "GULF LANDSCAPING",
    "GULF CONTRACTORS CO LLC": "GULF",
    "GULF CONTRACTORS COLLC": "GULF",
    "GULF CONTRACTORS COMPANY -LLC": "GULF",
    "GULF INDUSTRIAL SERVICES CO GISCO - L.L.C - S.P.C": "GULF",
    "GULF INDUSTRIAL SERVICES COMPANY - GISCO - LLC": "GULF INDUSTRIAL",
    "GULF INDUSTRIAL SERVICES COMPANY GISCO LLC": "GULF",
    "GULF PREACAST CONCRATE": "GULF",
    "GULF SNIPE ENGINEERING CONSTRUCTIONS LLC": "GULF SNIPE",
    "GULF TUNNELING COMPANY L.L.C.": "GULF",
    "GULF TUNNELING COMPANY LLC": "GULF",
    "HAFILAT GENERAL TRANSPORT - SOLE PROP. LLC": "HAFILAT GENERAL TRANSPORT - SOLE PROP. LLC",
    "HAFILAT GENERAL TRANSPORT-SOLE": "HAFILAT GENERAL TRANSPORT-SOLE",
    "HASSAN ALLAM CONSTRUCTION LLC": "HASSAN ALLAM CONTRUCTION",
    "HASSAN ALLAM CONTRUCTION SAE": "HASSAN ALLAM CONTRUCTION",
    "HEALTH FOR ALL MEDICAL CENTER": "HEALTH FOR ALL MEDICAL CENTER",
    "HFZA ARABIAN GULF STEEL INDUSTRIES": "GULF STEEL",
    "HILAL BIL BADI & PARTNERS CONTRACTING COMPANY - W L L - HILALCO": "HILAL BIL BADI",
    "HILAL BIL BADI AND PARTNERS CONT CO WLL HILALCO": "HILAL BIL BADI",
    "HYSSNA INTERNATIONAL L.L.C": "HYSSNA INTERNATIONAL L.L.C",
    "I G G FOR MILITARY AND FORMAL GERMENTS": "I G G FOR MILITARY AND FORMAL GERMENTS",
    "INNOVO BUILD L.L.C": "INNOVO",
    "INNOVO MEP ELECTROMECHANICAL WORKS LLC": "INNOVO",
    "INTERNATIONAL DECOR CO L.L.C - DUBAI BRANCH": "INTERNATIONAL DECOR CO L.L.C - DUBAI BRANCH",
    "INTERNATIONAL DEVELOPMENT COMPANY L.L.C": "INTERNATIONAL DEVELOPMENT COMPANY L.L.C",
    "J P M ASSOCIATES TECHNICAL SERVICES L.L.C": "J P M ASSOCIATES TECHNICAL SERVICES L.L.C",
    "JAZAL ENGINEERING & CONTRACTING (L.L.C)": "JAZAL ENGINEERING",
    "JEET CONSTRUCTION LLC": "JEET CONSTRUCTION LLC",
    "KAYAN ALMOSTQBAL CONTRACTING & GEN. MAINTENANCE": "KAYAN ALMOSTQBAL CONTRACTING & GEN. MAINTENANCE",
    "LAITH ELECTRO MECHANICAL LLC": "LAITH ELECTRO",
    "LAITH ELECTRO-MECHANICAL - L L C - DUBAI BRANCH": "LAITH ELECTRO",
    "LAITH ELECTRO-MECHANICAL SOLE": "LAITH ELECTRO",
    "LASSI TOP CAFE": "LASSI TOP CAFE",
    "M 4 CONTRACTING": "M4 CONTRACTING",
    "M 4 CONTRCTING ABU DHABI": "M4 CONTRACTING",
    "M 4 CONTRCTING S A R  I ABU DHABI": "M4 CONTRACTING",
    "M FOUR CONTRACTING A RI ABU DHABI": "M4 CONTRACTING",
    "M4 CONTRACTING S A R I ABU DHABI": "M4 CONTRACTING",
    "M4CONTRACTING S A R IABU DHABI": "M4 CONTRACTING",
    "MALABAR DAWATH REASTURANT AND GRILL": "MALABAR DAWATH REASTURANT AND GRILL",
    "MAZAYA ALMUTAHIDA CONSTRUCTIONS": "MAZAYA ALMUTAHIDA CONSTRUCTIONS",
    "MECHANICAL & CIVIL ENG.CONTRACTORS CO(MACE)LLC": "MECHANICAL & CIVIL ENG.CONTRACTORS CO(MACE)LLC",
    "MES SECURITY SERVICES": "MES SECURITY",
    "MFOUR BUILDING CONTRACTING - BRANCH OF ABU DHABI": "M4 CONTRACTING",
    "MFOUR BUILDING CONTRACTING LLC": "M4 CONTRACTING",
    "ML MAN POWER LLC": "ML MANPOWER",
    "ML MANPOWER": "ML MANPOWER",
    "MOBILE SOLUTIONS LLC": "MOBILE SOLUTIONS LLC",
    "MODERN BUILDING GENERAL CONTRACTING L L C": "MODERN BUILDING GENERAL CONTRACTING L L C",
    "MOSAYED BIN HAFEEZ CONT GEN TRANSPORT": "MOSAYED BIN HAFEEZ CONT GEN TRANSPORT",
    "MOUNTAIN GATE PROPRTY INVESTEMENT LLC": "MOUNTAIN GATE PROPRTY INVESTEMENT LLC",
    "MQ PEAEL ENGINEERING-L.L.C": "MQ PEAEL ENGINEERING-L.L.C",
    "NATIONAL CATERING COMPANY": "NATIONAL CATERING COMPANY",
    "NATIONAL CATERING COMPANY LIMITED-SOLE": "NATIONAL CATERING COMPANY",
    "NATIONAL INNOVATIVE GENERAL MAINTAINCE - SOLE": "NATIONAL INNOVATIVE",
    "NATIONAL INNOVATIVE GENERAL MAINTNANCE": "NATIONAL INNOVATIVE",
    "NAZIM MAINTENANC GENERAL CONTRACTING COMPANY": "NAZIM MAINTENANC GENERAL CONTRACTING COMPANY",
    "NMDC ENERGY P.J.S.C": "NMDC ENERGY P.J.S.C",
    "NOOR AL SAHARA GEN CONTRACTING-SOLE": "NOOR AL SAHARA",
    "NOOR AL SAHARA GEN. CONTRACTING LLC": "NOOR AL SAHARA",
    "NOOR AL SAHARA GENERAL TRANSPORTATION": "NOOR AL SAHARA",
    "NOOR AL SAHRA": "NOOR AL SAHARA",
    "NOOR AL SAHRAA INTERNATIONAL GENERAL": "NOOR AL SAHARA",
    "NOOR ALSAHRAA INTERNATIONAL GENERAL": "NOOR AL SAHARA",
    "NUROL LLC.": "NUROL LLC.",
    "OPTIMUM ENGINEERING  S A L ABU DHABI": "OPTIMUM ENGINEERING",
    "OPTIMUM ENGINEERING SAL ABU DHABI": "OPTIMUM ENGINEERING",
    "PERFECT STEP GENERAL CONTRACTING & MAINTENANCE": "PERFECT STEP GENERAL CONTRACTING & MAINTENANCE",
    "PIONEER PRECAST CONCRETE LLC": "PIONEER PRECAST CONCRETE LLC",
    "PIONEER PRECAST CONCRETE LLCC": "PIONEER PRECAST CONCRETE LLCC",
    "POLENSKY & ZOELLNER COMPANY ABU DHABI W L L": "POLENSKY & ZOELLNER COMPANY ABU DHABI W L L",
    "PRINCE INTERNATIONALGENERAL TRANSPORT - L.L.C": "PRINCE INTERNATIONALGENERAL TRANSPORT - L.L.C",
    "PROFILE RECRUITMENT": "PROFILE RECRUITMENT",
    "PROFILE RECRUITMENT-SOLE PROPRIETORSHIP": "PROFILE RECRUITMENT",
    "QAMARA ALUMINIUM WORK EST": "QAMRA",
    "QAMARA CARPENTRY": "QAMRA",
    "QAMARA ELECTROMECHANICAL CONTRACTING EST": "QAMRA",
    "QAMARA ELECTROMECHANICLA CONTRACTIN EST": "QAMRA",
    "QAMRA TRANSPORT AND GEN CONT EST": "QAMRA",
    "QAMRA TRANSPORT AND GEN. CONT. EST.": "QAMRA",
    "QMRA TRANSPORT AND GEN CONT EST": "QAMRA",
    "QUICK SERVICES GENERAL TRANSPORT": "QUICK SERVICES GENERAL TRANSPORT",
    "QUMRA FOR DECORATION AND INTERIOR DESIGN.": "QAMRA",
    "QUMRA FOR DECORATION AND INTERIOR DESIGN....": "QAMRA",
    "QUMRA TRANSPORT & GENERAL CONTRACTING EST.": "QAMRA",
    "QUMRA TRANSPORT AND GENERAL CONTRACTING - LLC-SPC": "QAMRA",
    "QURMA TRANSPORT AND GENERAL": "QAMRA",
    "SBK HOLDING (L.L.C)": "SBK HOLDING (L.L.C)",
    "SECURIGUARD MIDDLE EAST LLC": "SECURIGUARD MIDDLE EAST LLC",
    "SIBCA ELECTRONIC EQUIPMENT COMPANY LIMITED - SOLE PROPRIETORSHIP L.L.C": "SIBCA ELECTRONIC",
    "SILVER SCREEN GENERAL CONTRACTING L.L.C": "SILVER SCREEN GENERAL CONTRACTING L.L.C",
    "SIX SIGMA MIDDLE EAST CONSTRUCTIONS LLC": "SIX SIGMA",
    "SNIPE OIL AND GAS EQUIPMENT": "SNIPE OIL AND GAS EQUIPMENT",
    "STAR SECURITY": "STAR SERVICES",
    "STAR SERVICES  LLC": "STAR SERVICES",
    "STAR SERVICES LLC DUBAI BRANCH": "STAR SERVICES",
    "SWITCHGEAR ELECTRO MECHANICAL LLC": "SWITCHGEAR ELECTRO MECHANICAL LLC",
    "TANZIFCO EMIRATES": "TANZIFO",
    "TANZIFO EMIRATES LLC": "TANZIFO",
    "TECTON ENGINEEING & CONSTRUCTION": "TECTON ENGINEEING & CONSTRUCTION",
    "TOOLS MAN GENERAL MAINTENANCE": "TOOLS MAN GENERAL MAINTENANCE",
    "UNITED MAZAYA BLACKSMITH& REINFORNCE CARPENTRY": "UNITED MAZAYA",
    "UNITED MAZAYA GENERAL MAINTENANCE": "UNITED MAZAYA",
    "VOLTAS LIMITED": "VOLTAS LIMITED",
    "VOLTAS LIMITED ABU DHABI": "VOLTAS LIMITED",
    "WADE ADAMS CONTRACTING LLC": "WADE ADAMS CONTRACTING LLC",
    "ZUBLIN CONSTRUCTION LLC": "ZUBLIN CONSTRUCTION LLC"
}

# ---------------- Employer normalization helpers (mapping + cleaning) ----------------
def _clean_employer_key(x: str) -> str:
    s = str(x or '').strip().upper()
    s = re.sub(r'\s+', ' ', s)            # collapse spaces
    s = s.replace('.', '').replace(',', '')
    s = s.replace(' L L C', ' LLC')        # normalize spaced LLC
    s = s.replace('LLCC', 'LLC')           # common typo
    return s

def _norm_emp(x: str) -> str:
    """Canonical key for grouping employer names."""
    k = _clean_employer_key(x)
    canon = EMPLOYER_CANON_MAP.get(k, mapped:=None)
    if canon is None:
        canon = EMPLOYER_CANON_MAP.get(k, k)
    return _clean_employer_key(canon)

def _display_emp_from_norm(norm_key: str) -> str:
    """Display value for employer (prefer Check/canonical if any)."""
    # Try direct map lookup first
    v = EMPLOYER_CANON_MAP.get(norm_key, None)
    return str(v).strip() if v else str(norm_key).strip()


# Canon display mapping
EMPLOYER_DISPLAY_MAP = {
    "A.D.C ENERGY AND CONTRACTING": "A.D.C Energy And Contracting",
    "ABDULLA MOHSEN HADI HUSAIN AL HAMED": "ABDULLA MOHSEN HADI HUSAIN AL HAMED",
    "ABUDHABI BERKELEY": "ABUDHABI BERKELEY",
    "ADEEB ELECTRICAL AND ELECTRONICS SERVICES": "ADEEB ELECTRICAL AND ELECTRONICS SERVICES",
    "AG FACILITIES SOLUTIONS": "AG FACILITIES SOLUTIONS",
    "AGILITY": "AGILITY",
    "AL BAYADER IRRIGATION AND CONTRACTING LLC": "AL BAYADER IRRIGATION AND CONTRACTING LLC",
    "AL GEEMI": "AL GEEMI",
    "AL NASIYA": "AL NASIYA",
    "AL RAWAI CONTRACTING": "AL RAWAI CONTRACTING",
    "AL SAIF GRAPHICS L.L.C": "AL SAIF GRAPHICS L.L.C",
    "AL SHOUMOKH": "AL SHOUMOKH",
    "ALBA TEC": "ALBA TEC",
    "ALBA TEC EMPLOYMENT": "ALBA TEC EMPLOYMENT",
    "ALKALINE ELECTROMECHANICAL LLC.": "ALKALINE ELECTROMECHANICAL LLC.",
    "ALRYUM CONTRACTING": "ALRYUM CONTRACTING",
    "ALWATHBA CEMENT INDUSTRIES -SOLE PROPRIETORSHIP": "ALWATHBA CEMENT INDUSTRIES -SOLE PROPRIETORSHIP",
    "ARCO": "ARCO",
    "ARCO GENERAL": "ARCO GENERAL",
    "ARCO INTL": "ARCO INTL",
    "BGC INTERNATIONAL GENERAL CONTRACTING- LLC": "BGC INTERNATIONAL GENERAL CONTRACTING- LLC",
    "BHATTI GENTS SALOON": "BHATTI GENTS SALOON",
    "CLEANPRO FACILITIES MANAGEMENT SERVICES": "CLEANPRO FACILITIES MANAGEMENT SERVICES",
    "CYLINGAS COMPANY LLC": "CYLINGAS COMPANY LLC",
    "DELMON AUTOCLAVED AERATED CONCRETE": "DELMON AUTOCLAVED AERATED CONCRETE",
    "DOLPHIN OILFIELD EQUIPMENT": "DOLPHIN OILFIELD EQUIPMENT",
    "E F S FACILITEIS": "E F S FACILITEIS",
    "EDAN GARDREN": "EDAN GARDREN",
    "EMARAT ALOULA INDUSTRIES SOLE PROPRITORSHIP": "EMARAT ALOULA INDUSTRIES SOLE PROPRITORSHIP",
    "EMIRATES ELECTRICAL": "EMIRATES ELECTRICAL",
    "EMIRATES GATEWAY": "EMIRATES GATEWAY",
    "EXCEED PRECAST": "EXCEED PRECAST",
    "EXCELLENT MEDICAL CENTER": "EXCELLENT MEDICAL CENTER",
    "EXEED INDUSTRIES": "EXEED INDUSTRIES",
    "FALCON ZINC METAL INDUSTRIES LLC": "FALCON ZINC METAL INDUSTRIES LLC",
    "FATEMA ALI WIDOW MOHAMED K AL MANSOORI": "FATEMA ALI WIDOW MOHAMED K AL MANSOORI",
    "FIBREX L L C": "FIBREX L L C",
    "FOCUS SECURITY SERVICES": "FOCUS SECURITY SERVICES",
    "FURSAN SECURITY SERVICES L.L.C": "FURSAN SECURITY SERVICES L.L.C",
    "G4S SECURE SOLUTIONS L.L.C.": "G4S SECURE SOLUTIONS L.L.C.",
    "GIFT ACTION TRADING": "GIFT ACTION TRADING",
    "GISCO": "GISCO",
    "GREAT MART GENERAL TRADING LLC.": "GREAT MART GENERAL TRADING LLC.",
    "GULF": "GULF",
    "GULF INDUSTRIAL": "GULF INDUSTRIAL",
    "GULF LANDSCAPING": "GULF LANDSCAPING",
    "GULF SNIPE": "GULF SNIPE",
    "GULF STEEL": "GULF STEEL",
    "HAFILAT GENERAL TRANSPORT - SOLE PROP. LLC": "HAFILAT GENERAL TRANSPORT - SOLE PROP. LLC",
    "HAFILAT GENERAL TRANSPORT-SOLE": "HAFILAT GENERAL TRANSPORT-SOLE",
    "HASSAN ALLAM CONTRUCTION": "HASSAN ALLAM CONTRUCTION",
    "HEALTH FOR ALL MEDICAL CENTER": "HEALTH FOR ALL MEDICAL CENTER",
    "HILAL BIL BADI": "HILAL BIL BADI",
    "HYSSNA INTERNATIONAL L.L.C": "HYSSNA INTERNATIONAL L.L.C",
    "I G G FOR MILITARY AND FORMAL GERMENTS": "I G G FOR MILITARY AND FORMAL GERMENTS",
    "INNOVO": "INNOVO",
    "INTERNATIONAL DECOR CO L.L.C - DUBAI BRANCH": "INTERNATIONAL DECOR CO L.L.C - DUBAI BRANCH",
    "INTERNATIONAL DEVELOPMENT COMPANY L.L.C": "INTERNATIONAL DEVELOPMENT COMPANY L.L.C",
    "J P M ASSOCIATES TECHNICAL SERVICES L.L.C": "J P M ASSOCIATES TECHNICAL SERVICES L.L.C",
    "JAZAL ENGINEERING": "JAZAL ENGINEERING",
    "JEET CONSTRUCTION LLC": "JEET CONSTRUCTION LLC",
    "KAYAN ALMOSTQBAL CONTRACTING & GEN. MAINTENANCE": "KAYAN ALMOSTQBAL CONTRACTING & GEN. MAINTENANCE",
    "LAITH ELECTRO": "LAITH ELECTRO",
    "LASSI TOP CAFE": "LASSI TOP CAFE",
    "M4 CONTRACTING": "M4 CONTRACTING",
    "MALABAR DAWATH REASTURANT AND GRILL": "MALABAR DAWATH REASTURANT AND GRILL",
    "MAZAYA ALMUTAHIDA CONSTRUCTIONS": "MAZAYA ALMUTAHIDA CONSTRUCTIONS",
    "MECHANICAL & CIVIL ENG.CONTRACTORS CO(MACE)LLC": "MECHANICAL & CIVIL ENG.CONTRACTORS CO(MACE)LLC",
    "MES SECURITY": "MES SECURITY",
    "ML MANPOWER": "ML MANPOWER",
    "MOBILE SOLUTIONS LLC": "MOBILE SOLUTIONS LLC",
    "MODERN BUILDING GENERAL CONTRACTING L L C": "MODERN BUILDING GENERAL CONTRACTING L L C",
    "MOSAYED BIN HAFEEZ CONT GEN TRANSPORT": "MOSAYED BIN HAFEEZ CONT GEN TRANSPORT",
    "MOUNTAIN GATE PROPRTY INVESTEMENT LLC": "MOUNTAIN GATE PROPRTY INVESTEMENT LLC",
    "MQ PEAEL ENGINEERING-L.L.C": "MQ PEAEL ENGINEERING-L.L.C",
    "NATIONAL CATERING COMPANY": "NATIONAL CATERING COMPANY",
    "NATIONAL INNOVATIVE": "NATIONAL INNOVATIVE",
    "NAZIM MAINTENANC GENERAL CONTRACTING COMPANY": "NAZIM MAINTENANC GENERAL CONTRACTING COMPANY",
    "NMDC ENERGY P.J.S.C": "NMDC ENERGY P.J.S.C",
    "NOOR AL SAHARA": "NOOR AL SAHARA",
    "NUROL LLC.": "NUROL LLC.",
    "OPTIMUM ENGINEERING": "OPTIMUM ENGINEERING",
    "PERFECT STEP GENERAL CONTRACTING & MAINTENANCE": "PERFECT STEP GENERAL CONTRACTING & MAINTENANCE",
    "PIONEER PRECAST CONCRETE LLC": "PIONEER PRECAST CONCRETE LLC",
    "PIONEER PRECAST CONCRETE LLCC": "PIONEER PRECAST CONCRETE LLCC",
    "POLENSKY & ZOELLNER COMPANY ABU DHABI W L L": "POLENSKY & ZOELLNER COMPANY ABU DHABI W L L",
    "PRINCE INTERNATIONALGENERAL TRANSPORT - L.L.C": "PRINCE INTERNATIONALGENERAL TRANSPORT - L.L.C",
    "PROFILE RECRUITMENT": "PROFILE RECRUITMENT",
    "QAMRA": "QAMRA",
    "QUICK SERVICES GENERAL TRANSPORT": "QUICK SERVICES GENERAL TRANSPORT",
    "SBK HOLDING (L.L.C)": "SBK HOLDING (L.L.C)",
    "SECURIGUARD MIDDLE EAST LLC": "SECURIGUARD MIDDLE EAST LLC",
    "SIBCA ELECTRONIC": "SIBCA ELECTRONIC",
    "SILVER SCREEN GENERAL CONTRACTING L.L.C": "SILVER SCREEN GENERAL CONTRACTING L.L.C",
    "SIX SIGMA": "SIX SIGMA",
    "SNIPE OIL AND GAS EQUIPMENT": "SNIPE OIL AND GAS EQUIPMENT",
    "STAR SERVICES": "STAR SERVICES",
    "SWITCHGEAR ELECTRO MECHANICAL LLC": "SWITCHGEAR ELECTRO MECHANICAL LLC",
    "TANZIFO": "TANZIFO",
    "TECTON ENGINEEING & CONSTRUCTION": "TECTON ENGINEEING & CONSTRUCTION",
    "TOOLS MAN GENERAL MAINTENANCE": "TOOLS MAN GENERAL MAINTENANCE",
    "UNITED MAZAYA": "UNITED MAZAYA",
    "VOLTAS LIMITED": "VOLTAS LIMITED",
    "WADE ADAMS CONTRACTING LLC": "WADE ADAMS CONTRACTING LLC",
    "ZUBLIN CONSTRUCTION LLC": "ZUBLIN CONSTRUCTION LLC"
}

# --------------------
# CPT/ICD helper: safe DF pick + debug
# --------------------
def _pick_first_df(*candidates):
    """Return the first candidate that is a non-empty DataFrame."""
    for x in candidates:
        if isinstance(x, pd.DataFrame) and not x.empty:
            return x
    return pd.DataFrame()

def _summary_keys(dfs):
    try:
        return sorted(list(dfs.keys()))
    except Exception:
        return []

# Optional S3
try:
    import boto3
except Exception:
    boto3 = None



# ---------------------------
# Date formatting (management-friendly)
# ---------------------------
def fmt_day(ts) -> str:
    """Friendly day label with weekday for management views."""
    try:
        return pd.to_datetime(ts).strftime("%A, %d %b %Y")
    except Exception:
        return str(ts)

def fmt_dt(ts) -> str:
    try:
        return pd.to_datetime(ts).strftime("%d %b %Y %H:%M")
    except Exception:
        return str(ts)


def fmt_short_day(ts) -> str:
    """Short readable date without weekday."""
    try:
        return pd.to_datetime(ts).strftime("%d %b %Y")  # 03 Feb 2026
    except Exception:
        return str(ts)

def fmt_range(a, b) -> str:
    return f"{fmt_short_day(a)} → {fmt_short_day(b)}"


st.set_page_config(page_title="Registration Summary (View Only)", layout="wide", initial_sidebar_state="collapsed")


# ---------------------------
# Premium UI (management view)
# ---------------------------
st.markdown(
    """
    <style>
      :root{
        --card-bg: rgba(255,255,255,0.92);
        --card-border: rgba(16, 24, 40, 0.08);
        --shadow2: 0 6px 18px rgba(16,24,40,0.08);
        --text: #0f172a;
        --muted: #64748b;
      }
      .block-container{max-width: 100% !important; width: 100% !important; padding-top: 0.6rem; padding-bottom: 2.5rem; padding-left: 2rem; padding-right: 2rem; margin-left: 0 !important; margin-right: 0 !important;}
      /* --- HARD FULL-WIDTH OVERRIDE (Streamlit Cloud DOM variations) --- */
      [data-testid="stAppViewBlockContainer"],
      div[data-testid="stAppViewBlockContainer"]{
        max-width: 100% !important;
        width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
        margin-left: 0 !important;
        margin-right: 0 !important;
      }
      div[data-testid="stAppViewContainer"], .stApp{
        max-width: 100% !important;
        width: 100% !important;
      }
      section.main, div[data-testid="stAppViewContainer"] > div.main{
        max-width: 100% !important;
        width: 100% !important;
      }
      .main .block-container{
        max-width: 100% !important;
        width: 100% !important;
      }
      h1,h2,h3{letter-spacing:-0.02em; line-height: 1.35 !important; overflow: visible !important; white-space: normal !important;}
      h1{font-weight:800; padding-bottom: 4px;}
      h2{font-weight:800;}
      h3{font-weight:700;}
      div[data-baseweb="select"] span { line-height: 1.4 !important; }
      hr{border: none; border-top: 1px solid rgba(16,24,40,0.08); margin: 1.25rem 0;}
      .kpi-grid{
        display:grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 14px;
        margin: 0.25rem 0 0.75rem 0;
      }
      @media (max-width: 1200px){ .kpi-grid{grid-template-columns: repeat(2, minmax(0, 1fr));} }
      @media (max-width: 700px){ .kpi-grid{grid-template-columns: repeat(1, minmax(0, 1fr));} }

      .kpi-card{
        background: var(--card-bg);
        border: 1px solid var(--card-border);
        border-radius: 18px;
        box-shadow: var(--shadow2);
        padding: 14px 16px;
      }
      .kpi-label{
        font-size: 13px;
        color: var(--muted);
        font-weight: 700;
        margin-bottom: 6px;
      }
      .kpi-value{
        font-size: 28px;
        font-weight: 850;
        color: var(--text);
        line-height: 1.1;
      }
      .kpi-sub{
        font-size: 12px;
        color: var(--muted);
        margin-top: 6px;
      }

      div[data-testid="stDataFrame"]{
        background: rgba(255,255,255,0.92);
        border: 1px solid rgba(16,24,40,0.08);
        border-radius: 16px;
        box-shadow: var(--shadow2);
        padding: 8px 10px 2px 10px;
      }
      details{
        border-radius: 16px;
        border: 1px solid rgba(16,24,40,0.08);
        box-shadow: var(--shadow2);
        background: rgba(255,255,255,0.92);
        padding: 6px 10px;
      }
      button[data-baseweb="tab"]{ font-weight: 800 !important; }
      .stCaption{color: var(--muted);}
    
      .page-title{
        font-size: 2.35rem;
        font-weight: 900;
        color: var(--text);
        letter-spacing:-0.03em;
        white-space: nowrap;
        overflow: visible;
        line-height: 1.05;
      }
      @media (max-width: 1100px){
        .page-title{font-size: 1.9rem; white-space: normal;}
      }

    </style>
    """,
    unsafe_allow_html=True,
)

def _kpi_cards(items, subtitle: str = ""):
    """Render premium KPI cards. items = list of (label, value)."""
    cards_html = []
    for label, value in items:
        cards_html.append(
            f"""<div class='kpi-card'>
                  <div class='kpi-label'>{label}</div>
                  <div class='kpi-value'>{value}</div>
                </div>"""
        )
    sub_html = f"<div class='kpi-sub'>{subtitle}</div>" if subtitle else ""
    html = f"<div class='kpi-grid'>{''.join(cards_html)}</div>{sub_html}"
    st.markdown(html, unsafe_allow_html=True)




# ---------------------------
# Helpers
# ---------------------------
def s3_key(*parts: str) -> str:
    return "/".join([p.strip("/").strip() for p in parts if p is not None and str(p).strip() != ""])


def load_secrets() -> Dict[str, str]:
    def get_any(*keys):
        for k in keys:
            if k in st.secrets:
                v = st.secrets.get(k)
                if v is not None and str(v).strip() != "":
                    return str(v).strip()
            v = os.getenv(k)
            if v is not None and str(v).strip() != "":
                return str(v).strip()
        return ""

    return {
        "AWS_ACCESS_KEY_ID": get_any("AWS_ACCESS_KEY_ID"),
        "AWS_SECRET_ACCESS_KEY": get_any("AWS_SECRET_ACCESS_KEY"),
        "AWS_REGION": get_any("AWS_REGION", "AWS_DEFAULT_REGION"),
        "S3_BUCKET_NAME": get_any("S3_BUCKET_NAME", "S3_BUCKET"),
        "S3_BASE_PREFIX": get_any("S3_BASE_PREFIX", "S3_PREFIX"),  # optional (unused by default)
    }


def s3_enabled(cfg: Dict[str, str]) -> bool:
    return (
        bool(cfg.get("S3_BUCKET_NAME"))
        and bool(cfg.get("AWS_REGION"))
        and bool(cfg.get("AWS_ACCESS_KEY_ID"))
        and bool(cfg.get("AWS_SECRET_ACCESS_KEY"))
        and boto3 is not None
    )


@st.cache_resource(show_spinner=False)
def s3_client_cached(cfg: Dict[str, str]):
    if not s3_enabled(cfg):
        return None
    return boto3.client(
        "s3",
        region_name=cfg["AWS_REGION"],
        aws_access_key_id=cfg["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=cfg["AWS_SECRET_ACCESS_KEY"],
    )


def s3_get_bytes(s3, bucket: str, key: str) -> Optional[bytes]:
    try:
        obj = s3.get_object(Bucket=bucket, Key=key)
        return obj["Body"].read()
    except Exception:
        return None


def s3_key_exists(s3, bucket: str, key: str) -> bool:
    try:
        s3.head_object(Bucket=bucket, Key=key)
        return True
    except Exception:
        return False


def candidate_base_prefixes(cfg: Dict[str, str]) -> List[str]:
    """Try a few likely prefixes so the viewer works even if uploader/viewer prefixes differ."""
    prefs: List[str] = []
    p = (cfg.get("S3_BASE_PREFIX") or "").strip().strip("/")
    if p:
        prefs.append(p)
    # common fallbacks
    prefs.append("")  # root of bucket
    if "streamlit" not in prefs:
        prefs.append("streamlit")
    # de-dup while preserving order
    out: List[str] = []
    for x in prefs:
        x = (x or "").strip().strip("/")
        if x not in out:
            out.append(x)
    return out


def history_paths(center: str, base_prefix: str = "") -> Tuple[str, str]:
    """Return (root_prefix, history_csv_key) for this center.

    Expected uploader layout (based on your S3 screenshots):
      <base_prefix>/registration/<center>/history.csv
      <base_prefix>/registration/<center>/<YYYY-MM-DD>/summary.pkl
    """
    root = s3_key(base_prefix, "registration", center)
    return root, s3_key(root, "history.csv")


def resolve_center_root_from_s3(s3, cfg: Dict[str, str], center_key: str) -> Tuple[str, str]:
    """Return (root_prefix, history_csv_key) that actually exists in S3."""
    bucket = cfg["S3_BUCKET_NAME"]
    for pref in candidate_base_prefixes(cfg):
        root, hist_key = history_paths(center_key, pref)
        if s3_key_exists(s3, bucket, hist_key):
            return root, hist_key
    # default to the configured prefix path (even if missing), for clearer error messages
    root, hist_key = history_paths(center_key, (cfg.get("S3_BASE_PREFIX") or ""))
    return root, hist_key


def load_history_from_s3(s3, cfg: Dict[str, str], center_key: str) -> Tuple[pd.DataFrame, str]:
    root, hist_key = resolve_center_root_from_s3(s3, cfg, center_key)
    b = s3_get_bytes(s3, cfg["S3_BUCKET_NAME"], hist_key)
    if not b:
        return pd.DataFrame(), root
    try:
        df = pd.read_csv(io.BytesIO(b), parse_dates=["day"])
    except Exception:
        df = pd.read_csv(io.BytesIO(b))
    return df, root


def load_summary_from_s3(
    s3,
    cfg: Dict[str, str],
    root_prefix: str,
    day_ts: pd.Timestamp
) -> Optional[Dict[str, pd.DataFrame]]:
    day_str = pd.to_datetime(day_ts).date().isoformat()
    key = s3_key(root_prefix, day_str, "summary.pkl")
    b = s3_get_bytes(s3, cfg["S3_BUCKET_NAME"], key)
    if not b:
        return None
    try:
        return pickle.loads(b)
    except Exception:
        return None


def add_cumulative(hist: pd.DataFrame) -> pd.DataFrame:
    if hist is None or hist.empty:
        return pd.DataFrame()
    h = hist.sort_values("day").copy()
    for c in ["total_visits", "unique_emr", "unique_visitno", "cash_patients", "pending_patients"]:
        if c in h.columns:
            h[c] = h[c].fillna(0).astype(int)
            h[f"cum_{c}"] = h[c].cumsum()
    # show latest first
    return h.sort_values("day", ascending=False).reset_index(drop=True)


def render_summary(dfs: Dict[str, pd.DataFrame], day_ts: pd.Timestamp, heading: str = "header", label: str = "Current Day", picked_label_override: Optional[str] = None):
    # NOTE: day_ts is used only for display/keying; weekly/monthly uses latest saved day.
    title = picked_label_override or f"{label} ({fmt_day(day_ts)})"

    _hdr_col1, _hdr_col2 = st.columns([6, 2])
    with _hdr_col1:
        if heading == "subheader":
            st.subheader(title)
        else:
            st.header(title)
    with _hdr_col2:
        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)
        if st.button("📧 Email Income Analysis", key=f"email_income_top_{label}_{str(day_ts.date())}"):
            try:
                _report_dt = pd.to_datetime(day_ts)
                _fname = _report_dt.strftime("EMC - INCOME ANALYSIS REPORT - %d %B %Y.xlsx")
                _excel_bytes = _build_income_excel(dfs, title)
                _html_body = _dfs_to_html(dfs, "Income Analysis (Doctor Revenue)", title)
                _subject = f"EMC Income Analysis Report – {title}"
                _send_email_smtp(
                    subject=_subject,
                    html_body=_html_body,
                    attachment_bytes=_excel_bytes,
                    attachment_filename=_fname,
                )
                st.success(f"✅ Email sent with attachment: {_fname}")
            except Exception as _e:
                st.error(f"Email failed: {_e}")
    def _sort_with_total(df: pd.DataFrame, label_col: str, count_col: str = "Count", total_label: str = "TOTAL") -> pd.DataFrame:
        """Sort by count desc, keep TOTAL row at bottom if present."""
        if df is None or df.empty or count_col not in df.columns or label_col not in df.columns:
            return df
        d = df.copy()
        # Separate TOTAL row (case-insensitive) if exists
        lbl = d[label_col].astype(str).str.strip().str.upper()
        is_total = lbl.eq(total_label.upper())
        total = d[is_total]
        d = d[~is_total]
        d[count_col] = pd.to_numeric(d[count_col], errors="coerce").fillna(0)
        d = d.sort_values(count_col, ascending=False, kind="mergesort")
        if not total.empty:
            return pd.concat([d, total], ignore_index=True)
        return d

    def _sort_income(df: pd.DataFrame) -> pd.DataFrame:
        """Sort Income Analysis tables largest→smallest by visits (Total_Visit) when available.
        Falls back to amount-based sorting if visit column is missing.
        """
        if df is None or df.empty:
            return df
        d = df.copy()

        # 1) Prefer visit-based sorting (requested by user)
        def _norm(s: str) -> str:
            return re.sub(r"[^a-z0-9]+", "", str(s).lower())

        visit_col = None
        visit_norms = {
            "totalvisit", "totalvisits", "total_visit", "total_visits", "visits", "visit"
        }
        for c in d.columns:
            if _norm(c) in visit_norms:
                visit_col = c
                break
        # common exact header in your tables
        if visit_col is None and "Total_Visit" in d.columns:
            visit_col = "Total_Visit"

        if visit_col is not None:
            d[visit_col] = pd.to_numeric(d[visit_col], errors="coerce").fillna(0)
            return d.sort_values(visit_col, ascending=False, kind="mergesort")

        # 2) Fallback: amount-based sorting
        preferred = ["net_amount", "net amount", "total_amount", "total amount", "amount", "net", "paid"]
        num_cols = []
        for c in d.columns:
            if pd.api.types.is_numeric_dtype(d[c]):
                num_cols.append(c)
        for c in d.columns:
            if c not in num_cols and any(str(c).lower() == p for p in preferred):
                d[c] = pd.to_numeric(d[c], errors="coerce")
                if pd.api.types.is_numeric_dtype(d[c]):
                    num_cols.append(c)

        sort_col = None
        for p in preferred:
            for c in d.columns:
                if str(c).lower() == p:
                    sort_col = c
                    break
            if sort_col:
                break
        if sort_col is None and num_cols:
            sort_col = num_cols[-1]
        if sort_col:
            d[sort_col] = pd.to_numeric(d[sort_col], errors="coerce").fillna(0)
            d = d.sort_values(sort_col, ascending=False, kind="mergesort")
        return d
    kpi = dfs.get("KPI")
    if kpi is not None and not kpi.empty and "Metric" in kpi.columns and "Value" in kpi.columns:
        k = kpi.set_index("Metric")["Value"]
        # Premium KPI cards (management-friendly)
        subtitle = f"Generated: {fmt_dt(datetime.now())}"
        _kpi_cards([
            ("Total Visits", int(k.get("Total Visits", 0))),
            ("New Patients", int(k.get("New Patients", 0))),
            ("Established Patients", int(k.get("Established Patients", 0))),
            ("Follow Up", int(k.get("Follow Up", 0))),
            ("Pending Patients", int(k.get("Pending Patients", 0))),
        ], subtitle=subtitle)
    else:
        st.info("KPI is not available for this day.")

    st.subheader(f"Pending Status Wise (Day: {fmt_day(day_ts)})")
    st.dataframe(dfs.get("Pending Status Wise", pd.DataFrame()), use_container_width=True, hide_index=True)

    st.subheader("Insurance Wise Visits")
    _iw = dfs.get("Insurance Wise Visits", pd.DataFrame())
    _iw = _sort_with_total(_iw, label_col="Insurance", count_col="Count", total_label="TOTAL")
    st.dataframe(_iw, use_container_width=True, hide_index=True)

    st.subheader("Doctor Wise Visits")
    _dw = dfs.get("Doctor Wise Visits", pd.DataFrame())
    _dw = _sort_with_total(_dw, label_col="Doctor", count_col="Count", total_label="TOTAL")
    st.dataframe(_dw, use_container_width=True, hide_index=True)

    # -------------------- Income Analysis (Doctor Revenue) -------------------- (Doctor Revenue) --------------------
    income_keys = [k for k in dfs.keys() if str(k).startswith("Income | ")]
    if income_keys:
        st.markdown("---")
        st.header("Income Analysis (Doctor Revenue)")

        df_doc = dfs.get("Income | Doctor Wise Revenue")
        df_ins = dfs.get("Income | Insurance Wise Revenue")
        df_dx  = dfs.get("Income | Doctor x Insurance Revenue")

        def _add_proc_rad_per_visit(df: pd.DataFrame) -> pd.DataFrame:
            """Override Procedure/Radiology per-visit values for display (Procedure/Visits, Radiology/Visits).
            Keeps GRAND TOTAL intact and avoids misleading % columns.
            """
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                return df
            out = df.copy()

            # Identify visit column (usually Total_Visit)
            visit_col = "Total_Visit" if "Total_Visit" in out.columns else None
            if visit_col is None:
                for c in out.columns:
                    if str(c).strip().lower() in ["total visit", "total visits", "visits", "visit", "total_visit", "total_visits"]:
                        visit_col = c
                        break
            if visit_col is None:
                return out

            denom = pd.to_numeric(out[visit_col], errors="coerce").replace(0, pd.NA)

            # Procedure per visit
            if "Procedure" in out.columns:
                out["Procedure_Per_Visit"] = (pd.to_numeric(out["Procedure"], errors="coerce") / denom).fillna(0)

            # Radiology per visit
            if "Radiology" in out.columns:
                out["Radiology_Per_Visit"] = (pd.to_numeric(out["Radiology"], errors="coerce") / denom).fillna(0)

            # Remove misleading percentage columns if present
            for c in ["Procedure_%", "Radiology_%"]:
                if c in out.columns:
                    out = out.drop(columns=[c], errors="ignore")

            return out

        # Apply per-visit override for display
        df_doc = _add_proc_rad_per_visit(df_doc)
        df_ins = _add_proc_rad_per_visit(df_ins)
        df_dx  = _add_proc_rad_per_visit(df_dx)

        tabs = st.tabs(["Doctor Wise", "Insurance Wise", "Doctor x Insurance"])
        def _move_grand_total_bottom(df: pd.DataFrame) -> pd.DataFrame:
            """Keep GRAND TOTAL row(s) at the bottom for management tables."""
            if df is None or df.empty:
                return df
            d = df.copy()
            mask = pd.Series(False, index=d.index)
            for col in ["Doctor", "Insurance", "Department"]:
                if col in d.columns:
                    mask = mask | d[col].astype(str).str.upper().str.contains("GRAND TOTAL", na=False)
            total = d[mask]
            d = d[~mask]
            return pd.concat([d, total], ignore_index=True)



        def _round_income_display(df: pd.DataFrame) -> pd.DataFrame:
            if df is None or df.empty:
                return df
            x = _sort_income(df)
            x = _move_grand_total_bottom(x)
            x = x.copy()

            # Drop misleading % cols
            for c in ["Radiology_%", "Procedure_%"]:
                if c in x.columns:
                    x = x.drop(columns=[c])

            # Compute per-visit ratios if not already present
            visit_col = next((c for c in x.columns if str(c).strip().lower() in
                              ["total_visit", "total visit", "visits", "visit"]), None)
            if visit_col is not None:
                denom = pd.to_numeric(x[visit_col], errors="coerce").replace(0, pd.NA)
                if "Procedure" in x.columns and "Procedure_Per_Visit" not in x.columns:
                    x["Procedure_Per_Visit"] = (pd.to_numeric(x["Procedure"], errors="coerce") / denom).round(2).fillna(0)
                if "Radiology" in x.columns and "Radiology_Per_Visit" not in x.columns:
                    x["Radiology_Per_Visit"] = (pd.to_numeric(x["Radiology"], errors="coerce") / denom).round(2).fillna(0)

            # Smart rounding: integers for counts, 2dp for avg/%, 1dp for amounts
            _int_c = {"Consultation","Lab","Radiology","Procedure","Total_Visit","Visits"}
            _pct_c = {"Lab_%","Avg_Amount_Service","Avg_Amount_Insuance","Avg_Amount_Insurance",
                      "Procedure_Per_Visit","Radiology_Per_Visit"}
            for col in x.columns:
                if not pd.api.types.is_numeric_dtype(x[col]):
                    continue
                if col in _int_c:
                    x[col] = pd.to_numeric(x[col], errors="coerce").round(0).fillna(0).astype(int)
                elif col in _pct_c:
                    x[col] = pd.to_numeric(x[col], errors="coerce").round(2)
                else:
                    s = pd.to_numeric(x[col], errors="coerce").round(1)
                    if s.dropna().apply(lambda v: v == int(v)).all():
                        x[col] = s.fillna(0).astype(int)
                    else:
                        x[col] = s

            # Full rename to clean display names
            x = x.rename(columns={
                "Total_Amount_Service":   "Total Service",
                "Total_Amount_Insuance":  "Total Insurance",
                "Total_Amount_Insurance": "Total Insurance",
                "Avg_Amount_Service":     "Avg Service",
                "Avg_Amount_Insuance":    "Avg Insurance",
                "Avg_Amount_Insurance":   "Avg Insurance",
                "Avg.Amount":             "Avg Insurance",
                "Lab_%":                  "Lab %",
                "Procedure_Per_Visit":    "Procedure %",
                "Radiology_Per_Visit":    "Radiology %",
                "Total_Visit":            "Visits",
                "Department":             "Dept",
            })

            # Reorder columns
            _ORDER = [
                "Dept", "Doctor", "Insurance",
                "Consultation", "Lab", "Radiology", "Procedure",
                "Visits",
                "Total Service", "Total Insurance",
                "Avg Service", "Avg Insurance",
                "Lab %", "Procedure %", "Radiology %",
            ]
            ordered = [c for c in _ORDER if c in x.columns]
            remaining = [c for c in x.columns if c not in ordered]
            return x[ordered + remaining]


        with tabs[0]:
            if df_doc is None or df_doc.empty:
                st.info("No Doctor Wise revenue data for this day.")
            else:
                st.dataframe(_round_income_display(df_doc), use_container_width=True, hide_index=True)

        with tabs[1]:
            if df_ins is None or df_ins.empty:
                st.info("No Insurance Wise revenue data for this day.")
            else:
                st.dataframe(_round_income_display(df_ins), use_container_width=True, hide_index=True)

        with tabs[2]:
            if df_dx is None or df_dx.empty:
                st.info("No Doctor x Insurance revenue data for this day.")
            else:
                df_f = df_dx.copy()

                # Filter: pick doctor first
                if "Doctor" in df_f.columns:
                    doctors = sorted([
                        d for d in df_f["Doctor"].dropna().unique()
                        if str(d).strip().lower() not in ["", "none", "nan"]
                        and str(d).strip().upper() != "GRAND TOTAL"
                    ])
                    if doctors:
                        pick_doc = st.selectbox("Select Doctor", options=doctors, key=f"income_pick_doc_{str(day_ts)}")
                        df_f = df_f[df_f["Doctor"] == pick_doc].copy()

                # Filter: pick insurance (optional)
                if "Insurance" in df_f.columns:
                    ins_list = sorted([
                        i for i in df_f["Insurance"].dropna().unique()
                        if str(i).strip().lower() not in ["", "none", "nan"] and str(i).strip().upper() != "GRAND TOTAL"
                    ])
                    pick_ins = st.selectbox("Select Insurance", options=["All"] + ins_list, key=f"income_pick_ins_{str(day_ts)}")
                    if pick_ins != "All":
                        df_f = df_f[df_f["Insurance"] == pick_ins].copy()

                st.dataframe(_round_income_display(df_f), use_container_width=True, hide_index=True)



    

    # -------------------- CPT / ICD Analysis --------------------
    # NOTE: Never use `or` between DataFrames (pandas raises: "truth value of a DataFrame is ambiguous").
    # We support BOTH key styles:
    #   New (viewer-style): "CPTICD | ..."
    #   Old (uploader-style): "Doctor x Company | ..." / "CPT -> Top Principal ICD" / "Employer Expiry Tracker"
    def _pick_first_df(keys: List[str]) -> Optional[pd.DataFrame]:
        first_any: Optional[pd.DataFrame] = None
        for kk in keys:
            v = dfs.get(kk)
            if isinstance(v, pd.DataFrame):
                if first_any is None:
                    first_any = v
                if not v.empty:
                    return v
        return first_any

    has_cpticd = any(str(k).startswith("CPTICD | ") for k in dfs.keys()) or any(
        k in dfs for k in [
            "Doctor x Company | Principal DX (Top1)",
            "Doctor x Company | Secondary DX (Top1)",
            "Doctor x Insurance | Principal DX (Counts)",
            "Doctor x Insurance | Secondary DX (Counts)",
            "Doctor x Insurance | Visits",
            "Doctor x Insurance | Principal DX (Top1)",
            "Doctor x Insurance | Secondary DX (Top1)",
            "CPT -> Top Principal ICD",
            "Employer Expiry Tracker",
        ]
    )

    if has_cpticd:
        st.markdown("---")
        st.header("CPT / ICD Analysis")

        # Simplified display (Doctor + Insurance only)
        # Prefer VISIT-LEVEL Principal DX counts (totals intended to match visits)
        df_pri = _pick_first_df([
            "CPTICD | Doctor x Insurance | Principal DX (Counts)",
            "Doctor x Insurance | Principal DX (Counts)",
            # fallback: old Top1 keys
            "CPTICD | Doctor x Insurance | Principal DX (Top1)",
            "CPTICD | Doctor x Company | Principal DX (Top1)",
            "Doctor x Insurance | Principal DX (Top1)",
            "Doctor x Company | Principal DX (Top1)",
        ])
        df_sec = _pick_first_df([
            "CPTICD | Doctor x Insurance | Secondary DX (Counts)",
            "Doctor x Insurance | Secondary DX (Counts)",
            # fallback: old Top1 keys
            "CPTICD | Doctor x Insurance | Secondary DX (Top1)",
            "CPTICD | Doctor x Company | Secondary DX (Top1)",
            "Doctor x Insurance | Secondary DX (Top1)",
            "Doctor x Company | Secondary DX (Top1)",
        ])
        df_cpt_map = _pick_first_df([
            "CPTICD | CPT -> Top Principal ICD",
            "CPT -> Top Principal ICD",
        ])

        tabs = st.tabs(["Doctor x Insurance", "CPT Mapping"])

        def _clean_diag(df: Optional[pd.DataFrame]) -> pd.DataFrame:
            if df is None or df.empty:
                return pd.DataFrame()
            out = df.copy()

            # Drop unnamed/blank columns (common after Excel export)
            bad_cols = []
            for c in list(out.columns):
                sc = str(c).strip()
                if sc == "" or sc.lower().startswith("unnamed"):
                    bad_cols.append(c)
            if bad_cols:
                out = out.drop(columns=bad_cols, errors="ignore")

            # Drop employer/company columns if present
            for drop_c in ["Employer", "Company"]:
                if drop_c in out.columns:
                    out = out.drop(columns=[drop_c])

            # Fix common typo
            if "Insuance" in out.columns and "Insurance" not in out.columns:
                out = out.rename(columns={"Insuance": "Insurance"})

            # Keep only requested columns where available (Doctor/Insurance/ICD/Count/Desc)
            keep = [c for c in ["Doctor", "Insurance", "ICD", "Count", "ICD Description"] if c in out.columns]
            return out[keep] if keep else out

        with tabs[0]:
            st.subheader("Top Diagnosis (Doctor x Insurance)")

            pri_clean = _clean_diag(df_pri)
            sec_clean = _clean_diag(df_sec)

            # --- Filters (Doctor + Insurance) ---
            doctors = []
            ins_list = []
            for _d in [pri_clean, sec_clean]:
                if isinstance(_d, pd.DataFrame) and not _d.empty:
                    if "Doctor" in _d.columns:
                        doctors += _d["Doctor"].dropna().astype(str).tolist()
                    if "Insurance" in _d.columns:
                        ins_list += _d["Insurance"].dropna().astype(str).tolist()

            doctors = sorted({d.strip() for d in doctors if str(d).strip() != ""})
            ins_list = sorted({i.strip() for i in ins_list if str(i).strip() != ""})

            f1, f2 = st.columns(2)
            with f1:
                pick_doc = st.selectbox("Select Doctor", ["All"] + doctors, index=0, key="cpticd_pick_doc")
            with f2:
                pick_ins = st.selectbox("Select Insurance", ["All"] + ins_list, index=0, key="cpticd_pick_ins")

            def _filter_diag(df: pd.DataFrame) -> pd.DataFrame:
                if df is None or df.empty:
                    return pd.DataFrame()
                out = df.copy()
                if pick_doc != "All" and "Doctor" in out.columns:
                    out = out[out["Doctor"].astype(str) == str(pick_doc)]
                if pick_ins != "All" and "Insurance" in out.columns:
                    out = out[out["Insurance"].astype(str) == str(pick_ins)]
                return out

            # --- Principal DX (Counts) with TOTAL + match with visits ---
            pri_show = _filter_diag(pri_clean)

            # Detect if this is visit-level counts table
            is_counts_view = False
            if isinstance(df_pri, pd.DataFrame):
                is_counts_view = any(
                    str(k).strip() in ["CPTICD | Doctor x Insurance | Principal DX (Counts)", "Doctor x Insurance | Principal DX (Counts)"]
                    for k in dfs.keys()
                )

            total_dx = None
            if not pri_show.empty and "Count" in pri_show.columns:
                try:
                    total_dx = int(pd.to_numeric(pri_show["Count"], errors="coerce").fillna(0).sum())
                except Exception:
                    total_dx = None

            # Expected visits from Income Doctor x Insurance table (if available)
            # Expected visits (prefer CPT/ICD visit table; fallback to Income Doctor x Insurance)
            expected_visits = None

            df_vis = dfs.get("Doctor x Insurance | Visits")
            if isinstance(df_vis, pd.DataFrame) and not df_vis.empty and "Visits" in df_vis.columns:
                tmpv = df_vis.copy()
                if pick_doc != "All" and "Doctor" in tmpv.columns:
                    tmpv = tmpv[tmpv["Doctor"].astype(str) == str(pick_doc)]
                if pick_ins != "All" and "Insurance" in tmpv.columns:
                    tmpv = tmpv[tmpv["Insurance"].astype(str) == str(pick_ins)]
                try:
                    expected_visits = int(pd.to_numeric(tmpv["Visits"], errors="coerce").fillna(0).sum())
                except Exception:
                    expected_visits = None

            if expected_visits is None:
                df_income_dx = dfs.get("Income | Doctor x Insurance Revenue")
                if isinstance(df_income_dx, pd.DataFrame) and not df_income_dx.empty and "Total_Visit" in df_income_dx.columns:
                    tmp = df_income_dx.copy()
                    if pick_doc != "All" and "Doctor" in tmp.columns:
                        tmp = tmp[tmp["Doctor"].astype(str) == str(pick_doc)]
                    if pick_ins != "All" and "Insurance" in tmp.columns:
                        tmp = tmp[tmp["Insurance"].astype(str) == str(pick_ins)]
                    # Exclude GRAND TOTAL rows if present
                    for coln in ["Doctor", "Insurance"]:
                        if coln in tmp.columns:
                            tmp = tmp[tmp[coln].astype(str).str.upper() != "GRAND TOTAL"]
                    try:
                        expected_visits = int(pd.to_numeric(tmp["Total_Visit"], errors="coerce").fillna(0).sum())
                    except Exception:
                        expected_visits = None

            # Remove any existing TOTAL/GRAND TOTAL rows (avoid duplicate totals)
            if not pri_show.empty:
                for _c in ["ICD", "Doctor", "Insurance"]:
                    if _c in pri_show.columns:
                        pri_show = pri_show[~pri_show[_c].astype(str).str.strip().str.upper().isin(["TOTAL", "GRAND TOTAL"])].copy()

            # Sort by Count (largest → smallest)
            if not pri_show.empty and "Count" in pri_show.columns:
                pri_show["Count"] = pd.to_numeric(pri_show["Count"], errors="coerce").fillna(0)
                pri_show = pri_show.sort_values("Count", ascending=False)

            # TOTAL (for management): Principal DX total should match Visits (unique VisitID) on day-level.
            # If we don't have VisitID-level data here, we enforce display TOTAL = expected_visits when available.
            total_dx_display = total_dx
            if expected_visits is not None:
                total_dx_display = expected_visits

            # Append ONE TOTAL row at end
            if not pri_show.empty and total_dx_display is not None:
                total_row = {c: "" for c in pri_show.columns}
                if "ICD" in pri_show.columns:
                    total_row["ICD"] = "TOTAL"
                elif "Doctor" in pri_show.columns:
                    total_row["Doctor"] = "TOTAL"
                else:
                    total_row[pri_show.columns[0]] = "TOTAL"
                if "Count" in pri_show.columns:
                    total_row["Count"] = int(total_dx_display)
                pri_show = pd.concat([pri_show, pd.DataFrame([total_row])], ignore_index=True)


            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Principal DX (Visit-level Counts)**" if is_counts_view else "**Principal DX**")
                # Summary line
                if total_dx is not None:
                    if expected_visits is not None:
                        st.caption(f"Principal DX TOTAL: {expected_visits}  |  Visits: {expected_visits}")
                    else:
                        st.caption(f"Principal DX TOTAL: {total_dx_display}")
                st.dataframe(pri_show, use_container_width=True, hide_index=True)

            with c2:
                st.markdown("**Secondary DX (Counts)**")
                sec_show = _filter_diag(sec_clean)

                total_sec = None
                if not sec_show.empty and "Count" in sec_show.columns:
                    try:
                        total_sec = int(pd.to_numeric(sec_show["Count"], errors="coerce").fillna(0).sum())
                    except Exception:
                        total_sec = None

                # Remove any existing TOTAL/GRAND TOTAL rows
                if not sec_show.empty:
                    for _c in ["ICD", "Doctor", "Insurance"]:
                        if _c in sec_show.columns:
                            sec_show = sec_show[
                                ~sec_show[_c].astype(str).str.strip().str.upper().isin(["TOTAL", "GRAND TOTAL"])
                            ].copy()

                # Sort by count (largest -> smallest)
                if not sec_show.empty and "Count" in sec_show.columns:
                    sec_show["Count"] = pd.to_numeric(sec_show["Count"], errors="coerce").fillna(0).astype(int)
                    sec_show = sec_show.sort_values("Count", ascending=False)

                # Append ONE footer TOTAL row at the bottom (only once)
                if not sec_show.empty and total_sec is not None:
                    total_row = {c: "" for c in sec_show.columns}
                    # Prefer putting TOTAL under ICD (or Doctor if ICD not present)
                    if "ICD" in sec_show.columns:
                        total_row["ICD"] = "TOTAL"
                    elif "Doctor" in sec_show.columns:
                        total_row["Doctor"] = "TOTAL"
                    if "Count" in sec_show.columns:
                        total_row["Count"] = total_sec
                    sec_show = pd.concat([sec_show, pd.DataFrame([total_row])], ignore_index=True)

                # Caption (Visit-level expected count is shown for reference)
                if total_sec is not None:
                    if expected_visits is not None:
                        st.caption(f"Secondary DX TOTAL: {total_sec}  |  Visits: {expected_visits}")
                    else:
                        st.caption(f"Secondary DX TOTAL: {total_sec}")

                st.dataframe(sec_show, use_container_width=True, hide_index=True)
        with tabs[1]:
            # Prefer new visit-bundle summary (saved by Summary page)
            df_bundle = dfs.get("CPT Visit Bundle", pd.DataFrame())
            if isinstance(df_bundle, pd.DataFrame) and not df_bundle.empty:
                st.subheader("CPT Visit Bundle Summary")

                df_show = df_bundle.copy()

                # Filters (Doctor + Insurance) like ICD
                f1, f2 = st.columns(2)
                with f1:
                    doc_list = sorted([
                        x for x in df_show.get("Doctor", pd.Series([], dtype=str)).dropna().astype(str).unique().tolist()
                        if str(x).strip() not in ["", "nan", "None"] and str(x).strip().upper() != "GRAND TOTAL"
                    ])
                    pick_doc2 = st.selectbox("Select Doctor", ["All"] + doc_list, index=0, key="cptbundle_pick_doc")

                with f2:
                    ins_list2 = sorted([
                        x for x in df_show.get("Insurance", pd.Series([], dtype=str)).dropna().astype(str).unique().tolist()
                        if str(x).strip() not in ["", "nan", "None"] and str(x).strip().upper() != "GRAND TOTAL"
                    ])
                    pick_ins2 = st.selectbox("Select Insurance", ["All"] + ins_list2, index=0, key="cptbundle_pick_ins")

                if pick_doc2 != "All" and "Doctor" in df_show.columns:
                    df_show = df_show[df_show["Doctor"].astype(str) == str(pick_doc2)].copy()
                if pick_ins2 != "All" and "Insurance" in df_show.columns:
                    df_show = df_show[df_show["Insurance"].astype(str) == str(pick_ins2)].copy()

                # Sort by Visits desc
                if "Visits" in df_show.columns:
                    df_show["Visits"] = pd.to_numeric(df_show["Visits"], errors="coerce").fillna(0).astype(int)
                    df_show = df_show.sort_values("Visits", ascending=False)

                # Caption like ICD: CPT TOTAL VISITS + Day
                total_visits_cpt = None
                if "Visits" in df_show.columns and not df_show.empty:
                    try:
                        total_visits_cpt = int(df_show["Visits"].sum())
                    except Exception:
                        total_visits_cpt = None

                if total_visits_cpt is not None:
                    st.caption(f"CPT TOTAL VISITS: {total_visits_cpt}  |  Day: {fmt_day(day_ts)}")

                # Display only key columns first (if they exist)
                prefer_cols = [c for c in ["Doctor", "Insurance", "CPT Bundle", "Principal DX", "Secondary DX", "Visits"] if c in df_show.columns]
                st.dataframe(df_show[prefer_cols] if prefer_cols else df_show, use_container_width=True, hide_index=True)

            else:
                # Fallback: old CPT -> Top Principal ICD mapping
                st.subheader("CPT → Most Common Principal ICD")

                df_cpt = df_cpt_map if isinstance(df_cpt_map, pd.DataFrame) else pd.DataFrame()
                if df_cpt is None or df_cpt.empty:
                    st.info("No CPT mapping data for this day.")
                else:
                    df_show = df_cpt.copy()

                    # --- Optional filters (like ICD tab): Doctor + Insurance + CPT ---
                    f1, f2, f3 = st.columns(3)

                    with f1:
                        if "Doctor" in df_show.columns:
                            doc_list = sorted([x for x in df_show["Doctor"].dropna().astype(str).unique().tolist() if str(x).strip() not in ["", "nan", "None"]])
                            pick_doc2 = st.selectbox("Select Doctor", ["All"] + doc_list, index=0, key="cptmap_pick_doc")
                        else:
                            pick_doc2 = "All"

                    with f2:
                        if "Insurance" in df_show.columns:
                            ins_list2 = sorted([x for x in df_show["Insurance"].dropna().astype(str).unique().tolist() if str(x).strip() not in ["", "nan", "None"]])
                            pick_ins2 = st.selectbox("Select Insurance", ["All"] + ins_list2, index=0, key="cptmap_pick_ins")
                        else:
                            pick_ins2 = "All"

                    with f3:
                        if "CPT" in df_show.columns:
                            cpt_list = sorted([x for x in df_show["CPT"].dropna().astype(str).unique().tolist() if str(x).strip() not in ["", "nan", "None"]])
                            pick_cpt = st.selectbox("Select CPT", ["All"] + cpt_list, index=0, key="cpticd_pick_cpt")
                        else:
                            pick_cpt = "All"

                    # Apply filters
                    if pick_doc2 != "All" and "Doctor" in df_show.columns:
                        df_show = df_show[df_show["Doctor"].astype(str) == str(pick_doc2)].copy()

                    if pick_ins2 != "All" and "Insurance" in df_show.columns:
                        df_show = df_show[df_show["Insurance"].astype(str) == str(pick_ins2)].copy()

                    if pick_cpt != "All" and "CPT" in df_show.columns:
                        df_show = df_show[df_show["CPT"].astype(str) == str(pick_cpt)].copy()

                    # Sort by Count (largest → smallest) if available
                    if "Count" in df_show.columns:
                        df_show["Count"] = pd.to_numeric(df_show["Count"], errors="coerce").fillna(0)
                        df_show = df_show.sort_values("Count", ascending=False)

                    # TOTAL row at end (Count)
                    if not df_show.empty and "Count" in df_show.columns:
                        total_c = int(df_show["Count"].sum())
                        total_row = {c: "" for c in df_show.columns}
                        # Put TOTAL label in CPT if present, else first column
                        if "CPT" in df_show.columns:
                            total_row["CPT"] = "TOTAL"
                        else:
                            total_row[df_show.columns[0]] = "TOTAL"
                        total_row["Count"] = total_c
                        df_show = pd.concat([df_show, pd.DataFrame([total_row])], ignore_index=True)
                    else:
                        total_c = None

                    # Caption like ICD: CPT TOTAL + Visits (+ Day)
                    try:
                        _vis = expected_visits if "expected_visits" in locals() else None
                    except Exception:
                        _vis = None
                    try:
                        _day_label = fmt_day(day_ts) if "day_ts" in locals() else None
                    except Exception:
                        _day_label = None

                    if total_c is not None:
                        if _vis is not None and _day_label:
                            st.caption(f"CPT TOTAL: {int(total_c)}  |  Visits: {_vis}  |  Day: {_day_label}")
                        elif _vis is not None:
                            st.caption(f"CPT TOTAL: {int(total_c)}  |  Visits: {_vis}")
                        elif _day_label:
                            st.caption(f"CPT TOTAL: {int(total_c)}  |  Day: {_day_label}")
                        else:
                            st.caption(f"CPT TOTAL: {int(total_c)}")

                    st.dataframe(df_show, use_container_width=True, hide_index=True)

                    # Note for you (in UI) if Doctor/Insurance are missing in saved table
                    if "Doctor" not in df_cpt.columns or "Insurance" not in df_cpt.columns:
                        st.caption("Note: Doctor/Insurance filters will appear only if the saved CPT mapping table contains Doctor and Insurance columns.")

    st.subheader("Employer Wise")
    emp_df = dfs.get("Employer Wise", pd.DataFrame()).copy()

    # Apply employer canonicalization to avoid duplicates (QUMRA/QAMRA etc.)
    if not emp_df.empty and "Employer" in emp_df.columns:
        def _norm_emp_local(x: object) -> str:
            s = "" if pd.isna(x) else str(x)
            s = re.sub(r"\s+", " ", s).strip().upper()
            return EMPLOYER_CANON_MAP.get(s, s)

        emp_df["Employer_norm"] = emp_df["Employer"].apply(_norm_emp_local)
        # Display using canonical label if available
        emp_df["Employer"] = emp_df["Employer_norm"].map(EMPLOYER_DISPLAY_MAP).fillna(emp_df["Employer_norm"])

        group_cols = ["Employer"]
        if "Insurance" in emp_df.columns:
            group_cols.append("Insurance")

        if "Count" in emp_df.columns:
            emp_df["Count"] = pd.to_numeric(emp_df["Count"], errors="coerce").fillna(0)
            emp_df = emp_df.groupby(group_cols, as_index=False)["Count"].sum()
        else:
            emp_df = emp_df.drop_duplicates(subset=group_cols).copy()

        # Sort by Count desc, keep TOTAL at bottom
        if "Count" in emp_df.columns:
            emp_df = _sort_with_total(emp_df, label_col="Employer", count_col="Count", total_label="TOTAL")

        emp_df = emp_df.drop(columns=["Employer_norm"], errors="ignore")
    # --- Employer expiry summary (STRICT employer from Registration, expiry from CPT/ICD) ---
    # We expect the uploader to save a tracker table that includes at least:
    #   Employer (from RegistrationList "Employer Name") + Expiry Date (from CPT/ICD file)
    # But to be robust, we also fall back to any table that contains an Employer-like column and an Expiry column.
    def _pick_expiry_df(dfs_dict: dict) -> pd.DataFrame | None:
        # 1) Prefer explicit tracker key
        for k, v in dfs_dict.items():
            if isinstance(v, pd.DataFrame) and "expiry" in str(k).lower() and "tracker" in str(k).lower():
                return v
        # 2) Any key mentioning expiry
        for k, v in dfs_dict.items():
            if isinstance(v, pd.DataFrame) and "expiry" in str(k).lower():
                return v
        # 3) Any DF that has expiry+employer columns
        for _, v in dfs_dict.items():
            if not isinstance(v, pd.DataFrame) or v.empty:
                continue
            cols = [c.lower().strip() for c in v.columns]
            if any("expiry" in c for c in cols) and any(c in ("employer", "employer name") or "employer" in c for c in cols):
                return v
        return None

    df_exp_all = _pick_expiry_df(dfs)

    exp_display_map: dict[str, str] = {}
    exp_top_date_map: dict[str, date | None] = {}
    today = date.today()

    def _norm_emp(x: str) -> str:
        """Normalize employer for grouping (uses canon map if present)."""
        s = str(x or "").strip().upper()
        s = re.sub(r"\s+", " ", s)
        canon = EMPLOYER_CANON_MAP.get(s, s)
        canon = str(canon or "").strip()
        canon_u = canon.upper()
        canon_u = re.sub(r"\s+", " ", canon_u)
        return canon_u

    if df_exp_all is not None and not df_exp_all.empty and not emp_df.empty and "Employer" in emp_df.columns:
        exp = df_exp_all.copy()

        # Detect employer column in tracker (STRICT: should already be Employer from RegistrationList)
        emp_col = None
        for c in exp.columns:
            cl = str(c).strip().lower()
            if cl == "employer" or cl == "employer name" or "employer" in cl:
                emp_col = c
                break

        # Detect expiry column
        exp_col = None
        for c in exp.columns:
            cl = str(c).strip().lower()
            if "expiry" in cl and ("date" in cl or cl == "expiry"):
                exp_col = c
                break
        if exp_col is None:
            # fallback: any column containing 'expiry'
            for c in exp.columns:
                if "expiry" in str(c).strip().lower():
                    exp_col = c
                    break

        if emp_col and exp_col:
            exp[emp_col] = exp[emp_col].astype(str).map(_norm_emp)
            exp["_expiry_date"] = pd.to_datetime(exp[exp_col], errors="coerce").dt.date

            # Option B: base % only on valid (non-null) expiry dates
            for emp in emp_df["Employer"].dropna().astype(str).unique().tolist():
                emp_key = _norm_emp(emp)
                sub = exp.loc[exp[emp_col] == emp_key, "_expiry_date"].dropna()
                if sub.empty:
                    exp_display_map[emp_key] = ""
                    exp_top_date_map[emp_key] = None
                    continue

                vc = sub.value_counts()
                top_date = vc.index[0]
                top_count = int(vc.iloc[0])
                total_valid = int(vc.sum())
                pct = (top_count / total_valid) * 100.0 if total_valid else 0.0

                if pct >= 70.0:
                    display = top_date.strftime("%Y-%m-%d")
                else:
                    display = f"Mixed (Top: {top_date.strftime('%Y-%m-%d')} – {int(round(pct))}%)"

                exp_display_map[emp_key] = display
                exp_top_date_map[emp_key] = top_date
        else:
            # missing columns
            pass

    if emp_df is None or emp_df.empty:
        st.dataframe(emp_df if emp_df is not None else pd.DataFrame(), use_container_width=True, hide_index=True)
    else:
        # Attach Expiry Date / Days To Expiry (prefer saved columns if present)
        if "Expiry Date" not in emp_df.columns or emp_df["Expiry Date"].fillna("").astype(str).str.strip().eq("").all():
            if "Employer" in emp_df.columns:
                emp_df["Expiry Date"] = emp_df["Employer"].map(lambda e: exp_display_map.get(_norm_emp(e), ""))
            else:
                emp_df["Expiry Date"] = ""

        # Days To Expiry (prefer saved)
        if "Days To Expiry" not in emp_df.columns or emp_df["Days To Expiry"].isna().all():
            if "Employer" in emp_df.columns:
                def _days_from_emp(e):
                    d = exp_top_date_map.get(_norm_emp(e))
                    if d is None:
                        return ""
                    try:
                        return (d - today).days
                    except Exception:
                        return ""
                emp_df["Days To Expiry"] = emp_df["Employer"].map(_days_from_emp)
            else:
                emp_df["Days To Expiry"] = ""

        # Styling bands (today-based):
        #   expired (<0) -> dark red
        #   <=30 -> red
        #   31-60 -> yellow
        #   >60 -> normal
        def _style_exp_cell(emp_val, disp_val):
            if not disp_val:
                return ""
            top_date = exp_top_date_map.get(_norm_emp(emp_val))
            if not top_date:
                return ""
            diff = (top_date - today).days
            if diff < 0:
                return "background-color:#8B0000;color:white;font-weight:700;"
            if diff <= 30:
                return "background-color:red;color:white;font-weight:700;"
            if diff <= 60:
                return "background-color:yellow;color:black;font-weight:700;"
            return ""

        show_df = emp_df.copy()
        sty = show_df.style.apply(
            lambda row: [""] * (len(row) - 1) + [_style_exp_cell(row.get("Employer", ""), row.get("Expiry Date", ""))],
            axis=1,
        )
        st.dataframe(sty, use_container_width=True, hide_index=True)

        # ---- Expiry Detail List (Step 5) + Download ----
        # Defaults to avoid UnboundLocalError when expiry list is empty / not built
        win = "All"
        pick_ins = "All"
        pick_emp = "All"
        df_f = pd.DataFrame()

        with st.expander("Expiry Detail List (Step 5) — filter & download", expanded=False):
            df_detail = None
            # Prefer explicit saved expiry list/tracker from dfs
            for _k in ("Expiry_List", "Expiry List", "Employer Expiry Tracker", "Expiry_Tracker", "Expiry"):
                if _k in dfs and isinstance(dfs.get(_k), pd.DataFrame) and not dfs[_k].empty:
                    df_detail = dfs[_k].copy()
                    break
            if df_detail is None:
                df_detail = pd.DataFrame()

            # Fallback: pick any dataframe that has Expiry + Employer columns (in case key name differs)
            if df_detail is None or df_detail.empty:
                for _k, _v in (dfs or {}).items():
                    if not isinstance(_v, pd.DataFrame) or _v.empty:
                        continue
                    _cols = [str(c).strip().lower() for c in _v.columns]
                    if any("expiry" in c for c in _cols) and any("employer" in c for c in _cols):
                        df_detail = _v.copy()
                        break

        
            if df_detail is None or df_detail.empty:
                st.info("No expiry detail list found for this day/period.")
            else:
                # Normalize column names
                if "Insuance" in df_detail.columns and "Insurance" not in df_detail.columns:
                    df_detail = df_detail.rename(columns={"Insuance": "Insurance"})

                # --- Fix Expiry Date parsing & Days To Expiry (robust) ---
                if "Expiry Date" in df_detail.columns:
                    _exp_raw = df_detail["Expiry Date"]
                    _exp_dt = pd.to_datetime(_exp_raw, errors="coerce")
                    # fallback for dd/mm/yyyy style
                    if _exp_dt.notna().sum() == 0:
                        _exp_dt = pd.to_datetime(_exp_raw, errors="coerce", dayfirst=True)
                    df_detail["Expiry Date"] = _exp_dt

                    # Ensure Days To Expiry exists (or recompute if blank)
                    if "Days To Expiry" not in df_detail.columns or pd.to_numeric(df_detail.get("Days To Expiry"), errors="coerce").isna().all():
                        _today = pd.Timestamp.today().normalize()
                        df_detail["Days To Expiry"] = (df_detail["Expiry Date"] - _today).dt.days
                # Expected columns
                # Employer, Insurance, Name, EMR No, Visit ID, Doctor, Expiry Date, Days To Expiry
                # Filters
                f1, f2, f3 = st.columns([2, 2, 2])
                with f1:
                    win = st.selectbox(
                        "Expiry Window",
                        options=["All", "Expired", "Next 30 days", "Next 60 days", "Next 90 days"],
                        key=f"exp_win2_{str(day_ts)}",
                    )
                with f2:
                    ins_opts = []
                    if "Insurance" in df_detail.columns:
                        ins_opts = sorted([x for x in df_detail["Insurance"].dropna().unique() if str(x).strip() not in ["", "nan", "None"]])
                    pick_ins = st.selectbox("Insurance", options=["All"] + ins_opts, key=f"exp_ins2_{str(day_ts)}")
                with f3:
                    emp_opts = []
                    if "Employer" in df_detail.columns:
                        emp_opts = sorted([x for x in df_detail["Employer"].dropna().unique() if str(x).strip() not in ["", "nan", "None"]])
                    pick_emp = st.selectbox("Employer", options=["All"] + emp_opts, key=f"exp_emp2_{str(day_ts)}")
        
                df_f = df_detail.copy()
                if "Days To Expiry" in df_f.columns:
                    df_f["Days To Expiry"] = pd.to_numeric(df_f["Days To Expiry"], errors="coerce")
                    if win == "Expired":
                        df_f = df_f[df_f["Days To Expiry"] < 0]
                    elif win.startswith("Next"):
                        n = int(re.findall(r"\d+", win)[0])
                        df_f = df_f[(df_f["Days To Expiry"] >= 0) & (df_f["Days To Expiry"] <= n)]
        
                if pick_ins != "All" and "Insurance" in df_f.columns:
                    df_f = df_f[df_f["Insurance"] == pick_ins]
                if pick_emp != "All" and "Employer" in df_f.columns:
                    df_f = df_f[df_f["Employer"] == pick_emp]
        
                
                # ---- Summary counts (on-screen) ----
                grp_cols = [c for c in ["Employer", "Insurance"] if c in df_f.columns]
                if grp_cols:
                    df_counts = (
                        df_f.groupby(grp_cols, dropna=False)
                        .size()
                        .reset_index(name="Count")
                        .sort_values("Count", ascending=False)
                    )
                    # TOTAL row
                    total_n = int(df_counts["Count"].sum()) if "Count" in df_counts.columns else 0
                    total_row = {c: "" for c in df_counts.columns}
                    if "Employer" in total_row:
                        total_row["Employer"] = "TOTAL"
                    total_row["Count"] = total_n
                    df_counts = pd.concat([df_counts, pd.DataFrame([total_row])], ignore_index=True)
            
                    st.caption(f"Showing summary counts for: **{win}** | Rows: {len(df_counts)-1} | TOTAL: {total_n}")
                    st.dataframe(df_counts, use_container_width=True, hide_index=True)
                else:
                    st.info("Expiry list is missing Employer/Insurance columns, so summary counts cannot be built.")
                    df_counts = pd.DataFrame()
        
                # ---- Optional detailed list (only when needed) ----
                exp_key = f"{win}_{pick_ins}_{pick_emp}"
                show_details = st.checkbox(
                    "Show detailed patient list (only if you need to review before download)",
                    value=False,
                    key=f"exp_show_details_{exp_key}",
                )
                if show_details:
                    show_cols = [c for c in ["Employer","Insurance","Name","EMR No","Visit ID","Doctor","Expiry Date","Days To Expiry"] if c in df_f.columns]
                    st.dataframe(df_f[show_cols] if show_cols else df_f, use_container_width=True, hide_index=True)
        
                # ---- Downloads ----
                # Download counts
                try:
                    import io as _io
                    out_counts = _io.BytesIO()
                    with pd.ExcelWriter(out_counts, engine="openpyxl") as writer:
                        (df_counts if isinstance(df_counts, pd.DataFrame) else pd.DataFrame()).to_excel(writer, index=False, sheet_name="Expiry_Counts")
                    st.download_button(
                        "Download Counts (Excel)",
                        data=out_counts.getvalue(),
                        file_name="expiry_counts.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_exp_counts_{exp_key}",
                    )
                except Exception:
                    st.warning("Counts download is unavailable (Excel writer error).")
        
                # Download full list
                try:
                    import io as _io
                    out = _io.BytesIO()
                    with pd.ExcelWriter(out, engine="openpyxl") as writer:
                        (df_f[show_cols] if (show_details and show_cols) else df_f).to_excel(writer, index=False, sheet_name="Expiry_List")
                    st.download_button(
                        "Download Full List (Excel)",
                        data=out.getvalue(),
                        file_name="expiry_list.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_exp_full_{exp_key}",
                    )
                except Exception:
                    st.warning("Download is unavailable (Excel writer not found).")
# ---------------------------
# Top Header + Center selection (LOCKED if passed in URL)
# ---------------------------
CENTERS = {
    "easyhealth": "Easy Health Medical Clinic (MF8031)",
    "excellent": "Excellent Medical Center (MF4777)",
    "pharmacy": "Excellent Pharmacy (PF3205)",
}

# Streamlit new API: st.query_params is dict-like
qp_center = (st.query_params.get("center") or "").strip()
_locked_center = qp_center if qp_center in CENTERS else None

# Compact premium header row (title + center)
h1, h2 = st.columns([7, 3], vertical_alignment="center")
with h1:
    st.markdown("<div class='page-title'>📅 Registration Summary — Management View</div>", unsafe_allow_html=True)
with h2:
    st.caption("Center")
    if _locked_center:
        center_key = st.selectbox("Center", [_locked_center], format_func=lambda k: CENTERS[k], disabled=True, key="center_locked")
    else:
        center_opts = list(CENTERS.keys())
        default_center = "excellent" if "excellent" in center_opts else center_opts[0]
        center_key = st.selectbox(
            "Center",
            options=center_opts,
            index=center_opts.index(default_center),
            format_func=lambda k: CENTERS[k],
            key="center_pick",
        )

st.caption(f"Center: **{CENTERS.get(center_key, center_key)}**")

# ---------------------------
# S3 status
# ---------------------------
cfg = load_secrets()
s3_ok = s3_enabled(cfg)
s3 = s3_client_cached(cfg) if s3_ok else None

with st.expander("Storage Status (S3)", expanded=False):
    if s3_ok:
        st.success(f"S3 is configured ✅  Bucket: {cfg['S3_BUCKET_NAME']}  Region: {cfg['AWS_REGION']}")
        prefs = candidate_base_prefixes(cfg)
        st.caption('Viewer will look for: ' + '  |  '.join([((p + '/') if p else '') + 'registration/<center>/history.csv' for p in prefs]))
    else:
        st.error("S3 is NOT configured on this app, so View page cannot load saved results.")
        st.caption("Expected secrets: S3_BUCKET_NAME (or S3_BUCKET), AWS_REGION (or AWS_DEFAULT_REGION), AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY")

if not s3_ok:
    st.stop()

# ---------------------------
# Load history and auto-show latest result
# ---------------------------
hist, root_prefix = load_history_from_s3(s3, cfg, center_key)

if hist.empty or "day" not in hist.columns:
    hist_key = s3_key(root_prefix, 'history.csv')
    st.warning("No saved Daily Report found for this center yet.")
    st.write("✅ To fix:")
    st.markdown(
        "- Open **Registration Summary (Upload)** page\n"
        "- Upload Step 1/2/3\n"
        "- Click **Process & Save to S3**\n"
        "- Then come back here"
    )
    st.caption(f"Expected S3 key: {hist_key}")
    st.stop()

# normalize day
hist["day"] = pd.to_datetime(hist["day"], errors="coerce").dt.normalize()
hist = hist.dropna(subset=["day"]).sort_values("day")

days = list(hist["day"].unique())
latest_day = days[-1]


# ---------------------------
# View mode: Daily / Weekly / Monthly
# ---------------------------
st.markdown("<hr style='margin: 0.9rem 0 0.9rem 0; border: none; border-top: 1px solid rgba(16,24,40,0.10);'/>", unsafe_allow_html=True)

m1, m2 = st.columns([1.6, 4.4], vertical_alignment="center")
with m1:
    st.markdown("**View Mode**")
with m2:
    mode = st.radio(
        "View Mode",
        options=["Daily", "Weekly", "Monthly"],
        horizontal=True,
        index=0,
        label_visibility="collapsed",
    )

SS = st.session_state

SS = st.session_state
SS.setdefault("loaded_key", None)      # cache key (string) for loaded period
SS.setdefault("loaded_summary", None)  # dict of dfs
SS.setdefault("loaded_label", None)    # title label

def days_in_week(any_day: pd.Timestamp) -> List[pd.Timestamp]:
    d = pd.to_datetime(any_day).normalize()
    start = d - pd.Timedelta(days=int(d.weekday()))  # Monday
    end = start + pd.Timedelta(days=6)
    return [x for x in days if (x >= start) and (x <= end)]

def days_in_month(any_day: pd.Timestamp) -> List[pd.Timestamp]:
    d = pd.to_datetime(any_day).normalize()
    start = d.replace(day=1)
    # next month
    if start.month == 12:
        nxt = start.replace(year=start.year+1, month=1, day=1)
    else:
        nxt = start.replace(month=start.month+1, day=1)
    end = nxt - pd.Timedelta(days=1)
    return [x for x in days if (x >= start) and (x <= end)]

def aggregate_tables(frames: List[pd.DataFrame]) -> pd.DataFrame:
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)

    # Keep only real rows (drop total/grand total rows; we'll rebuild totals)
    def _is_total_row(x):
        s = str(x).strip().upper()
        return s in ["TOTAL", "GRAND TOTAL"]
    first_col = df.columns[0] if len(df.columns) else None
    if first_col:
        df = df[~df[first_col].astype(str).map(_is_total_row)].copy()

    # Group by non-numeric columns, sum numeric
    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    grp_cols = [c for c in df.columns if c not in num_cols]
    if num_cols and grp_cols:
        out = df.groupby(grp_cols, dropna=False, as_index=False)[num_cols].sum()
    elif "Count" in df.columns:
        grp_cols = [c for c in df.columns if c != "Count"]
        out = df.groupby(grp_cols, dropna=False, as_index=False)["Count"].sum()
    else:
        out = df

    # Re-add TOTAL / GRAND TOTAL
    if "Count" in out.columns and first_col:
        total = int(out["Count"].sum()) if not out.empty else 0
        out.loc[len(out)] = {first_col: "TOTAL", "Count": total}
    else:
        # If there is any numeric column, add GRAND TOTAL
        if num_cols and first_col:
            row = {c: "" for c in out.columns}
            row[first_col] = "GRAND TOTAL"
            for c in num_cols:
                row[c] = float(out[c].sum()) if not out.empty else 0.0
            out.loc[len(out)] = row

    return out


def aggregate_income(frames: List[pd.DataFrame]) -> pd.DataFrame:
    """Aggregate Income tables across many days.

    Rule:
    - Sum Consultation/Lab/Procedure/Total_Visit/Total_Amount_* across days (grouped by non-numeric columns)
    - Recompute Avg_Amount_* = Total_Amount_* / Total_Visit
    - Recompute Lab_% = (Lab / Total_Amount_Service) * 100
    - Rebuild GRAND TOTAL row
    """
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)

    # Remove any TOTAL rows; we rebuild totals after aggregation
    first_col = df.columns[0] if len(df.columns) else None
    if first_col:
        df = df[~df[first_col].astype(str).str.strip().str.upper().isin(["TOTAL", "GRAND TOTAL"])].copy()

    # Identify columns
    avg_cols = [c for c in df.columns if str(c).strip().lower().startswith("avg_") or str(c).strip().lower().startswith("avg ")]
    lab_pct_cols = [c for c in df.columns if str(c).strip().lower() in ["lab_%", "lab%", "lab pct", "lab_pct"]]
    ignore_sum = set(avg_cols + lab_pct_cols)

    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    sum_cols = [c for c in num_cols if c not in ignore_sum]

    grp_cols = [c for c in df.columns if c not in num_cols]
    if grp_cols and sum_cols:
        out = df.groupby(grp_cols, dropna=False, as_index=False)[sum_cols].sum()
    else:
        out = df.copy()

    # Normalize expected column names
    # (support both Total_Amount_Insurance and Total_Amount_Insuance)
    if "Total_Amount_Insurance" in out.columns and "Total_Amount_Insuance" not in out.columns:
        out["Total_Amount_Insuance"] = out["Total_Amount_Insurance"]
    if "Avg_Amount_Insurance" in out.columns and "Avg_Amount_Insuance" not in out.columns:
        out["Avg_Amount_Insuance"] = out["Avg_Amount_Insurance"]

    # Recompute averages (strictly by Total_Visit)
    if "Total_Visit" in out.columns:
        denom = out["Total_Visit"].replace(0, pd.NA)
        if "Total_Amount_Service" in out.columns:
            out["Avg_Amount_Service"] = out["Total_Amount_Service"] / denom
        if "Total_Amount_Insuance" in out.columns:
            out["Avg_Amount_Insuance"] = out["Total_Amount_Insuance"] / denom

    # Recompute Lab_% (service basis)
    if "Lab" in out.columns and "Total_Amount_Service" in out.columns:
        denom2 = out["Total_Amount_Service"].replace(0, pd.NA)
        out["Lab_%"] = (out["Lab"] / denom2) * 100

    # Rebuild GRAND TOTAL
    if first_col and any(c in out.columns for c in sum_cols):
        row = {c: "" for c in out.columns}
        row[first_col] = "GRAND TOTAL"
        for c in sum_cols:
            row[c] = float(out[c].sum()) if not out.empty else 0.0
        # Averages for grand total
        if "Total_Visit" in out.columns and row.get("Total_Visit", 0):
            tv = row["Total_Visit"] if row["Total_Visit"] else 0
            try:
                tv = float(tv)
            except Exception:
                tv = 0
            if tv:
                if "Total_Amount_Service" in out.columns:
                    row["Avg_Amount_Service"] = float(row.get("Total_Amount_Service", 0)) / tv
                if "Total_Amount_Insuance" in out.columns:
                    row["Avg_Amount_Insuance"] = float(row.get("Total_Amount_Insuance", 0)) / tv
                if "Lab" in out.columns and "Total_Amount_Service" in out.columns and float(row.get("Total_Amount_Service", 0)) != 0:
                    row["Lab_%"] = (float(row.get("Lab", 0)) / float(row.get("Total_Amount_Service", 0))) * 100
        out.loc[len(out)] = row

    return out

def load_and_aggregate(day_list: List[pd.Timestamp]) -> Optional[Dict[str, pd.DataFrame]]:
    if not day_list:
        return None
    loaded = []
    for d in day_list:
        dfs = load_summary_from_s3(s3, cfg, root_prefix, d)
        if dfs is not None:
            loaded.append(dfs)

    if not loaded:
        return None

    keys = sorted(set().union(*[set(x.keys()) for x in loaded]))
    agg: Dict[str, pd.DataFrame] = {}

    # KPI: sum across days (note: unique patients across a period is approximate because we only have daily aggregates)
    kpi_rows = []
    for d in loaded:
        k = d.get("KPI")
        if k is not None and not k.empty and "Metric" in k.columns and "Value" in k.columns:
            kpi_rows.append(k)
    if kpi_rows:
        kk = pd.concat(kpi_rows, ignore_index=True)
        kk["Value"] = pd.to_numeric(kk["Value"], errors="coerce").fillna(0)
        kpi_sum = kk.groupby("Metric", as_index=False)["Value"].sum()
        agg["KPI"] = kpi_sum

    for k in keys:
        if k == "KPI":
            continue
        frames = [d.get(k) for d in loaded if isinstance(d.get(k), pd.DataFrame)]
        if str(k).startswith("Income | "):
            agg[k] = aggregate_income(frames)
        else:
            agg[k] = aggregate_tables(frames)

    return agg


def _snap_to_saved(chosen: pd.Timestamp, saved: List[pd.Timestamp]) -> Tuple[pd.Timestamp, bool]:
    """Return (snapped_day, was_snapped). Picks the nearest saved day <= chosen, else the earliest."""
    if not saved:
        return chosen, False
    chosen = pd.to_datetime(chosen).normalize()
    saved_sorted = sorted(pd.to_datetime(saved).tolist())
    if chosen in saved_sorted:
        return chosen, False
    earlier = [d for d in saved_sorted if d <= chosen]
    if earlier:
        return earlier[-1], True
    return saved_sorted[0], True


min_day = pd.to_datetime(min(days)).normalize()
max_day = pd.to_datetime(max(days)).normalize()

if mode == "Daily":
    chosen = st.date_input(
        "Select day",
        value=max_day.date(),
        min_value=min_day.date(),
        max_value=max_day.date(),
    )
    picked, snapped = _snap_to_saved(pd.to_datetime(chosen), days)
    if snapped:
        st.info(f"No saved data for **{pd.to_datetime(chosen).strftime('%d %b %Y')}**. Showing nearest saved day: **{fmt_day(picked)}**")

    cache_key = f"daily:{picked.date().isoformat()}"
    if SS.get("loaded_key") != cache_key:
        SS["loaded_summary"] = load_summary_from_s3(s3, cfg, root_prefix, picked)
        SS["loaded_key"] = cache_key
        SS["loaded_label"] = f"Current Day ({fmt_day(picked)})"

    if SS.get("loaded_summary") is not None:
                render_summary(SS["loaded_summary"], picked, heading="header", label="Current Day")
    else:
        st.error("summary.pkl is missing for this day.")
        st.caption(f"Expected: {s3_key(root_prefix, picked.date().isoformat(), 'summary.pkl')}")

elif mode == "Weekly":
    c1, c2 = st.columns(2)
    with c1:
        s_in = st.date_input("Week Start", value=max_day.date(), min_value=min_day.date(), max_value=max_day.date(), key="wk_start")
    with c2:
        e_in = st.date_input("Week End", value=max_day.date(), min_value=min_day.date(), max_value=max_day.date(), key="wk_end")

    start_d, s_snap = _snap_to_saved(pd.to_datetime(s_in), days)
    end_d, e_snap = _snap_to_saved(pd.to_datetime(e_in), days)
    if start_d > end_d:
        start_d, end_d = end_d, start_d

    selected = [d for d in days if (d >= start_d) and (d <= end_d)]
    st.caption(f"Selected range: **{fmt_range(start_d, end_d)}**  (saved days: {len(selected)})")

    if not selected:
        st.warning("No saved days found in this range.")
    else:
        cache_key = f"range:{start_d.date().isoformat()}:{end_d.date().isoformat()}"
        if SS.get("loaded_key") != cache_key:
            SS["loaded_summary"] = load_and_aggregate(selected)
            SS["loaded_key"] = cache_key
            SS["loaded_label"] = f"Weekly Summary ({fmt_range(start_d, end_d)})"

        if SS.get("loaded_summary") is not None:
            st.header(SS.get("loaded_label", "Weekly Summary"))
            render_summary(SS["loaded_summary"], pd.to_datetime(max(selected)), heading="subheader", label="Latest Saved Day", picked_label_override=SS.get("loaded_label"))
        else:
            st.warning("No summary.pkl files found in this range.")

else:  # Monthly
    chosen = st.date_input(
        "Select any date in the month",
        value=max_day.date(),
        min_value=min_day.date(),
        max_value=max_day.date(),
        key="mo_pick",
    )
    d0 = pd.to_datetime(chosen).normalize()
    month_days = days_in_month(d0)

    if not month_days:
        st.warning("No saved days found for that month.")
    else:
        sel_month = d0.strftime("%Y-%m")
        start_m = min(month_days).date().isoformat()
        end_m = max(month_days).date().isoformat()
        st.caption(f"Month range: **{fmt_range(min(month_days), max(month_days))}**  (saved days: {len(month_days)})")

        cache_key = f"month:{sel_month}"
        if SS.get("loaded_key") != cache_key:
            SS["loaded_summary"] = load_and_aggregate(month_days)
            SS["loaded_key"] = cache_key
            SS["loaded_label"] = f"Monthly Summary ({pd.to_datetime(d0).strftime('%B %Y')})"

        if SS.get("loaded_summary") is not None:
            st.header(SS.get("loaded_label", "Monthly Summary"))
            render_summary(SS["loaded_summary"], pd.to_datetime(max(month_days)), heading="subheader", label="Latest Saved Day", picked_label_override=SS.get("loaded_label"))
        else:
            st.warning("No summary.pkl files found for that month.")
