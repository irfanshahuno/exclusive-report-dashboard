#!/usr/bin/env python3
# pages/5_Summary_Report.py
# Summary Report page — mirrors center selection flow.
# User selects a center → uploads a source file → runs exclusive_report_status_final.py
# logic in-memory → displays results (Insurance_Totals, Final_Bucket_Summary, Monthly_Totals,
# Balance_Aging_Summary, Balance_Status_Stage_Summary) + download button for full Excel.

import io
import sys
import hashlib
import re
import hmac
import base64
import json
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import boto3
from botocore.exceptions import BotoCoreError, ClientError

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Summary Report — Excellent Medical Group", layout="wide")
st.set_option("client.showErrorDetails", False)

# ─────────────────────────────────────────────────────────────────────────────
# AUTH — reuse same view password from main dashboard
# ─────────────────────────────────────────────────────────────────────────────
VIEW_PASSWORD = st.secrets.get("VIEW_PASSWORD", "Emc@2026")
TOKEN_SECRET  = st.secrets.get("TOKEN_SECRET", None)
TOKEN_TTL_SECONDS = int(st.secrets.get("TOKEN_TTL_SECONDS", 600))


def _b64url_decode(s: str) -> bytes:
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)


def verify_url_token(token: str) -> dict | None:
    if not TOKEN_SECRET:
        return None
    try:
        body_b64, sig_b64 = token.split(".", 1)
        body = _b64url_decode(body_b64)
        sig  = _b64url_decode(sig_b64)
        expected = hmac.new(TOKEN_SECRET.encode("utf-8"), body, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, expected):
            return None
        data = json.loads(body.decode("utf-8"))
        if int(time.time()) - int(data.get("iat", 0)) > TOKEN_TTL_SECONDS:
            return None
        return data
    except Exception:
        return None


def _auto_auth():
    tok = st.query_params.get("token")
    if tok:
        data = verify_url_token(tok)
        if data:
            st.session_state.is_view_auth = True
    auth_param = st.query_params.get("auth")
    if auth_param:
        _secret = st.secrets.get("TOKEN_SECRET", VIEW_PASSWORD)
        expected = hmac.new(_secret.encode("utf-8"), b"view_auth", hashlib.sha256).hexdigest()[:16]
        if auth_param == expected:
            st.session_state.is_view_auth = True

    # Auto-set year from URL param
    _year_param = st.query_params.get("year")
    if _year_param:
        try:
            _yr = int(_year_param)
            if _yr in (2024, 2025, 2026):
                st.session_state["rcm_year"] = _yr
        except Exception:
            pass

    # Auto-set center from URL param (skip selection screen)
    _center_param = st.query_params.get("center")
    if _center_param and _center_param in ("excellent", "easyhealth", "pharmacy"):
        if st.session_state.get("sum_center_key") != _center_param:
            st.session_state["sum_center_key"] = _center_param


_auto_auth()


def require_view_access():
    if st.session_state.get("is_view_auth", False):
        return
    st.title("🔒 Dashboard Access")
    st.info("Enter the view password to open the dashboard.")
    pwd = st.text_input("View Password", type="password", key="sum_view_pwd")
    if st.button("Enter Dashboard", use_container_width=True):
        if pwd == VIEW_PASSWORD:
            st.session_state.is_view_auth = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()


require_view_access()

# ─────────────────────────────────────────────────────────────────────────────
# PREMIUM CSS (same as main dashboard)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
.stApp{
  background: linear-gradient(145deg, #EDF2FB 0%, #F8FAFF 40%, #FAFCFF 100%) !important;
  font-family: 'Inter', sans-serif !important;
}
hr{ border:none!important;height:1px!important;background:linear-gradient(90deg,transparent,#C8D9F0,transparent)!important; }
div.stButton > button{
  width:100%!important;min-height:50px!important;padding:12px 20px!important;
  font-size:15px!important;font-weight:700!important;font-family:'Inter',sans-serif!important;
  background:linear-gradient(160deg,#FFFFFF 0%,#EEF4FF 100%)!important;
  color:#0A2647!important;border:1.5px solid #C5D8F5!important;border-radius:14px!important;
  box-shadow:0 2px 8px rgba(10,38,71,0.08),inset 0 1px 0 rgba(255,255,255,0.9)!important;
  transition:all 0.2s ease!important;
}
div.stButton > button:hover{
  background:linear-gradient(160deg,#E8F1FF 0%,#D6E8FF 100%)!important;
  border-color:#7DAAEE!important;box-shadow:0 6px 20px rgba(10,38,71,0.15)!important;
  transform:translateY(-1px)!important;
}
div.stButton > button:active{
  background:linear-gradient(160deg,#0A2647 0%,#154B8A 100%)!important;
  color:#fff!important;border-color:#0A2647!important;
}
.center-title{color:#0A2647!important;font-weight:900!important;font-family:'Inter',sans-serif!important;letter-spacing:-0.5px!important;margin-bottom:0!important;}
.kpi-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:14px;margin-top:10px;margin-bottom:10px;}
.kpi-card{background:rgba(255,255,255,0.85);backdrop-filter:blur(12px);-webkit-backdrop-filter:blur(12px);border:1.5px solid rgba(197,216,245,0.7);border-radius:18px;padding:16px 18px;box-shadow:0 4px 16px rgba(10,38,71,0.07),0 1px 3px rgba(10,38,71,0.05),inset 0 1px 0 rgba(255,255,255,0.95);min-width:0;transition:all 0.2s ease;}
.kpi-label{font-size:12px;color:#8A9BB5;font-weight:600;font-family:'Inter',sans-serif;letter-spacing:0.6px;text-transform:uppercase;margin-bottom:8px;}
.kpi-value{font-size:clamp(17px,2.1vw,28px);font-weight:800;color:#0D1B2E;letter-spacing:-0.5px;font-family:'Inter',sans-serif;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.kpi-card.balance{background:linear-gradient(145deg,rgba(10,38,71,0.96) 0%,rgba(15,56,110,0.96) 100%);border-color:rgba(180,210,255,0.25);}
.kpi-card.balance .kpi-label{color:rgba(180,205,255,0.75);}
.kpi-card.balance .kpi-value{color:#FFFFFF;}
@media(max-width:1100px){.kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr));}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CENTERS CONFIG (same as main dashboard)
# ─────────────────────────────────────────────────────────────────────────────
CENTERS = {
    "excellent":  {"key": "excellent",  "name": "Excellent Medical Center (MF4777)"},
    "pharmacy":   {"key": "pharmacy",   "name": "Excellent Pharmacy (PF3205)"},
    "easyhealth": {"key": "easyhealth", "name": "Easy Health Medical Clinic (MF8031)"},
}

# ─────────────────────────────────────────────────────────────────────────────
# S3 CONFIG — summary files ONLY (kept separate from main dashboard prefix)
# ─────────────────────────────────────────────────────────────────────────────
S3_BUCKET = st.secrets.get("S3_BUCKET", "emc-rcm-storage-2026")
SUMMARY_S3_PREFIX = st.secrets.get("SUMMARY_S3_PREFIX", "streamlit2")
AWS_REGION = st.secrets.get("AWS_REGION", "eu-north-1")


def get_s3_client():
    try:
        return boto3.client(
            "s3",
            aws_access_key_id=st.secrets.get("AWS_ACCESS_KEY_ID"),
            aws_secret_access_key=st.secrets.get("AWS_SECRET_ACCESS_KEY"),
            region_name=AWS_REGION,
        )
    except Exception:
        return None


def build_summary_s3_key(center_key: str, filename: str) -> str:
    """Fixed key per center — always replaces previous file."""
    year = st.session_state.get("rcm_year") or datetime.now().year
    ext = str(filename).rsplit(".", 1)[-1] if "." in str(filename) else "xlsx"
    return f"{SUMMARY_S3_PREFIX}/{year}/{center_key}/source_latest.{ext}"

def build_report_s3_key(center_key: str) -> str:
    """Fixed key for the processed report Excel — always replaces previous."""
    year = st.session_state.get("rcm_year") or datetime.now().year
    return f"{SUMMARY_S3_PREFIX}/{year}/{center_key}/report_latest.xlsx"


def upload_bytes_to_s3(file_bytes: bytes, key: str, content_type: str | None = None):
    s3 = get_s3_client()
    if s3 is None:
        return False, "S3 client not available. Check AWS secrets."
    extra = {}
    if content_type:
        extra["ContentType"] = content_type
    try:
        s3.put_object(Bucket=S3_BUCKET, Key=key, Body=file_bytes, **extra)
        return True, None
    except (BotoCoreError, ClientError, Exception) as e:
        return False, str(e)

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
h1, h2 = st.columns([8, 2])
with h1:
    st.title("📋 Summary Report")
    st.caption("Upload a source file per center — results are generated instantly in-memory.")
with h2:
    if st.button("🏠 Back to Dashboard", use_container_width=True, key="sum_back"):
        st.switch_page("exclusive_dashboard.py")

st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# SESSION HELPERS
# ─────────────────────────────────────────────────────────────────────────────
SUM_CENTER_KEY = "sum_center_key"

def reset_sum_center():
    st.session_state[SUM_CENTER_KEY] = None
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# KPI RENDERER (same visual as main dashboard, no links needed here)
# ─────────────────────────────────────────────────────────────────────────────
def render_kpi_cards(net, paid, bal, rej, acc):
    def fmt(x):
        try: return f"{float(x):,.2f}"
        except: return "—"
    html = f"""
    <div class="kpi-grid">
      <div class="kpi-card"><div class="kpi-label">Claimed Amount</div><div class="kpi-value">{fmt(net)}</div></div>
      <div class="kpi-card"><div class="kpi-label">Total Pay</div><div class="kpi-value">{fmt(paid)}</div></div>
      <div class="kpi-card balance"><div class="kpi-label">Under Process</div><div class="kpi-value">{fmt(bal)}</div></div>
      <div class="kpi-card"><div class="kpi-label">Final Rejn</div><div class="kpi-value">{fmt(rej)}</div></div>
      <div class="kpi-card"><div class="kpi-label">Rej. Accepted</div><div class="kpi-value">{fmt(acc)}</div></div>
    </div>"""
    st.markdown(html, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# INLINE ENGINE — all logic from exclusive_report_status_final.py
# ─────────────────────────────────────────────────────────────────────────────
MAX_RESUB_STAGE = 10


def _normalize_status(value) -> str:
    s = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", s)


def _extract_stage_info(status_value: str):
    s = _normalize_status(status_value)
    m = re.match(r"^(submitted|not submitted|approved)\s*(?:\(\s*resub\s*-\s*(\d+)\s*\))?$", s)
    if not m:
        return "Other", "Other", None
    base = m.group(1)
    stage_num = int(m.group(2)) if m.group(2) is not None else 0
    stage_label = "Initial" if stage_num == 0 else f"Resub-{stage_num}"
    base_label = {"submitted": "Submitted", "not submitted": "Not Submitted", "approved": "Approved"}[base]
    return base_label, stage_label, stage_num


def _classify_final_bucket(status_value: str) -> str:
    s = _normalize_status(status_value)
    if re.match(r"^rejected\s*(?:\(\s*resub\s*-\s*\d+\s*\))?$", s):
        return "Rejected"
    if re.match(r"^rejection accepted\s*(?:\(\s*resub\s*-\s*\d+\s*\))?$", s):
        return "Accepted"
    return "Balance"


def _choose_stage_date(row, df_columns):
    status_group = row.get("Balance Status Group", "")
    stage_no = row.get("Balance Submission No")
    if status_group == "Not Submitted":
        return pd.NaT
    candidates = []
    if stage_no is None or pd.isna(stage_no):
        candidates = []
    elif int(stage_no) == 0:
        candidates = ["SubDate"]
    else:
        candidates = [f"Resub{int(stage_no)}Date"]
    candidates += ["SubDate", "SubmissionDate", "ClaimDate", "VisitDate"]
    for c in candidates:
        if c in df_columns and pd.notna(row.get(c)):
            return row[c]
    return pd.NaT


def _add_grand_total_row(df: pd.DataFrame, key_col: str) -> pd.DataFrame:
    if df.empty:
        return df
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    total = {key_col: "Grand Total"}
    for c in df.columns:
        if c == key_col:
            continue
        total[c] = df[c].sum() if c in numeric_cols else ""
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)



def build_rcm_summary(df: pd.DataFrame, group_col: str, visit_days_series: pd.Series = None) -> pd.DataFrame:
    """
    Build the RCM Summary table grouped by any column (Insurance, DocName, etc.)
    Uses same logic as Engine 1 — SubInsShare based, no Difference column.

    Status rules:
      Submitted (plain)          → Sub Nt Rmtd
      Submitted (Resub-N)        → Rsub Nt Rmtd
      Not Submitted plain <90d   → Sub Nt Rmtd
      Not Submitted plain >90d   → Total Pay (extra_paid)
      Not Submitted (Resub-N)    → Rsub Nt Rmtd
      Rejection Accepted (any)   → Rejection Accepted
      Rejected (any)             → Final Rejn
      Approved (any)             → Total Pay (remit cols already capture)
      Everything else            → Sub Nt Rmtd (catch-all)
    """
    import re as _re

    d = df.copy()

    for c in ["SubInsShare","RemitInsShare","Resub1RemitInsShare",
              "Resub2RemitInsShare","Resub3RemitInsShare","UniqueID","Status"]:
        if c not in d.columns:
            d[c] = 0 if c not in ("UniqueID","Status") else ("" if c == "Status" else 0)

    for c in ["SubInsShare","RemitInsShare","Resub1RemitInsShare",
              "Resub2RemitInsShare","Resub3RemitInsShare"]:
        d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0.0)

    def _norm(v):
        return _re.sub(r"\s+", " ", str(v or "").strip().lower())

    d["_sn"] = d["Status"].apply(_norm)

    # Days series for Not Submitted >90 rule
    if visit_days_series is not None:
        _days = visit_days_series.reindex(d.index).fillna(0)
    else:
        _days = pd.Series(0, index=d.index)

    # ── Status masks ──────────────────────────────────────────────────────────
    mask_sub_init      = d["_sn"] == "submitted"
    mask_sub_resub     = d["_sn"].str.match(r"^submitted\s*\(\s*resub\s*-\s*\d+\s*\)$", na=False)
    mask_not_sub_plain = d["_sn"] == "not submitted"
    mask_not_sub_old   = mask_not_sub_plain & (_days > 90)
    mask_not_sub_fresh = mask_not_sub_plain & (_days <= 90)
    mask_not_sub_resub = d["_sn"].str.match(r"^not submitted\s*\(\s*resub\s*-\s*\d+\s*\)$", na=False)
    mask_rej_acc       = d["_sn"].str.match(r"^rejection accepted(\s*\(\s*resub\s*-\s*\d+\s*\))?$", na=False)
    mask_rejected      = d["_sn"].str.match(r"^rejected(\s*\(\s*resub\s*-\s*\d+\s*\))?$", na=False)
    mask_approved      = d["_sn"].str.match(r"^approved(\s*\(.+\))?$", na=False)

    # Catch-all unmatched
    mask_any_matched = (mask_sub_init | mask_sub_resub | mask_not_sub_plain |
                        mask_not_sub_resub | mask_rej_acc | mask_rejected | mask_approved)
    mask_unmatched   = (~mask_any_matched) & (d["SubInsShare"] > 0)

    # ── Per-row column assignments (all use SubInsShare) ──────────────────────
    # Sub Nt Rmtd: Submitted plain + Not Submitted fresh + unmatched
    d["_sub_nt_rmtd"] = 0.0
    d.loc[mask_sub_init,      "_sub_nt_rmtd"] = d.loc[mask_sub_init,      "SubInsShare"]
    d.loc[mask_not_sub_fresh, "_sub_nt_rmtd"] = d.loc[mask_not_sub_fresh, "SubInsShare"]
    d.loc[mask_unmatched,     "_sub_nt_rmtd"] = d.loc[mask_unmatched,     "SubInsShare"]

    # Rsub Nt Rmtd: Submitted Resub-N + Not Submitted Resub-N
    d["_rsub_nt_rmtd"] = 0.0
    d.loc[mask_sub_resub,     "_rsub_nt_rmtd"] = d.loc[mask_sub_resub,     "SubInsShare"]
    d.loc[mask_not_sub_resub, "_rsub_nt_rmtd"] = d.loc[mask_not_sub_resub, "SubInsShare"]

    # Rejection Accepted: uses SubInsShare
    d["_rej_accepted"] = 0.0
    d.loc[mask_rej_acc, "_rej_accepted"] = d.loc[mask_rej_acc, "SubInsShare"]

    # Extra Paid: Not Submitted old >90 days
    d["_extra_paid"] = 0.0
    d.loc[mask_not_sub_old, "_extra_paid"] = d.loc[mask_not_sub_old, "SubInsShare"]

    # ── Aggregate ─────────────────────────────────────────────────────────────
    agg = d.groupby(group_col, dropna=False, sort=True).agg(
        claim_count  = ("UniqueID",              "nunique"),
        claimed_amt  = ("SubInsShare",           "sum"),
        initial_pay  = ("RemitInsShare",         "sum"),
        resb1_pay    = ("Resub1RemitInsShare",   "sum"),
        resb2_pay    = ("Resub2RemitInsShare",   "sum"),
        resb3_pay    = ("Resub3RemitInsShare",   "sum"),
        sub_nt_rmtd  = ("_sub_nt_rmtd",         "sum"),
        rsub_nt_rmtd = ("_rsub_nt_rmtd",        "sum"),
        rej_accepted = ("_rej_accepted",         "sum"),
        extra_paid   = ("_extra_paid",           "sum"),
    ).reset_index()

    agg["total_pay"]   = agg["initial_pay"] + agg["resb1_pay"] + agg["resb2_pay"] + agg["resb3_pay"] + agg["extra_paid"]
    agg["remited_amt"] = agg["total_pay"]  # Remited Amt = Total pay in unified logic
    agg["final_rejn"]  = (agg["claimed_amt"]
                          - agg["total_pay"]
                          - agg["sub_nt_rmtd"]
                          - agg["rsub_nt_rmtd"]
                          - agg["rej_accepted"]).clip(lower=0)
    agg["rej_pct"]     = (agg["final_rejn"] / agg["claimed_amt"].replace(0, float("nan")) * 100).fillna(0.0)

    out = pd.DataFrame()
    out[group_col]                           = agg[group_col]
    out["Claim count"]                       = agg["claim_count"]
    out["Claimed Amount"]                    = agg["claimed_amt"]
    out["Remited Amt"]                       = agg["remited_amt"]
    out["Initial pay"]                       = agg["initial_pay"]
    out["Resb1 pay"]                         = agg["resb1_pay"]
    out["Resb2 pay"]                         = agg["resb2_pay"]
    out["Resb3 pay"]                         = agg["resb3_pay"]
    out["Total pay"]                         = agg["total_pay"]
    out["Sub Nt Rmtd (outstanding amount)"]  = agg["sub_nt_rmtd"]
    out["Rsub Nt Rmtd (outstanding amount)"] = agg["rsub_nt_rmtd"]
    out["Rejection Accepted"]                = agg["rej_accepted"]
    out["Final Rejn"]                        = agg["final_rejn"]
    out["Rej. %"]                            = agg["rej_pct"]

    # Grand Total row — always derive Final Rejn from components
    num_cols = out.select_dtypes(include="number").columns.tolist()
    gt = {c: out[c].sum() if c in num_cols else "Grand Total" for c in out.columns}
    gt[group_col]    = "Grand Total"
    gt_claimed       = out["Claimed Amount"].sum()
    gt_total_pay     = out["Total pay"].sum()
    gt_sub           = out["Sub Nt Rmtd (outstanding amount)"].sum()
    gt_rsub          = out["Rsub Nt Rmtd (outstanding amount)"].sum()
    gt_rej_acc       = out["Rejection Accepted"].sum()
    gt_final_rejn    = max(gt_claimed - gt_total_pay - gt_sub - gt_rsub - gt_rej_acc, 0)
    gt["Final Rejn"] = gt_final_rejn
    gt["Rej. %"]     = (gt_final_rejn / gt_claimed * 100) if gt_claimed else 0.0

    out = pd.concat([out, pd.DataFrame([gt])], ignore_index=True)
    return out

def run_summary_engine(uploaded_bytes: bytes, filename: str) -> dict:
    """
    Run the full exclusive_report_status_final.py logic in-memory.
    Returns a dict with all result DataFrames + metadata.
    """
    # ── Load ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO(uploaded_bytes)
    df = pd.read_excel(buf, engine="openpyxl")
    df.columns = df.columns.astype(str).str.strip()

    # ── Drop blank/footer rows ────────────────────────────────────────────────
    # Remove completely empty rows
    df = df.dropna(how="all").reset_index(drop=True)
    # Remove rows where entire row is blank string
    all_blank = df.fillna("").astype(str).apply(
        lambda r: "".join(r.values).strip() == "", axis=1
    )
    df = df[~all_blank].reset_index(drop=True)
    # Remove footer/total rows: Status is blank AND SubInsShare is 0 or null
    if "SubInsShare" in df.columns and "Status" in df.columns:
        sub_zero   = pd.to_numeric(df["SubInsShare"], errors="coerce").fillna(0) == 0
        stat_blank = df["Status"].astype(str).str.strip().str.lower().isin(
            ["", "nan", "none", "total", "grand total"]
        )
        df = df[~(sub_zero & stat_blank)].reset_index(drop=True)

    # Remove any row where UniqueID looks like a Grand Total label
    for _gc in ["UniqueID", "Insurance", "DocName", "Month"]:
        if _gc in df.columns:
            _gt_mask = df[_gc].astype(str).str.strip().str.lower().isin(
                ["grand total", "total"]
            )
            df = df[~_gt_mask].reset_index(drop=True)

    # ── Ensure numeric cols ───────────────────────────────────────────────────
    numeric_cols = ["SubInsShare", "RemitInsShare",
                    "Resub1RemitInsShare", "Resub2RemitInsShare",
                    "Resub3RemitInsShare", "Resub4RemitInsShare", "TakeBack"]
    for col in numeric_cols:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    if "Status" not in df.columns:
        df["Status"] = ""

    # ── Compute measures ─────────────────────────────────────────────────────
    df["Net Amount"] = df["SubInsShare"]

    # ── Not Submitted plain >90 days from VisitDate → treat as Paid ──────────
    _visit_col = next((c for c in ["VisitDate","SubDate","SubmissionDate","ClaimDate"] if c in df.columns), None)
    if _visit_col:
        df[_visit_col] = pd.to_datetime(df[_visit_col], errors="coerce", dayfirst=True)
        _today = pd.Timestamp(datetime.today().date())
        _days_since_visit = (_today - df[_visit_col]).dt.days
    else:
        _days_since_visit = pd.Series(0, index=df.index)

    _status_norm_e1   = df["Status"].apply(_normalize_status)
    _mask_not_sub_old = (
        (_status_norm_e1 == "not submitted") & (_days_since_visit > 90)
    )

    # Base Paid from remit columns
    df["Paid"] = df[["RemitInsShare","Resub1RemitInsShare","Resub2RemitInsShare",
                      "Resub3RemitInsShare","Resub4RemitInsShare","TakeBack"]].sum(axis=1)

    # Old Not Submitted: force Paid = Net Amount (fully written off to Paid)
    df.loc[_mask_not_sub_old, "Paid"] = df.loc[_mask_not_sub_old, "Net Amount"]

    df["Paid"] = df[["Paid","Net Amount"]].min(axis=1)
    df["Residual"] = (df["Net Amount"] - df["Paid"]).clip(lower=0)

    df["Rejected"] = 0.0
    df["Accepted"] = 0.0
    df["Balance"]  = 0.0

    df["Final Bucket"] = df["Status"].apply(_classify_final_bucket)
    mask_rej = df["Final Bucket"] == "Rejected"
    mask_acc = df["Final Bucket"] == "Accepted"
    # Balance = everything except Rejected, Accepted, and old Not Submitted (already in Paid)
    mask_bal = (df["Final Bucket"] == "Balance") & (~_mask_not_sub_old)

    df.loc[mask_rej, "Rejected"] = df.loc[mask_rej, "Residual"]
    df.loc[mask_acc, "Accepted"] = df.loc[mask_acc, "Residual"]
    df.loc[mask_bal, "Balance"]  = df.loc[mask_bal, "Residual"]

    df["Recon Total"] = df[["Paid","Balance","Rejected","Accepted"]].sum(axis=1)
    df["Recon Diff"]  = (df["Net Amount"] - df["Recon Total"]).round(2)

    stage_info = df["Status"].apply(_extract_stage_info)
    df["Balance Status Group"]      = stage_info.apply(lambda x: x[0])
    df["Balance Submission Stage"]  = stage_info.apply(lambda x: x[1])
    df["Balance Submission No"]     = stage_info.apply(lambda x: x[2])
    df.loc[~mask_bal, ["Balance Status Group","Balance Submission Stage","Balance Submission No"]] = ["","",None]

    # Bucket detail columns
    bucket_defs = [("Submitted",0,"Initial Submitted Balance"),
                   ("Approved", 0,"Initial Approved Balance"),
                   ("Not Submitted",0,"Initial Not Submitted Balance")]
    for n in range(1, MAX_RESUB_STAGE + 1):
        bucket_defs.extend([
            ("Submitted",n,f"Resub{n} Submitted Balance"),
            ("Approved", n,f"Resub{n} Approved Balance"),
            ("Not Submitted",n,f"Resub{n} Not Submitted Balance"),
        ])
    for _, _, col in bucket_defs:
        df[col] = 0.0
    for group, stage_no, col in bucket_defs:
        mask = (
            (df["Balance"] > 0)
            & (df["Balance Status Group"] == group)
            & (df["Balance Submission No"].fillna(-999).astype(int) == stage_no)
        )
        df.loc[mask, col] = df.loc[mask, "Balance"]

    # ── Dates & aging ─────────────────────────────────────────────────────────
    date_cols_all = ["SubDate","Resub1Date","Resub2Date","Resub3Date","Resub4Date","Resub5Date",
                     "Resub6Date","Resub7Date","Resub8Date","Resub9Date","Resub10Date",
                     "SubmissionDate","ClaimDate","VisitDate"]
    for c in [x for x in date_cols_all if x in df.columns]:
        df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)

    df_cols = list(df.columns)
    df["Balance RefDate"] = df.apply(lambda r: _choose_stage_date(r, df_cols), axis=1)
    today = pd.Timestamp(datetime.today().date())
    df["DaysDiff"] = (today - df["Balance RefDate"]).dt.days

    bins   = [-1, 30, 45, 60, 90, float("inf")]
    labels = ["0–30 Days","31–45 Days","46–60 Days","61–90 Days",">90 Days"]
    df["AgingBucket"] = pd.cut(df["DaysDiff"], bins=bins, labels=labels)

    # ── Insurance column ──────────────────────────────────────────────────────
    ins_col = next((c for c in ["Insurance","PayerName","Insurer","Plan"] if c in df.columns), None)
    if ins_col is None:
        df["Insurance"] = "Not Available"
    elif ins_col != "Insurance":
        df["Insurance"] = df[ins_col]
    df["Insurance"] = df["Insurance"].fillna("Not Available")

    # ── Build summary tables ──────────────────────────────────────────────────
    cols_base = ["Net Amount","Paid","Balance","Rejected","Accepted","Recon Diff"]

    # Insurance Totals
    ins_totals = df.groupby("Insurance", dropna=False)[cols_base].sum().reset_index()
    ins_totals = _add_grand_total_row(ins_totals, "Insurance")

    # Final Bucket Summary
    fb_summary = df.groupby("Final Bucket", dropna=False)[["Net Amount","Paid","Balance","Rejected","Accepted"]].sum().reset_index()
    fb_summary = _add_grand_total_row(fb_summary, "Final Bucket")

    # Monthly Totals
    date_col_m = next((c for c in ["VisitDate","SubDate","SubmissionDate","ClaimDate"] if c in df.columns), None)
    monthly = pd.DataFrame()
    if date_col_m:
        tmp = df.copy()
        tmp[date_col_m] = pd.to_datetime(tmp[date_col_m], errors="coerce", dayfirst=True)
        tmp = tmp.dropna(subset=[date_col_m])
        if not tmp.empty:
            tmp["Month"] = tmp[date_col_m].dt.to_period("M").dt.strftime("%B %Y")
            monthly = tmp.groupby("Month", observed=True)[cols_base].sum().reset_index()
            monthly = _add_grand_total_row(monthly, "Month")

    # Balance Aging Summary
    balance_df = df[(df["Balance"] > 0) & df["AgingBucket"].notna()].copy()
    if not balance_df.empty:
        aging_pivot = pd.pivot_table(
            balance_df, index="Insurance", columns="AgingBucket",
            values="Balance", aggfunc="sum", fill_value=0, observed=False,
        ).reindex(columns=labels, fill_value=0)
        aging_pivot["Grand Total"] = aging_pivot.sum(axis=1)
        aging_summary = aging_pivot.reset_index()
        aging_summary = _add_grand_total_row(aging_summary, "Insurance")
    else:
        aging_summary = pd.DataFrame(columns=["Insurance"] + labels + ["Grand Total"])

    # Balance Status Stage Summary
    bss_df = df[df["Balance"] > 0].copy()
    if not bss_df.empty:
        bss = (bss_df.groupby(["Balance Status Group","Balance Submission Stage"], dropna=False)["Balance"]
               .sum().reset_index()
               .sort_values(["Balance Status Group","Balance Submission Stage"]))
        gt_bss = pd.DataFrame([{"Balance Status Group":"Grand Total","Balance Submission Stage":"","Balance":bss["Balance"].sum()}])
        bss = pd.concat([bss, gt_bss], ignore_index=True)
    else:
        bss = pd.DataFrame(columns=["Balance Status Group","Balance Submission Stage","Balance"])

    # Recon check
    recon_diff_total = float(df["Recon Diff"].sum())

    # ── KPIs derived from rcm_insurance Grand Total (set after build below) ──
    # Placeholder — overwritten after rcm_insurance is built
    kpi_net = kpi_paid = kpi_bal = kpi_rej = kpi_acc = 0.0

    # ── RCM Summary (Insurance view) ─────────────────────────────────────────
    # Ensure Difference column exists
    if "Difference" not in df.columns:
        df["Difference"] = 0.0
    df["Difference"] = pd.to_numeric(df["Difference"], errors="coerce").fillna(0.0)
    if "UniqueID" not in df.columns:
        df["UniqueID"] = range(len(df))  # fallback sequential

    ins_col_rcm = next((c for c in ["Insurance","PayerName","Insurer","Plan"] if c in df.columns), None)
    rcm_group_col = ins_col_rcm if ins_col_rcm else "Insurance"
    if rcm_group_col not in df.columns:
        df[rcm_group_col] = "Not Available"
    df[rcm_group_col] = df[rcm_group_col].fillna("Not Available").astype(str).str.strip()

    # Drop rows where Insurance is blank/Not Available — these are embedded total/summary rows
    df_rcm = df[~df[rcm_group_col].str.lower().isin(
        ["not available", "", "nan", "none", "total", "grand total"]
    )].copy()

    rcm_insurance = build_rcm_summary(df_rcm, rcm_group_col, visit_days_series=_days_since_visit)

    # Doctor wise — try DocName, DoctorName, Doctor, PhysicianName
    rcm_doctor = pd.DataFrame()
    doc_col = next((c for c in ["DocName","DoctorName","Doctor","PhysicianName"] if c in df.columns), None)
    if doc_col:
        df_rcm[doc_col] = df_rcm[doc_col].fillna("Not Available").astype(str).str.strip()
        # Also filter out blank doctor names
        df_doc = df_rcm[~df_rcm[doc_col].str.lower().isin(
            ["not available","","nan","none","total","grand total"]
        )].copy()
        rcm_doctor = build_rcm_summary(df_doc, doc_col, visit_days_series=_days_since_visit)

    # Month wise — source file has a "Month" column directly
    rcm_month = pd.DataFrame()
    # Priority: direct "Month" column first, then parse from date columns
    date_col_rcm = next((c for c in ["VisitDate","SubDate","SubmissionDate","ClaimDate"] if c in df_rcm.columns), None)
    if "Month" in df_rcm.columns:
        import calendar as _cal
        _month_name_map = {m.lower(): i for i, m in enumerate(_cal.month_name) if m}
        _month_abbr_map = {m.lower(): i for i, m in enumerate(_cal.month_abbr) if m}

        def _to_month_label(row):
            """Convert Month + Year columns to Jan-22 format."""
            import re as _re
            s = str(row["Month"]).strip()

            # Get year — strip index names and try variants
            yr_val = None
            _row_idx_stripped = {str(k).strip(): k for k in row.index}
            for _ycol in ["Year", "year", "YEAR", "yr", "YR"]:
                _actual = _row_idx_stripped.get(_ycol)
                if _actual is not None:
                    try:
                        yr_val = int(float(str(row[_actual]))) % 100
                        break
                    except Exception:
                        pass
            # Also try to extract year from VisitDate or date columns
            if yr_val is None:
                for _dcol in ["VisitDate","SubDate","SubmissionDate","ClaimDate"]:
                    _actual = _row_idx_stripped.get(_dcol)
                    if _actual is not None:
                        try:
                            _dval = row[_actual]
                            # Handle pandas Timestamp directly
                            if hasattr(_dval, "year"):
                                yr_val = int(_dval.year) % 100
                                break
                            # Handle string "2022-01-15"
                            import re as _re2
                            m2 = _re2.match(r"(\d{4})", str(_dval).strip())
                            if m2:
                                yr_val = int(m2.group(1)) % 100
                                break
                        except Exception:
                            pass

            # datetime string: "2022-01-01 00:00:00" or "2022-01-01"
            dt_match = _re.match(r"(\d{4})-(\d{2})-\d{2}", s)
            if dt_match:
                yr = int(dt_match.group(1)) % 100
                mo = int(dt_match.group(2))
                return f"{_cal.month_abbr[mo]}-{yr:02d}"
            # Already Jan-22 format
            if _re.match(r"[A-Za-z]{3}-\d{2}", s):
                return s.title()[:3] + s[3:]
            # Full month name "April" + year
            lower = s.lower()
            if lower in _month_name_map:
                abbr = _cal.month_abbr[_month_name_map[lower]]
                if yr_val is not None:
                    return f"{abbr}-{yr_val:02d}"
                return abbr
            return s

        tmp_m = df_rcm.copy()
        tmp_m["Month"] = tmp_m.apply(_to_month_label, axis=1)
        tmp_m = tmp_m[tmp_m["Month"].str.lower() != "nan"]
        if not tmp_m.empty:
            rcm_month = build_rcm_summary(tmp_m, "Month", visit_days_series=_days_since_visit)
            # Sort chronologically
            def _month_sort_key(val):
                import re as _re
                try:
                    parts = str(val).strip().split("-")
                    if len(parts) == 2:
                        mon = parts[0].strip().title()[:3]
                        yr  = int(parts[1].strip())
                        return (yr, _month_abbr_map.get(mon.lower(), 0))
                except Exception:
                    pass
                return (9999, 99)
            gt_mask = rcm_month["Month"].astype(str).str.match(r"^\s*(grand\s*total|total)\s*$", case=False)
            body   = rcm_month[~gt_mask].copy()
            gt_row = rcm_month[gt_mask].copy()
            body["_sort"] = body["Month"].apply(_month_sort_key)
            body = body.sort_values("_sort").drop(columns=["_sort"])
            rcm_month = pd.concat([body, gt_row], ignore_index=True)
    elif date_col_rcm:
        tmp_m = df.copy()
        tmp_m[date_col_rcm] = pd.to_datetime(tmp_m[date_col_rcm], errors="coerce", dayfirst=True)
        tmp_m = tmp_m.dropna(subset=[date_col_rcm])
        if not tmp_m.empty:
            tmp_m["Month"] = tmp_m[date_col_rcm].dt.strftime("%b-%y")
            rcm_month = build_rcm_summary(tmp_m, "Month", visit_days_series=_days_since_visit)

    # ── KPIs from rcm_insurance Grand Total row ─────────────────────────────
    gt_pat_kpi = re.compile(r"^\s*(grand\s*total|total)\s*$", re.I)
    if rcm_insurance is not None and not rcm_insurance.empty:
        gt_row_kpi = rcm_insurance[rcm_insurance.iloc[:,0].astype(str).str.match(gt_pat_kpi)]
        if not gt_row_kpi.empty:
            def _kv(col):
                try: return float(gt_row_kpi[col].values[0])
                except: return 0.0
            kpi_net  = _kv("Claimed Amount")
            kpi_paid = _kv("Total pay")
            kpi_bal  = _kv("Sub Nt Rmtd (outstanding amount)") + _kv("Rsub Nt Rmtd (outstanding amount)")
            kpi_rej  = _kv("Final Rejn")
            kpi_acc  = _kv("Rejection Accepted")

    return {
        "df":            df,
        "ins_totals":    ins_totals,
        "fb_summary":    fb_summary,
        "monthly":       monthly,
        "aging_summary": aging_summary,
        "bss":           bss,
        "kpi":           (kpi_net, kpi_paid, kpi_bal, kpi_rej, kpi_acc),
        "recon_diff":    recon_diff_total,
        "row_count":     len(df),
        "filename":      filename,
        "generated_at":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rcm_insurance": rcm_insurance,
        "rcm_doctor":    rcm_doctor,
        "rcm_month":     rcm_month,
        "rcm_group_col": rcm_group_col,
        "claim_detail":  _build_claim_detail(df, _days_since_visit),
    }


def _build_claim_detail(df: pd.DataFrame, days_series: pd.Series) -> pd.DataFrame:
    """
    Build per-claim detail report — identical logic to build_rcm_summary.
    Uses SubInsShare for all columns. No Difference column.
    """
    d = df.copy()

    for c in ["SubInsShare","RemitInsShare","Resub1RemitInsShare",
              "Resub2RemitInsShare","Resub3RemitInsShare"]:
        if c not in d.columns:
            d[c] = 0.0
        d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0.0)

    def _norm(v):
        return re.sub(r"\s+", " ", str(v or "").strip().lower())

    d["_sn"]  = d["Status"].apply(_norm)
    _days     = days_series.reindex(d.index).fillna(0)

    # ── Same masks as build_rcm_summary ───────────────────────────────────────
    mask_sub_init      = d["_sn"] == "submitted"
    mask_sub_resub     = d["_sn"].str.match(r"^submitted\s*\(\s*resub\s*-\s*\d+\s*\)$", na=False)
    mask_not_sub_plain = d["_sn"] == "not submitted"
    mask_not_sub_old   = mask_not_sub_plain & (_days > 90)
    mask_not_sub_fresh = mask_not_sub_plain & (_days <= 90)
    mask_not_sub_resub = d["_sn"].str.match(r"^not submitted\s*\(\s*resub\s*-\s*\d+\s*\)$", na=False)
    mask_rej_acc       = d["_sn"].str.match(r"^rejection accepted(\s*\(\s*resub\s*-\s*\d+\s*\))?$", na=False)
    mask_rejected      = d["_sn"].str.match(r"^rejected(\s*\(\s*resub\s*-\s*\d+\s*\))?$", na=False)
    mask_approved      = d["_sn"].str.match(r"^approved(\s*\(.+\))?$", na=False)
    mask_any_matched   = (mask_sub_init | mask_sub_resub | mask_not_sub_plain |
                          mask_not_sub_resub | mask_rej_acc | mask_rejected | mask_approved)
    mask_unmatched     = (~mask_any_matched) & (d["SubInsShare"] > 0)

    # ── Pay columns ───────────────────────────────────────────────────────────
    d["Initial pay"]  = d["RemitInsShare"]
    d["Resb1 pay"]    = d["Resub1RemitInsShare"]
    d["Resb2 pay"]    = d["Resub2RemitInsShare"]
    d["Resb3 pay"]    = d["Resub3RemitInsShare"]
    d["_extra_paid"]  = d["SubInsShare"].where(mask_not_sub_old, 0.0)
    d["Total pay"]    = d["Initial pay"] + d["Resb1 pay"] + d["Resb2 pay"] + d["Resb3 pay"] + d["_extra_paid"]

    # ── Sub Nt Rmtd ───────────────────────────────────────────────────────────
    d["Sub Nt Rmtd"] = 0.0
    d.loc[mask_sub_init,      "Sub Nt Rmtd"] = d.loc[mask_sub_init,      "SubInsShare"]
    d.loc[mask_not_sub_fresh, "Sub Nt Rmtd"] = d.loc[mask_not_sub_fresh, "SubInsShare"]
    d.loc[mask_unmatched,     "Sub Nt Rmtd"] = d.loc[mask_unmatched,     "SubInsShare"]

    # ── Rsub Nt Rmtd ──────────────────────────────────────────────────────────
    d["Rsub Nt Rmtd"] = 0.0
    d.loc[mask_sub_resub,     "Rsub Nt Rmtd"] = d.loc[mask_sub_resub,     "SubInsShare"]
    d.loc[mask_not_sub_resub, "Rsub Nt Rmtd"] = d.loc[mask_not_sub_resub, "SubInsShare"]

    # ── Rejection Accepted ────────────────────────────────────────────────────
    d["Rejection Accepted"] = 0.0
    d.loc[mask_rej_acc, "Rejection Accepted"] = d.loc[mask_rej_acc, "SubInsShare"]

    # ── Final Rejn ────────────────────────────────────────────────────────────
    d["Final Rejn"] = (
        d["SubInsShare"]
        - d["Total pay"]
        - d["Sub Nt Rmtd"]
        - d["Rsub Nt Rmtd"]
        - d["Rejection Accepted"]
    ).clip(lower=0)

    id_cols = [c for c in ["UniqueID","Insurance","DocName","Status","VisitDate","Month","SubInsShare"] if c in d.columns]
    report_cols = id_cols + [
        "Initial pay","Resb1 pay","Resb2 pay","Resb3 pay","Total pay",
        "Sub Nt Rmtd","Rsub Nt Rmtd","Rejection Accepted","Final Rejn"
    ]
    return d[report_cols].reset_index(drop=True)


def build_claim_detail_excel(df_detail: pd.DataFrame) -> bytes:
    """Build styled Excel for claim detail report — fast bulk pandas write."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Claim Detail"

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")
    def _border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    cols = df_detail.columns.tolist()
    num_cols = set(df_detail.select_dtypes(include="number").columns.tolist())
    text_cols = {"UniqueID","Insurance","DocName","Status"}

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = _fill("0A2647")
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Bulk write data using pandas values ───────────────────────────────────
    data = df_detail.values.tolist()
    for ri, row_data in enumerate(data, 2):
        bg = "FFFFFF" if ri % 2 == 0 else "F2F2F2"
        fill = _fill(bg)
        for ci, (col, val) in enumerate(zip(cols, row_data), 1):
            # Convert NaT/NaN to None for cleaner Excel
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = fill
            cell.border = _border()
            cell.alignment = Alignment(
                horizontal="left" if col in text_cols else "right",
                vertical="center"
            )
            if col == "Final Rejn":
                cell.font = _font(bold=True, color="C0392B", size=10)
            else:
                cell.font = _font(size=10)
            if col in num_cols:
                cell.number_format = "#,##0.00"

    # ── Grand Total row ───────────────────────────────────────────────────────
    gt_row = len(data) + 2
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=gt_row, column=ci)
        cell.fill   = _fill("D9D9D9")
        cell.font   = _font(bold=True, size=10)
        cell.border = _border()
        if col in num_cols:
            cell.value          = float(df_detail[col].sum())
            cell.number_format  = "#,##0.00"
            cell.alignment      = Alignment(horizontal="right", vertical="center")
        elif ci == 1:
            cell.value     = "Grand Total"
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        if col in ("Insurance","DocName","Status"):
            ws.column_dimensions[get_column_letter(ci)].width = 30
        elif col == "UniqueID":
            ws.column_dimensions[get_column_letter(ci)].width = 20
        elif col == "VisitDate":
            ws.column_dimensions[get_column_letter(ci)].width = 14
        else:
            ws.column_dimensions[get_column_letter(ci)].width = 16

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_output(result: dict) -> bytes:
    """
    Build a single-sheet Excel workbook matching the reference format:
    - Light grey title row with auto date range
    - Dark green column headers, green text for pay columns
    - White/light grey alternating rows
    - Green Grand Total row
    - All 3 sections (Insurance / Doctor / Month) in one sheet
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Colour palette (matching screenshot) ─────────────────────────────────
    C_TITLE_BG   = "D9D9D9"   # light grey  — title bg
    C_TITLE_FG   = "000000"   # black       — title text
    C_HDR_BG     = "595959"   # dark grey   — column header bg
    C_HDR_FG     = "FFFFFF"   # white       — header text
    C_GREEN_FG   = "00B050"   # green       — Initial/Resb pay col text
    C_SUBHDR_BG  = "BFBFBF"   # mid grey    — section label row
    C_SUBHDR_FG  = "000000"   # black       — section label text
    C_GT_BG      = "D9D9D9"   # grey        — grand total row
    C_GT_FG      = "000000"   # black bold  — grand total text
    C_ALT1       = "FFFFFF"   # white       — odd rows
    C_ALT2       = "F2F2F2"   # light grey  — even rows
    C_REJ_FG     = "FF0000"   # red         — Rej.% > 0
    C_BORDER     = "BFBFBF"   # grey border

    NUM_COLS = 13   # total columns (same as reference)

    def _thin():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic,
                    name="Calibri")

    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = Workbook()
    ws = wb.active
    ws.title = "RCM SUMMARY"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [42, 12, 16, 14, 13, 11, 11, 11, 13, 18, 18, 14, 9]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    cur = 1  # current row pointer

    def _write_section(df_sec, section_label, start_row):
        """Write one section (Insurance / Doctor / Month) starting at start_row.
        Returns next available row."""
        r = start_row

        # Section label row (e.g. "Insurance Name")
        ws.row_dimensions[r].height = 20
        for c in range(1, NUM_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill    = _fill(C_SUBHDR_BG)
            cell.font    = _font(bold=True, color=C_SUBHDR_FG, size=11)
            cell.border  = _thin()
            cell.alignment = _align("center")
        ws.cell(row=r, column=1, value=section_label)
        ws.cell(row=r, column=1).alignment = _align("left")
        r += 1

        # Column header row
        ws.row_dimensions[r].height = 36
        GREEN_COLS = {"Initial pay", "Resb1 pay", "Resb2 pay", "Resb3 pay"}
        for ci, col in enumerate(df_sec.columns, 1):
            cell = ws.cell(row=r, column=ci, value=col)
            cell.fill      = _fill(C_HDR_BG)
            cell.font      = _font(bold=True,
                                   color=("90EE90" if col in GREEN_COLS else C_HDR_FG),
                                   size=10)
            cell.border    = _thin()
            cell.alignment = _align("center", wrap=True)
        r += 1

        # Data rows
        for row_idx, (_, row) in enumerate(df_sec.iterrows()):
            is_gt = str(row.iloc[0]).strip().lower() in ("grand total", "total")
            ws.row_dimensions[r].height = 18

            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=ci)

                # Format numbers
                if ci > 1 and not is_gt:
                    try:
                        val = float(val)
                        col_name = df_sec.columns[ci - 1]
                        if col_name == "Rej. %":
                            cell.number_format = "0.00%"
                            val = val / 100
                        elif col_name == "Claim count":
                            cell.number_format = "0"
                            val = int(val)
                        else:
                            cell.number_format = "#,##0.00"
                    except (ValueError, TypeError):
                        pass
                elif ci > 1 and is_gt:
                    try:
                        val = float(val)
                        col_name = df_sec.columns[ci - 1]
                        if col_name == "Rej. %":
                            cell.number_format = "0.00%"
                            val = val / 100
                        elif col_name == "Claim count":
                            cell.number_format = "0"
                            val = int(val)
                        else:
                            cell.number_format = "#,##0.00"
                    except (ValueError, TypeError):
                        pass

                cell.value = val

                if is_gt:
                    cell.fill      = _fill(C_GT_BG)
                    cell.font      = _font(bold=True, color=C_GT_FG, size=10)
                else:
                    bg = C_ALT1 if row_idx % 2 == 0 else C_ALT2
                    cell.fill = _fill(bg)
                    # Red Rej.%
                    col_name = df_sec.columns[ci - 1]
                    if col_name == "Rej. %" and not is_gt:
                        try:
                            if float(row.iloc[ci - 1]) > 0:
                                cell.font = _font(bold=True, color=C_REJ_FG, size=10)
                            else:
                                cell.font = _font(color="000000", size=10)
                        except Exception:
                            cell.font = _font(size=10)
                    elif col_name in ("Initial pay","Resb1 pay","Resb2 pay","Resb3 pay"):
                        cell.font = _font(color="1E8449", size=10)
                    else:
                        cell.font = _font(size=10)
                    cell.alignment = _align("right" if ci > 1 else "left")

                cell.border = _thin()
                if is_gt:
                    cell.alignment = _align("right" if ci > 1 else "left")
            r += 1

        return r  # next row after this section

    # ── Build title from date range in data ──────────────────────────────────
    def _get_date_range():
        """Get start/end month-year from rcm_month rows (e.g. Jan-22, Feb-22)."""
        try:
            import calendar as _cal
            _abbr_map = {m.lower(): i for i, m in enumerate(_cal.month_abbr) if m}

            def _mk(v):
                """Sort key for Mon-YY labels."""
                p = str(v).strip().split("-")
                if len(p) == 2:
                    try:
                        return (int(p[1]), _abbr_map.get(p[0].lower()[:3], 0))
                    except Exception:
                        pass
                return (9999, 99)

            def _expand(lbl):
                """Jan-22 -> JAN 2022"""
                p = str(lbl).strip().split("-")
                if len(p) == 2:
                    mon = p[0].upper()
                    yr  = int(p[1])
                    full_yr = (2000 + yr) if yr < 50 else (1900 + yr)
                    return f"{mon} {full_yr}"
                return lbl.upper()

            rcm_m = result.get("rcm_month")
            if rcm_m is not None and not rcm_m.empty:
                gt_p = re.compile(r"^\s*(grand\s*total|total)\s*$", re.I)
                rows = [str(v).strip() for v in rcm_m.iloc[:, 0]
                        if not gt_p.match(str(v))]
                # Only use rows that have Mon-YY format (contain "-")
                valid = [r for r in rows if "-" in r]
                if valid:
                    rows_sorted = sorted(valid, key=_mk)
                    return f"EMC - RCM SUMMARY - {_expand(rows_sorted[0])} - {_expand(rows_sorted[-1])}"
        except Exception:
            pass
        return "EMC - RCM SUMMARY"

    title_text = _get_date_range()

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=NUM_COLS)
    title_cell = ws.cell(row=cur, column=1)
    title_cell.value     = title_text
    title_cell.fill      = _fill(C_TITLE_BG)
    title_cell.font      = _font(bold=True, color=C_TITLE_FG, size=14)
    title_cell.alignment = _align("center")
    ws.row_dimensions[cur].height = 36
    cur += 1

    # blank row
    ws.row_dimensions[cur].height = 8
    cur += 1

    # ── Insurance section ─────────────────────────────────────────────────────
    rcm_ins = result.get("rcm_insurance")
    if rcm_ins is not None and not rcm_ins.empty:
        cur = _write_section(rcm_ins, "Insurance Name", cur)
        ws.row_dimensions[cur].height = 8
        cur += 1   # blank separator

    # ── Doctor section ────────────────────────────────────────────────────────
    rcm_doc = result.get("rcm_doctor")
    if rcm_doc is not None and not rcm_doc.empty:
        cur = _write_section(rcm_doc, "Doctor Name", cur)
        ws.row_dimensions[cur].height = 8
        cur += 1

    # ── Month section ─────────────────────────────────────────────────────────
    rcm_mon = result.get("rcm_month")
    if rcm_mon is not None and not rcm_mon.empty:
        cur = _write_section(rcm_mon, "Month", cur)

    # freeze panes below title
    ws.freeze_panes = "A3"


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# DISPLAY HELPER — styled dataframe (Grand Total last, highlighted)
# ─────────────────────────────────────────────────────────────────────────────
GT_PAT = re.compile(r'^\s*(grand\s*total|total)\s*$', re.I)


def _move_gt_last(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    first = df.columns[0]
    mask = df[first].astype(str).str.match(GT_PAT)
    return pd.concat([df.loc[~mask], df.loc[mask]], ignore_index=True)


def _style_gt(df: pd.DataFrame):
    """Apply Grand Total bold + orange highlight."""
    def _highlight(row):
        if GT_PAT.match(str(row.iloc[0])):
            return ["background-color:#FCE4D6;font-weight:bold"] * len(row)
        return [""] * len(row)
    return df.style.apply(_highlight, axis=1)


def _fmt_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """Format numeric columns to 2 decimal places for display."""
    num_cols = df.select_dtypes(include="number").columns.tolist()
    return df.style.format({c: "{:,.2f}" for c in num_cols})


def show_table(df: pd.DataFrame, key: str):
    if df is None or df.empty:
        st.info("No data available.")
        return
    df = _move_gt_last(df)
    num_cols = df.select_dtypes(include="number").columns.tolist()
    styled = df.style.apply(
        lambda row: ["background-color:#FCE4D6;font-weight:bold" if GT_PAT.match(str(row.iloc[0])) else "" for _ in row],
        axis=1
    ).format({c: "{:,.2f}" for c in num_cols})
    st.dataframe(styled, use_container_width=True, hide_index=True, key=key)




def _result_cache_key(center_key: str) -> str:
    """S3 key for the cached result JSON for a center."""
    year = st.session_state.get("rcm_year") or datetime.now().year
    return f"{SUMMARY_S3_PREFIX}/{year}/{center_key}/_latest_result_cache.json"


def _save_result_to_s3(result: dict, center_key: str):
    """Serialize the summary tables to JSON and save to S3."""
    import json
    try:
        # Only serialize the display DataFrames — skip raw df (too large) and bytes
        cache = {
            "filename":     result.get("filename", ""),
            "generated_at": result.get("generated_at", ""),
            "row_count":    result.get("row_count", 0),
            "recon_diff":   result.get("recon_diff", 0.0),
            "kpi":          list(result.get("kpi", [0,0,0,0,0])),
            "rcm_insurance": result["rcm_insurance"].to_json(orient="split") if result.get("rcm_insurance") is not None and not result["rcm_insurance"].empty else None,
            "rcm_doctor":    result["rcm_doctor"].to_json(orient="split")    if result.get("rcm_doctor")    is not None and not result["rcm_doctor"].empty    else None,
            "rcm_month":     result["rcm_month"].to_json(orient="split")     if result.get("rcm_month")     is not None and not result["rcm_month"].empty     else None,
        }
        key = _result_cache_key(center_key)
        s3 = get_s3_client()
        if s3 is None: return
        s3.put_object(
            Bucket=S3_BUCKET, Key=key,
            Body=json.dumps(cache, ensure_ascii=False).encode("utf-8"),
            ContentType="application/json"
        )
    except Exception:
        pass  # cache save failure is non-fatal


def _load_result_from_s3(center_key: str) -> dict | None:
    """Load cached result JSON from S3. Returns None if not found."""
    import json
    try:
        key = _result_cache_key(center_key)
        s3 = get_s3_client()
        if s3 is None: return None
        obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
        cache = json.loads(obj["Body"].read().decode("utf-8"))
        # Reconstruct DataFrames
        def _from_json(j):
            if j is None: return pd.DataFrame()
            return pd.read_json(j, orient="split")
        # ── 3-day TTL check ──────────────────────────────────────────────
        generated_at = cache.get("generated_at", "")
        if generated_at:
            try:
                age = datetime.now() - datetime.strptime(generated_at, "%Y-%m-%d %H:%M:%S")
                if age.total_seconds() > 3 * 24 * 3600:
                    return None  # expired — force fresh upload
            except Exception:
                pass
        # ─────────────────────────────────────────────────────────────────────
        cache["rcm_insurance"] = _from_json(cache.get("rcm_insurance"))
        cache["rcm_doctor"]    = _from_json(cache.get("rcm_doctor"))
        cache["rcm_month"]     = _from_json(cache.get("rcm_month"))
        cache["xl_bytes"]      = None

        # Always recompute KPIs from rcm_insurance Grand Total — never trust stored kpi
        import re as _re2
        _gt_p = _re2.compile(r"^\s*(grand\s*total|total)\s*$", _re2.I)
        _ins = cache["rcm_insurance"]
        if _ins is not None and not _ins.empty:
            _gt = _ins[_ins.iloc[:,0].astype(str).str.match(_gt_p)]
            if not _gt.empty:
                def _kv(col):
                    try: return float(_gt[col].values[0])
                    except: return 0.0
                cache["kpi"] = (
                    _kv("Claimed Amount"),
                    _kv("Total pay"),
                    _kv("Sub Nt Rmtd (outstanding amount)") + _kv("Rsub Nt Rmtd (outstanding amount)"),
                    _kv("Final Rejn"),
                    _kv("Rejection Accepted"),
                )
            else:
                cache["kpi"] = tuple(cache.get("kpi", [0,0,0,0,0]))
        else:
            cache["kpi"] = tuple(cache.get("kpi", [0,0,0,0,0]))
        return cache
    except Exception:
        return None

# =============================================================================
# EMAIL — Outlook/Office365 SMTP
# =============================================================================
def send_rcm_email(excel_bytes: bytes, excel_filename: str, result: dict, center_name: str) -> tuple:
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders
    import calendar as _cal, re as _re

    # Read secrets — handle both flat and nested [email] section formats
    def _get_secret(*keys):
        for k in keys:
            # flat: EMAIL_SENDER = "..."
            try:
                v = st.secrets[k]
                if v: return str(v).strip()
            except Exception:
                pass
            # nested: [email] section
            try:
                v = st.secrets["email"][k]
                if v: return str(v).strip()
            except Exception:
                pass
            # nested: [EMAIL] section
            try:
                v = st.secrets["EMAIL"][k]
                if v: return str(v).strip()
            except Exception:
                pass
        return ""

    sender    = _get_secret("SMTP_USER",   "EMAIL_SENDER",    "sender")
    password  = _get_secret("SMTP_PASS",   "EMAIL_PASSWORD",  "password")
    recipient = _get_secret("EMAIL_TO",    "EMAIL_RECIPIENT", "recipient")
    cc        = _get_secret("EMAIL_CC",    "EMAIL_RECIPIENT_CC", "cc")
    smtp_host = _get_secret("SMTP_HOST") or "smtp.office365.com"
    smtp_port = int(_get_secret("SMTP_PORT") or 587)

    missing = []
    if not sender:    missing.append("EMAIL_SENDER")
    if not password:  missing.append("EMAIL_PASSWORD")
    if not recipient: missing.append("EMAIL_RECIPIENT")
    if missing:
        available = list(st.secrets.keys())
        return False, f"Missing secrets: {', '.join(missing)}. Available keys: {available}"

    net, paid, bal, rej, acc = result.get("kpi", (0,0,0,0,0))
    generated_at  = result.get("generated_at", "")
    filename_orig = result.get("filename", "")

    # Build title
    title_line = "EMC - RCM SUMMARY"
    try:
        _abbr_map = {m.lower(): i for i, m in enumerate(_cal.month_abbr) if m}
        def _mk(v):
            p = str(v).strip().split("-")
            if len(p) == 2:
                try: return (int(p[1]), _abbr_map.get(p[0].lower()[:3], 0))
                except: pass
            return (9999, 99)
        def _expand(lbl):
            p = str(lbl).strip().split("-")
            if len(p) == 2:
                yr = int(p[1])
                return f"{p[0].upper()} {(2000+yr) if yr<50 else (1900+yr)}"
            return lbl.upper()
        gt_p = _re.compile(r"^\s*(grand\s*total|total)\s*$", _re.I)
        rcm_m = result.get("rcm_month")
        if rcm_m is not None and not rcm_m.empty:
            rows = [str(v).strip() for v in rcm_m.iloc[:,0] if not gt_p.match(str(v))]
            valid = [r for r in rows if "-" in r]
            if valid:
                s = sorted(valid, key=_mk)
                title_line = f"EMC - RCM SUMMARY - {_expand(s[0])} - {_expand(s[-1])}"
    except Exception:
        pass

    def _fmt(v):
        try: return f"{float(v):,.2f}"
        except: return "0.00"

    gt_p2 = re.compile(r"^\s*(grand\s*total|total)\s*$", re.I)

    # ── Exact Excel colors ────────────────────────────────────────────────────
    E_TITLE_BG  = "#D9D9D9"
    E_HDR_BG    = "#595959"
    E_HDR_FG    = "#FFFFFF"
    E_SUBHDR_BG = "#BFBFBF"
    E_GT_BG     = "#D9D9D9"
    E_ALT1      = "#FFFFFF"
    E_ALT2      = "#F2F2F2"
    E_GREEN     = "#1E8449"
    E_RED       = "#C0392B"
    E_BORDER    = "#BFBFBF"
    GREEN_COLS  = {"Initial pay", "Resb1 pay", "Resb2 pay", "Resb3 pay"}

    def _df_to_html(df, caption=""):
        if df is None or df.empty:
            return ""
        cols = list(df.columns)

        # Section header banner (matching Excel section label row)
        sec_banner = (
            f"<tr><td colspan='{len(cols)}' style='background:{E_SUBHDR_BG};"
            f"font-weight:700;font-size:13px;padding:8px 12px;"
            f"border:1px solid {E_BORDER};color:#000;text-align:left;'>"
            f"{caption}</td></tr>"
        ) if caption else ""

        # Column headers
        ths = "".join(
            f"<th style='padding:7px 10px;background:{E_HDR_BG};color:{E_HDR_FG};"
            f"border:1px solid #888;text-align:{'left' if i==0 else 'right'};"
            f"font-size:11px;white-space:nowrap;"
            f"{'color:#90EE90;' if c in GREEN_COLS else ''}'>{c}</th>"
            for i, c in enumerate(cols)
        )

        rows_html = ""
        for ri, (_, row) in enumerate(df.iterrows()):
            is_gt = bool(gt_p2.match(str(row.iloc[0])))
            bg = E_GT_BG if is_gt else (E_ALT1 if ri % 2 == 0 else E_ALT2)
            fw = "font-weight:700;" if is_gt else ""
            cells = ""
            for ci, val in enumerate(row):
                col = cols[ci]
                align = "left" if ci == 0 else "right"
                extra = ""
                if col in GREEN_COLS and not is_gt:
                    extra = f"color:{E_GREEN};"
                if col == "Rej. %" and not is_gt:
                    try:
                        if float(val) > 0:
                            extra = f"color:{E_RED};font-weight:600;"
                    except: pass
                # Format values
                if col == "Rej. %":
                    try: val = f"{float(val):.2f}%"
                    except: pass
                elif col == "Claim count":
                    try: val = f"{int(float(val)):,}"
                    except: pass
                elif ci > 0:
                    try: val = f"{float(val):,.2f}"
                    except: pass
                cells += (
                    f"<td style='padding:6px 10px;border:1px solid {E_BORDER};"
                    f"text-align:{align};background:{bg};{fw}{extra}"
                    f"font-size:11px;font-family:Calibri,Arial,sans-serif;'>{val}</td>"
                )
            rows_html += f"<tr>{cells}</tr>"

        return (
            f"<table style='border-collapse:collapse;width:100%;"
            f"font-size:11px;font-family:Calibri,Arial,sans-serif;margin-bottom:20px;'>"
            f"<thead><tr>{sec_banner}</tr><tr>{ths}</tr></thead>"
            f"<tbody>{rows_html}</tbody></table>"
        )

    ins_html = _df_to_html(result.get("rcm_insurance"), "Insurance Name")
    doc_html = _df_to_html(result.get("rcm_doctor"),    "Doctor Name")
    mon_html = _df_to_html(result.get("rcm_month"),     "Month")

    # ── KPI cards (matching Streamlit UI style) ───────────────────────────────
    def _kpi_card(label, value, dark=False):
        bg   = "#0A2647" if dark else "#FFFFFF"
        lc   = "rgba(180,205,255,0.75)" if dark else "#8A9BB5"
        vc   = "#FFFFFF" if dark else "#0D1B2E"
        bord = "rgba(180,210,255,0.25)" if dark else "#C5D8F5"
        return (
            f"<td style='padding:6px;'>"
            f"<div style='background:{bg};border:1.5px solid {bord};"
            f"border-radius:14px;padding:14px 18px;"
            f"box-shadow:0 4px 12px rgba(10,38,71,0.08);min-width:130px;'>"
            f"<div style='font-size:11px;color:{lc};font-weight:600;"
            f"text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;'>{label}</div>"
            f"<div style='font-size:22px;font-weight:800;color:{vc};"
            f"font-family:Inter,Arial,sans-serif;'>{_fmt(value)}</div>"
            f"</div></td>"
        )

    html_body = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#EDF2FB;font-family:Calibri,Arial,sans-serif;color:#222;">
<div style="max-width:1000px;margin:20px auto;background:#fff;border-radius:14px;
     box-shadow:0 8px 24px rgba(10,38,71,0.10);overflow:hidden;">

  <!-- Title bar -->
  <div style="background:{E_HDR_BG};padding:18px 24px;">
    <div style="color:{E_HDR_FG};font-size:18px;font-weight:700;letter-spacing:-0.3px;">
      {title_line}
    </div>
    <div style="color:#ccc;font-size:12px;margin-top:4px;">
      {center_name} &nbsp;·&nbsp; {filename_orig} &nbsp;·&nbsp; Generated: {generated_at}
    </div>
  </div>

  <div style="padding:20px 24px;">

    <!-- KPI Cards -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
      <tr>
        {_kpi_card("Claimed Amount", net)}
        {_kpi_card("Total Pay", paid)}
        {_kpi_card("Under Process", bal, dark=True)}
        {_kpi_card("Final Rejn", rej)}
        {_kpi_card("Rej. Accepted", acc)}
      </tr>
    </table>

    <!-- Tables -->
    {ins_html}
    {doc_html}
    {mon_html}

    <div style="color:#999;font-size:10px;border-top:1px solid #e0e0e0;
                padding-top:10px;margin-top:8px;">
      Auto-generated by Excellent Medical Group RCM Dashboard. Excel report attached.
    </div>
  </div>
</div>
</body></html>"""

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"{title_line} — {center_name}"
    msg["From"]    = sender
    msg["To"]      = recipient
    if cc:
        msg["Cc"] = cc
    msg.attach(MIMEText(html_body, "html"))

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{excel_filename}"')
    msg.attach(part)

    all_to = [r.strip() for r in ([recipient] + ([cc] if cc else [])) if r.strip()]
    errors = []

    # Try 1: SSL on configured port (default 465) — matches mail.emc-uae.com
    try:
        import ssl
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx, timeout=20) as srv:
            srv.login(sender, password)
            srv.sendmail(sender, all_to, msg.as_string())
        return True, f"Email sent to {recipient}"
    except Exception as e1:
        errors.append(f"{smtp_port}/SSL: {e1}")

    # Try 2: STARTTLS on port 587
    try:
        with smtplib.SMTP(smtp_host, 587, timeout=20) as srv:
            srv.ehlo(); srv.starttls(); srv.ehlo()
            srv.login(sender, password)
            srv.sendmail(sender, all_to, msg.as_string())
        return True, f"Email sent to {recipient}"
    except Exception as e2:
        errors.append(f"587/STARTTLS: {e2}")

    # Try 3: STARTTLS on port 25
    try:
        with smtplib.SMTP(smtp_host, 25, timeout=20) as srv:
            srv.ehlo()
            try: srv.starttls(); srv.ehlo()
            except Exception: pass
            srv.login(sender, password)
            srv.sendmail(sender, all_to, msg.as_string())
        return True, f"Email sent to {recipient}"
    except Exception as e3:
        errors.append(f"25: {e3}")

    return False, " | ".join(errors)

# ─────────────────────────────────────────────────────────────────────────────
# CENTER SELECTION
# ─────────────────────────────────────────────────────────────────────────────
ck = st.session_state.get(SUM_CENTER_KEY)

if ck not in CENTERS:
    sel_year = st.session_state.get("rcm_year") or 2026
    if sel_year not in (2024, 2025, 2026):
        sel_year = 2026

    st.subheader(f"Choose a center — {sel_year}")

    # ── Year selector (only shown when no year passed via URL) ───────────────
    if not st.query_params.get("year"):
        yr_col, _ = st.columns([2, 5])
        with yr_col:
            chosen_year = st.selectbox(
                "Select Year",
                options=[2026, 2025, 2024],
                index=[2026, 2025, 2024].index(sel_year),
                key="sum_year_select",
            )
            if chosen_year != sel_year:
                st.session_state["rcm_year"] = chosen_year
                st.rerun()
        st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        if st.container(border=True).button(CENTERS["excellent"]["name"], use_container_width=True, key="sum_exc"):
            st.session_state[SUM_CENTER_KEY] = "excellent"
            st.rerun()
    with col2:
        if st.container(border=True).button(CENTERS["easyhealth"]["name"], use_container_width=True, key="sum_easy"):
            st.session_state[SUM_CENTER_KEY] = "easyhealth"
            st.rerun()

    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CENTER DETAIL + UPLOAD + RESULTS
# ─────────────────────────────────────────────────────────────────────────────
center_cfg = CENTERS[ck]

st.markdown(
    f"""
    <div style="
        background:#F5FAFF;border:1.5px solid #CFE3FF;padding:14px 18px;border-radius:16px;
        margin-bottom:10px;box-shadow:0 6px 18px rgba(11,45,92,0.08);
    ">
      <div style="font-size:24px;font-weight:900;color:#0B2D5C;">{center_cfg['name']}</div>
      <div style="font-size:13px;color:#334155;margin-top:2px;font-weight:600;">Summary Report — upload a source file to generate results</div>
    </div>
    """,
    unsafe_allow_html=True,
)

if st.button("◀ Choose another center", key="sum_back_center"):
    st.session_state[SUM_CENTER_KEY] = None
    st.rerun()

st.markdown("---")

# ── File uploader ────────────────────────────────────────────────────────────
RESULT_KEY = f"sum_result_{ck}"

up = st.file_uploader(
    f"Upload source Excel for **{center_cfg['name']}** (.xlsx)",
    type=["xlsx"],
    key=f"sum_uploader_{ck}",
    help="Upload the raw claims source file. The summary engine will process it and display results below.",
)

up_rcm = None  # placeholder — no second uploader in this version

if up is not None:
    if st.button("⚙️ Process File", use_container_width=True, key=f"sum_process_btn_{ck}"):
        with st.spinner("Processing file — please wait..."):
            try:
                file_bytes = up.read()

                source_s3_key = build_summary_s3_key(center_cfg["key"], up.name)
                ok_src, err_src = upload_bytes_to_s3(
                    file_bytes,
                    source_s3_key,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                result = run_summary_engine(file_bytes, up.name)
                result["source_s3_key"] = source_s3_key if ok_src else ""
                result["source_s3_saved"] = ok_src
                result["source_s3_error"] = err_src or ""
                result["rcm_summary_bytes"] = up_rcm.read() if up_rcm is not None else None
                result["rcm_summary_name"]  = up_rcm.name  if up_rcm is not None else None
                st.session_state[RESULT_KEY] = result
                _save_result_to_s3(result, ck)  # persist for refresh
                st.rerun()
            except Exception as e:
                st.error(f"Processing failed: {e}")
                st.session_state.pop(RESULT_KEY, None)

# ── Show results — restore from S3 if session_state is empty ────────────────
if RESULT_KEY not in st.session_state:
    cached = _load_result_from_s3(ck)
    if cached:
        st.session_state[RESULT_KEY] = cached

result = st.session_state.get(RESULT_KEY)

if result:
    st.success(f"✅ Processed **{result['filename']}** — {result['row_count']:,} rows · Generated at {result['generated_at']}")

    recon = result["recon_diff"]
    if abs(recon) > 0.01:
        st.warning(f"⚠️ Recon Diff: {recon:,.2f} — Net Amount does not fully reconcile with Paid + Balance + Rejected + Accepted.")
    else:
        st.info("✅ Reconciliation check passed — Net = Paid + Balance + Rejected + Accepted.")

    # KPI cards
    net, paid, bal, rej, acc = result["kpi"]
    render_kpi_cards(net, paid, bal, rej, acc)

    st.markdown("---")

    # Download full Excel — rebuild if loaded from S3 cache (xl_bytes not stored)
    if result.get("xl_bytes"):
        excel_bytes = result["xl_bytes"]
    else:
        excel_bytes = build_excel_output(result)
    safe_name = re.sub(r"[^\w\-.]", "_", center_cfg["key"])
    # Build clean filename: EMC - RCM SUMMARY - JAN 2025 - DEC 2025.xlsx
    def _build_excel_filename():
        try:
            import calendar as _cal
            _abbr_map = {m.lower(): i for i, m in enumerate(_cal.month_abbr) if m}
            def _mk(v):
                p = str(v).strip().split("-")
                if len(p) == 2:
                    try: return (int(p[1]), _abbr_map.get(p[0].lower()[:3], 0))
                    except: pass
                return (9999, 99)
            def _expand_full(lbl):
                p = str(lbl).strip().split("-")
                if len(p) == 2:
                    yr = int(p[1])
                    full_yr = (2000 + yr) if yr < 50 else (1900 + yr)
                    return f"{p[0].upper()} {full_yr}"
                return lbl.upper()
            rcm_m = result.get("rcm_month")
            if rcm_m is not None and not rcm_m.empty:
                gt_p_ = re.compile(r"^\s*(grand\s*total|total)\s*$", re.I)
                rows_ = [str(v).strip() for v in rcm_m.iloc[:,0] if not gt_p_.match(str(v))]
                valid_ = [r for r in rows_ if "-" in r]
                if valid_:
                    s_ = sorted(valid_, key=_mk)
                    return f"EMC - RCM SUMMARY - {_expand_full(s_[0])} - {_expand_full(s_[-1])}.xlsx"
        except Exception:
            pass
        return f"EMC - RCM SUMMARY - {datetime.now().strftime('%b %Y').upper()}.xlsx"
    dl_name = _build_excel_filename()

    report_s3_key = build_report_s3_key(center_cfg["key"])
    ok_rep, err_rep = upload_bytes_to_s3(
        excel_bytes,
        report_s3_key,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if result.get("source_s3_saved"):
        st.caption(f"Source saved to S3: {result.get('source_s3_key','')}")
    elif result.get("source_s3_error"):
        st.warning(f"Source file could not be saved to S3: {result['source_s3_error']}")

    if ok_rep:
        st.caption(f"Summary saved to S3: {report_s3_key}")
    else:
        st.warning(f"Summary report could not be saved to S3: {err_rep}")

    col_dl, col_em = st.columns(2)
    with col_dl:
        st.download_button(
            "⬇️ Download Full Summary Report (Excel)",
            data=excel_bytes,
            file_name=dl_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"sum_dl_{ck}",
        )
    with col_em:
        if st.button("📧 Send to Management", use_container_width=True, key=f"sum_email_{ck}"):
            with st.spinner("Sending email..."):
                ok, msg = send_rcm_email(
                    excel_bytes=excel_bytes,
                    excel_filename=dl_name,
                    result=result,
                    center_name=center_cfg["name"],
                )
                if ok:
                    st.success(f"✅ {msg}")
                else:
                    st.error(f"❌ Email failed: {msg}")

    # ── Claim Detail Download ─────────────────────────────────────────────────
    df_detail = result.get("claim_detail")
    if df_detail is not None and not df_detail.empty:
        detail_name = dl_name.replace(".xlsx", " - Claim Detail.xlsx")
        DETAIL_KEY  = f"sum_detail_bytes_{ck}"

        if st.button(
            f"⚙️ Prepare Claim Detail Report ({len(df_detail):,} claims)",
            use_container_width=True,
            key=f"sum_prepare_detail_{ck}",
        ):
            with st.spinner("Building claim detail Excel..."):
                st.session_state[DETAIL_KEY] = build_claim_detail_excel(df_detail)

        if DETAIL_KEY in st.session_state:
            st.download_button(
                f"⬇️ Download Claim Detail Report ({len(df_detail):,} claims)",
                data=st.session_state[DETAIL_KEY],
                file_name=detail_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"sum_dl_detail_{ck}",
            )

    st.markdown("---")

    # ── Tabs ──────────────────────────────────────────────────────────────────
    # ── 3 RCM tabs ────────────────────────────────────────────────────────────
    tabs = st.tabs([
        "🏥 Insurance",
        "👨‍⚕️ Doctor",
        "📅 Month",
    ])

    def _show_rcm(df_rcm, key):
        """Shared styled display for all 3 RCM summary views."""
        if df_rcm is None or (hasattr(df_rcm, "empty") and df_rcm.empty):
            st.info("No data available.")
            return
        gt_p = re.compile(r'^\s*(grand\s*total|total)\s*$', re.I)
        nums = df_rcm.select_dtypes(include="number").columns.tolist()
        fmt  = {c: ("{:.2f}%" if c == "Rej. %" else "{:,.2f}") for c in nums}

        def _style(row):
            if gt_p.match(str(row.iloc[0])):
                return ["background-color:#FCE4D6;font-weight:bold"] * len(row)
            styles = [""] * len(row)
            if "Rej. %" in row.index:
                try:
                    if float(row["Rej. %"]) > 0:
                        styles[list(row.index).index("Rej. %")] = "color:#C0392B;font-weight:600"
                except Exception:
                    pass
            return styles

        st.dataframe(
            df_rcm.style.apply(_style, axis=1).format(fmt),
            use_container_width=True, hide_index=True, key=key
        )

    with tabs[0]:
        st.subheader("RCM Summary — by Insurance")
        _show_rcm(result.get("rcm_insurance"), "rcm_ins")

    with tabs[1]:
        st.subheader("RCM Summary — by Doctor")
        _show_rcm(result.get("rcm_doctor"), "rcm_doc")

    with tabs[2]:
        st.subheader("RCM Summary — by Month")
        _show_rcm(result.get("rcm_month"), "rcm_mon")

else:
    st.info("👆 Upload a source Excel file above to generate the summary report.")
