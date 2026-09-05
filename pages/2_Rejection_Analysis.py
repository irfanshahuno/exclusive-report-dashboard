# pages/2_Rejection_Analysis.py

import boto3
from botocore.exceptions import ClientError
import io
import hashlib
import re
import smtplib
import ssl
from email.message import EmailMessage
from datetime import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================
# PAGE CONFIG (wide + clean)
# =========================================
st.set_page_config(page_title="RCM Denial & Recovery Intelligence Dashboard", layout="wide")

# ✅ Premium Deep Crimson + Warm Cream (CSS only)
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

    /* ---- Page Background ---- */
    .stApp{
      background: linear-gradient(145deg, #FDF6F6 0%, #FFF9F9 50%, #FFFBFB 100%) !important;
      font-family: 'Inter', sans-serif !important;
    }

    .block-container {
      max-width: 100% !important;
      padding-top: 1.0rem;
      padding-left: 1.2rem;
      padding-right: 1.2rem;
    }

    /* ---- Cards ---- */
    .card{
      background: rgba(255, 255, 255, 0.90);
      backdrop-filter: blur(10px);
      -webkit-backdrop-filter: blur(10px);
      border: 1.5px solid rgba(220, 170, 170, 0.35);
      border-left: 4px solid #9B1C1C;
      border-radius: 16px;
      padding: 14px 16px 12px 16px;
      box-shadow: 0 4px 18px rgba(155, 28, 28, 0.07), 0 1px 3px rgba(0,0,0,0.04), inset 0 1px 0 rgba(255,255,255,0.95);
      transition: box-shadow 0.2s ease, transform 0.2s ease;
    }
    .card:hover{
      box-shadow: 0 8px 28px rgba(155, 28, 28, 0.12), 0 2px 6px rgba(0,0,0,0.05);
      transform: translateY(-1px);
    }

    /* ---- Card Title ---- */
    .card-title{
      color: #9B1C1C;
      font-size: 11px;
      font-weight: 700;
      font-family: 'Inter', sans-serif;
      letter-spacing: 0.8px;
      text-transform: uppercase;
      margin-bottom: 8px;
    }

    /* ---- Card Value ---- */
    .card-value{
      color: #1A0A0A;
      font-size: 24px;
      font-weight: 900;
      font-family: 'Inter', sans-serif;
      line-height: 1.15;
      letter-spacing: -0.5px;
    }

    /* ---- Card Sub ---- */
    .card-sub{
      color: #9CA3AF;
      font-size: 11.5px;
      font-family: 'Inter', sans-serif;
      margin-top: 7px;
      font-weight: 500;
    }

    /* ---- Section Headings ---- */
    h3{
      font-size: 22px !important;
      font-weight: 800 !important;
      font-family: 'Inter', sans-serif !important;
      margin-top: 24px !important;
      margin-bottom: 10px !important;
      color: #1A0A0A !important;
      letter-spacing: -0.3px !important;
    }

    /* ---- HR Dividers ---- */
    hr{
      border: none !important;
      height: 1px !important;
      background: linear-gradient(90deg, transparent, #E8C5C5, transparent) !important;
    }

    /* ---- DataFrames ---- */
    div[data-testid="stDataFrame"] {
      border: 1px solid rgba(220,170,170,0.3);
      border-radius: 14px;
      overflow: hidden;
      box-shadow: 0 2px 10px rgba(155,28,28,0.05);
    }

    /* ---- Buttons ---- */
    div.stButton > button,
    div.stDownloadButton > button{
      background: linear-gradient(160deg, #9B1C1C 0%, #7F1D1D 100%) !important;
      color: #ffffff !important;
      border: 1px solid rgba(255,255,255,0.15) !important;
      border-radius: 12px !important;
      padding: 0.38rem 0.90rem !important;
      font-size: 0.875rem !important;
      font-weight: 700 !important;
      font-family: 'Inter', sans-serif !important;
      min-height: 2.25rem !important;
      box-shadow: 0 4px 14px rgba(127, 29, 29, 0.30), inset 0 1px 0 rgba(255,255,255,0.12) !important;
      letter-spacing: 0.1px !important;
      transition: all 0.2s ease !important;
    }

    div.stButton > button:hover,
    div.stDownloadButton > button:hover{
      background: linear-gradient(160deg, #B91C1C 0%, #9B1C1C 100%) !important;
      box-shadow: 0 6px 20px rgba(127, 29, 29, 0.38) !important;
      transform: translateY(-1px) !important;
    }

    div.stButton > button:active,
    div.stDownloadButton > button:active{
      background: linear-gradient(160deg, #7F1D1D 0%, #6B1919 100%) !important;
      transform: translateY(0px) !important;
      box-shadow: 0 2px 8px rgba(127, 29, 29, 0.25) !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# =========================================
# CONFIG
# =========================================
S3_BUCKET = "emc-rcm-storage-2026"
SOURCE_FILENAME = "source.xlsx"
DEFAULT_YEAR_OPTIONS = ["2024", "2025", "2026"]

# =========================================
# EMAIL CONFIG (reused across the dashboard — same secrets as other pages)
# =========================================
def _get_secret(key: str, default: str = "") -> str:
    try:
        return str(st.secrets.get(key, default))
    except Exception:
        return default

SMTP_HOST = _get_secret("SMTP_HOST")
SMTP_PORT = int(_get_secret("SMTP_PORT", "587") or "587")
SMTP_USER = _get_secret("SMTP_USER")
SMTP_PASSWORD = _get_secret("SMTP_PASSWORD")
SMTP_SENDER = _get_secret("SMTP_SENDER", SMTP_USER)
DEFAULT_MANAGEMENT_RECIPIENTS = _get_secret("MANAGEMENT_EMAIL_RECIPIENTS")

# ✅ Persistent cache (so results stay even after refresh / clicking again)
REJ_CACHE_PREFIX = "rejection_cache"
# Bump this whenever analytical rules change so an old cached workbook is NEVER reused.
ANALYSIS_VERSION = "2026-09-06-v8.4-owner-email-preview"
REJ_CACHE_FILENAME = f"rejection_{ANALYSIS_VERSION}.xlsx"

# =========================================
# CENTER NORMALIZATION (MUST be BEFORE use)
# =========================================
CENTER_ALIASES = {
    "excellent medical center": "excellent",
    "excellent pharmacy": "pharmacy",
    "easyhealth clinic": "easyhealth",
    "easy health medical clinic": "easyhealth",
    "easy health clinic": "easyhealth",
    "easyhealth": "easyhealth",
    "excellent": "excellent",
    "pharmacy": "pharmacy",
}

def normalize_center_for_s3(center_value: str) -> str:
    c = str(center_value).strip().lower()
    c = " ".join(c.split())
    return CENTER_ALIASES.get(c, c)

# =========================================
# S3 HELPERS
# =========================================
def s3_client():
    return boto3.client("s3")

def s3_exists(bucket: str, key: str) -> bool:
    try:
        s3_client().head_object(Bucket=bucket, Key=key)
        return True
    except ClientError:
        return False

def load_file_from_s3(bucket: str, key: str) -> bytes:
    obj = s3_client().get_object(Bucket=bucket, Key=key)
    return obj["Body"].read()

def save_file_to_s3(bucket: str, key: str, data: bytes) -> None:
    s3_client().put_object(Bucket=bucket, Key=key, Body=data)

def delete_file_from_s3(bucket: str, key: str) -> None:
    try:
        s3_client().delete_object(Bucket=bucket, Key=key)
    except Exception:
        pass

# =========================================
# REJECTION ANALYSIS ENGINE
# =========================================
def sha1_short_bytes(b: bytes) -> str:
    return hashlib.sha1(b).hexdigest()[:12]

def normalize_source_schema(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Normalize exports and select the business Status column used for analysis.

    IMPORTANT:
      - The user's rejection journey is defined by the column named `Status`.
      - If `Status` is unavailable, fall back to ActivityStatus, then CurrentActivityStatus.
      - FinalPaidAmount and FinalBalance are retained only as raw source columns; they are
        NOT used in rejection/recovery calculations.
      - Initial rejection = ActivityIns - actRemitInsShare.
    """
    info = {
        "format": "old",
        "status_source": "missing",
        "denial_source": "DenialCode",
        "paid_source": "remit/resub fields + TKBK override",
        "amount_source": "ActivityIns - actRemitInsShare",
        "analysis_version": ANALYSIS_VERSION,
    }

    # CRITICAL: use the actual business Status column first.
    if "Status" in df.columns:
        df["AnalysisStatus"] = df["Status"]
        info["status_source"] = "Status"
        info["format"] = "new"
    elif "ActivityStatus" in df.columns:
        df["AnalysisStatus"] = df["ActivityStatus"]
        info["status_source"] = "ActivityStatus"
    elif "CurrentActivityStatus" in df.columns:
        df["AnalysisStatus"] = df["CurrentActivityStatus"]
        info["status_source"] = "CurrentActivityStatus"
        info["format"] = "new"
    else:
        df["AnalysisStatus"] = ""

    # Keep ActivityStatus for backward-compatible display/export, but do not let it
    # override AnalysisStatus when a real Status column exists.
    if "ActivityStatus" not in df.columns:
        df["ActivityStatus"] = df["AnalysisStatus"]

    # Prefer the original denial code for initial rejection analysis.
    # FinalDenialCode is only a fallback when original DenialCode is blank.
    if "FinalDenialCode" in df.columns:
        info["format"] = "new"
        final_code = df["FinalDenialCode"].astype(str).fillna("").str.strip()
        final_code = final_code.mask(final_code.str.lower().isin(["nan", "none", "null"]), "")
        if "DenialCode" in df.columns:
            base = df["DenialCode"].astype(str).fillna("").str.strip()
            base = base.mask(base.str.lower().isin(["nan", "none", "null"]), "")
            df["DenialCode"] = base.where(base.ne(""), final_code)
            info["denial_source"] = "DenialCode (fallback FinalDenialCode)"
        else:
            df["DenialCode"] = final_code
            info["denial_source"] = "FinalDenialCode fallback"

    if "FinalPaidAmount" in df.columns or "FinalBalance" in df.columns:
        info["format"] = "new"

    return df, info


def ensure_numeric(df: pd.DataFrame) -> pd.DataFrame:
    required_num_cols = [
        "ActivityIns",
        "actRemitInsShare", "actResub1RemitInsShare",
        "actResub2RemitInsShare", "actResub3RemitInsShare",
        "TKBKAmountAct",
    ]
    for c in required_num_cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # New-export numeric fields are optional; do not create them for old files,
    # because their presence is used to detect which calculation method to use.
    for c in ["FinalPaidAmount", "FinalBalance"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df


def _distinct_positive_sum(values) -> float:
    """Sum positive amounts once; exact duplicate amounts are counted only once."""
    seen = set()
    total = 0.0
    for value in values:
        try:
            v = float(value)
        except Exception:
            v = 0.0
        if v <= 0:
            continue
        # Currency-safe key so 21.4 and 21.40 are treated as the same amount.
        key = round(v, 2)
        if key not in seen:
            seen.add(key)
            total += v
    return total


def _distinct_positive_sum_vectorized(v1: pd.Series, v2: pd.Series, v3: pd.Series) -> pd.Series:
    """Vectorized equivalent of calling _distinct_positive_sum([v1, v2, v3]) row by row:
    sum positive amounts, counting an exact duplicate (rounded to cents) only once,
    checked in order v1 -> v2 -> v3 exactly like the original per-row loop.

    This replaces what used to be a Python-level df.apply(axis=1) call — the single
    biggest reason the page was slow/crashing on large exports (a row-wise Python
    loop over every activity, sometimes hundreds of thousands of rows).
    """
    a1 = v1.clip(lower=0)
    a2 = v2.clip(lower=0)
    a3 = v3.clip(lower=0)
    k1, k2, k3 = a1.round(2), a2.round(2), a3.round(2)

    inc1 = a1 > 0
    inc2 = (a2 > 0) & ~(inc1 & (k2 == k1))
    inc3 = (a3 > 0) & ~((inc1 & (k3 == k1)) | (inc2 & (k3 == k2)))

    return a1.where(inc1, 0) + a2.where(inc2, 0) + a3.where(inc3, 0)


def compute_paid(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate Paid without using FinalPaidAmount. Fully vectorized (no per-row
    Python loop) so it stays fast regardless of how many rows the export has.

    Normal calculation:
      initial remit + distinct positive resub remit amounts.

    Special TKBK rule supplied by the user:
      when TKBKAmountAct contains a non-zero amount, it OVERRIDES the calculated
      paid amount and Paid becomes ABS(TKBKAmountAct) exactly.
    """
    initial = pd.to_numeric(df["actRemitInsShare"], errors="coerce").fillna(0)
    r1 = pd.to_numeric(df["actResub1RemitInsShare"], errors="coerce").fillna(0)
    r2 = pd.to_numeric(df["actResub2RemitInsShare"], errors="coerce").fillna(0)
    r3 = pd.to_numeric(df["actResub3RemitInsShare"], errors="coerce").fillna(0)
    takeback = pd.to_numeric(df["TKBKAmountAct"], errors="coerce").fillna(0)

    resub_total = _distinct_positive_sum_vectorized(r1, r2, r3)
    normal_paid = initial.clip(lower=0) + resub_total

    df["Paid"] = takeback.abs().where(takeback != 0, normal_paid)
    df["InitialPaid"] = initial.clip(lower=0)
    return df


def _status_parts(value) -> tuple[str, int]:
    """Return (canonical_base_status, resub_stage) using tolerant matching.

    Billing exports can vary only in formatting, for example:
      Approved(Resub- 1)
      Approved (Resub -1)
      Approved(Resub 1)
      Approved ( RESUB-01 )

    We normalize those formatting differences WITHOUT broadening the business
    rule to unrelated statuses.
    """
    raw = str(value or "").strip().lower()
    if raw in {"", "nan", "none", "null", "<na>"}:
        return "", 0

    # Normalize common punctuation/spacing variants but retain the words.
    s = raw.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s).strip()

    # Stage is taken from any explicit Resub marker; accepts resub-1, resub 1,
    # resub - 01, etc. Only stages 1-3 are valid for this analysis.
    stage = 0
    m_stage = re.search(r"\bresub\b\s*[-:]?\s*0*([123])\b", s, flags=re.IGNORECASE)
    if m_stage:
        stage = int(m_stage.group(1))

    # Remove the resub suffix/parenthetical only for identifying the base status.
    base_text = re.sub(r"\(\s*resub\b[^)]*\)", "", s, flags=re.IGNORECASE).strip()
    base_text = re.sub(r"\bresub\b\s*[-:]?\s*0*[123]\b", "", base_text, flags=re.IGNORECASE).strip(" -:()")
    base_text = re.sub(r"\s+", " ", base_text).strip()

    # Order matters: 'not submitted' before 'submitted'; 'rejection accepted'
    # before 'rejected'. These are the exact business-status families agreed.
    if re.fullmatch(r"rejection\s+accepted", base_text):
        base = "rejection accepted"
    elif re.fullmatch(r"not\s+submitted", base_text):
        base = "not submitted"
    elif re.fullmatch(r"approved", base_text):
        base = "approved"
    elif re.fullmatch(r"rejected", base_text):
        base = "rejected"
    elif re.fullmatch(r"submitted", base_text):
        base = "submitted"
    else:
        return base_text, stage

    return base, stage


def _is_initial_rejection_status(value) -> bool:
    """Exact analytical status logic agreed with the user."""
    base, stage = _status_parts(value)
    if base in {"rejected", "rejection accepted"}:
        return True
    if stage >= 1 and base in {"approved", "submitted", "not submitted"}:
        return True
    return False


def attach_status_parts(df: pd.DataFrame) -> pd.DataFrame:
    """Parse AnalysisStatus into StatusBaseRaw / ResubStage / IsInitialRejection
    ONCE PER DISTINCT STATUS VALUE, then map that back onto every row.

    A billing export can have hundreds of thousands of rows but almost always
    only a few dozen distinct status strings (e.g. "Rejected", "Approved(Resub-1)").
    The regex parsing in _status_parts is the same cost either way, so doing it
    per unique value instead of per row turns an O(rows) regex workload into an
    O(unique statuses) one — this was a major contributor to the slow/crashing
    behavior on large files. Downstream functions (build_rejected_df,
    build_status_audit, build_recovery_detail) now reuse these columns instead
    of re-parsing.
    """
    status_col = df["AnalysisStatus"].astype(str)
    unique_vals = status_col.unique()
    parts_map = {v: _status_parts(v) for v in unique_vals}
    included_map = {v: _is_initial_rejection_status(v) for v in unique_vals}

    df["StatusBaseRaw"] = status_col.map({v: p[0] for v, p in parts_map.items()})
    df["ResubStage"] = status_col.map({v: p[1] for v, p in parts_map.items()}).astype(int)
    df["IsInitialRejection"] = status_col.map(included_map).astype(bool)
    return df


def _recovery_from_resub_fields(row, stage: int) -> float:
    """Recovery after rejection from resub remit fields.

    Resub-1 -> actResub1RemitInsShare
    Resub-2 -> Resub1 + Resub2, but duplicate equal amounts count once
    Resub-3 -> Resub1 + Resub2 + Resub3, duplicate equal amounts count once
    TKBKAmountAct rule -> if present/non-zero, use its absolute exact value instead.
    """
    takeback = float(row.get("TKBKAmountAct", 0) or 0)
    if takeback != 0:
        return abs(takeback)
    if stage <= 0:
        return 0.0
    vals = []
    if stage >= 1:
        vals.append(row.get("actResub1RemitInsShare", 0))
    if stage >= 2:
        vals.append(row.get("actResub2RemitInsShare", 0))
    if stage >= 3:
        vals.append(row.get("actResub3RemitInsShare", 0))
    return _distinct_positive_sum(vals)

def ensure_insurance_column(df: pd.DataFrame) -> pd.DataFrame:
    insurance_col = next(
        (c for c in ["Insurance", "PayerName", "Insurer", "Plan"] if c in df.columns),
        "Insurance",
    )
    if insurance_col not in df.columns:
        df["Insurance"] = "Not Available"
    elif insurance_col != "Insurance":
        df["Insurance"] = df[insurance_col]
    df["Insurance"] = df["Insurance"].astype(str).fillna("").str.strip()
    df.loc[df["Insurance"].eq(""), "Insurance"] = "Not Available"
    return df


def add_refdate_and_aging(df: pd.DataFrame) -> pd.DataFrame:
    # Both supplied exports use SubDate. Keep legacy alternatives as fallbacks.
    date_candidates = [
        c for c in ["SubDate", "SubmissionDate", "ClaimDate", "VisitDate"]
        if c in df.columns
    ]
    if date_candidates:
        for c in date_candidates:
            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
        df["RefDate"] = df[date_candidates].bfill(axis=1).iloc[:, 0]
    else:
        df["RefDate"] = pd.NaT

    today = pd.Timestamp(dt.today().date())
    df["DaysDiff"] = (today - df["RefDate"]).dt.days

    bins = [-1, 30, 45, 60, 90, float("inf")]
    labels = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]
    df["AgingBucket"] = pd.cut(df["DaysDiff"], bins=bins, labels=labels)
    return df


def normalize_denial_code(df: pd.DataFrame) -> pd.DataFrame:
    if "DenialCode" not in df.columns:
        df["DenialCode"] = ""
    df["DenialCode"] = df["DenialCode"].astype(str).fillna("").str.strip()
    df.loc[df["DenialCode"].str.lower().isin(["nan", "none", "null"]), "DenialCode"] = ""
    return df


def split_denial_code_levels(df: pd.DataFrame) -> pd.DataFrame:
    """Split the denial code into a level-2 category and the full level-3 code.

    Payer denial codes are typically written as CATEGORY-NUMBER, e.g. MNEC-006.
    Management reporting should show the granular level-3 code (MNEC-006), not
    just the level-2 category (MNEC) — this preserves both:
        DenialCodeLevel2 -> "MNEC"       (category, for grouping)
        DenialCodeLevel3 -> "MNEC-006"   (full/granular code, for detail)
    If a code has no "-" in it, both levels are set to the same value.
    """
    code = df["DenialCode"].astype(str).fillna("").str.strip()
    df["DenialCodeLevel3"] = code
    level2 = code.str.split("-").str[0].str.strip()
    df["DenialCodeLevel2"] = level2.where(level2.ne(""), code)
    return df


def build_status_audit(df: pd.DataFrame) -> pd.DataFrame:
    """Transparent audit of every raw ActivityStatus and how the parser treats it.

    This sheet is intentionally included in the output workbook so a status
    formatting variation can never silently disappear from the rejection total.
    """
    tmp = df.copy()
    tmp["RawStatus"] = tmp["AnalysisStatus"].astype(str).fillna("").str.strip()
    # Reuse the columns attach_status_parts() already computed once per unique
    # status value — no need to re-run the regex parser over every row here.
    tmp["ParsedBase"] = tmp["StatusBaseRaw"]
    tmp["ParsedResubStage"] = tmp["ResubStage"]
    tmp["IncludedAsInitialRejection"] = tmp["IsInitialRejection"]

    activity = pd.to_numeric(tmp.get("ActivityIns", 0), errors="coerce").fillna(0)
    initial_paid = pd.to_numeric(tmp.get("actRemitInsShare", 0), errors="coerce").fillna(0)
    tmp["AnalyticalInitialRejection"] = (activity - initial_paid).clip(lower=0).round(2)

    rows = []
    for keys, g in tmp.groupby(
        ["RawStatus", "ParsedBase", "ParsedResubStage", "IncludedAsInitialRejection"],
        dropna=False,
    ):
        raw, base, stage, included = keys
        rows.append({
            "RawStatus": raw,
            "ParsedBase": base,
            "ParsedResubStage": int(stage or 0),
            "IncludedAsInitialRejection": bool(included),
            "ActivityRows": int(len(g)),
            "ActivityIns": pd.to_numeric(g.get("ActivityIns", 0), errors="coerce").fillna(0).sum(),
            "InitialPaid": pd.to_numeric(g.get("actRemitInsShare", 0), errors="coerce").fillna(0).sum(),
            "AnalyticalInitialRejection": pd.to_numeric(g["AnalyticalInitialRejection"], errors="coerce").fillna(0).sum(),
        })
    if not rows:
        return pd.DataFrame(columns=[
            "RawStatus", "ParsedBase", "ParsedResubStage", "IncludedAsInitialRejection",
            "ActivityRows", "ActivityIns", "InitialPaid", "AnalyticalInitialRejection"
        ])
    return pd.DataFrame(rows).sort_values(
        ["IncludedAsInitialRejection", "AnalyticalInitialRejection"], ascending=[False, False]
    ).reset_index(drop=True)


def build_rejected_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build the INITIAL rejection population from the agreed Status values.

    RejectedAmount is analytical and always:
        ActivityIns - actRemitInsShare

    It does NOT depend on final/current payment status, FinalPaidAmount or FinalBalance.
    """
    status_mask = df["IsInitialRejection"]
    rej = df.loc[status_mask].copy()

    activity_ins = pd.to_numeric(rej["ActivityIns"], errors="coerce").fillna(0)
    initial_paid = pd.to_numeric(rej["actRemitInsShare"], errors="coerce").fillna(0)
    rej["RejectedAmount"] = (activity_ins - initial_paid).clip(lower=0).round(2)

    # Status journey fields used throughout the dashboard (parsed once per
    # unique status value in attach_status_parts(); ResubStage already present).
    rej["StatusBase"] = rej["StatusBaseRaw"].str.title()

    # Remove rows where the analytical rejection amount is zero.
    rej = rej.loc[rej["RejectedAmount"] > 0].copy()

    # Keep every rejected activity row for amount/detail analysis.
    # Claim counting remains deduplicated separately by UniqueID.
    return rej

def _unique_claim_count(series: pd.Series) -> int:
    """Count unique claim IDs without changing any amount calculation.

    Blank/missing UniqueID values are counted row-by-row instead of being
    collapsed into one claim, so we never undercount because of missing IDs.
    """
    s = series.astype("string").str.strip()
    valid = s.notna() & (~s.str.lower().isin(["", "nan", "none", "null", "<na>"]))
    return int(s[valid].nunique(dropna=True) + (~valid).sum())

def pivot_by_insurance(rej: pd.DataFrame) -> pd.DataFrame:
    # IMPORTANT: RejectedAmount remains the SUM of all rejected activities.
    # Only RejectedCount is deduplicated by claim UniqueID.
    if "UniqueID" in rej.columns:
        out = (
            rej.groupby("Insurance", dropna=False)
               .agg(
                   RejectedAmount=("RejectedAmount", "sum"),
                   RejectedCount=("UniqueID", _unique_claim_count),
               )
               .reset_index()
               .sort_values("RejectedAmount", ascending=False)
        )
        grand_count = _unique_claim_count(rej["UniqueID"])
    else:
        # Fallback for an unexpected export without UniqueID.
        out = (
            rej.groupby("Insurance", dropna=False)
               .agg(
                   RejectedAmount=("RejectedAmount", "sum"),
                   RejectedCount=("RejectedAmount", "size"),
               )
               .reset_index()
               .sort_values("RejectedAmount", ascending=False)
        )
        grand_count = int(len(rej))

    total_row = {
        "Insurance": "Grand Total",
        "RejectedAmount": out["RejectedAmount"].sum(),
        "RejectedCount": grand_count,
    }
    return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

def pivot_by_denialcode(rej: pd.DataFrame) -> pd.DataFrame:
    # Same principle here: amounts are untouched; claim count is UniqueID-based.
    if "UniqueID" in rej.columns:
        out = (
            rej.groupby("DenialCode", dropna=False)
               .agg(
                   RejectedAmount=("RejectedAmount", "sum"),
                   RejectedCount=("UniqueID", _unique_claim_count),
               )
               .reset_index()
               .sort_values("RejectedAmount", ascending=False)
        )
        grand_count = _unique_claim_count(rej["UniqueID"])
    else:
        out = (
            rej.groupby("DenialCode", dropna=False)
               .agg(
                   RejectedAmount=("RejectedAmount", "sum"),
                   RejectedCount=("RejectedAmount", "size"),
               )
               .reset_index()
               .sort_values("RejectedAmount", ascending=False)
        )
        grand_count = int(len(rej))

    total_row = {
        "DenialCode": "Grand Total",
        "RejectedAmount": out["RejectedAmount"].sum(),
        "RejectedCount": grand_count,
    }
    return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

def pivot_insurance_x_denialcode(rej: pd.DataFrame) -> pd.DataFrame:
    pv = pd.pivot_table(
        rej,
        index="Insurance",
        columns="DenialCode",
        values="RejectedAmount",
        aggfunc="sum",
        fill_value=0,
        observed=False,
    )
    pv["Grand Total"] = pv.sum(axis=1)
    pv.loc["Grand Total"] = pv.sum(axis=0)
    pv.reset_index(inplace=True)
    return pv

def pivot_rejection_aging(rej: pd.DataFrame) -> pd.DataFrame:
    labels = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]
    pv = pd.pivot_table(
        rej,
        index="Insurance",
        columns="AgingBucket",
        values="RejectedAmount",
        aggfunc="sum",
        fill_value=0,
        observed=False,
    ).reindex(columns=labels)
    pv["Grand Total"] = pv.sum(axis=1)
    pv.loc["Grand Total"] = pv.sum(axis=0)
    pv.reset_index(inplace=True)
    return pv


# =========================================
# REJECTION RECOVERY / RESUBMISSION TRACKER
# =========================================
def build_recovery_detail(df: pd.DataFrame) -> pd.DataFrame:
    """Trace the rejection journey using AnalysisStatus (Status column first).

    OriginalRejectedAmount = ActivityIns - actRemitInsShare.
    RecoveredAmount = actual additional resubmission payment only.
    The current journey bucket is determined by Status, not by payment amount.
    """
    hist = build_rejected_df(df)
    if hist.empty:
        return hist

    hist["CurrentStatus"] = hist["AnalysisStatus"].astype(str).fillna("").str.strip()
    hist["OriginalRejectedAmount"] = pd.to_numeric(hist["RejectedAmount"], errors="coerce").fillna(0)

    # Actual additional payment received after initial rejection (vectorized —
    # this used to be 5 separate row-wise .apply(axis=1) passes over `hist`,
    # which is a major reason large exports were slow/crashing).
    stage = hist["ResubStage"].fillna(0).astype(int)
    takeback = pd.to_numeric(hist["TKBKAmountAct"], errors="coerce").fillna(0)
    r1 = pd.to_numeric(hist["actResub1RemitInsShare"], errors="coerce").fillna(0)
    r2 = pd.to_numeric(hist["actResub2RemitInsShare"], errors="coerce").fillna(0)
    r3 = pd.to_numeric(hist["actResub3RemitInsShare"], errors="coerce").fillna(0)

    v1 = r1.where(stage >= 1, 0.0)
    v2 = r2.where(stage >= 2, 0.0)
    v3 = r3.where(stage >= 3, 0.0)
    resub_total = _distinct_positive_sum_vectorized(v1, v2, v3)

    hist["RecoveredAmount"] = takeback.abs().where(takeback != 0, resub_total)
    hist["RecoveredAmount"] = hist[["RecoveredAmount", "OriginalRejectedAmount"]].min(axis=1).clip(lower=0).round(2)
    hist["OutstandingAmount"] = (
        hist["OriginalRejectedAmount"] - hist["RecoveredAmount"]
    ).clip(lower=0).round(2)

    # Payment status for management reporting: distinguishes a claim that
    # received SOME money back (Partially Paid) from one that recovered
    # nothing at all (Fully Rejected), rather than lumping both under one
    # denial reason. A 1-fils/cent tolerance avoids float rounding noise.
    orig_amt = hist["OriginalRejectedAmount"]
    rec_amt = hist["RecoveredAmount"]
    payment_status = pd.Series("Fully Rejected", index=hist.index)
    payment_status = payment_status.mask(rec_amt > 0.01, "Partially Paid")
    payment_status = payment_status.mask(rec_amt >= (orig_amt - 0.01), "Fully Recovered")
    hist["PaymentStatus"] = payment_status

    # Amount that reached each resubmission stage. These are journey amounts, not payments.
    hist["Resub1Amount"] = hist["OriginalRejectedAmount"].where(stage >= 1, 0.0)
    hist["Resub2Amount"] = hist["OriginalRejectedAmount"].where(stage >= 2, 0.0)
    hist["Resub3Amount"] = hist["OriginalRejectedAmount"].where(stage >= 3, 0.0)

    base_lower = hist["StatusBase"].astype(str).str.strip().str.lower()
    stage_str = stage.astype(str)

    bucket = pd.Series("Other Initial Rejection", index=hist.index)
    bucket = bucket.mask((base_lower == "approved") & (stage >= 1), "Approved (Resub-" + stage_str + ")")
    bucket = bucket.mask((base_lower == "submitted") & (stage >= 1), "Submitted / Pending (Resub-" + stage_str + ")")
    bucket = bucket.mask((base_lower == "rejected") & (stage >= 1), "Still Rejected (Resub-" + stage_str + ")")
    bucket = bucket.mask((base_lower == "rejected") & (stage < 1), "Still Rejected (Initial)")
    bucket = bucket.mask((base_lower == "rejection accepted") & (stage >= 1), "Rejection Accepted (Resub-" + stage_str + ")")
    bucket = bucket.mask((base_lower == "rejection accepted") & (stage < 1), "Rejection Accepted")
    bucket = bucket.mask((base_lower == "not submitted") & (stage >= 1), "Not Submitted (Resub-" + stage_str + ")")
    hist["RecoveryBucket"] = bucket

    # Reconciliation flag: still fully rejected (zero recovered) after being
    # resubmitted TWICE or more. Two resubmissions with no money back is a
    # signal the claim likely needs reconciliation/write-off review rather
    # than a third resubmission attempt.
    hist["ReconciliationEligible"] = (
        (base_lower == "rejected") & (stage >= 2) & (hist["RecoveredAmount"] <= 0.01)
    )

    # Keep all useful denial/comment fields for later management justification work.
    if "FinalDenialCode" in hist.columns:
        final_code = hist["FinalDenialCode"].astype(str).fillna("").str.strip()
        base_code = hist["DenialCode"].astype(str).fillna("").str.strip() if "DenialCode" in hist.columns else ""
        if isinstance(base_code, pd.Series):
            hist["RecoveryDenialCode"] = base_code.where(~base_code.str.lower().isin(["", "nan", "none", "null"]), final_code)
        else:
            hist["RecoveryDenialCode"] = final_code
    elif "DenialCode" in hist.columns:
        hist["RecoveryDenialCode"] = hist["DenialCode"]
    else:
        hist["RecoveryDenialCode"] = ""

    return hist


def build_recovery_summary(recovery: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "RecoveryBucket", "OriginalRejectedAmount", "RecoveredAmount",
        "OutstandingAmount", "ActivityRows", "UniqueClaims"
    ]
    if recovery.empty:
        return pd.DataFrame(columns=cols)

    rows = []
    for bucket, g in recovery.groupby("RecoveryBucket", dropna=False):
        rows.append({
            "RecoveryBucket": bucket,
            "OriginalRejectedAmount": pd.to_numeric(g["OriginalRejectedAmount"], errors="coerce").fillna(0).sum(),
            "RecoveredAmount": pd.to_numeric(g["RecoveredAmount"], errors="coerce").fillna(0).sum(),
            "OutstandingAmount": pd.to_numeric(g["OutstandingAmount"], errors="coerce").fillna(0).sum(),
            "ActivityRows": int(len(g)),
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
        })
    out = pd.DataFrame(rows).sort_values("OriginalRejectedAmount", ascending=False)
    total = {
        "RecoveryBucket": "Grand Total",
        "OriginalRejectedAmount": out["OriginalRejectedAmount"].sum(),
        "RecoveredAmount": out["RecoveredAmount"].sum(),
        "OutstandingAmount": out["OutstandingAmount"].sum(),
        "ActivityRows": int(out["ActivityRows"].sum()),
        "UniqueClaims": _unique_claim_count(recovery["UniqueID"]) if "UniqueID" in recovery.columns else int(len(recovery)),
    }
    return pd.concat([out, pd.DataFrame([total])], ignore_index=True)


def build_recovery_by_insurance(recovery: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Insurance", "OriginalRejectedAmount", "RecoveredAmount",
        "OutstandingAmount", "UniqueClaims", "RecoveryRatePct"
    ]
    if recovery.empty:
        return pd.DataFrame(columns=cols)

    rows = []
    for insurance, g in recovery.groupby("Insurance", dropna=False):
        orig = pd.to_numeric(g["OriginalRejectedAmount"], errors="coerce").fillna(0).sum()
        rec = pd.to_numeric(g["RecoveredAmount"], errors="coerce").fillna(0).sum()
        outst = pd.to_numeric(g["OutstandingAmount"], errors="coerce").fillna(0).sum()
        rows.append({
            "Insurance": insurance,
            "OriginalRejectedAmount": orig,
            "RecoveredAmount": rec,
            "OutstandingAmount": outst,
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
            "RecoveryRatePct": (rec / orig * 100) if orig > 0 else 0,
        })
    out = pd.DataFrame(rows).sort_values("OriginalRejectedAmount", ascending=False)
    orig = out["OriginalRejectedAmount"].sum()
    rec = out["RecoveredAmount"].sum()
    total = {
        "Insurance": "Grand Total",
        "OriginalRejectedAmount": orig,
        "RecoveredAmount": rec,
        "OutstandingAmount": out["OutstandingAmount"].sum(),
        "UniqueClaims": _unique_claim_count(recovery["UniqueID"]) if "UniqueID" in recovery.columns else int(len(recovery)),
        "RecoveryRatePct": (rec / orig * 100) if orig > 0 else 0,
    }
    return pd.concat([out, pd.DataFrame([total])], ignore_index=True)


def build_resub_stage_summary(recovery: pd.DataFrame) -> pd.DataFrame:
    """Management view of how much initial rejection reached each resubmission stage."""
    rows = []
    if recovery.empty:
        return pd.DataFrame(columns=["Stage", "RejectedAmount", "RecoveredAmount", "OutstandingAmount", "UniqueClaims"])
    for stage in [1, 2, 3]:
        g = recovery[pd.to_numeric(recovery["ResubStage"], errors="coerce").fillna(0) >= stage].copy()
        rejected = pd.to_numeric(g["OriginalRejectedAmount"], errors="coerce").fillna(0).sum()
        recovered = pd.to_numeric(g["RecoveredAmount"], errors="coerce").fillna(0).sum()
        rows.append({
            "Stage": f"Resub-{stage}",
            "RejectedAmount": rejected,
            "RecoveredAmount": recovered,
            "OutstandingAmount": max(rejected - recovered, 0),
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
        })
    return pd.DataFrame(rows)


def build_resub_stage_summary_exclusive(recovery: pd.DataFrame) -> pd.DataFrame:
    """Non-cumulative version of the stage funnel: each claim is counted in
    EXACTLY ONE row (the highest stage it currently sits at), so the rows
    add up to the Total Rejected Amount instead of overlapping.

    Use this table when you need the numbers to sum correctly; use
    Resub_Stage_Summary (cumulative funnel) when you want "how much reached
    at least this stage."
    """
    cols = ["Stage", "RejectedAmount", "RecoveredAmount", "OutstandingAmount", "UniqueClaims"]
    if recovery.empty:
        return pd.DataFrame(columns=cols)

    stage = pd.to_numeric(recovery["ResubStage"], errors="coerce").fillna(0).astype(int)
    labels = {0: "Initial (Never Resubmitted)", 1: "Resub-1", 2: "Resub-2", 3: "Resub-3"}

    rows = []
    for s in [0, 1, 2, 3]:
        g = recovery.loc[stage == s]
        rows.append({
            "Stage": labels[s],
            "RejectedAmount": pd.to_numeric(g["OriginalRejectedAmount"], errors="coerce").fillna(0).sum(),
            "RecoveredAmount": pd.to_numeric(g["RecoveredAmount"], errors="coerce").fillna(0).sum(),
            "OutstandingAmount": pd.to_numeric(g["OutstandingAmount"], errors="coerce").fillna(0).sum(),
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
        })
    out = pd.DataFrame(rows)
    total = {
        "Stage": "Grand Total",
        "RejectedAmount": out["RejectedAmount"].sum(),
        "RecoveredAmount": out["RecoveredAmount"].sum(),
        "OutstandingAmount": out["OutstandingAmount"].sum(),
        "UniqueClaims": _unique_claim_count(recovery["UniqueID"]) if "UniqueID" in recovery.columns else int(len(recovery)),
    }
    return pd.concat([out, pd.DataFrame([total])], ignore_index=True)


def build_denial_management_summary(recovery: pd.DataFrame) -> pd.DataFrame:
    """Insurance + granular denial code + payment status, for management.

    Shows the level-3 (granular) denial code alongside the level-2 category,
    and splits amounts by PaymentStatus so management can see, for example,
    how much of MNEC-006 for a payer was fully rejected vs partially paid —
    instead of one blended number under the level-2 category.
    """
    cols = [
        "Insurance", "DenialCodeLevel2", "DenialCodeLevel3", "PaymentStatus",
        "OriginalRejectedAmount", "RecoveredAmount", "OutstandingAmount", "UniqueClaims"
    ]
    if recovery.empty or "DenialCodeLevel3" not in recovery.columns:
        return pd.DataFrame(columns=cols)

    group_cols = ["Insurance", "DenialCodeLevel2", "DenialCodeLevel3", "PaymentStatus"]
    rows = []
    for keys, g in recovery.groupby(group_cols, dropna=False):
        ins, l2, l3, pstatus = keys
        rows.append({
            "Insurance": ins,
            "DenialCodeLevel2": l2,
            "DenialCodeLevel3": l3,
            "PaymentStatus": pstatus,
            "OriginalRejectedAmount": pd.to_numeric(g["OriginalRejectedAmount"], errors="coerce").fillna(0).sum(),
            "RecoveredAmount": pd.to_numeric(g["RecoveredAmount"], errors="coerce").fillna(0).sum(),
            "OutstandingAmount": pd.to_numeric(g["OutstandingAmount"], errors="coerce").fillna(0).sum(),
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
        })
    return pd.DataFrame(rows).sort_values("OriginalRejectedAmount", ascending=False).reset_index(drop=True)


def build_reconciliation_candidates(recovery: pd.DataFrame) -> pd.DataFrame:
    """Claim-level list of activities flagged ReconciliationEligible:
    resubmitted 2+ times with zero money recovered. This is the working
    list for the reconciliation team, not a summary."""
    if recovery.empty or "ReconciliationEligible" not in recovery.columns:
        return pd.DataFrame()

    cand = recovery.loc[recovery["ReconciliationEligible"]].copy()
    wanted = [
        "UniqueID", "Insurance", "VisitNo", "VisitDate", "Code", "Description",
        "DenialCodeLevel2", "DenialCodeLevel3", "RecoveryDenialCode",
        "CurrentStatus", "RecoveryBucket", "PaymentStatus",
        "OriginalRejectedAmount", "RecoveredAmount", "OutstandingAmount",
        "ResubStage", "RefDate", "DaysDiff",
        "Resub1Comments/Accpt Comments",
    ]
    cols = [c for c in wanted if c in cand.columns]
    return cand[cols].sort_values("OriginalRejectedAmount", ascending=False).reset_index(drop=True)




# =========================================
# BUSINESS DISPOSITION / RECOVERABILITY RULE ENGINE
# =========================================
# Rule priority:
#   1) insurance-specific rule
#   2) global denial rule
#   3) current status / generic reconciliation logic
#   4) Needs Review fallback
#
# IMPORTANT: rules classify ONLY the remaining outstanding balance. Money already
# recovered remains Recovered / Paid even if the original denial reason is normally
# non-recoverable.

def _norm_rule_text(value) -> str:
    s = str(value or "").strip().upper()
    if s in {"NAN", "NONE", "NULL", "<NA>"}:
        return ""
    return re.sub(r"\s+", " ", s)


def _insurance_family(value) -> str:
    """Return a stable payer family keyword using tolerant substring matching."""
    s = _norm_rule_text(value)
    if "AAFIYA" in s:
        return "AAF IYA".replace(" ", "")
    if "DAMAN" in s:
        return "DAMAN"
    if "INAYAH" in s:
        return "INAYAH"
    # NAS can appear as NAS ADMINISTRATION / NAS TPA etc. Avoid matching random words.
    if re.search(r"(^|[^A-Z])NAS([^A-Z]|$)", s) or s.startswith("NAS "):
        return "NAS"
    if "OMAN" in s:
        return "OMAN"
    if re.search(r"(^|[^A-Z])FMC([^A-Z]|$)", s) or s.startswith("FMC "):
        return "FMC"
    return s


def apply_disposition_rules(recovery: pd.DataFrame) -> pd.DataFrame:
    """Attach auditable business-rule classification to every rejected activity.

    Added columns:
      RuleDisposition     - Non-Recoverable / Recoverable / Reconciliation Eligible /
                            Needs Review / Unclassified
      ManagementReason    - plain-language reason for management
      RecommendedAction   - operational next step
      RuleApplied         - exact rule identifier for auditability
      RuleSource          - Insurance Specific / Global / Fallback
      RulePriority        - numeric priority (1 = insurance-specific, 2 = global, 9 = fallback)
      RuleConflict        - reserved flag; True only if future overlapping rules are detected
      ManagementOutstandingBucket - current financial position of remaining balance
    """
    if recovery.empty:
        return recovery.copy()

    d = recovery.copy()
    idx = d.index

    denial_col = "DenialCodeLevel3" if "DenialCodeLevel3" in d.columns else (
        "RecoveryDenialCode" if "RecoveryDenialCode" in d.columns else "DenialCode"
    )
    denial = d.get(denial_col, pd.Series("", index=idx)).map(_norm_rule_text)
    insurance = d.get("Insurance", pd.Series("", index=idx)).map(_insurance_family)
    cpt = d.get("Code", pd.Series("", index=idx)).map(_norm_rule_text)
    base = d.get("StatusBase", pd.Series("", index=idx)).astype(str).str.strip().str.lower()
    stage = pd.to_numeric(d.get("ResubStage", 0), errors="coerce").fillna(0).astype(int)
    generic_recon = d.get("ReconciliationEligible", pd.Series(False, index=idx)).fillna(False).astype(bool)

    d["RuleDisposition"] = "Unclassified"
    d["ManagementReason"] = "Unclassified Denial"
    d["RecommendedAction"] = "Review and classify"
    d["RuleApplied"] = "FALLBACK-UNCLASSIFIED"
    d["RuleSource"] = "Fallback"
    d["RulePriority"] = 9
    d["RuleConflict"] = False

    def set_rule(mask, disposition, reason, action, rule_name, source="Insurance Specific", priority=1):
        # A higher-priority rule wins. This keeps future exceptions auditable.
        eligible = mask & (pd.to_numeric(d["RulePriority"], errors="coerce").fillna(99) >= priority)
        d.loc[eligible, "RuleDisposition"] = disposition
        d.loc[eligible, "ManagementReason"] = reason
        d.loc[eligible, "RecommendedAction"] = action
        d.loc[eligible, "RuleApplied"] = rule_name
        d.loc[eligible, "RuleSource"] = source
        d.loc[eligible, "RulePriority"] = priority

    # ---------------- INSURANCE-SPECIFIC RULES (priority 1) ----------------
    # Aafiya
    is_aafiya = insurance.eq("AAFIYA")
    set_rule(is_aafiya & denial.eq("MNEC-003") & cpt.isin(["99213", "99203"]),
             "Non-Recoverable", "E&M Level 3 Adjustment", "Adjust / no further recovery",
             "AAF IYA:MNEC-003:99213/99203".replace(" ", ""))
    set_rule(is_aafiya & denial.eq("MNEC-006"),
             "Non-Recoverable", "E&M Level 3 Adjustment", "Adjust / no further recovery",
             "AAF IYA:MNEC-006:LEVEL3".replace(" ", ""))

    # Daman
    is_daman = insurance.eq("DAMAN")
    for code in ["AUTH-001", "AUTH-003", "ELIG-006"]:
        set_rule(is_daman & denial.eq(code), "Needs Review", "Daman System Glitch",
                 "Review payer/system issue", f"DAMAN:{code}:SYSTEM-GLITCH")
    set_rule(is_daman & denial.eq("AUTH-007"), "Needs Review", "Daman System Glitch",
             "Review resubmission / payer response", "DAMAN:AUTH-007:SYSTEM-GLITCH")
    set_rule(is_daman & denial.eq("CLAI-012"), "Recoverable", "New vs Established Patient",
             "Correct and resubmit", "DAMAN:CLAI-012:NEW-ESTABLISHED")
    set_rule(is_daman & denial.eq("CLAI-016"), "Recoverable", "Modifier / Malaffi Issue",
             "Correct supporting issue and resubmit", "DAMAN:CLAI-016:MODIFIER-MALAFI")
    set_rule(is_daman & denial.eq("ELIG-007"), "Recoverable", "Daman System Glitch",
             "Resubmit / follow payer", "DAMAN:ELIG-007:SYSTEM-GLITCH")
    set_rule(is_daman & denial.eq("PRCE-010"), "Non-Recoverable", "Contractual / Bundled Service",
             "Adjust / no further recovery", "DAMAN:PRCE-010:BUNDLED")
    set_rule(is_daman & denial.eq("NCOV-003"), "Non-Recoverable", "Non-Covered Service",
             "Adjust and review billed CPT/service", "DAMAN:NCOV-003:NON-COVERED")
    for code in ["MNEC-004", "MNEC-005"]:
        set_rule(is_daman & denial.eq(code), "Needs Review", "Medical Necessity",
                 "Clinical/coding review for recovery", f"DAMAN:{code}:MEDICAL-NECESSITY")

    # FMC
    is_fmc = insurance.eq("FMC")
    set_rule(is_fmc & denial.eq("MNEC-006"), "Non-Recoverable", "E&M Level 3 Adjustment",
             "Adjust / no further recovery", "FMC:MNEC-006:LEVEL3")

    # Inayah
    is_inayah = insurance.eq("INAYAH")
    set_rule(is_inayah & denial.eq("CODE-012"), "Needs Review", "Coding Issue – Potentially Recoverable",
             "Review for correction / recovery", "INAYAH:CODE-012:REVIEW")
    for code in ["MNEC-003", "MNEC-004"]:
        set_rule(is_inayah & denial.eq(code), "Needs Review", "Medical Necessity – Potentially Recoverable",
                 "Review for recovery", f"INAYAH:{code}:REVIEW")

    # NAS
    is_nas = insurance.eq("NAS")
    set_rule(is_nas & denial.eq("CLAI-016"), "Non-Recoverable", "E&M Level 3 Adjustment",
             "Adjust / no further recovery", "NAS:CLAI-016:LEVEL3")
    for code in ["MNEC-003", "MNEC-004", "MNEC-006"]:
        set_rule(is_nas & denial.eq(code), "Reconciliation Eligible", "Service Payment Dispute",
                 "Payer reconciliation / escalation", f"NAS:{code}:RECONCILIATION")

    # Oman
    is_oman = insurance.eq("OMAN")
    set_rule(is_oman & denial.eq("MNEC-006"), "Non-Recoverable", "E&M Level 3 Adjustment",
             "Adjust / no further recovery", "OMAN:MNEC-006:LEVEL3")

    # ---------------- GLOBAL RULES (priority 2) ----------------
    # These apply to all insurers unless a payer-specific exception is defined above.
    set_rule(denial.eq("COPY-001"), "Non-Recoverable", "Copay / Patient Responsibility",
             "Patient responsibility / adjust insurer receivable", "GLOBAL:COPY-001:COPAY", "Global", 2)
    for code in ["PRCE-001", "PRCE-002"]:
        set_rule(denial.eq(code), "Non-Recoverable", "Contractual / Price Adjustment",
                 "Contractual adjustment", f"GLOBAL:{code}:CONTRACTUAL", "Global", 2)
    set_rule(denial.eq("PRCE-006"), "Non-Recoverable", "Follow-up Consultation Adjustment",
             "Contractual follow-up adjustment", "GLOBAL:PRCE-006:FOLLOW-UP", "Global", 2)
    for code in ["NCOV-001", "NCOV-0026"]:
        set_rule(denial.eq(code), "Non-Recoverable", "Non-Covered Service",
                 "Adjust and review billed CPT/service", f"GLOBAL:{code}:NON-COVERED", "Global", 2)
    set_rule(denial.str.startswith("TIME-", na=False), "Non-Recoverable", "Time / Timely Filing Issue",
             "No further recovery / adjust", "GLOBAL:TIME-*:TIME-ISSUE", "Global", 2)

    # Current financial position of the OUTSTANDING portion.
    bucket = pd.Series("Needs Review", index=idx, dtype="object")
    disposition = d["RuleDisposition"].astype(str)

    # Explicit business disposition wins over generic status.
    bucket = bucket.mask(disposition.eq("Non-Recoverable"), "Non-Recoverable / Adjustment")
    bucket = bucket.mask(disposition.eq("Reconciliation Eligible"), "Reconciliation / Escalation")

    # Explicit recoverable items: pending if already with payer, otherwise action required.
    recoverable = disposition.eq("Recoverable")
    bucket = bucket.mask(recoverable & (base.eq("submitted")) & (stage >= 1), "Pending with Payer")
    bucket = bucket.mask(recoverable & ~((base.eq("submitted")) & (stage >= 1)), "Action Required")

    # Explicit Needs Review items can still be currently pending with payer.
    needs_review = disposition.eq("Needs Review")
    bucket = bucket.mask(needs_review & base.eq("submitted") & (stage >= 1), "Pending with Payer")
    bucket = bucket.mask(needs_review & ~(base.eq("submitted") & (stage >= 1)), "Needs Review")

    # Unclassified: preserve submitted claims as pending, generic 2+ resub/zero recovery
    # as reconciliation, and everything else as Needs Review (safe fallback).
    unclassified = disposition.eq("Unclassified")
    bucket = bucket.mask(unclassified & base.eq("submitted") & (stage >= 1), "Pending with Payer")
    bucket = bucket.mask(unclassified & generic_recon, "Reconciliation / Escalation")
    bucket = bucket.mask(unclassified & ~(base.eq("submitted") & (stage >= 1)) & ~generic_recon, "Needs Review")

    d["ManagementOutstandingBucket"] = bucket
    d["RuleDenialCode"] = denial
    d["RuleInsuranceFamily"] = insurance
    d["RuleCPT"] = cpt
    return d


def build_rule_master_table() -> pd.DataFrame:
    """Human-readable rule register exported with every analysis workbook."""
    rows = [
        # Global
        ["Global", "All", "COPY-001", "", "Non-Recoverable", "Copay / Patient Responsibility", "All insurers"],
        ["Global", "All", "PRCE-001", "", "Non-Recoverable", "Contractual / Price Adjustment", "All insurers"],
        ["Global", "All", "PRCE-002", "", "Non-Recoverable", "Contractual / Price Adjustment", "All insurers"],
        ["Global", "All", "PRCE-006", "", "Non-Recoverable", "Follow-up Consultation Adjustment", "All insurers"],
        ["Global", "All", "NCOV-001", "", "Non-Recoverable", "Non-Covered Service", "Highlight CPT/service"],
        ["Global", "All", "NCOV-0026", "", "Non-Recoverable", "Non-Covered Service", "Highlight CPT/service"],
        ["Global", "All", "TIME-*", "", "Non-Recoverable", "Time / Timely Filing Issue", "All TIME codes"],
        # Aafiya
        ["Insurance Specific", "Aafiya", "MNEC-003", "99213 / 99203", "Non-Recoverable", "E&M Level 3 Adjustment", "CPT-specific"],
        ["Insurance Specific", "Aafiya", "MNEC-006", "", "Non-Recoverable", "E&M Level 3 Adjustment", "Level 3 billing"],
        # Daman
        ["Insurance Specific", "Daman", "AUTH-001", "", "Needs Review", "Daman System Glitch", ""],
        ["Insurance Specific", "Daman", "AUTH-003", "", "Needs Review", "Daman System Glitch", ""],
        ["Insurance Specific", "Daman", "AUTH-007", "", "Needs Review", "Daman System Glitch", "Resubmitted; recovery uncertain"],
        ["Insurance Specific", "Daman", "CLAI-012", "", "Recoverable", "New vs Established Patient", ""],
        ["Insurance Specific", "Daman", "CLAI-016", "", "Recoverable", "Modifier / Malaffi Issue", ""],
        ["Insurance Specific", "Daman", "ELIG-006", "", "Needs Review", "Daman System Glitch", ""],
        ["Insurance Specific", "Daman", "ELIG-007", "", "Recoverable", "Daman System Glitch", ""],
        ["Insurance Specific", "Daman", "PRCE-010", "", "Non-Recoverable", "Contractual / Bundled Service", ""],
        ["Insurance Specific", "Daman", "NCOV-003", "", "Non-Recoverable", "Non-Covered Service", "Highlight CPT/service"],
        ["Insurance Specific", "Daman", "MNEC-004", "", "Needs Review", "Medical Necessity", "Sometimes recoverable"],
        ["Insurance Specific", "Daman", "MNEC-005", "", "Needs Review", "Medical Necessity", "Sometimes recoverable"],
        # FMC
        ["Insurance Specific", "FMC", "MNEC-006", "", "Non-Recoverable", "E&M Level 3 Adjustment", ""],
        # Inayah
        ["Insurance Specific", "Inayah", "CODE-012", "", "Needs Review", "Coding Issue – Potentially Recoverable", ""],
        ["Insurance Specific", "Inayah", "MNEC-003", "", "Needs Review", "Medical Necessity – Potentially Recoverable", ""],
        ["Insurance Specific", "Inayah", "MNEC-004", "", "Needs Review", "Medical Necessity – Potentially Recoverable", ""],
        # NAS
        ["Insurance Specific", "NAS", "CLAI-016", "", "Non-Recoverable", "E&M Level 3 Adjustment", ""],
        ["Insurance Specific", "NAS", "MNEC-003", "", "Reconciliation Eligible", "Service Payment Dispute", ""],
        ["Insurance Specific", "NAS", "MNEC-004", "", "Reconciliation Eligible", "Service Payment Dispute", ""],
        ["Insurance Specific", "NAS", "MNEC-006", "", "Reconciliation Eligible", "Service Payment Dispute", ""],
        # Oman
        ["Insurance Specific", "Oman", "MNEC-006", "", "Non-Recoverable", "E&M Level 3 Adjustment", ""],
    ]
    return pd.DataFrame(rows, columns=[
        "RuleSource", "Insurance", "DenialCode", "CPTCondition", "Disposition", "ManagementReason", "Notes"
    ])

def _management_outstanding_bucket(recovery: pd.DataFrame) -> pd.Series:
    """Return the mutually-exclusive current bucket for the remaining balance."""
    if recovery.empty:
        return pd.Series(dtype="object")
    if "ManagementOutstandingBucket" in recovery.columns:
        return recovery["ManagementOutstandingBucket"].fillna("Needs Review").astype(str)
    return apply_disposition_rules(recovery)["ManagementOutstandingBucket"]


def build_management_outcome_summary(recovery: pd.DataFrame) -> pd.DataFrame:
    """Executive summary that fully partitions the original rejected amount."""
    cols = ["Outcome", "Amount", "UniqueClaims"]
    if recovery.empty:
        return pd.DataFrame(columns=cols)

    tmp = recovery.copy()
    if "ManagementOutstandingBucket" not in tmp.columns:
        tmp = apply_disposition_rules(tmp)
    tmp["ManagementOutstandingBucket"] = _management_outstanding_bucket(tmp)

    recovered_num = pd.to_numeric(tmp.get("RecoveredAmount", 0), errors="coerce").fillna(0)
    rows = [{
        "Outcome": "Recovered / Paid",
        "Amount": recovered_num.sum(),
        "UniqueClaims": _unique_claim_count(tmp.loc[recovered_num > 0.01, "UniqueID"])
            if "UniqueID" in tmp.columns else int((recovered_num > 0.01).sum()),
    }]

    for label in [
        "Pending with Payer",
        "Action Required",
        "Reconciliation / Escalation",
        "Non-Recoverable / Adjustment",
        "Needs Review",
    ]:
        g = tmp.loc[tmp["ManagementOutstandingBucket"] == label]
        rows.append({
            "Outcome": label,
            "Amount": pd.to_numeric(g.get("OutstandingAmount", 0), errors="coerce").fillna(0).sum(),
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
        })

    out = pd.DataFrame(rows)
    total = {
        "Outcome": "Grand Total",
        "Amount": out["Amount"].sum(),
        "UniqueClaims": _unique_claim_count(tmp["UniqueID"]) if "UniqueID" in tmp.columns else int(len(tmp)),
    }
    return pd.concat([out, pd.DataFrame([total])], ignore_index=True)


def _management_split_for_group(g: pd.DataFrame) -> dict:
    """Return management financial split for one insurance / denial group."""
    g = g.copy()
    if "ManagementOutstandingBucket" not in g.columns:
        g = apply_disposition_rules(g)
    bucket = _management_outstanding_bucket(g)
    recovered = pd.to_numeric(g["RecoveredAmount"], errors="coerce").fillna(0)
    outstanding = pd.to_numeric(g["OutstandingAmount"], errors="coerce").fillna(0)
    original = pd.to_numeric(g["OriginalRejectedAmount"], errors="coerce").fillna(0)

    return {
        "InitialRejected": original.sum(),
        "Recovered": recovered.sum(),
        "Pending": outstanding.where(bucket == "Pending with Payer", 0).sum(),
        "ActionRequired": outstanding.where(bucket == "Action Required", 0).sum(),
        "Reconciliation": outstanding.where(bucket == "Reconciliation / Escalation", 0).sum(),
        "AdjustmentClosed": outstanding.where(bucket == "Non-Recoverable / Adjustment", 0).sum(),
        "NeedsReview": outstanding.where(bucket == "Needs Review", 0).sum(),
        "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
    }


def build_management_by_insurance(recovery: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Insurance", "InitialRejected", "Recovered", "Pending",
        "ActionRequired", "Reconciliation", "AdjustmentClosed", "NeedsReview",
        "UniqueClaims", "RecoveredPct"
    ]
    if recovery.empty:
        return pd.DataFrame(columns=cols)

    rows = []
    for insurance, g in recovery.groupby("Insurance", dropna=False):
        d = _management_split_for_group(g)
        d["Insurance"] = insurance
        d["RecoveredPct"] = (d["Recovered"] / d["InitialRejected"] * 100) if d["InitialRejected"] > 0 else 0
        rows.append(d)

    return pd.DataFrame(rows)[cols].sort_values("InitialRejected", ascending=False).reset_index(drop=True)


def build_management_by_denial(recovery: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "DenialCode", "InitialRejected", "Recovered", "Pending",
        "ActionRequired", "Reconciliation", "AdjustmentClosed", "NeedsReview",
        "UniqueClaims"
    ]
    if recovery.empty:
        return pd.DataFrame(columns=cols)

    code_col = "DenialCodeLevel3" if "DenialCodeLevel3" in recovery.columns else "RecoveryDenialCode"
    tmp = recovery.copy()
    tmp[code_col] = tmp.get(code_col, "").astype(str).fillna("").str.strip()
    tmp.loc[tmp[code_col].isin(["", "nan", "None", "none"]), code_col] = "No Code"

    rows = []
    for code, g in tmp.groupby(code_col, dropna=False):
        d = _management_split_for_group(g)
        d["DenialCode"] = code
        rows.append(d)

    return pd.DataFrame(rows)[cols].sort_values("InitialRejected", ascending=False).reset_index(drop=True)


def build_management_by_reason(recovery: pd.DataFrame) -> pd.DataFrame:
    """Management reason summary used for clickable/expandable drill-down."""
    if recovery.empty:
        return pd.DataFrame(columns=["ManagementReason", "InitialRejected", "Recovered", "Outstanding", "Activities", "UniqueClaims"])
    tmp = recovery if "ManagementReason" in recovery.columns else apply_disposition_rules(recovery)
    rows = []
    for reason, g in tmp.groupby("ManagementReason", dropna=False):
        rows.append({
            "ManagementReason": str(reason),
            "InitialRejected": pd.to_numeric(g.get("OriginalRejectedAmount", 0), errors="coerce").fillna(0).sum(),
            "Recovered": pd.to_numeric(g.get("RecoveredAmount", 0), errors="coerce").fillna(0).sum(),
            "Outstanding": pd.to_numeric(g.get("OutstandingAmount", 0), errors="coerce").fillna(0).sum(),
            "Activities": int(len(g)),
            "UniqueClaims": _unique_claim_count(g["UniqueID"]) if "UniqueID" in g.columns else int(len(g)),
        })
    return pd.DataFrame(rows).sort_values("Outstanding", ascending=False).reset_index(drop=True)


def _dynamic_detail_summaries(detail: pd.DataFrame):
    """Build period-aware detail summaries directly from row-level classified data."""
    if detail.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    d = detail.copy()
    amount = pd.to_numeric(d.get("OriginalRejectedAmount", 0), errors="coerce").fillna(0)
    d["_Amt"] = amount

    by_ins = d.groupby("Insurance", dropna=False).agg(
        RejectedAmount=("_Amt", "sum"),
        RejectedCount=("UniqueID", _unique_claim_count) if "UniqueID" in d.columns else ("_Amt", "size"),
    ).reset_index().sort_values("RejectedAmount", ascending=False)

    code_col = "DenialCodeLevel3" if "DenialCodeLevel3" in d.columns else "RuleDenialCode"
    by_code = d.groupby(code_col, dropna=False).agg(
        RejectedAmount=("_Amt", "sum"),
        RejectedCount=("UniqueID", _unique_claim_count) if "UniqueID" in d.columns else ("_Amt", "size"),
    ).reset_index().rename(columns={code_col: "DenialCode"}).sort_values("RejectedAmount", ascending=False)

    matrix = pd.pivot_table(d, index="Insurance", columns=code_col, values="_Amt", aggfunc="sum", fill_value=0, margins=True, margins_name="Grand Total").reset_index()

    if "AgingBucket" in d.columns:
        aging = pd.pivot_table(d, index="Insurance", columns="AgingBucket", values="_Amt", aggfunc="sum", fill_value=0, margins=True, margins_name="Grand Total").reset_index()
    else:
        aging = pd.DataFrame()
    return by_ins, by_code, matrix, aging


# -------------------- excel styling --------------------
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TOTAL_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
# Highlight color for claims flagged eligible for reconciliation (resubmitted
# 2+ times, zero recovered) — distinct amber so it stands out from totals.
RECON_FILL  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

def style_headers(ws):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def highlight_grand_total_rows(ws, label_col=1, label_value="Grand Total"):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=label_col).value == label_value:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)

def highlight_all_data_rows(ws, fill):
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def highlight_last_col(ws):
    last_col = ws.max_column
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=last_col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)

def apply_styling_to_bytes(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    for ws in wb.worksheets:
        style_headers(ws)
        if ws.title in [
            "Rejected_By_Insurance",
            "Rejected_By_DenialCode",
            "Rejected_Ins_x_DenialCode",
            "Rejected_Aging_Summary",
            "Recovery_Summary",
            "Recovery_By_Insurance",
            "Resub_Stage_Summary",
        ]:
            highlight_grand_total_rows(ws, label_col=1, label_value="Grand Total")
            if ws.title in ["Rejected_Ins_x_DenialCode", "Rejected_Aging_Summary"]:
                highlight_last_col(ws)
    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()

def build_rejection_workbook_bytes(input_bytes: bytes, input_name: str = "source.xlsx") -> tuple[bytes, dict]:
    df = pd.read_excel(io.BytesIO(input_bytes), engine="openpyxl")
    df.columns = df.columns.str.strip()

    # Normalize old/new billing-software column layouts before analysis.
    df, source_info = normalize_source_schema(df)
    df = ensure_numeric(df)
    df = compute_paid(df)
    df = normalize_denial_code(df)
    df = split_denial_code_levels(df)
    df = ensure_insurance_column(df)
    df = add_refdate_and_aging(df)

    # Parse every DISTINCT status value once and map it back onto all rows
    # (see attach_status_parts docstring) instead of re-parsing per row.
    df = attach_status_parts(df)

    # Status audit makes parser inclusion/exclusion fully visible.
    status_audit = build_status_audit(df)
    rejected_df = build_rejected_df(df)

    by_ins = pivot_by_insurance(rejected_df) if len(rejected_df) else pd.DataFrame(
        [{"Insurance": "Grand Total", "RejectedAmount": 0.0, "RejectedCount": 0}]
    )
    by_code = pivot_by_denialcode(rejected_df) if len(rejected_df) else pd.DataFrame(
        [{"DenialCode": "Grand Total", "RejectedAmount": 0.0, "RejectedCount": 0}]
    )
    ins_x_code = pivot_insurance_x_denialcode(rejected_df) if len(rejected_df) else pd.DataFrame(
        [{"Insurance": "Grand Total", "Grand Total": 0.0}]
    )
    aging_sum = pivot_rejection_aging(rejected_df) if len(rejected_df) else pd.DataFrame(
        [{"Insurance": "Grand Total", "Grand Total": 0.0}]
    )

    # Full analytical rejection lifecycle from the exact ActivityStatus values.
    recovery_detail = build_recovery_detail(df)
    recovery_detail = apply_disposition_rules(recovery_detail)
    rule_master = build_rule_master_table()
    recovery_summary = build_recovery_summary(recovery_detail)
    recovery_by_insurance = build_recovery_by_insurance(recovery_detail)
    resub_stage_summary = build_resub_stage_summary(recovery_detail)
    resub_stage_summary_exclusive = build_resub_stage_summary_exclusive(recovery_detail)
    denial_management_summary = build_denial_management_summary(recovery_detail)
    reconciliation_candidates = build_reconciliation_candidates(recovery_detail)
    management_outcome = build_management_outcome_summary(recovery_detail)
    management_by_insurance = build_management_by_insurance(recovery_detail)
    management_by_denial = build_management_by_denial(recovery_detail)

    stats = {
        "rejected_rows": int(len(rejected_df)),
        "sha1": sha1_short_bytes(input_bytes),
        "source_format": source_info["format"],
        "recovery_rows": int(len(recovery_detail)),
        "recovery_available": bool(not recovery_detail.empty),
    }

    meta = pd.DataFrame([{
        "InputFile": input_name,
        "InputSHA1": stats["sha1"],
        "GeneratedAt": dt.now().strftime("%Y-%m-%d %H:%M:%S"),
        "AnalysisVersion": ANALYSIS_VERSION,
        "SourceFormat": source_info["format"],
        "StatusSource": source_info["status_source"],
        "DenialSource": source_info["denial_source"],
        "PaidSource": source_info["paid_source"],
        "RejectedAmountSource": source_info["amount_source"],
        "RejectedRule": "Agreed rejection/resub statuses; RejectedAmount = ActivityIns - actRemitInsShare",
        "RejectedRows": int(len(rejected_df)),
        "RecoveryTrackingAvailable": bool(not recovery_detail.empty),
        "HistoricalRejectedRows": int(len(recovery_detail)),
        "RecoveryRule": "Status-driven lifecycle; resub recovery from actResub remit fields with duplicate-equal amounts counted once",
        "DispositionRule": "Insurance-specific rules override global rules; unmatched outstanding defaults safely to Needs Review",
    }])

    out_buf = io.BytesIO()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        by_ins.to_excel(writer, sheet_name="Rejected_By_Insurance", index=False)
        by_code.to_excel(writer, sheet_name="Rejected_By_DenialCode", index=False)
        ins_x_code.to_excel(writer, sheet_name="Rejected_Ins_x_DenialCode", index=False)
        aging_sum.to_excel(writer, sheet_name="Rejected_Aging_Summary", index=False)
        rejected_df.to_excel(writer, sheet_name="Rejected_Detail", index=False)
        recovery_summary.to_excel(writer, sheet_name="Recovery_Summary", index=False)
        recovery_by_insurance.to_excel(writer, sheet_name="Recovery_By_Insurance", index=False)
        resub_stage_summary.to_excel(writer, sheet_name="Resub_Stage_Summary", index=False)
        resub_stage_summary_exclusive.to_excel(writer, sheet_name="Resub_Stage_Summary_Exclusive", index=False)
        denial_management_summary.to_excel(writer, sheet_name="Denial_Reason_Management", index=False)
        reconciliation_candidates.to_excel(writer, sheet_name="Reconciliation_Candidates", index=False)
        management_outcome.to_excel(writer, sheet_name="Management_Outcome", index=False)
        management_by_insurance.to_excel(writer, sheet_name="Management_By_Insurance", index=False)
        management_by_denial.to_excel(writer, sheet_name="Management_By_Denial", index=False)
        build_management_by_reason(recovery_detail).to_excel(writer, sheet_name="Management_By_Reason", index=False)
        rule_master.to_excel(writer, sheet_name="Rule_Master", index=False)
        recovery_detail.to_excel(writer, sheet_name="Rule_Classified_Detail", index=False)
        recovery_detail.to_excel(writer, sheet_name="Recovery_Detail", index=False)
        status_audit.to_excel(writer, sheet_name="Status_Audit", index=False)
        meta.to_excel(writer, sheet_name="Meta", index=False)

        # ✅ Style directly on the worksheet objects we already have open here,
        # instead of saving the whole workbook and then calling
        # apply_styling_to_bytes() to reload it a second time with openpyxl.
        # That second load_workbook() call deserialized EVERY cell of EVERY
        # sheet (including the large Rejected_Detail / Recovery_Detail sheets)
        # into Python objects all over again — on a big export that doubled
        # both the time and the memory used, and was the main reason this page
        # was slow or crashing. Styling in-place avoids that entirely.
        summary_sheets = {
            "Rejected_By_Insurance",
            "Rejected_By_DenialCode",
            "Rejected_Ins_x_DenialCode",
            "Rejected_Aging_Summary",
            "Recovery_Summary",
            "Recovery_By_Insurance",
            "Resub_Stage_Summary",
            "Resub_Stage_Summary_Exclusive",
        }
        for name, ws in writer.sheets.items():
            style_headers(ws)
            if name in summary_sheets:
                highlight_grand_total_rows(ws, label_col=1, label_value="Grand Total")
                if name in ("Rejected_Ins_x_DenialCode", "Rejected_Aging_Summary"):
                    highlight_last_col(ws)
            if name == "Reconciliation_Candidates":
                highlight_all_data_rows(ws, fill=RECON_FILL)

    return out_buf.getvalue(), stats

# =========================================
# UI HELPERS
# =========================================
def _card(title: str, value: str, sub: str = ""):
    st.markdown(
        f"""
        <div class="card">
          <div class="card-title">{title}</div>
          <div class="card-value">{value}</div>
          <div class="card-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

def _fmt_aed(x):
    try:
        return f"AED {float(x):,.2f}"
    except Exception:
        return f"AED {x}"


def build_email_summary_html(
    center: str, year: str,
    total_amount: float, total_claims: int,
    top_ins: pd.DataFrame,
    df_recovery_summary: pd.DataFrame,
    df_resub_stage_summary_excl: pd.DataFrame,
    recon_amount: float, recon_claims: int,
) -> str:
    """Build a compact HTML email body: headline numbers + the tables
    management actually reads at a glance. Full row-level detail stays in
    the attached Excel workbook, not in the email body."""

    def _df_to_html(df: pd.DataFrame, money_cols=()) -> str:
        if df is None or df.empty:
            return "<p><i>No data.</i></p>"
        d = df.copy()
        for c in money_cols:
            if c in d.columns:
                d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0).map(lambda v: f"AED {v:,.2f}")
        return d.to_html(index=False, border=0, justify="left")

    top_ins_html = _df_to_html(top_ins[["Insurance", "RejectedAmount"]] if not top_ins.empty else top_ins, money_cols=["RejectedAmount"])
    recovery_html = _df_to_html(
        df_recovery_summary[df_recovery_summary.get("RecoveryBucket", "") != "Grand Total"] if not df_recovery_summary.empty else df_recovery_summary,
        money_cols=["OriginalRejectedAmount", "RecoveredAmount", "OutstandingAmount"],
    )
    stage_html = _df_to_html(
        df_resub_stage_summary_excl[df_resub_stage_summary_excl.get("Stage", "") != "Grand Total"] if not df_resub_stage_summary_excl.empty else df_resub_stage_summary_excl,
        money_cols=["RejectedAmount", "RecoveredAmount", "OutstandingAmount"],
    )

    style = (
        "font-family:Arial,Helvetica,sans-serif;font-size:13px;color:#1A0A0A;"
    )
    table_style = "border-collapse:collapse;margin-bottom:16px;"
    css = f"""
    <style>
      table {{ {table_style} }}
      th, td {{ border:1px solid #ddd; padding:6px 10px; text-align:left; font-size:12.5px; }}
      th {{ background:#BDD7EE; }}
    </style>
    """

    html = f"""
    <html><body style="{style}">
      {css}
      <h2 style="color:#9B1C1C;">Rejection & Recovery Summary — {center.title()} ({year})</h2>
      <p>
        <b>Total Rejected Amount:</b> {_fmt_aed(total_amount)}<br/>
        <b>Total Rejected Claims:</b> {total_claims:,}<br/>
        <b>Reconciliation-Eligible (2+ resubs, zero recovered):</b> {_fmt_aed(recon_amount)} across {recon_claims:,} claims
      </p>

      <h3>Top Insurers by Rejected Amount</h3>
      {top_ins_html}

      <h3>Recovery Status Summary</h3>
      {recovery_html}

      <h3>Resubmission Stage Summary (Exclusive — adds up to total)</h3>
      {stage_html}

      <p style="color:#9CA3AF;font-size:11.5px;">
        Full claim-level detail, denial reason breakdown, and the reconciliation candidate list
        are in the attached Excel workbook.
      </p>
    </body></html>
    """
    return html



def build_owner_email_html(
    center: str,
    period_label: str,
    total_amount: float,
    total_claims: int,
    recovered_amount: float,
    pending_amount: float,
    action_amount: float,
    recon_amount: float,
    adjustment_amount: float,
    needs_review_amount: float,
    top_insurance: pd.DataFrame,
    top_denials: pd.DataFrame,
) -> str:
    """Outlook-friendly executive HTML email using inline styles only."""

    def card(title, value, note, accent):
        return f"""<td style='width:33.33%;padding:7px;vertical-align:top;'>
          <div style='border:1px solid #E7D8D8;border-left:5px solid {accent};border-radius:12px;padding:14px 16px;background:#FFFFFF;min-height:96px;'>
            <div style='font-size:11px;letter-spacing:.6px;font-weight:700;color:{accent};text-transform:uppercase;'>{title}</div>
            <div style='font-size:24px;line-height:1.25;font-weight:800;color:#201010;margin-top:7px;'>{_fmt_aed(value)}</div>
            <div style='font-size:11px;color:#8A8A98;margin-top:6px;'>{note}</div>
          </div>
        </td>"""

    def mini_cards(df, label_col, amount_col, extra_fn=None):
        if df is None or df.empty:
            return "<p style='color:#888;'>No data for the selected period.</p>"
        cells = []
        for _, r in df.head(5).iterrows():
            label = str(r.get(label_col, '')).strip()
            amt = float(pd.to_numeric(pd.Series([r.get(amount_col, 0)]), errors='coerce').fillna(0).iloc[0])
            extra = extra_fn(r) if extra_fn else ''
            cells.append(f"""<td style='width:20%;padding:6px;vertical-align:top;'>
              <div style='border:1px solid #E7D8D8;border-left:4px solid #A62020;border-radius:10px;padding:12px;background:#fff;min-height:90px;'>
                <div style='font-size:11px;font-weight:700;color:#A62020;text-transform:uppercase;'>{label}</div>
                <div style='font-size:19px;font-weight:800;color:#201010;margin-top:6px;'>{_fmt_aed(amt)}</div>
                <div style='font-size:10.5px;color:#8A8A98;margin-top:5px;'>{extra}</div>
              </div>
            </td>""")
        return "<table role='presentation' width='100%' cellspacing='0' cellpadding='0'><tr>" + ''.join(cells) + "</tr></table>"

    top_ins_html = mini_cards(
        top_insurance, 'Insurance', 'InitialRejected',
        lambda r: f"{int(r.get('UniqueClaims', 0)):,} claims"
    )
    top_den_html = mini_cards(
        top_denials, 'DenialCode', 'InitialRejected',
        lambda r: f"{int(r.get('UniqueClaims', 0)):,} claims"
    )

    nonrec_note = (
        "Confirmed business-rule adjustments such as contractual/price adjustments, "
        "E&amp;M level adjustments, non-covered services, follow-up adjustments and patient responsibility."
    )

    return f"""<!doctype html>
<html><body style='margin:0;padding:0;background:#FAF4F4;font-family:Arial,Helvetica,sans-serif;color:#201010;'>
  <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#FAF4F4;'>
    <tr><td align='center' style='padding:22px 10px;'>
      <table role='presentation' width='920' cellspacing='0' cellpadding='0' style='width:920px;max-width:100%;background:#FAF4F4;'>
        <tr><td style='padding:4px 8px 18px 8px;'>
          <div style='font-size:28px;font-weight:800;color:#261414;'>RCM Denial &amp; Recovery Summary</div>
          <div style='font-size:13px;color:#7F7A83;margin-top:6px;'>{center.title()} • {period_label}</div>
        </td></tr>
        <tr><td>
          <table role='presentation' width='100%' cellspacing='0' cellpadding='0'><tr>
            {card('Initial Rejected', total_amount, f'{total_claims:,} unique claims', '#A62020')}
            {card('Recovered / Paid', recovered_amount, 'Cash recovered after initial rejection', '#2F7D4A')}
            {card('Pending with Payer', pending_amount, 'Already submitted; awaiting payer', '#B7791F')}
          </tr></table>
          <table role='presentation' width='100%' cellspacing='0' cellpadding='0'><tr>
            {card('Recoverable / Action Required', action_amount, 'Correct / resubmit / investigate', '#C2410C')}
            {card('Reconciliation / Escalation', recon_amount, 'Payer dispute / repeated rejection', '#2563A6')}
            {card('Non-Recoverable / Adjustment', adjustment_amount, 'Confirmed business-rule adjustment', '#6B1F1F')}
          </tr></table>
          <table role='presentation' width='100%' cellspacing='0' cellpadding='0'><tr>
            {card('Needs Review', needs_review_amount, 'Not automatically written off', '#7C3A8C')}
            <td style='width:66.66%;padding:7px;vertical-align:top;'>
              <div style='border:1px solid #E7D8D8;border-radius:12px;padding:14px 16px;background:#FFF9F9;min-height:96px;'>
                <div style='font-size:12px;font-weight:700;color:#6B1F1F;'>Management note</div>
                <div style='font-size:12px;line-height:1.55;color:#5D555B;margin-top:6px;'>{nonrec_note}</div>
              </div>
            </td>
          </tr></table>
        </td></tr>
        <tr><td style='padding:22px 8px 5px 8px;font-size:19px;font-weight:800;'>Top 5 Insurances</td></tr>
        <tr><td>{top_ins_html}</td></tr>
        <tr><td style='padding:22px 8px 5px 8px;font-size:19px;font-weight:800;'>Top 5 Denial Codes</td></tr>
        <tr><td>{top_den_html}</td></tr>
        <tr><td style='padding:20px 8px 4px 8px;'>
          <div style='background:#EEF5FF;border:1px solid #D5E4F8;border-radius:10px;padding:13px 15px;font-size:12px;line-height:1.55;color:#2C4E75;'>
            RCM focus: active recovery opportunities, payer-pending cases and reconciliation/escalation. The attached Excel provides insurance, denial-code, CPT/service and claim-level support for the selected period.
          </div>
        </td></tr>
        <tr><td style='padding:20px 8px 6px 8px;font-size:12px;color:#5F5960;'>
          Best Regards,<br/><b>Irfan Shah</b><br/>RCM / Insurance Department
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""


def build_owner_workbook_bytes(
    period_label: str,
    filtered_detail: pd.DataFrame,
    management_by_insurance: pd.DataFrame,
    management_by_denial: pd.DataFrame,
    total_amount: float,
    total_claims: int,
    recovered_amount: float,
    pending_amount: float,
    action_amount: float,
    recon_amount: float,
    adjustment_amount: float,
    needs_review_amount: float,
) -> bytes:
    """Create a management-ready workbook for the currently selected period."""
    detail = filtered_detail.copy()
    bucket_col = "ManagementOutstandingBucket"

    def subset(bucket):
        if detail.empty or bucket_col not in detail.columns:
            return pd.DataFrame()
        return detail.loc[detail[bucket_col].astype(str).eq(bucket)].copy()

    executive = pd.DataFrame([
        ["Selected Period", period_label, ""],
        ["Initial Rejected", total_amount, total_claims],
        ["Recovered / Paid", recovered_amount, ""],
        ["Pending with Payer", pending_amount, ""],
        ["Recoverable / Action Required", action_amount, ""],
        ["Reconciliation / Escalation", recon_amount, ""],
        ["Non-Recoverable / Adjustment", adjustment_amount, ""],
        ["Needs Review", needs_review_amount, ""],
    ], columns=["KPI", "Amount / Value", "Unique Claims"])

    preferred_cols = [
        "UniqueID", "Insurance", "VisitNo", "VisitDate", "DenialCodeLevel3", "RecoveryDenialCode",
        "Code", "Description", "OriginalRejectedAmount", "RecoveredAmount", "OutstandingAmount",
        "ManagementOutstandingBucket", "ManagementReason", "RuleApplied", "RuleSource", "RecommendedAction"
    ]
    cols = [c for c in preferred_cols if c in detail.columns]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        executive.to_excel(writer, sheet_name="Executive Summary", index=False)
        management_by_insurance.to_excel(writer, sheet_name="Insurance Summary", index=False)
        management_by_denial.to_excel(writer, sheet_name="Denial Codes", index=False)
        subset("Non-Recoverable / Adjustment")[cols].to_excel(writer, sheet_name="Non-Recoverable", index=False)
        subset("Action Required")[cols].to_excel(writer, sheet_name="Action Required", index=False)
        subset("Reconciliation / Escalation")[cols].to_excel(writer, sheet_name="Reconciliation", index=False)
        subset("Needs Review")[cols].to_excel(writer, sheet_name="Needs Review", index=False)
        detail[cols].to_excel(writer, sheet_name="Claim Detail", index=False)

    wb = load_workbook(io.BytesIO(buf.getvalue()))
    header_fill = PatternFill("solid", fgColor="A62020")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="F8EAEA")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = 0
            for cell in col_cells[:250]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(1, cell.column).value or '').lower()
                if isinstance(cell.value, (int, float)) and any(k in header for k in ['amount','rejected','recovered','pending','action','reconciliation','adjustment','review']):
                    cell.number_format = '#,##0.00'

    ws = wb["Executive Summary"]
    ws.freeze_panes = None
    ws.auto_filter.ref = None
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).fill = sub_fill
        ws.cell(r, 1).font = Font(bold=True, color="6B1F1F")
        if r >= 3 and isinstance(ws.cell(r, 2).value, (int, float)):
            ws.cell(r, 2).number_format = 'AED #,##0.00'
            ws.cell(r, 2).font = Font(bold=True, size=12)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def send_email_with_attachment(
    recipients: list[str], subject: str, html_body: str,
    attachment_bytes: bytes, attachment_filename: str,
) -> None:
    """Send via SMTP using credentials from Streamlit secrets. Works with
    Gmail/Outlook app passwords, a corporate SMTP relay, or AWS SES's SMTP
    interface — whichever this dashboard's other pages already use."""
    if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_SENDER):
        raise RuntimeError(
            "Email is not configured yet. Add SMTP_HOST, SMTP_PORT, SMTP_USER, "
            "SMTP_PASSWORD and SMTP_SENDER to Streamlit secrets (the same values "
            "used for email on the other dashboard pages)."
        )
    if not recipients:
        raise RuntimeError("Add at least one recipient email address.")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SMTP_SENDER
    msg["To"] = ", ".join(recipients)
    msg.set_content("This email requires an HTML-capable mail client to view the summary.")
    msg.add_alternative(html_body, subtype="html")
    msg.add_attachment(
        attachment_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=attachment_filename,
    )

    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls(context=context)
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)

def load_result_from_workbook_bytes(xlsx_bytes: bytes, center: str, year: str, s3_key: str) -> dict:
    xls = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")

    df_by_ins = pd.read_excel(xls, sheet_name="Rejected_By_Insurance")
    df_by_code = pd.read_excel(xls, sheet_name="Rejected_By_DenialCode")
    df_ins_x_code = pd.read_excel(xls, sheet_name="Rejected_Ins_x_DenialCode")
    df_aging = pd.read_excel(xls, sheet_name="Rejected_Aging_Summary")

    try:
        df_recovery_summary = pd.read_excel(xls, sheet_name="Recovery_Summary")
        df_recovery_by_insurance = pd.read_excel(xls, sheet_name="Recovery_By_Insurance")
        df_resub_stage_summary = pd.read_excel(xls, sheet_name="Resub_Stage_Summary")
        recovery_header = pd.read_excel(xls, sheet_name="Recovery_Detail", nrows=0).columns.tolist()
        recovery_preview_cols = [c for c in [
            "UniqueID", "Insurance", "VisitNo", "VisitDate", "Code", "Description",
            "ActivityStatus", "CurrentActivityStatus", "InitialActivityStatus", "CurrentStatus",
            "RecoveryBucket", "RecoveryDenialCode", "DenialCodeLevel2", "DenialCodeLevel3",
            "PaymentStatus", "ReconciliationEligible", "OriginalRejectedAmount",
            "RecoveredAmount", "OutstandingAmount", "ResubStage",
            "actRemitInsShare", "actResub1RemitInsShare", "actResub2RemitInsShare",
            "actResub3RemitInsShare", "TKBKAmountAct", "Paid",
            "Resub1Comments/Accpt Comments",
            "Resub1ActivityStatus", "Resub2ActivityStatus", "Resub3ActivityStatus",
            "Resub1Date", "Resub2Date", "Resub3Date"
        ] if c in recovery_header]
        df_recovery_preview = pd.read_excel(
            xls, sheet_name="Recovery_Detail", usecols=recovery_preview_cols, nrows=2000
        )
    except Exception:
        df_recovery_summary = pd.DataFrame()
        df_recovery_by_insurance = pd.DataFrame()
        df_resub_stage_summary = pd.DataFrame()
        df_recovery_preview = pd.DataFrame()

    try:
        df_resub_stage_summary_excl = pd.read_excel(xls, sheet_name="Resub_Stage_Summary_Exclusive")
    except Exception:
        df_resub_stage_summary_excl = pd.DataFrame()

    try:
        df_denial_management = pd.read_excel(xls, sheet_name="Denial_Reason_Management")
    except Exception:
        df_denial_management = pd.DataFrame()

    try:
        df_reconciliation = pd.read_excel(xls, sheet_name="Reconciliation_Candidates")
    except Exception:
        df_reconciliation = pd.DataFrame()

    try:
        df_management_outcome = pd.read_excel(xls, sheet_name="Management_Outcome")
    except Exception:
        df_management_outcome = pd.DataFrame()

    try:
        df_management_by_insurance = pd.read_excel(xls, sheet_name="Management_By_Insurance")
    except Exception:
        df_management_by_insurance = pd.DataFrame()

    try:
        df_management_by_denial = pd.read_excel(xls, sheet_name="Management_By_Denial")
    except Exception:
        df_management_by_denial = pd.DataFrame()

    try:
        df_rule_detail = pd.read_excel(xls, sheet_name="Rule_Classified_Detail")
    except Exception:
        # Backward compatibility only; v8 cache version normally guarantees the new sheet.
        try:
            df_rule_detail = pd.read_excel(xls, sheet_name="Recovery_Detail")
            df_rule_detail = apply_disposition_rules(df_rule_detail)
        except Exception:
            df_rule_detail = pd.DataFrame()

    try:
        df_rule_master = pd.read_excel(xls, sheet_name="Rule_Master")
    except Exception:
        df_rule_master = build_rule_master_table()

    PREVIEW_ROWS = 2000
    detail_header = pd.read_excel(xls, sheet_name="Rejected_Detail", nrows=0).columns.tolist()
    wanted_cols = [
        "UniqueID", "Insurance", "DenialCode", "FinalDenialCode",
        "DenialCodeLevel2", "DenialCodeLevel3",
        "ActivityStatus", "CurrentActivityStatus", "InitialActivityStatus",
        "ActivityIns", "actRemitInsShare", "actResub1RemitInsShare",
        "actResub2RemitInsShare", "actResub3RemitInsShare", "TKBKAmountAct",
        "Paid", "RejectedAmount", "StatusBase", "ResubStage",
        "AgingBucket", "DaysDiff", "RefDate", "Resub1Comments/Accpt Comments"
    ]
    usecols = [c for c in wanted_cols if c in detail_header]
    df_preview = pd.read_excel(xls, sheet_name="Rejected_Detail", usecols=usecols, nrows=PREVIEW_ROWS)

    rejected_rows = 0
    source_format = "unknown"
    try:
        df_meta = pd.read_excel(xls, sheet_name="Meta")
        if "RejectedRows" in df_meta.columns and len(df_meta):
            rejected_rows = int(pd.to_numeric(df_meta.loc[0, "RejectedRows"], errors="coerce") or 0)
        if "SourceFormat" in df_meta.columns and len(df_meta):
            source_format = str(df_meta.loc[0, "SourceFormat"])
    except Exception:
        rejected_rows = 0

    stats = {
        "sha1": sha1_short_bytes(xlsx_bytes),
        "rejected_rows": rejected_rows,
        "source_format": source_format,
    }

    return {
        "center": center,
        "year": year,
        "s3_key": s3_key,
        "out_bytes": xlsx_bytes,
        "stats": stats,
        "df_by_ins": df_by_ins,
        "df_by_code": df_by_code,
        "df_ins_x_code": df_ins_x_code,
        "df_aging": df_aging,
        "df_recovery_summary": df_recovery_summary,
        "df_recovery_by_insurance": df_recovery_by_insurance,
        "df_resub_stage_summary": df_resub_stage_summary,
        "df_resub_stage_summary_excl": df_resub_stage_summary_excl,
        "df_denial_management": df_denial_management,
        "df_reconciliation": df_reconciliation,
        "df_management_outcome": df_management_outcome,
        "df_management_by_insurance": df_management_by_insurance,
        "df_management_by_denial": df_management_by_denial,
        "df_rule_detail": df_rule_detail,
        "df_rule_master": df_rule_master,
        "df_recovery_preview": df_recovery_preview,
        "df_preview": df_preview,
        "preview_rows": PREVIEW_ROWS,
    }

# =========================================
# APP
# =========================================
def run_rejection_app():
    st.markdown("## RCM Denial & Recovery Intelligence Dashboard")
    st.caption("Management-first view with payer-specific recoverability rules, month filtering, reconciliation logic, and CPT/service drill-down.")

    # --- session init ---
    if "rej_result" not in st.session_state:
        st.session_state.rej_result = None
    if "rej_prev_sel" not in st.session_state:
        st.session_state.rej_prev_sel = None

    # ✅ detect from URL (when clicking Rejected card)
    qp = st.query_params
    if qp.get("center"):
        st.session_state["selected_center"] = qp.get("center")
    if qp.get("year"):
        st.session_state["selected_year"] = qp.get("year")

    detected_center = st.session_state.get("selected_center")
    detected_year = st.session_state.get("selected_year")

    # ---- Sidebar controls ----
    with st.sidebar:
        st.subheader("Controls")

        if detected_center is None or detected_year is None:
            st.warning("Center/Year not detected. Select manually.")
            center = st.selectbox(
                "Center",
                ["Excellent Medical Center", "Excellent Pharmacy", "Easyhealth Clinic"],
                key="rej_center_manual",
            )
            year = st.selectbox("Year", DEFAULT_YEAR_OPTIONS, key="rej_year_manual")
        else:
            center = str(detected_center).lower().strip()
            year = str(detected_year).strip()

            st.success("Detected from dashboard ✅")
            st.selectbox(
                "Center",
                ["excellent", "pharmacy", "easyhealth"],
                index=["excellent", "pharmacy", "easyhealth"].index(center),
                disabled=True,
            )
            st.selectbox(
                "Year",
                DEFAULT_YEAR_OPTIONS,
                index=DEFAULT_YEAR_OPTIONS.index(year),
                disabled=True,
            )

        center_raw = center
        center = normalize_center_for_s3(center_raw)
        year = str(year)

        s3_key = f"streamlit/{center}/{year}/{SOURCE_FILENAME}"
        rej_cache_key = f"{REJ_CACHE_PREFIX}/{center}/{year}/{REJ_CACHE_FILENAME}"

        # ✅ When year/center changes: drop current in-memory result (then auto-load saved if exists)
        current_sel = f"{center}|{year}"
        if st.session_state.rej_prev_sel != current_sel:
            st.session_state.rej_prev_sel = current_sel
            st.session_state.rej_result = None

        st.write("**Source**")
        st.code(f"s3://{S3_BUCKET}/{s3_key}", language="text")

        # ✅ Direct upload from Rejection Analysis page
        st.write("**Upload / Replace Source File**")
        uploaded_source = st.file_uploader(
            "Upload Excel file",
            type=["xlsx"],
            key=f"rej_source_upload_{center}_{year}",
            help="Uploads this workbook directly as the source file for the selected center/year.",
        )
        st.caption("Select the Excel file first. Processing will start only after you click Generate.")

        # IMPORTANT: selecting a file must NOT trigger any upload, parsing, or analysis.
        # The file remains pending in the Streamlit uploader until Generate is clicked.
        pending_upload = uploaded_source is not None
        if pending_upload:
            st.info(f"Ready to generate: {uploaded_source.name}")

        # Auto-load a previously saved result only when there is NO new file waiting to be generated.
        # This avoids showing an old dashboard underneath a newly selected (but not yet processed) file.
        if (not pending_upload) and st.session_state.rej_result is None and s3_exists(S3_BUCKET, rej_cache_key):
            try:
                cached_bytes = load_file_from_s3(S3_BUCKET, rej_cache_key)
                st.session_state.rej_result = load_result_from_workbook_bytes(cached_bytes, center, year, s3_key)
                st.success("Loaded saved result ✅")
            except Exception:
                st.warning("Saved result found but could not be loaded. Click Generate once.")

        cA, cB = st.columns(2)
        with cA:
            generate = st.button("Generate", type="primary", width="stretch")
        with cB:
            clear = st.button("Clear", width="stretch")

        if clear:
            # ✅ Clear BOTH session + saved cache (so it won't reappear on refresh)
            st.session_state.rej_result = None
            delete_file_from_s3(S3_BUCKET, rej_cache_key)
            st.rerun()

        if generate:
            # If the user selected a new file, use that file ONLY now, when Generate is clicked.
            if uploaded_source is not None:
                input_bytes = uploaded_source.getvalue()
                input_name = uploaded_source.name

                with st.spinner("Uploading source and building rejection analysis..."):
                    # Replace the source workbook in S3 only after explicit Generate click.
                    save_file_to_s3(S3_BUCKET, s3_key, input_bytes)

                    # The source changed, so remove the previous cached analysis before rebuilding.
                    delete_file_from_s3(S3_BUCKET, rej_cache_key)

                    out_xlsx_bytes, _stats = build_rejection_workbook_bytes(input_bytes, input_name)
                    save_file_to_s3(S3_BUCKET, rej_cache_key, out_xlsx_bytes)
                    st.session_state.rej_result = load_result_from_workbook_bytes(
                        out_xlsx_bytes, center, year, s3_key
                    )

                st.success(f"Generated from: {input_name} ✅")

            else:
                # No new upload selected: regenerate from the source already saved in S3.
                if not s3_exists(S3_BUCKET, s3_key):
                    st.error("Source file not found in S3. Upload an Excel file above first.")
                    st.stop()

                with st.spinner("Building rejection analysis..."):
                    input_bytes = load_file_from_s3(S3_BUCKET, s3_key)
                    out_xlsx_bytes, _stats = build_rejection_workbook_bytes(input_bytes, SOURCE_FILENAME)
                    save_file_to_s3(S3_BUCKET, rej_cache_key, out_xlsx_bytes)
                    st.session_state.rej_result = load_result_from_workbook_bytes(
                        out_xlsx_bytes, center, year, s3_key
                    )

                st.success("Done ✅")

    # ---- Main UI ----
    if st.session_state.rej_result is None:
        st.info("No saved result for this year yet. Click Generate once.")
        return

    R = st.session_state.rej_result
    out_xlsx_bytes = R["out_bytes"]
    stats = R["stats"]

    df_by_ins = R["df_by_ins"]
    df_by_code = R["df_by_code"]
    df_ins_x_code = R["df_ins_x_code"]
    df_aging = R["df_aging"]
    df_recovery_summary = R.get("df_recovery_summary", pd.DataFrame())
    df_recovery_by_insurance = R.get("df_recovery_by_insurance", pd.DataFrame())
    df_resub_stage_summary = R.get("df_resub_stage_summary", pd.DataFrame())
    df_resub_stage_summary_excl = R.get("df_resub_stage_summary_excl", pd.DataFrame())
    df_denial_management = R.get("df_denial_management", pd.DataFrame())
    df_reconciliation = R.get("df_reconciliation", pd.DataFrame())
    df_recovery_preview = R.get("df_recovery_preview", pd.DataFrame())
    df_preview = R["df_preview"]
    PREVIEW_ROWS = R["preview_rows"]

    # New management-first datasets
    df_management_outcome = R.get("df_management_outcome", pd.DataFrame())
    df_management_by_insurance = R.get("df_management_by_insurance", pd.DataFrame())
    df_management_by_denial = R.get("df_management_by_denial", pd.DataFrame())
    df_rule_detail = R.get("df_rule_detail", pd.DataFrame())
    df_rule_master = R.get("df_rule_master", pd.DataFrame())

    # ===== FLEXIBLE PERIOD / DATE FILTER =====
    # Supports all data, one month, multiple months (e.g. Jul + Aug), or an exact date/date range.
    filtered_rule_detail = df_rule_detail.copy()
    date_candidates = [c for c in ["VisitDate", "ServiceDate", "EncounterDate", "ClaimDate", "RefDate"] if c in filtered_rule_detail.columns]

    if date_candidates:
        f1, f2, f3 = st.columns([1.1, 1.2, 2.7])
        default_date_idx = date_candidates.index("VisitDate") if "VisitDate" in date_candidates else 0
        with f1:
            period_date_col = st.selectbox("Performance date", date_candidates, index=default_date_idx, key="period_date_source")
        parsed_dates = pd.to_datetime(filtered_rule_detail[period_date_col], errors="coerce")
        period_values = sorted(parsed_dates.dropna().dt.to_period("M").unique().tolist())
        period_labels = [p.strftime("%B %Y") for p in period_values]

        with f2:
            filter_mode = st.selectbox(
                "Period filter",
                ["All Data", "Single Month", "Multiple Months", "Specific Date / Range"],
                key="management_filter_mode",
            )

        selected_period_label = "All Data"
        if filter_mode == "Single Month":
            with f3:
                chosen_month = st.selectbox("Select month", period_labels, key="management_single_month") if period_labels else None
            if chosen_month:
                chosen_period = period_values[period_labels.index(chosen_month)]
                filtered_rule_detail = filtered_rule_detail.loc[parsed_dates.dt.to_period("M") == chosen_period].copy()
                selected_period_label = chosen_month

        elif filter_mode == "Multiple Months":
            with f3:
                chosen_months = st.multiselect(
                    "Select two or more months",
                    period_labels,
                    default=period_labels[-2:] if len(period_labels) >= 2 else period_labels,
                    key="management_multi_months",
                )
            if chosen_months:
                chosen_periods = {period_values[period_labels.index(m)] for m in chosen_months}
                filtered_rule_detail = filtered_rule_detail.loc[parsed_dates.dt.to_period("M").isin(chosen_periods)].copy()
                selected_period_label = " + ".join(chosen_months)
            else:
                filtered_rule_detail = filtered_rule_detail.iloc[0:0].copy()
                selected_period_label = "No months selected"

        elif filter_mode == "Specific Date / Range":
            valid_dates = parsed_dates.dropna()
            if not valid_dates.empty:
                min_d = valid_dates.min().date()
                max_d = valid_dates.max().date()
                with f3:
                    chosen_dates = st.date_input(
                        "Select one date or a date range",
                        value=(min_d, max_d),
                        min_value=min_d,
                        max_value=max_d,
                        key="management_date_range",
                    )
                if isinstance(chosen_dates, (list, tuple)):
                    if len(chosen_dates) == 2:
                        start_d, end_d = chosen_dates
                    elif len(chosen_dates) == 1:
                        start_d = end_d = chosen_dates[0]
                    else:
                        start_d, end_d = min_d, max_d
                else:
                    start_d = end_d = chosen_dates
                normalized = parsed_dates.dt.date
                filtered_rule_detail = filtered_rule_detail.loc[(normalized >= start_d) & (normalized <= end_d)].copy()
                selected_period_label = start_d.strftime("%d %b %Y") if start_d == end_d else f"{start_d.strftime('%d %b %Y')} – {end_d.strftime('%d %b %Y')}"
            else:
                with f3:
                    st.caption("No valid dates found in the selected date column.")
                filtered_rule_detail = filtered_rule_detail.iloc[0:0].copy()
                selected_period_label = "No valid dates"
        else:
            with f3:
                st.caption(f"Showing the complete uploaded report using **{period_date_col}**.")
    else:
        period_date_col = None
        selected_period_label = "All Data"
        st.info("No usable service/visit date column was found, so date filtering is unavailable.")

    # Rebuild every management summary from the period-filtered activity-level detail.
    df_management_outcome = build_management_outcome_summary(filtered_rule_detail)
    df_management_by_insurance = build_management_by_insurance(filtered_rule_detail)
    df_management_by_denial = build_management_by_denial(filtered_rule_detail)
    df_management_by_reason = build_management_by_reason(filtered_rule_detail)
    df_by_ins, df_by_code, df_ins_x_code, df_aging = _dynamic_detail_summaries(filtered_rule_detail)
    df_by_ins_nogt = df_by_ins.copy()
    df_recovery_summary = build_recovery_summary(filtered_rule_detail)
    df_recovery_by_insurance = build_recovery_by_insurance(filtered_rule_detail)
    df_resub_stage_summary = build_resub_stage_summary(filtered_rule_detail)
    df_resub_stage_summary_excl = build_resub_stage_summary_exclusive(filtered_rule_detail)
    df_preview = filtered_rule_detail.copy()
    df_recovery_preview = filtered_rule_detail.copy()
    df_reconciliation = filtered_rule_detail.loc[
        filtered_rule_detail.get("ManagementOutstandingBucket", pd.Series("", index=filtered_rule_detail.index)).astype(str).eq("Reconciliation / Escalation")
    ].copy() if not filtered_rule_detail.empty else pd.DataFrame()
    df_denial_management = filtered_rule_detail.copy()

    # ===== Headline values =====
    total_amount = float(pd.to_numeric(filtered_rule_detail.get("OriginalRejectedAmount", 0), errors="coerce").fillna(0).sum()) if not filtered_rule_detail.empty else 0.0
    total_claims = _unique_claim_count(filtered_rule_detail["UniqueID"]) if (not filtered_rule_detail.empty and "UniqueID" in filtered_rule_detail.columns) else int(len(filtered_rule_detail))

    def _outcome_amount(label: str) -> float:
        if df_management_outcome.empty or "Outcome" not in df_management_outcome.columns:
            return 0.0
        s = df_management_outcome.loc[
            df_management_outcome["Outcome"].astype(str) == label, "Amount"
        ]
        return float(pd.to_numeric(s, errors="coerce").fillna(0).sum())

    recovered_amount = _outcome_amount("Recovered / Paid")
    pending_amount = _outcome_amount("Pending with Payer")
    action_amount = _outcome_amount("Action Required")
    recon_amount = _outcome_amount("Reconciliation / Escalation")
    adjustment_amount = _outcome_amount("Non-Recoverable / Adjustment")
    needs_review_amount = _outcome_amount("Needs Review")

    # A more meaningful rate than recovered / all historical denials:
    # recovery among money that has reached a closed financial outcome.
    matured_base = recovered_amount + adjustment_amount
    matured_recovery_rate = (recovered_amount / matured_base * 100) if matured_base > 0 else 0.0

    # ===== Top toolbar =====
    top_left, top_download, top_email = st.columns([3.4, 1, 1])
    with top_left:
        st.caption(
            f"Executive view first • {selected_period_label} • Rule engine {ANALYSIS_VERSION}. Detailed tracing stays under RCM Detail."
        )
    with top_download:
        st.download_button(
            "Download Full Excel",
            data=out_xlsx_bytes,
            file_name=f"Rejection_Analysis_{R['center']}_{R['year']}_{stats['sha1']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    with top_email:
        if st.button("📧 Email Owner", width="stretch", key="toggle_owner_email_top"):
            st.session_state["show_owner_email"] = not st.session_state.get("show_owner_email", False)

    # ===== Main navigation =====
    tab_exec, tab_ins, tab_denial, tab_work, tab_detail = st.tabs([
        "Executive Summary",
        "Insurance Analysis",
        "Denial Reasons",
        "Recovery Worklist",
        "RCM Detail",
    ])

    # ------------------------------------------------------------------
    # TAB 1 — EXECUTIVE SUMMARY
    # ------------------------------------------------------------------
    with tab_exec:
        # Owner email workflow on Executive Summary first page.
        if st.session_state.get("show_owner_email", False):
            st.markdown("### 📧 Owner Email — Preview & Send")
            st.caption(
                f"The email and attachment use the CURRENT dashboard filter: **{selected_period_label}**. "
                "Nothing is sent until you click Send Email."
            )

            owner_xlsx_bytes = build_owner_workbook_bytes(
                period_label=selected_period_label,
                filtered_detail=filtered_rule_detail,
                management_by_insurance=df_management_by_insurance,
                management_by_denial=df_management_by_denial,
                total_amount=total_amount,
                total_claims=total_claims,
                recovered_amount=recovered_amount,
                pending_amount=pending_amount,
                action_amount=action_amount,
                recon_amount=recon_amount,
                adjustment_amount=adjustment_amount,
                needs_review_amount=needs_review_amount,
            )

            owner_top_ins = df_management_by_insurance.sort_values("InitialRejected", ascending=False).head(5).copy() if not df_management_by_insurance.empty else pd.DataFrame()
            owner_top_den = df_management_by_denial.sort_values("InitialRejected", ascending=False).head(5).copy() if not df_management_by_denial.empty else pd.DataFrame()
            owner_html = build_owner_email_html(
                center=R["center"],
                period_label=selected_period_label,
                total_amount=total_amount,
                total_claims=total_claims,
                recovered_amount=recovered_amount,
                pending_amount=pending_amount,
                action_amount=action_amount,
                recon_amount=recon_amount,
                adjustment_amount=adjustment_amount,
                needs_review_amount=needs_review_amount,
                top_insurance=owner_top_ins,
                top_denials=owner_top_den,
            )

            em1, em2 = st.columns([2, 1])
            with em1:
                owner_recipients_raw = st.text_input(
                    "To (comma-separated)",
                    value=DEFAULT_MANAGEMENT_RECIPIENTS,
                    key="owner_email_recipients",
                )
            with em2:
                owner_subject = st.text_input(
                    "Subject",
                    value=f"RCM Denial & Recovery Summary — {selected_period_label}",
                    key="owner_email_subject",
                )

            st.markdown("##### Email Preview")
            st.html(owner_html, width="stretch")

            safe_period = re.sub(r"[^A-Za-z0-9_-]+", "_", selected_period_label).strip("_") or "Selected_Period"
            owner_attach_name = f"RCM_Denial_Recovery_{safe_period}.xlsx"
            dl_col, send_col = st.columns([1, 1])
            with dl_col:
                st.download_button(
                    "📎 Download Owner Excel",
                    data=owner_xlsx_bytes,
                    file_name=owner_attach_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                    key="owner_email_attachment_download",
                )
            with send_col:
                send_clicked = st.button(
                    "📧 Send Email",
                    type="primary",
                    width="stretch",
                    key="owner_send_email_btn",
                    disabled=not bool(SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_SENDER),
                )

            if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_SENDER):
                st.info(
                    "Email preview and Owner Excel are ready. To enable Send Email, configure "
                    "SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD and SMTP_SENDER in Streamlit secrets."
                )

            if send_clicked:
                recipients = [r.strip() for r in owner_recipients_raw.split(",") if r.strip()]
                try:
                    with st.spinner("Sending owner email..."):
                        send_email_with_attachment(
                            recipients=recipients,
                            subject=owner_subject,
                            html_body=owner_html,
                            attachment_bytes=owner_xlsx_bytes,
                            attachment_filename=owner_attach_name,
                        )
                    st.success(f"Owner email sent to {len(recipients)} recipient(s) ✅")
                except Exception as e:
                    st.error(f"Could not send email: {e}")


        st.markdown("### Financial Position of Denials")

        k1, k2, k3 = st.columns(3)
        with k1:
            _card("Initial Rejected", _fmt_aed(total_amount), f"{total_claims:,} unique claims • {selected_period_label}")
        with k2:
            _card("Recovered / Paid", _fmt_aed(recovered_amount), "Cash recovered after initial rejection")
        with k3:
            _card("Pending with Payer", _fmt_aed(pending_amount), "Already submitted; awaiting payer")

        k4, k5, k6, k7 = st.columns(4)
        with k4:
            _card("Recoverable / Action Required", _fmt_aed(action_amount), "Correct / resubmit / investigate")
        with k5:
            _card("Reconciliation / Escalation", _fmt_aed(recon_amount), "Payer dispute / repeated rejection")
        with k6:
            _card("Non-Recoverable / Adjustment", _fmt_aed(adjustment_amount), "Confirmed business-rule adjustment")
        with k7:
            _card("Needs Review", _fmt_aed(needs_review_amount), "Not automatically written off")

        st.caption(
            "Financial control: Recovered + Pending + Action Required + Reconciliation + Non-Recoverable + Needs Review "
            "must reconcile to Initial Rejected. Unknown cases default to Needs Review rather than being written off."
        )

        with st.expander(f"🔎 Non-Recoverable / Adjustment Drill-Down — {_fmt_aed(adjustment_amount)}", expanded=False):
            nonrec = filtered_rule_detail.loc[
                filtered_rule_detail.get("ManagementOutstandingBucket", pd.Series("", index=filtered_rule_detail.index)).astype(str).eq("Non-Recoverable / Adjustment")
            ].copy() if not filtered_rule_detail.empty else pd.DataFrame()

            if nonrec.empty:
                st.info("No rule-confirmed non-recoverable outstanding amount for the selected period.")
            else:
                reason_summary = build_management_by_reason(nonrec)
                st.markdown("##### Why is it non-recoverable? — reason, denial code and service")

                denial_col = "DenialCodeLevel3" if "DenialCodeLevel3" in nonrec.columns else ("DenialCode" if "DenialCode" in nonrec.columns else None)
                service_col_all = "Code" if "Code" in nonrec.columns else ("RuleCPT" if "RuleCPT" in nonrec.columns else None)
                desc_col_all = "Description" if "Description" in nonrec.columns else None
                summary_group_cols = ["ManagementReason"]
                # Always keep payer visible in the management drill-down so the same
                # denial/service can be traced back to the responsible insurance.
                if "Insurance" in nonrec.columns:
                    summary_group_cols.append("Insurance")
                if denial_col:
                    summary_group_cols.append(denial_col)
                if service_col_all:
                    summary_group_cols.append(service_col_all)
                if desc_col_all:
                    summary_group_cols.append(desc_col_all)

                detailed_reason_summary = nonrec.groupby(summary_group_cols, dropna=False).agg(
                    InitialRejected=("OriginalRejectedAmount", "sum"),
                    Recovered=("RecoveredAmount", "sum"),
                    Outstanding=("OutstandingAmount", "sum"),
                    Activities=("OutstandingAmount", "size"),
                    UniqueClaims=("UniqueID", _unique_claim_count) if "UniqueID" in nonrec.columns else ("OutstandingAmount", "size"),
                ).reset_index().sort_values("Outstanding", ascending=False)
                rename_map = {}
                if denial_col:
                    rename_map[denial_col] = "Denial Code"
                if service_col_all:
                    rename_map[service_col_all] = "Service / CPT"
                if desc_col_all:
                    rename_map[desc_col_all] = "Service Description"
                st.dataframe(detailed_reason_summary.rename(columns=rename_map), width="stretch", hide_index=True)

                reasons = reason_summary["ManagementReason"].dropna().astype(str).tolist()
                selected_reason = st.selectbox("Select reason to see services/CPT", reasons, key="nonrec_reason_drill") if reasons else None
                reason_rows = nonrec.loc[nonrec["ManagementReason"].astype(str) == selected_reason].copy() if selected_reason else nonrec.iloc[0:0]

                if not reason_rows.empty:
                    service_col = "Code" if "Code" in reason_rows.columns else "RuleCPT"
                    desc_col = "Description" if "Description" in reason_rows.columns else None
                    group_cols = [service_col] + ([desc_col] if desc_col else [])
                    service_summary = reason_rows.groupby(group_cols, dropna=False).agg(
                        NonRecoverableAmount=("OutstandingAmount", "sum"),
                        Activities=("OutstandingAmount", "size"),
                        Claims=("UniqueID", _unique_claim_count) if "UniqueID" in reason_rows.columns else ("OutstandingAmount", "size"),
                    ).reset_index().sort_values("NonRecoverableAmount", ascending=False)
                    st.markdown(f"##### Services / CPT — {selected_reason}")
                    st.dataframe(service_summary, width="stretch", hide_index=True)

                    cpt_options = service_summary[service_col].dropna().astype(str).tolist()
                    selected_cpt = st.selectbox("Select CPT/service for claim detail", ["All"] + cpt_options, key="nonrec_cpt_drill")
                    claim_rows = reason_rows.copy()
                    if selected_cpt != "All":
                        claim_rows = claim_rows.loc[claim_rows[service_col].astype(str) == selected_cpt]
                    detail_cols = [c for c in [
                        "UniqueID", "VisitNo", "VisitDate", "Insurance", "DenialCodeLevel3",
                        "Code", "Description", "DocName", "OriginalRejectedAmount", "RecoveredAmount",
                        "OutstandingAmount", "CurrentStatus", "RuleDisposition", "ManagementReason",
                        "RecommendedAction", "RuleApplied"
                    ] if c in claim_rows.columns]
                    st.markdown("##### Claim / Activity Detail")
                    st.dataframe(claim_rows[detail_cols].sort_values("OutstandingAmount", ascending=False), width="stretch", hide_index=True)

        st.markdown("### Top 5 Insurances")
        if not df_management_by_insurance.empty:
            top5 = df_management_by_insurance.head(5).copy()
            cols = st.columns(5)
            for i, (_, row) in enumerate(top5.iterrows()):
                ins_name = str(row.get("Insurance", "Unknown"))
                rejected = float(pd.to_numeric(pd.Series([row.get("InitialRejected", 0)]), errors="coerce").fillna(0).iloc[0])
                open_amt = sum(float(pd.to_numeric(pd.Series([row.get(c, 0)]), errors="coerce").fillna(0).iloc[0]) for c in ["Pending", "ActionRequired", "Reconciliation", "AdjustmentClosed", "NeedsReview"])
                claims = int(pd.to_numeric(pd.Series([row.get("UniqueClaims", 0)]), errors="coerce").fillna(0).iloc[0])
                with cols[i]:
                    _card(ins_name, _fmt_aed(rejected), f"{claims:,} claims • Open/adjusted {_fmt_aed(open_amt)}")

        st.markdown("### Top 5 Denial Codes")
        if not df_management_by_denial.empty:
            topd = df_management_by_denial.head(5).copy()
            dcols = st.columns(5)
            denial_detail_col = "DenialCodeLevel3" if "DenialCodeLevel3" in filtered_rule_detail.columns else ("DenialCode" if "DenialCode" in filtered_rule_detail.columns else None)
            for i, (_, row) in enumerate(topd.iterrows()):
                code = str(row.get("DenialCode", "Unknown"))
                rejected = float(pd.to_numeric(pd.Series([row.get("InitialRejected", 0)]), errors="coerce").fillna(0).iloc[0])
                claims = int(pd.to_numeric(pd.Series([row.get("UniqueClaims", 0)]), errors="coerce").fillna(0).iloc[0])

                top_payer = ""
                payer_amount = 0.0
                if denial_detail_col and not filtered_rule_detail.empty and "Insurance" in filtered_rule_detail.columns:
                    code_rows = filtered_rule_detail.loc[
                        filtered_rule_detail[denial_detail_col].astype(str).str.strip().str.upper().eq(code.strip().upper())
                    ].copy()
                    if not code_rows.empty:
                        payer_summary = (
                            code_rows.groupby("Insurance", dropna=False)["OriginalRejectedAmount"]
                            .sum()
                            .sort_values(ascending=False)
                        )
                        if not payer_summary.empty:
                            top_payer = str(payer_summary.index[0])
                            payer_amount = float(payer_summary.iloc[0])

                subtitle = f"{claims:,} claims"
                if top_payer:
                    subtitle += f" • Top payer: {top_payer} ({_fmt_aed(payer_amount)})"
                with dcols[i]:
                    _card(code, _fmt_aed(rejected), subtitle)

        st.info(
            "Management reading: Initial Rejected is the historical picture. The operational focus is Action Required + "
            "Reconciliation / Escalation. Pending is already with the payer. Non-Recoverable is rule-confirmed only; Needs Review should not be "
            "presented as active recoverable denial inventory."
        )

    # ------------------------------------------------------------------
    # TAB 2 — INSURANCE ANALYSIS
    # ------------------------------------------------------------------
    with tab_ins:
        st.markdown("### Denial & Recovery by Insurance")
        st.caption("One table answers: which payer denied, how much, what was recovered, and what is still actionable.")
        if df_management_by_insurance.empty:
            st.info("No insurance management summary available.")
        else:
            ins_view = df_management_by_insurance.copy()
            st.dataframe(
                ins_view.rename(columns={
                    "InitialRejected": "Initial Rejected",
                    "Recovered": "Recovered",
                    "Pending": "Pending with Payer",
                    "ActionRequired": "Action Required",
                    "Reconciliation": "Reconciliation / Escalation",
                    "AdjustmentClosed": "Non-Recoverable / Adjustment",
                    "NeedsReview": "Needs Review",
                    "UniqueClaims": "Claims",
                    "RecoveredPct": "Recovered %",
                }),
                width="stretch",
                hide_index=True,
            )

            ins_options = ["All"] + sorted(ins_view["Insurance"].dropna().astype(str).unique().tolist())
            selected_ins = st.selectbox("Focus on one insurance", ins_options, key="management_insurance_focus")

            if selected_ins != "All" and not df_denial_management.empty:
                sub = df_denial_management[
                    df_denial_management["Insurance"].astype(str) == selected_ins
                ].copy()
                if not sub.empty:
                    code_summary = (
                        sub.groupby("DenialCodeLevel3", dropna=False)
                        .agg(
                            InitialRejected=("OriginalRejectedAmount", "sum"),
                            Recovered=("RecoveredAmount", "sum"),
                            Outstanding=("OutstandingAmount", "sum"),
                            Claims=("UniqueID", _unique_claim_count),
                        )
                        .reset_index()
                        .sort_values("InitialRejected", ascending=False)
                    )
                    st.markdown(f"#### Main denial codes — {selected_ins}")
                    st.dataframe(code_summary.head(15), width="stretch", hide_index=True)

    # ------------------------------------------------------------------
    # TAB 3 — DENIAL REASONS
    # ------------------------------------------------------------------
    with tab_denial:
        st.markdown("### Denial Reasons — Financial Impact")
        st.caption(
            "Prioritize by amount first, then check whether the amount is pending, actionable, escalated, recovered or closed."
        )
        if df_management_by_denial.empty:
            st.info("No denial reason summary available.")
        else:
            st.dataframe(
                df_management_by_denial.rename(columns={
                    "DenialCode": "Denial Code",
                    "InitialRejected": "Initial Rejected",
                    "Recovered": "Recovered",
                    "Pending": "Pending with Payer",
                    "ActionRequired": "Action Required",
                    "Reconciliation": "Reconciliation / Escalation",
                    "AdjustmentClosed": "Non-Recoverable / Adjustment",
                    "NeedsReview": "Needs Review",
                    "UniqueClaims": "Claims",
                }),
                width="stretch",
                hide_index=True,
            )

            codes = ["All"] + df_management_by_denial["DenialCode"].dropna().astype(str).tolist()
            focus_code = st.selectbox("Drill into denial code", codes, key="executive_denial_focus")
            if focus_code != "All" and not df_denial_management.empty:
                code_rows = df_denial_management[
                    df_denial_management["DenialCodeLevel3"].astype(str) == focus_code
                ].copy()
                if not code_rows.empty:
                    by_payer = (
                        code_rows.groupby("Insurance", dropna=False)
                        .agg(
                            InitialRejected=("OriginalRejectedAmount", "sum"),
                            Recovered=("RecoveredAmount", "sum"),
                            Outstanding=("OutstandingAmount", "sum"),
                            Claims=("UniqueID", _unique_claim_count),
                        )
                        .reset_index()
                        .sort_values("InitialRejected", ascending=False)
                    )
                    st.markdown(f"#### {focus_code} by Insurance")
                    st.dataframe(by_payer, width="stretch", hide_index=True)

    # ------------------------------------------------------------------
    # TAB 4 — RECOVERY WORKLIST
    # ------------------------------------------------------------------
    with tab_work:
        st.markdown("### Recovery Worklist")
        st.caption("Operational view for RCM: focus on money that still needs intervention.")

        w1, w2, w3 = st.columns(3)
        with w1:
            _card("Action Required", _fmt_aed(action_amount), "Correct / resubmit / investigate")
        with w2:
            _card("Pending with Payer", _fmt_aed(pending_amount), "Follow payer turnaround")
        with w3:
            _card("Escalation", _fmt_aed(recon_amount), "Reconciliation / complaint / payer meeting")

        st.markdown("#### Needs Review")
        needs_rows = filtered_rule_detail.loc[
            filtered_rule_detail.get("ManagementOutstandingBucket", pd.Series("", index=filtered_rule_detail.index)).astype(str).eq("Needs Review")
        ].copy() if not filtered_rule_detail.empty else pd.DataFrame()
        if needs_rows.empty:
            st.success("No outstanding activities currently require manual classification review.")
        else:
            needs_summary = build_management_by_reason(needs_rows)
            st.dataframe(needs_summary, width="stretch", hide_index=True)

        st.markdown("#### Reconciliation / Escalation Candidates")
        if df_reconciliation.empty:
            st.success("No claims currently meet the 2+ resubmissions / zero-recovery escalation rule.")
        else:
            recon_claims = _unique_claim_count(df_reconciliation["UniqueID"]) if "UniqueID" in df_reconciliation.columns else len(df_reconciliation)
            st.caption(f"{recon_claims:,} unique claims currently flagged.")
            st.dataframe(df_reconciliation, width="stretch", hide_index=True)

        st.markdown("#### Trace Initial Rejection → Current Status")
        if not df_recovery_preview.empty:
            ins_options = sorted([
                str(x) for x in df_recovery_preview.get("Insurance", pd.Series(dtype=str)).dropna().unique()
                if str(x).strip()
            ])
            bucket_options = sorted([
                str(x) for x in df_recovery_preview.get("RecoveryBucket", pd.Series(dtype=str)).dropna().unique()
                if str(x).strip()
            ])

            rf1, rf2, rf3 = st.columns([1, 1, 1])
            with rf1:
                rec_ins = st.selectbox("Insurance", ["All"] + ins_options, key="recovery_insurance_filter")
            with rf2:
                rec_bucket = st.selectbox("Status", ["All"] + bucket_options, key="recovery_bucket_filter")
            with rf3:
                rec_claim = st.text_input("UniqueID contains", key="recovery_uniqueid_filter")

            rec_filt = df_recovery_preview.copy()
            if rec_ins != "All" and "Insurance" in rec_filt.columns:
                rec_filt = rec_filt[rec_filt["Insurance"].astype(str) == rec_ins]
            if rec_bucket != "All" and "RecoveryBucket" in rec_filt.columns:
                rec_filt = rec_filt[rec_filt["RecoveryBucket"].astype(str) == rec_bucket]
            if rec_claim.strip() and "UniqueID" in rec_filt.columns:
                rec_filt = rec_filt[
                    rec_filt["UniqueID"].astype(str).str.contains(rec_claim.strip(), case=False, na=False)
                ]

            st.dataframe(rec_filt, width="stretch", hide_index=True)

    # ------------------------------------------------------------------
    # TAB 5 — RCM DETAIL
    # ------------------------------------------------------------------
    with tab_detail:
        st.markdown("### RCM Detailed Analysis")
        with st.expander("View Active Denial Rule Master", expanded=False):
            st.caption("Insurance-specific rules take priority over global rules. This table is exported in the Excel workbook too.")
            st.dataframe(df_rule_master, width="stretch", hide_index=True)
        st.caption(
            "The detailed analytical tools are kept here so management is not overloaded, while RCM still has full drill-down."
        )

        dtab1, dtab2, dtab3, dtab4, dtab5 = st.tabs([
            "By Insurance",
            "By Denial Code",
            "Insurance × Denial",
            "Aging",
            "Rejected Detail",
        ])

        with dtab1:
            st.dataframe(df_by_ins, width="stretch")

        with dtab2:
            st.dataframe(df_by_code, width="stretch")

        with dtab3:
            st.dataframe(df_ins_x_code, width="stretch")

        with dtab4:
            st.dataframe(df_aging, width="stretch")

        with dtab5:
            ins_list = sorted([
                x for x in df_preview["Insurance"].dropna().unique().tolist()
                if str(x).strip() != ""
            ]) if "Insurance" in df_preview.columns else []
            _detail_code_col = "DenialCode" if "DenialCode" in df_preview.columns else ("DenialCodeLevel3" if "DenialCodeLevel3" in df_preview.columns else None)
            code_list = sorted([
                x for x in df_preview[_detail_code_col].dropna().unique().tolist()
                if str(x).strip() != ""
            ]) if _detail_code_col else []

            c1, c2, c3 = st.columns([1, 1, 1])
            with c1:
                sel_ins = st.selectbox("Insurance", ["All"] + ins_list, key="rej_filter_ins")
            with c2:
                sel_code = st.selectbox("Denial Code", ["All"] + code_list, key="rej_filter_code")
            with c3:
                show_top = st.number_input(
                    "Preview rows", min_value=50, max_value=2000,
                    value=500, step=50, key="rej_preview_rows"
                )

            filt = df_preview.copy()
            if sel_ins != "All" and "Insurance" in filt.columns:
                filt = filt[filt["Insurance"].astype(str) == str(sel_ins)]
            if sel_code != "All" and _detail_code_col and _detail_code_col in filt.columns:
                filt = filt[filt[_detail_code_col].astype(str) == str(sel_code)]

            st.dataframe(filt.head(int(show_top)), width="stretch", hide_index=True)

            if st.button("Build Filtered Detail Excel", type="primary", key="rej_dl_btn"):
                with st.spinner("Preparing full filtered detail..."):
                    xls_full = pd.ExcelFile(io.BytesIO(out_xlsx_bytes), engine="openpyxl")
                    df_full = pd.read_excel(xls_full, sheet_name="Rejected_Detail")

                    if sel_ins != "All" and "Insurance" in df_full.columns:
                        df_full = df_full[df_full["Insurance"].astype(str) == str(sel_ins)]
                    if sel_code != "All":
                        _full_code_col = "DenialCode" if "DenialCode" in df_full.columns else ("DenialCodeLevel3" if "DenialCodeLevel3" in df_full.columns else None)
                        if _full_code_col:
                            df_full = df_full[df_full[_full_code_col].astype(str) == str(sel_code)]

                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        df_full.to_excel(writer, sheet_name="Rejected_Detail_Filtered", index=False)

                    safe_name = f"Rejected_Detail_{R['center']}_{R['year']}_{sel_ins}_{sel_code}_{stats['sha1']}.xlsx"
                    safe_name = (
                        safe_name.replace(" ", "_")
                        .replace("/", "_")
                        .replace("\\", "_")
                        .replace(":", "_")
                    )

                    st.download_button(
                        "Download Filtered Detail Excel",
                        data=buf.getvalue(),
                        file_name=safe_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    st.success(f"Filtered rows: {len(df_full):,} ✅")

        # Keep the technical recovery tables available, but collapsed.
        with st.expander("Technical recovery / resubmission tables"):
            if not df_recovery_summary.empty:
                st.write("**Recovery Status Summary**")
                st.dataframe(df_recovery_summary, width="stretch")
            if not df_recovery_by_insurance.empty:
                st.write("**Recovery by Insurance**")
                st.dataframe(df_recovery_by_insurance, width="stretch")
            if not df_resub_stage_summary.empty:
                st.write("**Resubmission Funnel — Cumulative**")
                st.dataframe(df_resub_stage_summary, width="stretch")
            if not df_resub_stage_summary_excl.empty:
                st.write("**Resubmission Stages — Exclusive**")
                st.dataframe(df_resub_stage_summary_excl, width="stretch")


run_rejection_app()
