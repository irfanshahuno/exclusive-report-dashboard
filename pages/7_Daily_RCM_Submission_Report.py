#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Streamlit Page: Daily RCM Submission Report

Purpose
-------
Upload the daily claim-status report and calculate:
- NOT ASSIGNED = not coded yet / within 48-hour coding window
- CLOSED       = done / already submitted
- OPEN         = pending for query (doctor, nursing/lab, reception, etc.)
- PROCESSED    = complete / ready to submit

Amount basis:
- "Ins Share" = net insurance amount

Main analysis:
- Status-wise claim count + Ins Share amount
- Open-query department classification using User remark
- Insurance-wise status/count/amount
- Doctor-wise status/count/amount
- Not Assigned >48 hours alert
- Saved results in S3 so the latest report remains after Streamlit reopens

The report uses Visit No as the claim key where available.
"""

import io
import os
import re
import pickle
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
import html

# Optional S3
try:
    import boto3
except Exception:
    boto3 = None


# =========================================================
# PAGE
# =========================================================
st.set_page_config(
    page_title="Daily RCM Submission Report",
    layout="wide",
    initial_sidebar_state="collapsed",
)

SS = st.session_state


# =========================================================
# CONFIG / CENTER / S3
# =========================================================
CENTERS = {
    "easyhealth": "Easy Health Medical Clinic (MF8031)",
    "excellent": "Excellent Medical Center (MF4777)",
    "pharmacy": "Excellent Pharmacy (PF3205)",
}

SS.setdefault("center_key", "excellent")
_q_center = st.query_params.get("center")
if _q_center in CENTERS:
    SS["center_key"] = _q_center

center_key = SS.get("center_key", "excellent")
if center_key not in CENTERS:
    center_key = "excellent"
    SS["center_key"] = center_key


def load_secrets() -> Dict[str, str]:
    def get_any(*keys):
        for k in keys:
            if k in st.secrets:
                v = st.secrets.get(k)
                if v is not None and str(v).strip():
                    return str(v).strip()
            v = os.getenv(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ""

    return {
        "AWS_ACCESS_KEY_ID": get_any("AWS_ACCESS_KEY_ID"),
        "AWS_SECRET_ACCESS_KEY": get_any("AWS_SECRET_ACCESS_KEY"),
        "AWS_REGION": get_any("AWS_REGION", "AWS_DEFAULT_REGION"),
        "S3_BUCKET_NAME": get_any("S3_BUCKET_NAME", "S3_BUCKET"),
        "S3_BASE_PREFIX": get_any("S3_BASE_PREFIX", "S3_PREFIX"),
    }


def s3_enabled(cfg: Dict[str, str]) -> bool:
    return bool(
        cfg.get("S3_BUCKET_NAME")
        and cfg.get("AWS_REGION")
        and cfg.get("AWS_ACCESS_KEY_ID")
        and cfg.get("AWS_SECRET_ACCESS_KEY")
        and boto3 is not None
    )


@st.cache_resource(show_spinner=False)
def s3_client_cached(cfg: Dict[str, str]):
    if not s3_enabled(cfg):
        return None
    return boto3.client(
        "s3",
        region_name=cfg["AWS_REGION"],
        aws_access_key_id=cfg["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=cfg["AWS_SECRET_ACCESS_KEY"],
    )


def s3_key(*parts: str) -> str:
    return "/".join(
        [str(p).strip("/").strip() for p in parts if p is not None and str(p).strip()]
    )


def daily_root(cfg: Dict[str, str], center: str) -> str:
    return s3_key(cfg.get("S3_BASE_PREFIX", ""), "daily_rcm", center)


def s3_put_bytes(s3, bucket: str, key: str, data: bytes, content_type="application/octet-stream"):
    s3.put_object(Bucket=bucket, Key=key, Body=data, ContentType=content_type)


def s3_get_bytes(s3, bucket: str, key: str) -> Optional[bytes]:
    try:
        obj = s3.get_object(Bucket=bucket, Key=key)
        return obj["Body"].read()
    except Exception:
        return None


cfg = load_secrets()
s3_ok = s3_enabled(cfg)
s3 = s3_client_cached(cfg) if s3_ok else None

# Respect center already selected in the main dashboard, while still allowing manual change.
center_key = st.selectbox(
    "Center",
    options=list(CENTERS.keys()),
    index=list(CENTERS.keys()).index(center_key),
    format_func=lambda x: CENTERS[x],
    key="daily_rcm_center_selector",
)
SS["center_key"] = center_key


# =========================================================
# STYLE
# =========================================================
st.markdown(
    """
<style>
.block-container{
    max-width:100% !important;
    padding-top:0.7rem !important;
    padding-left:2rem !important;
    padding-right:2rem !important;
}
h1,h2,h3{letter-spacing:-0.02em;}
.rcm-kpi-grid{
    display:grid;
    grid-template-columns:repeat(3,minmax(0,1fr));
    gap:10px;
    margin:.25rem 0 .75rem 0;
}
.rcm-card{
    min-height:98px;
    padding:12px 16px;
    border-radius:16px;
    border:1px solid rgba(15,23,42,.08);
    box-shadow:0 6px 16px rgba(15,23,42,.055);
    display:flex;
    align-items:center;
    gap:13px;
}
.rcm-blue{background:#EEF6FF;border-color:#B9D5F3;}
.rcm-green{background:#ECF9F1;border-color:#B8DFC6;}
.rcm-yellow{background:#FFF4D9;border-color:#EBCB72;}
.rcm-red{background:#FFF0F0;border-color:#EFB1B1;}
.rcm-purple{background:#F1F0FF;border-color:#C9C3F2;}
.rcm-white{background:#EEF9F8;border:1px solid #B8DDD8;}
.rcm-icon{font-size:22px;min-width:34px;text-align:center;line-height:1;color:#17335F;font-weight:900;}
.rcm-label{font-size:12px;font-weight:800;color:#4B607A;margin-bottom:4px;text-transform:uppercase;letter-spacing:.25px;}
.rcm-value{font-size:28px;font-weight:950;color:#0B2342;line-height:1.05;letter-spacing:-.35px;}
.rcm-sub{font-size:11px;font-weight:700;color:#64748B;margin-top:5px;}
.premium-header{
    display:flex;
    justify-content:space-between;
    align-items:center;
    gap:16px;
    margin:.25rem 0 .65rem 0;
    padding:16px 20px;
    border-radius:18px;
    background:linear-gradient(135deg,#0B2342 0%,#153A63 100%);
    box-shadow:0 9px 24px rgba(7,26,93,.16);
}
.premium-header-title{
    color:white;
    font-size:26px;
    font-weight:900;
    letter-spacing:-.02em;
}
.premium-header-sub{
    color:#cbdaf5;
    font-size:12px;
    font-weight:650;
    margin-top:3px;
}
.premium-table-wrap{
    border:1px solid #dfe6ef;
    border-radius:15px;
    overflow:hidden;
    box-shadow:0 6px 18px rgba(15,23,42,.05);
    margin-bottom:.8rem;
}
.premium-table{
    width:100%;
    border-collapse:collapse;
    background:white;
}
.premium-table th{
    background:#0b2342;
    color:white;
    text-align:left;
    padding:11px 14px;
    font-size:12px;
    font-weight:800;
}
.premium-table td{
    padding:11px 14px;
    border-bottom:1px solid #edf1f5;
    font-size:13px;
    color:#263447;
}
.premium-table td.num{
    text-align:right;
    font-variant-numeric:tabular-nums;
}
.premium-table tr:nth-child(even):not(.total-row){
    background:#f7faff;
}
.premium-table tr.total-row{
    background:#ff7a00;
}
.premium-table tr.total-row td{
    color:#FFFFFF !important;
    font-weight:900 !important;
    border-bottom:none;
}
.premium-table tr.total-row td *{
    color:#FFFFFF !important;
}
div[data-testid="stButton"] > button{
    border-radius:11px !important;
    font-weight:800 !important;
}
div[data-testid="stDataFrame"]{
    border-radius:14px;
    overflow:hidden;
    box-shadow:0 5px 15px rgba(15,23,42,.045);
}


.exec-strip{
    display:grid;
    grid-template-columns:repeat(3,minmax(0,1fr));
    gap:10px;
    margin:.15rem 0 .85rem 0;
}
.exec-item{
    background:#FFFFFF;
    border:1px solid #E2E8F0;
    border-radius:12px;
    padding:10px 14px;
    box-shadow:0 3px 10px rgba(15,23,42,.035);
}
.exec-label{
    color:#64748B;
    font-size:11px;
    font-weight:800;
    text-transform:uppercase;
    letter-spacing:.3px;
}
.exec-value{
    color:#0B2342;
    font-size:20px;
    font-weight:900;
    margin-top:2px;
}
.exec-good{color:#16784A;}
.exec-warn{color:#A66B00;}
.exec-bad{color:#B42318;}
@media(max-width:800px){.exec-strip{grid-template-columns:1fr;}}

.rcm-section{
    margin-top:.8rem;
    margin-bottom:.35rem;
    font-size:1.45rem;
    font-weight:850;
    color:#202939;
}
@media(max-width:1000px){
  .rcm-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
}
@media(max-width:650px){
  .rcm-kpi-grid{grid-template-columns:repeat(1,minmax(0,1fr));}
}
</style>
""",
    unsafe_allow_html=True,
)


def money(v) -> str:
    try:
        return f"AED {float(v):,.2f}"
    except Exception:
        return "AED 0.00"


def kpi_cards(items: List[Tuple[str, str, str, str, str]]):
    """item = (label, value, subtitle, icon, color_class)"""
    cards = []
    for label, value, subtitle, icon, cls in items:
        cards.append(
            f'<div class="rcm-card {cls}">'
            f'<div class="rcm-icon">{icon}</div>'
            f'<div>'
            f'<div class="rcm-label">{label}</div>'
            f'<div class="rcm-value">{value}</div>'
            f'<div class="rcm-sub">{subtitle}</div>'
            f'</div>'
            f'</div>'
        )
    html = '<div class="rcm-kpi-grid">' + ''.join(cards) + '</div>'
    st.markdown(html, unsafe_allow_html=True)


# =========================================================
# COLUMN HELPERS
# =========================================================
def norm_col(v: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(v).strip().lower())


def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cmap = {norm_col(c): c for c in df.columns}
    for cand in candidates:
        n = norm_col(cand)
        if n in cmap:
            return cmap[n]

    # cautious contains fallback
    for cand in candidates:
        n = norm_col(cand)
        if not n:
            continue
        for nc, original in cmap.items():
            if n == nc or (len(n) >= 5 and n in nc):
                return original
    return None


def clean_amount(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype=float)
    s = series.astype(str).str.replace(",", "", regex=False)
    s = s.str.replace("AED", "", case=False, regex=False)
    s = s.str.replace(r"[^\d\.\-]", "", regex=True)
    return pd.to_numeric(s, errors="coerce").fillna(0.0)


def parse_date_series(series: pd.Series) -> pd.Series:
    s1 = pd.to_datetime(series, errors="coerce", dayfirst=False)
    s2 = pd.to_datetime(series, errors="coerce", dayfirst=True)
    return s2 if s2.notna().sum() >= s1.notna().sum() else s1


# =========================================================
# EXCEL LOADER
# =========================================================
EXPECTED_HEADERS = [
    "Visit No",
    "Visit Date",
    "Ins Share",
    "Doctor Name",
    "Ins. Company",
    "Status",
    "User remark",
]


def read_daily_report(file_obj, filename: str) -> pd.DataFrame:
    """
    Reads .xls/.xlsx and also scans the first rows if the true header is not row 1.
    For legacy .xls, Streamlit environment should include xlrd>=2.0.1.
    """
    data = file_obj.getvalue() if hasattr(file_obj, "getvalue") else file_obj.read()
    bio = io.BytesIO(data)

    def header_score(columns) -> int:
        norms = {norm_col(c) for c in columns}
        return sum(1 for h in EXPECTED_HEADERS if norm_col(h) in norms)

    # First attempt
    try:
        bio.seek(0)
        df = pd.read_excel(bio)
        if header_score(df.columns) >= 4:
            df.columns = [str(c).strip() for c in df.columns]
            return df.dropna(how="all").reset_index(drop=True)
    except ImportError as exc:
        if str(filename).lower().endswith(".xls"):
            raise RuntimeError(
                "Legacy .xls support is missing. Add `xlrd==2.0.1` to requirements.txt."
            ) from exc
    except Exception:
        pass

    # Header scan fallback
    try:
        bio.seek(0)
        raw = pd.read_excel(bio, header=None)
    except ImportError as exc:
        raise RuntimeError(
            "Legacy .xls support is missing. Add `xlrd==2.0.1` to requirements.txt."
        ) from exc

    best_row = None
    best_score = -1
    for r in range(min(60, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[r].tolist()]
        score = header_score(vals)
        if score > best_score:
            best_score = score
            best_row = r

    if best_row is None or best_score < 4:
        raise ValueError(
            "Could not detect the daily report header. Expected columns such as "
            "Visit No, Ins Share, Doctor Name, Ins. Company, Status and User remark."
        )

    header = [str(v).strip() for v in raw.iloc[best_row].tolist()]
    df = raw.iloc[best_row + 1 :].copy()
    df.columns = header
    df = df.dropna(how="all").reset_index(drop=True)
    return df



# =========================================================
# REGISTRATION + DAILY REVENUE HELPERS
# =========================================================
def _file_bytes(file_obj) -> bytes:
    if hasattr(file_obj, "getvalue"):
        return file_obj.getvalue()
    data = file_obj.read()
    try:
        file_obj.seek(0)
    except Exception:
        pass
    return data


def _read_excel_header_scan(file_obj, expected_headers: List[str], sheet_name=None, min_score: int = 2) -> pd.DataFrame:
    """Generic Excel loader that finds the true header row in the first 60 rows."""
    data = _file_bytes(file_obj)
    bio = io.BytesIO(data)

    def score(cols):
        norms = {norm_col(c) for c in cols}
        return sum(1 for h in expected_headers if norm_col(h) in norms)

    # Direct read first.
    try:
        bio.seek(0)
        df = pd.read_excel(bio, sheet_name=sheet_name if sheet_name is not None else 0)
        if score(df.columns) >= min_score:
            df.columns = [str(c).strip() for c in df.columns]
            return df.dropna(how="all").reset_index(drop=True)
    except Exception:
        pass

    # Header scan.
    bio.seek(0)
    raw = pd.read_excel(bio, sheet_name=sheet_name if sheet_name is not None else 0, header=None)
    best_row, best_score = None, -1
    for r in range(min(60, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[r].tolist()]
        sc = score(vals)
        if sc > best_score:
            best_row, best_score = r, sc
    if best_row is None or best_score < min_score:
        raise ValueError("Could not detect the report header.")
    header = [str(v).strip() for v in raw.iloc[best_row].tolist()]
    df = raw.iloc[best_row + 1:].copy()
    df.columns = header
    return df.dropna(how="all").reset_index(drop=True)


def read_revenue_report(file_obj) -> pd.DataFrame:
    """Load Daily Collection Details using the same logic as the previous Registration Summary revenue module."""
    data = _file_bytes(file_obj)
    bio = io.BytesIO(data)
    try:
        return _read_excel_header_scan(
            bio,
            ["Visit Date", "Visit No", "Insurance Name", "Department", "Doctor", "Consultation", "Lab", "Procedure", "Insuance"],
            sheet_name="Daily Collection Details",
            min_score=6,
        )
    except Exception:
        bio.seek(0)
        return _read_excel_header_scan(
            bio,
            ["Visit Date", "Visit No", "Insurance Name", "Department", "Doctor", "Consultation", "Lab", "Procedure", "Insuance"],
            sheet_name=0,
            min_score=6,
        )


def read_registration_report(file_obj) -> pd.DataFrame:
    return _read_excel_header_scan(
        file_obj,
        ["Visit No", "Reg:Date", "Doctor"],
        sheet_name=0,
        min_score=2,
    )


def _date_bounds(selected_day):
    # Streamlit date_input returns either one date or a (start, end) tuple.
    if isinstance(selected_day, (tuple, list)):
        if len(selected_day) == 0:
            return None, None
        start = pd.to_datetime(selected_day[0]).normalize()
        end = pd.to_datetime(selected_day[-1]).normalize()
    else:
        start = end = pd.to_datetime(selected_day).normalize()
    if start > end:
        start, end = end, start
    return start, end


def _period_label(selected_day) -> str:
    start, end = _date_bounds(selected_day)
    if start is None:
        return "All available dates"
    if start == end:
        return start.strftime("%A, %d %b %Y")
    return f"{start.strftime('%d %b %Y')} – {end.strftime('%d %b %Y')}"


def _date_set_in_period(df: pd.DataFrame, candidates: List[str], selected_day) -> set:
    if df is None or df.empty:
        return set()
    c = find_col(df, candidates)
    if not c:
        return set()
    d = parse_date_series(df[c]).dt.normalize()
    start, end = _date_bounds(selected_day)
    if start is not None:
        d = d[d.between(start, end, inclusive="both")]
    return {pd.Timestamp(x).date() for x in d.dropna().unique()}


def _filter_by_day(df: pd.DataFrame, date_col: str, selected_day) -> pd.DataFrame:
    if df is None or df.empty or not date_col or date_col not in df.columns:
        return pd.DataFrame(columns=df.columns if isinstance(df, pd.DataFrame) else None)
    d = parse_date_series(df[date_col]).dt.normalize()
    start, end = _date_bounds(selected_day)
    if start is None:
        return df.copy().reset_index(drop=True)
    return df.loc[d.between(start, end, inclusive="both")].copy().reset_index(drop=True)


def registration_patient_count(reg_df: pd.DataFrame, selected_day) -> int:
    if reg_df is None or reg_df.empty:
        return 0
    c_visit = find_col(reg_df, ["Visit No", "VisitNo", "Visit ID", "VisitID"])
    c_date = find_col(reg_df, ["Reg:Date", "Reg Date", "Registration Date", "RegistrationDate", "Date"])
    if not c_visit or not c_date:
        raise ValueError("Registration report must contain Visit No and Reg:Date.")
    d = _filter_by_day(reg_df, c_date, selected_day)
    visits = d[c_visit].fillna("").astype(str).str.strip()
    visits = visits[~visits.str.lower().isin(["", "nan", "none"])]
    return int(visits.nunique())


def revenue_analysis(rev_df: pd.DataFrame, selected_day) -> Dict[str, object]:
    """
    Previous daily-revenue logic:
    - unique Visit No = visit count
    - Service Revenue = Consultation + Lab + Radiology + Procedure
    - Insurance Amount = strict 'Insuance' column
    - service counts = unique visits where that service amount > 0
    """
    if rev_df is None or rev_df.empty:
        return {"daily": pd.DataFrame(), "doctor": pd.DataFrame(), "service_counts": {}, "totals": {}}

    c_date = find_col(rev_df, ["Visit Date", "VisitDate"])
    c_visit = find_col(rev_df, ["Visit No", "VisitNo", "Visit ID", "VisitID"])
    c_doc = find_col(rev_df, ["Doctor"])
    c_dept = find_col(rev_df, ["Department"])
    c_cons = find_col(rev_df, ["Consultation"])
    c_lab = find_col(rev_df, ["Lab"])
    c_rad = find_col(rev_df, ["Radiology", "Radiology Amount", "X-Ray", "Xray", "Ultrasound", "USG"])
    c_proc = find_col(rev_df, ["Procedure"])
    # Preserve the prior script's strict typo-based insurance amount logic.
    c_insu = next((c for c in rev_df.columns if norm_col(c) == "insuance"), None)

    required = {"Visit Date": c_date, "Visit No": c_visit, "Doctor": c_doc,
                "Consultation": c_cons, "Lab": c_lab, "Procedure": c_proc, "Insuance": c_insu}
    missing = [k for k,v in required.items() if v is None]
    if missing:
        raise ValueError("Daily Revenue report missing required column(s): " + ", ".join(missing))

    d = _filter_by_day(rev_df, c_date, selected_day)
    if d.empty:
        return {"daily": d, "doctor": pd.DataFrame(), "service_counts": {"Consultation":0,"Lab":0,"Radiology":0,"Procedure":0},
                "totals": {"visits":0,"service_revenue":0.0,"insurance_amount":0.0,"avg_service":0.0,"avg_insurance":0.0}}

    d[c_visit] = d[c_visit].fillna("").astype(str).str.strip()
    d[c_doc] = d[c_doc].fillna("UNKNOWN").astype(str).str.strip().replace("", "UNKNOWN")
    for c in [c_cons, c_lab, c_proc, c_insu]:
        d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0.0)
    if c_rad:
        d[c_rad] = pd.to_numeric(d[c_rad], errors="coerce").fillna(0.0)
    else:
        d["_Radiology"] = 0.0
        c_rad = "_Radiology"

    d = d[d[c_visit] != ""].copy()
    d["_ServiceRevenue"] = d[c_cons] + d[c_lab] + d[c_rad] + d[c_proc]
    d["_InsuranceAmount"] = d[c_insu]

    group_cols = ([c_dept] if c_dept else []) + [c_doc]
    # Doctor table: service columns are COUNTS of unique visits, not AED amounts.
    base = d.groupby(group_cols, dropna=False).agg(
        Visits=(c_visit, pd.Series.nunique),
        Total_Service_Revenue=("_ServiceRevenue", "sum"),
        Insurance_Amount=("_InsuranceAmount", "sum"),
    ).reset_index()

    doctor = base.copy()
    for label, amount_col in [("Consultation", c_cons), ("Lab", c_lab), ("Procedure", c_proc), ("Radiology", c_rad)]:
        positive = d[pd.to_numeric(d[amount_col], errors="coerce").fillna(0) > 0]
        cnt = positive.groupby(group_cols, dropna=False)[c_visit].nunique().rename(label).reset_index()
        doctor = doctor.merge(cnt, on=group_cols, how="left")
        doctor[label] = pd.to_numeric(doctor[label], errors="coerce").fillna(0).astype(int)
    rename = {c_doc: "Doctor"}
    if c_dept:
        rename[c_dept] = "Department"
    doctor = doctor.rename(columns=rename)
    denom = doctor["Visits"].replace(0, pd.NA)
    doctor["Avg_Service_Per_Visit"] = (doctor["Total_Service_Revenue"] / denom).fillna(0.0)
    doctor["Avg_Insurance_Per_Visit"] = (doctor["Insurance_Amount"] / denom).fillna(0.0)

    # Unique-visit service counts, identical to the previous logic.
    def svc_count(c):
        mask = pd.to_numeric(d[c], errors="coerce").fillna(0) > 0
        return int(d.loc[mask, c_visit].nunique())
    counts = {
        "Consultation": svc_count(c_cons),
        "Lab": svc_count(c_lab),
        "Radiology": svc_count(c_rad),
        "Procedure": svc_count(c_proc),
    }
    visits = int(d[c_visit].nunique())
    service_total = float(d["_ServiceRevenue"].sum())
    ins_total = float(d["_InsuranceAmount"].sum())
    totals = {
        "visits": visits,
        "service_revenue": service_total,
        "insurance_amount": ins_total,
        "avg_service": service_total / visits if visits else 0.0,
        "avg_insurance": ins_total / visits if visits else 0.0,
    }
    doctor = doctor.sort_values("Total_Service_Revenue", ascending=False).reset_index(drop=True)
    return {"daily": d, "doctor": doctor, "service_counts": counts, "totals": totals}


def _available_dates_from_report(df: pd.DataFrame, candidates: List[str]) -> List[pd.Timestamp]:
    if df is None or df.empty:
        return []
    c = find_col(df, candidates)
    if not c:
        return []
    s = parse_date_series(df[c]).dropna().dt.normalize().drop_duplicates().sort_values()
    return list(s)

# =========================================================
# QUERY OWNER CLASSIFICATION
# =========================================================
def classify_query_owner(remark: object) -> str:
    s = "" if pd.isna(remark) else str(remark).strip().lower()
    s = re.sub(r"\s+", " ", s)

    if not s:
        return "Unspecified"

    # Explicit doctor wording takes priority even if the sentence later mentions lab.
    doctor_terms = [
        "dear doctor", "dear dr", "dear doc", "doctor please", "dr please",
        "chief complaint", "chief complains", "laterality", "diagnosis",
        "clinical note", "medical note", "specify diagnosis", "add diagnosis",
    ]
    if any(x in s for x in doctor_terms):
        return "Doctor"

    # User rule: if remark says lab -> Nursing / Lab department.
    nursing_lab_terms = [
        "lab", "laboratory", "sample", "specimen", "nurse", "nursing",
        "vital", "temperature", "bp reading", "blood pressure",
    ]
    if any(x in s for x in nursing_lab_terms):
        return "Nursing / Lab"

    reception_terms = [
        "reception", "registration", "card no", "card number", "emirates id",
        "patient id", "mobile", "phone", "demographic", "visit type",
        "eligibility", "member id", "policy no", "policy number",
    ]
    if any(x in s for x in reception_terms):
        return "Reception"

    approval_terms = [
        "approval", "authorization", "authorisation", "pre approval",
        "preapproval", "insurance confirmation", "benefit", "coverage",
        "network", "portal",
    ]
    if any(x in s for x in approval_terms):
        return "Insurance / Approval"

    rcm_terms = [
        "coding", "coder", "cpt", "icd", "modifier", "billing", "claim",
        "resubmit", "resubmission",
    ]
    if any(x in s for x in rcm_terms):
        return "RCM / Coding"

    return "Other"


# =========================================================
# PROCESSING
# =========================================================
STATUS_ORDER = ["CLOSED", "PROCESSED", "OPEN", "NOT ASSIGNED"]


def process_report(raw: pd.DataFrame, selected_day=None) -> Dict[str, object]:
    df = raw.copy()
    df.columns = [str(c).strip() for c in df.columns]

    col_visit = find_col(df, ["Visit No", "VisitNo", "Visit Number", "Visit ID", "VisitID"])
    col_visit_date = find_col(df, ["Visit Date", "VisitDate", "Enc Date", "Encounter Date"])
    col_amount = find_col(df, ["Ins Share", "Insurance Share", "InsShare"])
    col_status = find_col(df, ["Status"])
    col_claim_status = find_col(df, ["ClaimStatus", "Claim Status"])
    col_remark = find_col(df, ["User remark", "User Remark", "Remark", "Remarks"])
    col_ins = find_col(df, ["Ins. Company", "Ins Company", "Insurance Company", "Payer"])
    col_ins_type = find_col(df, ["Ins. Type", "Ins Type", "Insurance Type", "Payer Type"])
    col_doc = find_col(df, ["Doctor Name", "Doctor"])
    col_patient = find_col(df, ["Patient Name", "Name"])
    col_opened = find_col(df, ["Opened Date", "Open Date"])
    col_processed = find_col(df, ["Processed Date"])
    col_closed = find_col(df, ["Closed Date"])
    col_submission = find_col(df, ["Submission Date"])
    col_assigned = find_col(df, ["Assigned Date"])

    required = {
        "Visit No": col_visit,
        "Ins Share": col_amount,
        "Status": col_status,
        "Ins. Company": col_ins,
        "Doctor Name": col_doc,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError("Missing required column(s): " + ", ".join(missing))

    # Canonical fields
    df["_VisitNo"] = df[col_visit].fillna("").astype(str).str.strip()
    df["_Amount"] = clean_amount(df[col_amount])
    df["_Status"] = (
        df[col_status]
        .fillna("")
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.upper()
    )
    df["_Insurance"] = df[col_ins].fillna("UNKNOWN").astype(str).str.strip()
    df["_Doctor"] = df[col_doc].fillna("UNKNOWN").astype(str).str.strip()
    df["_Remark"] = df[col_remark].fillna("").astype(str).str.strip() if col_remark else ""
    df["_ClaimStatus"] = (
        df[col_claim_status].fillna("").astype(str).str.strip()
        if col_claim_status else ""
    )

    if col_visit_date:
        df["_VisitDate"] = parse_date_series(df[col_visit_date])
    else:
        df["_VisitDate"] = pd.NaT

    if col_opened:
        df["_OpenedDate"] = parse_date_series(df[col_opened])
    else:
        df["_OpenedDate"] = pd.NaT

    if col_processed:
        df["_ProcessedDate"] = parse_date_series(df[col_processed])
    else:
        df["_ProcessedDate"] = pd.NaT

    if col_closed:
        df["_ClosedDate"] = parse_date_series(df[col_closed])
    else:
        df["_ClosedDate"] = pd.NaT

    if col_submission:
        df["_SubmissionDate"] = parse_date_series(df[col_submission])
    else:
        df["_SubmissionDate"] = pd.NaT

    if col_assigned:
        df["_AssignedDate"] = parse_date_series(df[col_assigned])
    else:
        df["_AssignedDate"] = pd.NaT

    # Remove clearly empty rows
    df = df[(df["_VisitNo"] != "") | (df["_Status"] != "")].copy()

    # RCM submission is INSURANCE ONLY — exclude cash/self-pay patients completely.
    # Cash can be identified either from Ins. Type or Ins. Company depending on the export.
    cash_terms = r"\b(CASH|SELF[ -]?PAY|SELF[ -]?PAYMENT|PRIVATE[ -]?PAY|CASH[ -]?PATIENT)\b"
    cash_mask = pd.Series(False, index=df.index)
    if col_ins_type:
        cash_mask = cash_mask | df[col_ins_type].fillna("").astype(str).str.upper().str.contains(cash_terms, regex=True, na=False)
    if col_ins:
        cash_mask = cash_mask | df[col_ins].fillna("").astype(str).str.upper().str.contains(cash_terms, regex=True, na=False)
    df = df.loc[~cash_mask].copy()

    # Calendar/range filter: all three reports use the same selected period.
    if selected_day is not None and df["_VisitDate"].notna().any():
        _start, _end = _date_bounds(selected_day)
        df = df[df["_VisitDate"].dt.normalize().between(_start, _end, inclusive="both")].copy()

    # One claim = one Visit No. Keep last report row when duplicates exist.
    # Blank Visit No rows remain as separate rows.
    with_visit = df[df["_VisitNo"] != ""].drop_duplicates(subset=["_VisitNo"], keep="last")
    without_visit = df[df["_VisitNo"] == ""]
    claims = pd.concat([with_visit, without_visit], ignore_index=True)

    claims["_QueryOwner"] = claims["_Remark"].apply(classify_query_owner)

    # Reporting day: selected calendar day when supplied, otherwise most common Visit Date.
    if selected_day is not None:
        _start, _end = _date_bounds(selected_day)
        report_day = _end
    else:
        valid_days = claims["_VisitDate"].dropna().dt.normalize()
        if not valid_days.empty:
            report_day = valid_days.value_counts().index[0]
        else:
            report_day = pd.Timestamp.now(tz=ZoneInfo("Asia/Dubai")).tz_localize(None).normalize()

    # SLA age for NOT ASSIGNED
    now_dubai = pd.Timestamp.now(tz=ZoneInfo("Asia/Dubai")).tz_localize(None)
    claims["_AgeHours"] = (
        (now_dubai - claims["_VisitDate"]).dt.total_seconds() / 3600.0
    )
    claims["_NotAssignedOver48h"] = (
        claims["_Status"].eq("NOT ASSIGNED")
        & claims["_AgeHours"].notna()
        & claims["_AgeHours"].gt(48)
    )

    # Status summary
    status_rows = []
    for status in STATUS_ORDER:
        part = claims[claims["_Status"] == status]
        status_rows.append({
            "Status": status,
            "Claims": int(len(part)),
            "Ins Share": float(part["_Amount"].sum()),
        })
    status_summary = pd.DataFrame(status_rows)

    # Any other statuses
    other_status = claims[~claims["_Status"].isin(STATUS_ORDER)].copy()
    if not other_status.empty:
        extra = (
            other_status.groupby("_Status", dropna=False)
            .agg(Claims=("_VisitNo", "size"), **{"Ins Share": ("_Amount", "sum")})
            .reset_index()
            .rename(columns={"_Status": "Status"})
        )
        status_summary = pd.concat([status_summary, extra], ignore_index=True)

    # Query owner breakdown only OPEN
    open_df = claims[claims["_Status"] == "OPEN"].copy()
    if open_df.empty:
        query_summary = pd.DataFrame(columns=["Query Department", "Claims", "Ins Share"])
    else:
        query_summary = (
            open_df.groupby("_QueryOwner", dropna=False)
            .agg(Claims=("_VisitNo", "size"), **{"Ins Share": ("_Amount", "sum")})
            .reset_index()
            .rename(columns={"_QueryOwner": "Query Department"})
            .sort_values(["Claims", "Ins Share"], ascending=[False, False])
        )

    def build_group_summary(group_col: str, display_name: str) -> pd.DataFrame:
        rows = []
        for grp, gdf in claims.groupby(group_col, dropna=False):
            row = {
                display_name: grp if str(grp).strip() else "UNKNOWN",
                "Total Claims": int(len(gdf)),
                "Total Ins Share": float(gdf["_Amount"].sum()),
            }
            for st in STATUS_ORDER:
                p = gdf[gdf["_Status"] == st]
                row[f"{st} Claims"] = int(len(p))
                row[f"{st} Amount"] = float(p["_Amount"].sum())
            rows.append(row)
        out = pd.DataFrame(rows)
        if not out.empty:
            out = out.sort_values("Total Ins Share", ascending=False)
            out = out.rename(columns={
                "CLOSED Claims": "Already Submitted Claims",
                "CLOSED Amount": "Already Submitted Amount",
                "PROCESSED Claims": "Ready to Submit Claims",
                "PROCESSED Amount": "Ready to Submit Amount",
                "OPEN Claims": "Pending Resolution Claims",
                "OPEN Amount": "Pending Resolution Amount",
                "NOT ASSIGNED Claims": "Within Coding TAT Claims",
                "NOT ASSIGNED Amount": "Within Coding TAT Amount",
            })
        return out

    insurance_summary = build_group_summary("_Insurance", "Insurance")
    doctor_summary = build_group_summary("_Doctor", "Doctor")

    # ClaimStatus optional
    if col_claim_status:
        claim_status_summary = (
            claims.groupby("_ClaimStatus", dropna=False)
            .agg(Claims=("_VisitNo", "size"), **{"Ins Share": ("_Amount", "sum")})
            .reset_index()
            .rename(columns={"_ClaimStatus": "Claim Status"})
            .sort_values("Claims", ascending=False)
        )
    else:
        claim_status_summary = pd.DataFrame()

    _period_start, _period_end = _date_bounds(selected_day) if selected_day is not None else (pd.to_datetime(report_day), pd.to_datetime(report_day))
    return {
        "report_day": pd.to_datetime(report_day),
        "report_start": pd.to_datetime(_period_start if _period_start is not None else report_day),
        "report_end": pd.to_datetime(_period_end if _period_end is not None else report_day),
        "claims": claims,
        "status_summary": status_summary,
        "query_summary": query_summary,
        "insurance_summary": insurance_summary,
        "doctor_summary": doctor_summary,
        "claim_status_summary": claim_status_summary,
        "columns": {
            "visit": col_visit,
            "patient": col_patient,
            "visit_date": col_visit_date,
            "amount": col_amount,
            "status": col_status,
            "claim_status": col_claim_status,
            "remark": col_remark,
            "insurance": col_ins,
            "doctor": col_doc,
            "opened": col_opened,
            "processed": col_processed,
            "closed": col_closed,
            "submission": col_submission,
            "assigned": col_assigned,
        },
    }


# =========================================================
# S3 SAVE / LOAD
# =========================================================
def save_analysis_to_s3(result: Dict[str, object], raw_bytes: bytes, raw_name: str):
    if not s3_ok:
        return False, "S3 is not configured."

    bucket = cfg["S3_BUCKET_NAME"]
    root = daily_root(cfg, center_key)
    day = pd.to_datetime(result["report_day"]).strftime("%Y-%m-%d")
    day_root = s3_key(root, day)

    # Save processed analysis
    payload = pickle.dumps(result, protocol=pickle.HIGHEST_PROTOCOL)
    s3_put_bytes(
        s3, bucket, s3_key(day_root, "analysis.pkl"), payload,
        "application/octet-stream"
    )

    # Save raw report with original extension
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", raw_name or "daily_report.xls")
    s3_put_bytes(
        s3, bucket, s3_key(day_root, safe_name), raw_bytes,
        "application/vnd.ms-excel"
    )

    # History
    hist_key = s3_key(root, "history.csv")
    hist_b = s3_get_bytes(s3, bucket, hist_key)
    if hist_b:
        try:
            hist = pd.read_csv(io.BytesIO(hist_b))
        except Exception:
            hist = pd.DataFrame(columns=["day", "saved_at"])
    else:
        hist = pd.DataFrame(columns=["day", "saved_at"])

    new_row = pd.DataFrame([{
        "day": day,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }])
    hist = pd.concat([hist, new_row], ignore_index=True)
    hist = hist.drop_duplicates(subset=["day"], keep="last").sort_values("day")

    s3_put_bytes(
        s3, bucket, hist_key,
        hist.to_csv(index=False).encode("utf-8"),
        "text/csv",
    )
    return True, day


def load_history() -> pd.DataFrame:
    if not s3_ok:
        return pd.DataFrame(columns=["day", "saved_at"])
    key = s3_key(daily_root(cfg, center_key), "history.csv")
    b = s3_get_bytes(s3, cfg["S3_BUCKET_NAME"], key)
    if not b:
        return pd.DataFrame(columns=["day", "saved_at"])
    try:
        h = pd.read_csv(io.BytesIO(b))
        h["day"] = pd.to_datetime(h["day"], errors="coerce")
        return h.dropna(subset=["day"]).sort_values("day")
    except Exception:
        return pd.DataFrame(columns=["day", "saved_at"])


def load_saved_day(day) -> Optional[Dict[str, object]]:
    if not s3_ok:
        return None
    day_s = pd.to_datetime(day).strftime("%Y-%m-%d")
    key = s3_key(daily_root(cfg, center_key), day_s, "analysis.pkl")
    b = s3_get_bytes(s3, cfg["S3_BUCKET_NAME"], key)
    if not b:
        return None
    try:
        return pickle.loads(b)
    except Exception:
        return None



def _bundle_key() -> str:
    return s3_key(daily_root(cfg, center_key), "latest_source_bundle.pkl")


def save_source_bundle_to_s3(bundle: Dict[str, object]) -> bool:
    """Persist the latest 3 uploaded source reports so date changes never require re-upload."""
    if not s3_ok:
        return False
    try:
        payload = pickle.dumps(bundle, protocol=pickle.HIGHEST_PROTOCOL)
        s3_put_bytes(
            s3,
            cfg["S3_BUCKET_NAME"],
            _bundle_key(),
            payload,
            "application/octet-stream",
        )
        return True
    except Exception:
        return False


def load_source_bundle_from_s3() -> Optional[Dict[str, object]]:
    if not s3_ok:
        return None
    b = s3_get_bytes(s3, cfg["S3_BUCKET_NAME"], _bundle_key())
    if not b:
        return None
    try:
        bundle = pickle.loads(b)
        if isinstance(bundle, dict) and all(k in bundle for k in ["registration", "revenue", "submission"]):
            return bundle
    except Exception:
        pass
    return None


def _named_bytes(data: bytes, name: str):
    bio = io.BytesIO(data)
    bio.name = name
    return bio


@st.cache_data(show_spinner=False)
def _cached_registration(data: bytes, name: str) -> pd.DataFrame:
    return read_registration_report(_named_bytes(data, name))


@st.cache_data(show_spinner=False)
def _cached_revenue(data: bytes, name: str) -> pd.DataFrame:
    return read_revenue_report(_named_bytes(data, name))


@st.cache_data(show_spinner=False)
def _cached_submission(data: bytes, name: str) -> pd.DataFrame:
    return read_daily_report(_named_bytes(data, name), name)


def read_bundle(bundle: Dict[str, object]):
    reg = bundle["registration"]
    rev = bundle["revenue"]
    sub = bundle["submission"]
    reg_df = _cached_registration(reg["bytes"], reg["name"])
    rev_df = _cached_revenue(rev["bytes"], rev["name"])
    sub_df = _cached_submission(sub["bytes"], sub["name"])
    return reg_df, rev_df, sub_df


def build_result_from_bundle(bundle: Dict[str, object], selected_period):
    reg_df, rev_df, sub_raw = read_bundle(bundle)
    patient_count = registration_patient_count(reg_df, selected_period)
    rev_result = revenue_analysis(rev_df, selected_period)
    result = process_report(sub_raw, selected_day=selected_period)
    result["registration_count"] = patient_count
    result["revenue"] = rev_result
    result["selected_period"] = selected_period

    reg_days = _date_set_in_period(reg_df, ["Reg:Date", "Reg Date", "Registration Date", "Date"], selected_period)
    rev_days = _date_set_in_period(rev_df, ["Visit Date", "VisitDate"], selected_period)
    sub_days = _date_set_in_period(sub_raw, ["Visit Date", "VisitDate", "Enc Date", "Encounter Date"], selected_period)
    result["date_coverage"] = {
        "Registration": sorted(reg_days),
        "Revenue": sorted(rev_days),
        "Submission": sorted(sub_days),
    }
    return result



# =========================================================
# EMAIL
# =========================================================
def _email_recipients():
    to_addr = str(st.secrets.get("EMAIL_TO", "") or "").strip()
    cc_addr = str(st.secrets.get("EMAIL_CC", "") or "").strip()
    return to_addr, cc_addr


def _build_daily_rcm_email(result: Dict[str, object]) -> str:
    claims = result["claims"]
    report_day = pd.to_datetime(result["report_day"])
    report_start = pd.to_datetime(result.get("report_start", report_day)).normalize()
    report_end = pd.to_datetime(result.get("report_end", report_day)).normalize()
    if report_start == report_end:
        email_period = report_start.strftime("%d %b %Y")
    else:
        email_period = f"{report_start.strftime('%d %b %Y')} – {report_end.strftime('%d %b %Y')}"

    closed_n, closed_a = status_value(result, "CLOSED")
    proc_n, proc_a = status_value(result, "PROCESSED")
    open_n, open_a = status_value(result, "OPEN")
    na_n, na_a = status_value(result, "NOT ASSIGNED")
    total_n = int(len(claims))
    total_a = float(claims["_Amount"].sum())
    over48_n = int(claims["_NotAssignedOver48h"].sum())
    over48_a = float(claims.loc[claims["_NotAssignedOver48h"], "_Amount"].sum())

    status_rows = [
        ("Already Submitted", closed_n, closed_a),
        ("Ready to Submit", proc_n, proc_a),
        ("Pending Resolution", open_n, open_a),
        ("Within Coding TAT (≤48h)", na_n, na_a),
    ]

    def row_html(label, count, amount, total=False):
        _status_colors = {
            "Already Submitted": "#ECF9F1",
            "Ready to Submit": "#EEF6FF",
            "Pending Resolution": "#FFF7E6",
            "Within Coding TAT (≤48h)": "#F1F0FF",
        }
        bg = "#0B2342" if total else _status_colors.get(label, "#FFFFFF")
        color = "#FFFFFF" if total else "#263447"
        fw = "900" if total else "700"
        return (
            f"<tr style='background:{bg};color:{color};font-weight:{fw};'>"
            f"<td style='padding:9px 12px;border-bottom:1px solid #e8eef5;'>{html.escape(str(label))}</td>"
            f"<td style='padding:9px 12px;border-bottom:1px solid #e8eef5;text-align:right;'>{int(count):,}</td>"
            f"<td style='padding:9px 12px;border-bottom:1px solid #e8eef5;text-align:right;'>AED {float(amount):,.2f}</td>"
            f"</tr>"
        )

    status_html = "".join(row_html(*r) for r in status_rows)
    status_html += row_html("TOTAL", total_n, total_a, total=True)

    q = result.get("query_summary", pd.DataFrame())
    query_html = ""
    if isinstance(q, pd.DataFrame) and not q.empty:
        q2 = q.copy()
        q2["Ins Share"] = pd.to_numeric(q2["Ins Share"], errors="coerce").fillna(0)
        rows = []
        for _, r in q2.iterrows():
            rows.append(
                f"<tr>"
                f"<td style='padding:8px 12px;border-bottom:1px solid #e8eef5;'>{html.escape(str(r['Query Department']))}</td>"
                f"<td style='padding:8px 12px;border-bottom:1px solid #e8eef5;text-align:right;'>{int(r['Claims']):,}</td>"
                f"<td style='padding:8px 12px;border-bottom:1px solid #e8eef5;text-align:right;'>AED {float(r['Ins Share']):,.2f}</td>"
                f"</tr>"
            )
        query_html = f"""
        <div style="margin-top:20px;font-weight:900;color:#0B2342;font-size:15px;">Open Query Breakdown</div>
        <table style="width:100%;border-collapse:collapse;margin-top:8px;">
          <tr style="background:#0B2342;color:white;">
            <th style="padding:8px 12px;text-align:left;">Department</th>
            <th style="padding:8px 12px;text-align:right;">Claims</th>
            <th style="padding:8px 12px;text-align:right;">Ins Share</th>
          </tr>
          {''.join(rows)}
        </table>
        """

    doctor_email_html = ""
    rev = result.get("revenue", {}) or {}
    doc = rev.get("doctor") if isinstance(rev, dict) else None
    if isinstance(doc, pd.DataFrame) and not doc.empty:
        preferred=["Department","Doctor","Visits","Lab","Procedure","Insurance_Amount","Avg_Insurance_Per_Visit"]
        d=doc[[c for c in preferred if c in doc.columns]].copy()
        header_cells=''.join(f"<th style='padding:8px 10px;text-align:{'left' if c in ['Department','Doctor'] else 'right'};'>{html.escape(c)}</th>" for c in d.columns)
        body_rows=[]; col_colors={"Visits":"#EEF6FF","Lab":"#ECF9F1","Procedure":"#F1F0FF","Insurance_Amount":"#EAF8F7","Avg_Insurance_Per_Visit":"#F7F9FC"}
        for _,rr in d.iterrows():
            cells=[]
            for c in d.columns:
                bg=col_colors.get(c,"#FFFFFF"); v=rr[c]
                if c in ["Visits","Lab","Procedure"]:
                    num=pd.to_numeric(v,errors="coerce"); val=f"{int(0 if pd.isna(num) else num):,}"; align="right"
                elif c in ["Insurance_Amount","Avg_Insurance_Per_Visit"]:
                    num=pd.to_numeric(v,errors="coerce"); val=f"{float(0 if pd.isna(num) else num):,.2f}"; align="right"
                else:
                    val=html.escape(str(v)); align="left"
                cells.append(f"<td style='padding:7px 10px;border-bottom:1px solid #e8eef5;background:{bg};text-align:{align};'>{val}</td>")
            body_rows.append('<tr>'+''.join(cells)+'</tr>')
        doctor_email_html=f"<div style='margin-top:20px;font-weight:900;color:#0B2342;font-size:15px;'>Doctor Revenue — Daily Collection Details</div><table style='width:100%;border-collapse:collapse;margin-top:8px;'><tr style='background:#0B2342;color:white;'>{header_cells}</tr>{''.join(body_rows)}</table>"

    return f"""
    <html>
    <body style="font-family:Segoe UI,Arial,sans-serif;background:#f4f7fb;margin:0;padding:20px;">
      <div style="max-width:850px;margin:auto;background:white;border-radius:14px;overflow:hidden;box-shadow:0 7px 25px rgba(15,23,42,.10);">
        <div style="background:#0B2342;color:white;padding:20px 22px;">
          <div style="font-size:20px;font-weight:900;">Daily RCM Submission Report</div>
          <div style="font-size:26px;line-height:1.2;font-weight:900;color:#ffffff;margin-top:8px;letter-spacing:.2px;">{email_period}</div>
          <div style="font-size:13px;font-weight:700;color:#a8c1df;margin-top:6px;">{html.escape(CENTERS.get(center_key, center_key))}</div>
        </div>
        <div style="padding:18px 22px;">
          <table style="width:100%;border-collapse:separate;border-spacing:8px;">
            <tr>
              <td style="background:#eef8ff;padding:14px;border-radius:10px;"><div style="font-size:12px;font-weight:800;color:#4b6380;text-transform:uppercase;">Total Claims</div><div style="font-size:27px;font-weight:900;color:#071a5d;margin-top:5px;">AED {total_a:,.2f}</div><div style="font-size:13px;font-weight:700;color:#64748b;margin-top:4px;">{total_n:,} claims</div></td>
              <td style="background:#f0fcf4;padding:14px;border-radius:10px;"><div style="font-size:12px;font-weight:800;color:#4b6380;text-transform:uppercase;">Already Submitted</div><div style="font-size:27px;font-weight:900;color:#071a5d;margin-top:5px;">AED {closed_a:,.2f}</div><div style="font-size:13px;font-weight:700;color:#64748b;margin-top:4px;">{closed_n:,} claims</div></td>
              <td style="background:#eef6ff;padding:14px;border-radius:10px;"><div style="font-size:12px;font-weight:800;color:#4b6380;text-transform:uppercase;">Ready to Submit</div><div style="font-size:27px;font-weight:900;color:#071a5d;margin-top:5px;">AED {proc_a:,.2f}</div><div style="font-size:13px;font-weight:700;color:#64748b;margin-top:4px;">{proc_n:,} claims</div></td>
            </tr>
            <tr>
              <td style="background:#fff9e9;padding:14px;border-radius:10px;"><div style="font-size:12px;font-weight:800;color:#4b6380;text-transform:uppercase;">Pending Resolution</div><div style="font-size:27px;font-weight:900;color:#071a5d;margin-top:5px;">AED {open_a:,.2f}</div><div style="font-size:13px;font-weight:700;color:#64748b;margin-top:4px;">{open_n:,} claims</div></td>
              <td style="background:#f7f1ff;padding:14px;border-radius:10px;"><div style="font-size:12px;font-weight:800;color:#4b6380;text-transform:uppercase;">Within Coding TAT</div><div style="font-size:27px;font-weight:900;color:#071a5d;margin-top:5px;">AED {na_a:,.2f}</div><div style="font-size:13px;font-weight:700;color:#64748b;margin-top:4px;">{na_n:,} claims</div></td>
              <td style="background:#fff1f3;padding:14px;border-radius:10px;"><div style="font-size:12px;font-weight:800;color:#4b6380;text-transform:uppercase;">Coding TAT Breach &gt;48h</div><div style="font-size:27px;font-weight:900;color:#071a5d;margin-top:5px;">AED {over48_a:,.2f}</div><div style="font-size:13px;font-weight:700;color:#64748b;margin-top:4px;">{over48_n:,} claims</div></td>
            </tr>
          </table>

          <div style="margin-top:18px;font-weight:900;color:#0B2342;font-size:15px;">Status Summary</div>
          <table style="width:100%;border-collapse:collapse;margin-top:8px;">
            <tr style="background:#0B2342;color:white;">
              <th style="padding:9px 12px;text-align:left;">Status</th>
              <th style="padding:9px 12px;text-align:right;">Claims</th>
              <th style="padding:9px 12px;text-align:right;">Ins Share</th>
            </tr>
            {status_html}
          </table>
          {query_html}
          {doctor_email_html}
          <div style="margin-top:18px;font-size:11px;color:#8492a6;">Auto-generated by EMC RCM Dashboard.</div>
        </div>
      </div>
    </body>
    </html>
    """


def _build_colored_excel_attachment(result: Dict[str, object]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb=Workbook(); ws=wb.active; ws.title="Status Summary"
    navy="0B2342"; white="FFFFFF"; thin=Side(style="thin",color="D9E2EC")
    fills={"Already Submitted":"ECF9F1","Ready to Submit":"EEF6FF","Pending Resolution":"FFF7E6","Within Coding TAT (≤48h)":"F1F0FF"}
    status=result.get("status_summary",pd.DataFrame()).copy(); labels={"CLOSED":"Already Submitted","PROCESSED":"Ready to Submit","OPEN":"Pending Resolution","NOT ASSIGNED":"Within Coding TAT (≤48h)"}
    if not status.empty:
        status["Status"]=status["Status"].astype(str).str.upper().map(lambda x:labels.get(x,x.title())); status["Ins Share"]=pd.to_numeric(status["Ins Share"],errors="coerce").fillna(0)
    ws.append(["Status","Claims","Net Insurance Amount (AED)"])
    for c in ws[1]: c.fill=PatternFill("solid",fgColor=navy); c.font=Font(color=white,bold=True); c.alignment=Alignment(horizontal="center")
    for _,r in status.iterrows():
        ws.append([r["Status"],int(r["Claims"]),float(r["Ins Share"])])
        fill=PatternFill("solid",fgColor=fills.get(str(r["Status"]),"FFFFFF"))
        for c in ws[ws.max_row]: c.fill=fill; c.border=Border(bottom=thin)
    tc=int(pd.to_numeric(status.get("Claims",pd.Series(dtype=float)),errors="coerce").fillna(0).sum()) if not status.empty else 0; ta=float(pd.to_numeric(status.get("Ins Share",pd.Series(dtype=float)),errors="coerce").fillna(0).sum()) if not status.empty else 0.0
    ws.append(["TOTAL",tc,ta])
    for c in ws[ws.max_row]: c.fill=PatternFill("solid",fgColor=navy); c.font=Font(color=white,bold=True)
    ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=15; ws.column_dimensions["C"].width=28
    for row in range(2,ws.max_row+1):
        ws.cell(row,3).number_format='AED #,##0.00'
        ws.cell(row,3).font=Font(bold=True,color="0B2342") if row < ws.max_row else Font(bold=True,color=white)
        if row < ws.max_row:
            ws.cell(row,2).font=Font(color="64748B")
    ws.freeze_panes="A2"
    rev=result.get("revenue",{}) or {}; doc=rev.get("doctor") if isinstance(rev,dict) else None
    if isinstance(doc,pd.DataFrame) and not doc.empty:
        preferred=["Department","Doctor","Visits","Lab","Procedure","Insurance_Amount","Avg_Insurance_Per_Visit"]; d=doc[[c for c in preferred if c in doc.columns]].copy(); wd=wb.create_sheet("Doctor Revenue"); wd.append(list(d.columns))
        for c in wd[1]: c.fill=PatternFill("solid",fgColor=navy); c.font=Font(color=white,bold=True); c.alignment=Alignment(horizontal="center")
        col_fill={"Visits":"EEF6FF","Lab":"ECF9F1","Procedure":"F1F0FF","Insurance_Amount":"EAF8F7","Avg_Insurance_Per_Visit":"F7F9FC"}
        for _,rr in d.iterrows():
            vals=[]
            for cname in d.columns:
                v=rr[cname]
                if cname in ["Visits","Lab","Procedure"]:
                    n=pd.to_numeric(v,errors="coerce"); v=int(0 if pd.isna(n) else n)
                elif cname in ["Insurance_Amount","Avg_Insurance_Per_Visit"]:
                    n=pd.to_numeric(v,errors="coerce"); v=float(0 if pd.isna(n) else n)
                vals.append(v)
            wd.append(vals)
            for j,cname in enumerate(d.columns,1):
                cell=wd.cell(wd.max_row,j); cell.fill=PatternFill("solid",fgColor=col_fill.get(cname,"FFFFFF")); cell.border=Border(bottom=thin)
                if cname in ["Insurance_Amount","Avg_Insurance_Per_Visit"]:
                    cell.number_format='AED #,##0.00'
                    cell.font=Font(bold=True,color="0B2342")
                elif cname in ["Visits","Lab","Procedure"]:
                    cell.font=Font(color="64748B")
        for col,w in {"A":32,"B":38,"C":12,"D":10,"E":12,"F":22,"G":24}.items(): wd.column_dimensions[col].width=w
        wd.freeze_panes="A2"
    q=result.get("query_summary",pd.DataFrame())
    if isinstance(q,pd.DataFrame) and not q.empty:
        wq=wb.create_sheet("Pending Resolution"); wq.append(["Query Department","Claims","Net Insurance Amount (AED)"])
        for c in wq[1]: c.fill=PatternFill("solid",fgColor=navy); c.font=Font(color=white,bold=True)
        qfills=["FFF7E6","EEF6FF","F1F0FF","F7F9FC","ECF9F1","FFF1F3"]
        for i,(_,r) in enumerate(q.iterrows()):
            wq.append([r.get("Query Department","Unspecified"),int(r.get("Claims",0) or 0),float(r.get("Ins Share",0) or 0)]); fill=PatternFill("solid",fgColor=qfills[i%len(qfills)])
            for c in wq[wq.max_row]: c.fill=fill; c.border=Border(bottom=thin)
            wq.cell(wq.max_row,3).number_format='AED #,##0.00'
            wq.cell(wq.max_row,3).font=Font(bold=True,color="0B2342")
            wq.cell(wq.max_row,2).font=Font(color="64748B")
        wq.column_dimensions["A"].width=30; wq.column_dimensions["B"].width=14; wq.column_dimensions["C"].width=28; wq.freeze_panes="A2"
    bio=io.BytesIO(); wb.save(bio); return bio.getvalue()


def _send_daily_rcm_email(result: Dict[str, object]) -> None:
    host=str(st.secrets.get("SMTP_HOST","") or "").strip(); port=int(st.secrets.get("SMTP_PORT",465)); user=str(st.secrets.get("SMTP_USER","") or "").strip(); pwd=str(st.secrets.get("SMTP_PASS","") or "").strip(); to_addr,cc_addr=_email_recipients()
    if not all([host,user,pwd,to_addr]): raise ValueError("Missing SMTP settings. Required: SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS and EMAIL_TO.")
    report_day=pd.to_datetime(result["report_day"]); report_start=pd.to_datetime(result.get("report_start",report_day)).normalize(); report_end=pd.to_datetime(result.get("report_end",report_day)).normalize(); period_subject=report_start.strftime("%d %b %Y") if report_start==report_end else f"{report_start.strftime('%d %b %Y')} - {report_end.strftime('%d %b %Y')}"; msg=MIMEMultipart("mixed"); msg["Subject"]=f"Daily RCM Submission Report - {period_subject}"; msg["From"]=user; msg["To"]=to_addr
    if cc_addr: msg["Cc"]=cc_addr
    alt=MIMEMultipart("alternative"); alt.attach(MIMEText(_build_daily_rcm_email(result),"html")); msg.attach(alt)
    xlsx_bytes=_build_colored_excel_attachment(result); attachment=MIMEApplication(xlsx_bytes,_subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"); file_period=report_start.strftime("%Y-%m-%d") if report_start==report_end else f"{report_start.strftime('%Y-%m-%d')}_to_{report_end.strftime('%Y-%m-%d')}"; attachment.add_header("Content-Disposition","attachment",filename=f"Daily_RCM_Report_{file_period}.xlsx"); msg.attach(attachment)
    recipients=[x.strip() for x in (to_addr.split(",")+(cc_addr.split(",") if cc_addr else [])) if x.strip()]
    with smtplib.SMTP_SSL(host,port) as server: server.login(user,pwd); server.sendmail(user,recipients,msg.as_string())


def _render_premium_status_table(status_show: pd.DataFrame) -> None:
    total_claims = int(pd.to_numeric(status_show["Claims"], errors="coerce").fillna(0).sum())
    total_amount = float(pd.to_numeric(status_show["Ins Share"], errors="coerce").fillna(0).sum())
    status_colors = {"Already Submitted":"#ECF9F1","Ready to Submit":"#EEF6FF","Pending Resolution":"#FFF7E6","Within Coding TAT (≤48h)":"#F1F0FF"}
    rows=[]
    for _,r in status_show.iterrows():
        label=str(r["Status"]); bg=status_colors.get(label,"#FFFFFF")
        rows.append(f"<tr style='background:{bg};'><td style='font-weight:700;'>{html.escape(label)}</td><td class='num'>{int(r['Claims']):,}</td><td class='num'>AED {float(r['Ins Share']):,.2f}</td></tr>")
    rows.append(f"<tr class='total-row' style='background:#0B2342 !important;font-weight:900;'><td style='color:#FFFFFF !important;'><span style='color:#FFFFFF !important;'>TOTAL</span></td><td class='num' style='color:#FFFFFF !important;'><span style='color:#FFFFFF !important;'>{total_claims:,}</span></td><td class='num' style='color:#FFFFFF !important;'><span style='color:#FFFFFF !important;'>AED {total_amount:,.2f}</span></td></tr>")
    st.markdown("<div class='premium-table-wrap'><table class='premium-table'><thead><tr><th>Status</th><th style='text-align:right;'>Claims</th><th style='text-align:right;'>Net Insurance Amount (AED)</th></tr></thead><tbody>"+''.join(rows)+"</tbody></table></div>", unsafe_allow_html=True)


def _render_doctor_revenue_table(df: pd.DataFrame) -> None:
    if df is None or df.empty:
        return
    col_colors={"Visits":"#EEF6FF","Lab":"#ECF9F1","Procedure":"#F1F0FF","Insurance_Amount":"#EAF8F7","Avg_Insurance_Per_Visit":"#F7F9FC"}
    headers=list(df.columns)
    th=''.join(f"<th>{html.escape(str(c))}</th>" for c in headers)
    rows=[]
    for _,r in df.iterrows():
        tds=[]
        for c in headers:
            bg=col_colors.get(c,"#FFFFFF"); v=r[c]
            if c in ["Visits","Lab","Procedure"]:
                num=pd.to_numeric(v,errors="coerce"); value=f"{int(0 if pd.isna(num) else num):,}"; align="right"
            elif c in ["Insurance_Amount","Avg_Insurance_Per_Visit"]:
                num=pd.to_numeric(v,errors="coerce"); value=f"{float(0 if pd.isna(num) else num):,.2f}"; align="right"
            else:
                value=html.escape(str(v)); align="left"
            tds.append(f"<td style='background:{bg};text-align:{align};'>{value}</td>")
        rows.append('<tr>'+''.join(tds)+'</tr>')

    # Grand total row for doctor revenue table.
    total_visits = int(pd.to_numeric(df.get("Visits", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if "Visits" in df.columns else 0
    total_lab = int(pd.to_numeric(df.get("Lab", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if "Lab" in df.columns else 0
    total_proc = int(pd.to_numeric(df.get("Procedure", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if "Procedure" in df.columns else 0
    total_ins = float(pd.to_numeric(df.get("Insurance_Amount", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if "Insurance_Amount" in df.columns else 0.0
    total_avg = (total_ins / total_visits) if total_visits else 0.0
    total_cells=[]
    for c in headers:
        if c == "Department":
            value="TOTAL"; align="left"
        elif c == "Doctor":
            value=""; align="left"
        elif c == "Visits":
            value=f"{total_visits:,}"; align="right"
        elif c == "Lab":
            value=f"{total_lab:,}"; align="right"
        elif c == "Procedure":
            value=f"{total_proc:,}"; align="right"
        elif c == "Insurance_Amount":
            value=f"{total_ins:,.2f}"; align="right"
        elif c == "Avg_Insurance_Per_Visit":
            value=f"{total_avg:,.2f}"; align="right"
        else:
            value=""; align="left"
        total_cells.append(f"<td style='background:#0B2342 !important;color:#FFFFFF !important;font-weight:900 !important;text-align:{align};'><span style='color:#FFFFFF !important;font-weight:900 !important;'>{value}</span></td>")
    rows.append("<tr class='total-row' style='background:#0B2342 !important;'>"+''.join(total_cells)+"</tr>")
    st.markdown("<div class='premium-table-wrap'><table class='premium-table'><thead><tr>"+th+"</tr></thead><tbody>"+''.join(rows)+"</tbody></table></div>", unsafe_allow_html=True)


def status_value(result: Dict[str, object], status: str) -> Tuple[int, float]:
    """Return claim count and insurance amount for one submission status."""
    s = result.get("status_summary")
    if s is None or getattr(s, "empty", True):
        return 0, 0.0
    row = s[s["Status"].astype(str).str.upper() == str(status).upper()]
    if row.empty:
        return 0, 0.0
    return int(row.iloc[0]["Claims"]), float(row.iloc[0]["Ins Share"])


def render_result(result: Dict[str, object]):
    claims = result["claims"]
    report_day = pd.to_datetime(result["report_day"])
    report_start = pd.to_datetime(result.get("report_start", report_day))
    report_end = pd.to_datetime(result.get("report_end", report_day))
    period_label = report_start.strftime("%A, %d %b %Y") if report_start.normalize() == report_end.normalize() else f"{report_start.strftime('%d %b %Y')} – {report_end.strftime('%d %b %Y')}"

    total_claims = int(len(claims))
    total_amount = float(claims["_Amount"].sum())

    closed_n, closed_a = status_value(result, "CLOSED")
    proc_n, proc_a = status_value(result, "PROCESSED")
    open_n, open_a = status_value(result, "OPEN")
    na_n, na_a = status_value(result, "NOT ASSIGNED")

    h1, h2 = st.columns([4.8, 1.2], vertical_alignment="center")
    with h1:
        st.markdown(
            f"""
            <div class="premium-header">
                <div>
                    <div class="premium-header-title">Daily RCM Submission Report</div>
                    <div class="premium-header-sub">{period_label} · {CENTERS.get(center_key, center_key)}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with h2:
        if st.button("✉️ Email Daily Report", use_container_width=True, key="send_daily_rcm_email"):
            try:
                _send_daily_rcm_email(result)
                to_addr, cc_addr = _email_recipients()
                st.success(f"Email sent to {to_addr}" + (f" · CC: {cc_addr}" if cc_addr else ""))
            except Exception as exc:
                st.error(f"Email could not be sent: {exc}")

    # Combined 3-report management snapshot
    _patients = int(result.get("registration_count", 0) or 0)
    _rev = result.get("revenue", {}) or {}
    _rt = _rev.get("totals", {}) or {}
    _daily_service_revenue = float(_rt.get("service_revenue", 0.0) or 0.0)
    _daily_rev_visits = int(_rt.get("visits", 0) or 0)
    _daily_avg = float(_rt.get("avg_service", 0.0) or 0.0)

    st.markdown('<div class="rcm-section">Patient Footfall & Insurance Value</div>', unsafe_allow_html=True)
    kpi_cards([
        ("Total Patients", f"{_patients:,}", "Registration report · unique Visit No", "P", "rcm-blue"),
        ("Submission Net Insurance", money(total_amount), f"{total_claims:,} claims in submission report", "Σ", "rcm-purple"),
    ])

    # Submission KPI cards: AED is primary, claim volume is secondary
    st.markdown('<div class="rcm-section">RCM Submission Pipeline</div>', unsafe_allow_html=True)
    kpi_cards([
        ("Total Claims", money(total_amount),
         f"{total_claims:,} claims", "Σ", "rcm-blue"),

        ("Already Submitted", money(closed_a),
         f"{closed_n:,} claims · {(closed_n / total_claims * 100 if total_claims else 0):.1f}%",
         "✓", "rcm-green"),

        ("Ready to Submit", money(proc_a),
         f"{proc_n:,} claims · {(proc_n / total_claims * 100 if total_claims else 0):.1f}%",
         "↑", "rcm-white"),

        ("Pending Resolution", money(open_a),
         f"{open_n:,} claims · {(open_n / total_claims * 100 if total_claims else 0):.1f}%",
         "?", "rcm-yellow"),

        ("Within Coding TAT (≤48h)", money(na_a),
         f"{na_n:,} claims · {(na_n / total_claims * 100 if total_claims else 0):.1f}%",
         "TAT", "rcm-purple"),

        (
            "Coding TAT Breach (>48h)",
            money(claims.loc[claims["_NotAssignedOver48h"], "_Amount"].sum()),
            f"{int(claims['_NotAssignedOver48h'].sum()):,} claims",
            "!",
            "rcm-red",
        ),
    ])

    # Executive financial snapshot
    _breach_n = int(claims["_NotAssignedOver48h"].sum())
    _breach_a = float(claims.loc[claims["_NotAssignedOver48h"], "_Amount"].sum())
    _pending_total_a = open_a + _breach_a
    _pending_total_n = open_n + _breach_n

    st.markdown(
        f"""
        <div class="exec-strip">
          <div class="exec-item">
            <div class="exec-label">AED Already Submitted</div>
            <div class="exec-value exec-good">{money(closed_a)}</div>
            <div class="rcm-sub">{closed_n:,} claims</div>
          </div>

          <div class="exec-item">
            <div class="exec-label">AED Ready to Submit</div>
            <div class="exec-value">{money(proc_a)}</div>
            <div class="rcm-sub">{proc_n:,} claims</div>
          </div>

          <div class="exec-item">
            <div class="exec-label">AED Pending Resolution</div>
            <div class="exec-value {'exec-good' if _pending_total_a == 0 else 'exec-warn'}">{money(_pending_total_a)}</div>
            <div class="rcm-sub">{_pending_total_n:,} claims pending query / resolution</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Submission pipeline
    st.markdown('<div class="rcm-section">Status Summary</div>', unsafe_allow_html=True)
    status_show = result["status_summary"].copy()
    _status_labels = {
        "CLOSED": "Already Submitted",
        "PROCESSED": "Ready to Submit",
        "OPEN": "Pending Resolution",
        "NOT ASSIGNED": "Within Coding TAT (≤48h)",
    }
    status_show["Status"] = status_show["Status"].astype(str).str.upper().map(
        lambda x: _status_labels.get(x, x.title())
    )
    status_show["Ins Share"] = pd.to_numeric(
        status_show["Ins Share"], errors="coerce"
    ).fillna(0).round(2)
    _render_premium_status_table(status_show)

    # OPEN query analysis as management KPI cards
    st.markdown('<div class="rcm-section">Pending Resolution Breakdown</div>', unsafe_allow_html=True)
    q = result["query_summary"].copy()
    if q.empty:
        st.success("No OPEN claims found.")
    else:
        q["Ins Share"] = pd.to_numeric(q["Ins Share"], errors="coerce").fillna(0).round(2)
        q["Claims"] = pd.to_numeric(q["Claims"], errors="coerce").fillna(0).astype(int)

        _query_cards = []
        _query_colors = ["rcm-yellow", "rcm-blue", "rcm-purple", "rcm-white", "rcm-green", "rcm-red"]
        for _i, _row in q.reset_index(drop=True).iterrows():
            _owner = str(_row.get("Query Department", "Unspecified"))
            _claims_n = int(_row.get("Claims", 0) or 0)
            _amount = float(_row.get("Ins Share", 0) or 0)
            _icon = "LAB" if "lab" in _owner.lower() or "nursing" in _owner.lower() else ("DR" if "doctor" in _owner.lower() else "Q")
            _query_cards.append((_owner, money(_amount), f"{_claims_n:,} claims pending", _icon, _query_colors[_i % len(_query_colors)]))
        kpi_cards(_query_cards)

        owner_options = ["All"] + sorted(q["Query Department"].dropna().astype(str).unique().tolist())
        owner_pick = st.selectbox("Pending Query Department", owner_options, key="daily_rcm_query_owner")
        odf = claims[claims["_Status"] == "OPEN"].copy()
        if owner_pick != "All":
            odf = odf[odf["_QueryOwner"] == owner_pick].copy()

        cols = result["columns"]
        detail_cols = []
        for c in [
            cols.get("visit"),
            cols.get("patient"),
            cols.get("visit_date"),
            cols.get("doctor"),
            cols.get("insurance"),
            cols.get("amount"),
            cols.get("remark"),
            cols.get("opened"),
        ]:
            if c and c in odf.columns and c not in detail_cols:
                detail_cols.append(c)

        odf["Query Department"] = odf["_QueryOwner"]
        if "Query Department" not in detail_cols:
            detail_cols.append("Query Department")

        if detail_cols:
            with st.expander("View pending query claim details", expanded=False):
                st.dataframe(odf[detail_cols], use_container_width=True, hide_index=True)

    # Doctor Revenue from Daily Collection Details — placed after RCM status summary
    st.markdown('<div class="rcm-section">Doctor Revenue — Daily Collection Details</div>', unsafe_allow_html=True)
    _docrev = (_rev.get("doctor") if isinstance(_rev, dict) else None)
    _svc_counts = (_rev.get("service_counts", {}) if isinstance(_rev, dict) else {}) or {}
    if isinstance(_docrev, pd.DataFrame) and not _docrev.empty:
        _doctor_visits = int(pd.to_numeric(_docrev.get("Visits", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
        _avg_insurance = float(_rt.get("avg_insurance", 0.0) or 0.0)
        kpi_cards([
            ("Visits", f"{_doctor_visits:,}", "Unique revenue visits", "V", "rcm-blue"),
            ("Lab Visits", f"{int(_svc_counts.get('Lab',0)):,}", "Unique visits with lab", "LAB", "rcm-green"),
            ("Procedure Visits", f"{int(_svc_counts.get('Procedure',0)):,}", "Unique visits with procedure", "P", "rcm-purple"),
            ("Avg Insurance / Visit", money(_avg_insurance), "Insurance amount ÷ visits", "AVG", "rcm-white"),
        ])

        _show = _docrev.copy()
        for _c in ["Lab","Procedure"]:
            if _c in _show.columns:
                _show[_c] = pd.to_numeric(_show[_c], errors="coerce").fillna(0).astype(int)
        for _c in ["Insurance_Amount","Avg_Insurance_Per_Visit"]:
            if _c in _show.columns:
                _show[_c] = pd.to_numeric(_show[_c], errors="coerce").fillna(0).round(2)
        preferred = ["Department","Doctor","Visits","Lab","Procedure","Insurance_Amount","Avg_Insurance_Per_Visit"]
        _show = _show[[c for c in preferred if c in _show.columns]]
        _render_doctor_revenue_table(_show)
    else:
        st.info("No Daily Collection Details revenue found for the selected date range.")

    # Insurance / Doctor tabs
    st.markdown('<div class="rcm-section">Performance Breakdown</div>', unsafe_allow_html=True)
    t1, t2, t3 = st.tabs(["Insurance Wise", "Doctor Wise", "Claim Status"])

    with t1:
        ins = result["insurance_summary"].copy()
        amount_cols = [c for c in ins.columns if c.endswith("Amount") or c == "Total Ins Share"]
        for c in amount_cols:
            ins[c] = pd.to_numeric(ins[c], errors="coerce").fillna(0).round(2)
        st.dataframe(ins, use_container_width=True, hide_index=True)

    with t2:
        doc = result["doctor_summary"].copy()
        amount_cols = [c for c in doc.columns if c.endswith("Amount") or c == "Total Ins Share"]
        for c in amount_cols:
            doc[c] = pd.to_numeric(doc[c], errors="coerce").fillna(0).round(2)
        st.dataframe(doc, use_container_width=True, hide_index=True)

    with t3:
        cs = result.get("claim_status_summary", pd.DataFrame())
        if isinstance(cs, pd.DataFrame) and not cs.empty:
            cs = cs.copy()
            cs["Ins Share"] = pd.to_numeric(cs["Ins Share"], errors="coerce").fillna(0).round(2)
            st.dataframe(cs, use_container_width=True, hide_index=True)
        else:
            st.info("ClaimStatus column is not available or has no values.")

    # Detailed filterable claim list
    st.markdown('<div class="rcm-section">Claim Detail</div>', unsafe_allow_html=True)
    f1, f2, f3 = st.columns(3)
    with f1:
        statuses = sorted([x for x in claims["_Status"].dropna().astype(str).unique() if x])
        _friendly_status = {
            "CLOSED": "Already Submitted",
            "PROCESSED": "Ready to Submit",
            "OPEN": "Pending Resolution",
            "NOT ASSIGNED": "Within Coding TAT (≤48h)",
        }
        pick_status = st.selectbox(
            "Status",
            ["All"] + statuses,
            format_func=lambda x: "All" if x == "All" else _friendly_status.get(x, x.title()),
            key="daily_rcm_status_filter",
        )
    with f2:
        doctors = sorted([x for x in claims["_Doctor"].dropna().astype(str).unique() if x])
        pick_doc = st.selectbox("Doctor", ["All"] + doctors, key="daily_rcm_doc_filter")
    with f3:
        insurers = sorted([x for x in claims["_Insurance"].dropna().astype(str).unique() if x])
        pick_ins = st.selectbox("Insurance", ["All"] + insurers, key="daily_rcm_ins_filter")

    fd = claims.copy()
    if pick_status != "All":
        fd = fd[fd["_Status"] == pick_status]
    if pick_doc != "All":
        fd = fd[fd["_Doctor"] == pick_doc]
    if pick_ins != "All":
        fd = fd[fd["_Insurance"] == pick_ins]

    cols = result["columns"]
    display_cols = []
    for c in [
        cols.get("visit"),
        cols.get("patient"),
        cols.get("visit_date"),
        cols.get("doctor"),
        cols.get("insurance"),
        cols.get("status"),
        cols.get("claim_status"),
        cols.get("amount"),
        cols.get("remark"),
        cols.get("processed"),
        cols.get("closed"),
        cols.get("submission"),
    ]:
        if c and c in fd.columns and c not in display_cols:
            display_cols.append(c)

    fd["Query Department"] = fd["_QueryOwner"]
    if "Query Department" not in display_cols:
        display_cols.append("Query Department")

    if display_cols:
        st.dataframe(fd[display_cols], use_container_width=True, hide_index=True)
    else:
        st.dataframe(fd, use_container_width=True, hide_index=True)

    # CSV download avoids any Excel writer dependency.
    csv_bytes = fd[display_cols].to_csv(index=False).encode("utf-8-sig") if display_cols else fd.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "⬇️ Download Filtered Claim List (CSV)",
        data=csv_bytes,
        file_name=f"daily_rcm_{report_day.strftime('%Y-%m-%d')}.csv",
        mime="text/csv",
    )


# =========================================================
# TOP UI
# =========================================================
st.title("Daily RCM Management Report")
st.caption(
    "3-report dashboard: Registration = patient count · Daily Collection Details = doctor revenue · "
    "Submission = submitted / ready / pending / coding TAT. The exact same From/To period is applied to all three reports."
)

SS.setdefault("daily_rcm_show_setup", False)

# ---------------------------------------------------------
# Restore saved sources and latest processed result
# ---------------------------------------------------------
if SS.get("daily_rcm_source_bundle") is None:
    _saved_bundle = load_source_bundle_from_s3()
    if _saved_bundle is not None:
        SS["daily_rcm_source_bundle"] = _saved_bundle

# Restore latest processed result from S3 when the app/session restarts.
if SS.get("daily_rcm_result") is None:
    _hist_boot = load_history()
    if not _hist_boot.empty:
        _latest_boot = _hist_boot["day"].dropna().sort_values().iloc[-1]
        _loaded_boot = load_saved_day(_latest_boot)
        if _loaded_boot is not None:
            SS["daily_rcm_result"] = _loaded_boot

bundle = SS.get("daily_rcm_source_bundle")
current = SS.get("daily_rcm_result")

# ---------------------------------------------------------
# Upload/source panel: hidden by default; date controls stay visible
# ---------------------------------------------------------
setup_left, setup_right = st.columns([1.2, 4.8])
with setup_left:
    if not SS.get("daily_rcm_show_setup", False):
        if st.button("📂 Upload / Replace Reports", use_container_width=True, key="daily_rcm_open_setup"):
            SS["daily_rcm_show_setup"] = True
            st.rerun()
    else:
        if st.button("✕ Close Uploads", use_container_width=True, key="daily_rcm_close_setup"):
            SS["daily_rcm_show_setup"] = False
            st.rerun()

with setup_right:
    if bundle is not None:
        _saved_at = bundle.get("saved_at", "")
        st.caption(f"✅ Source reports saved{' in S3' if s3_ok else ' for this session'}" + (f" · {_saved_at}" if _saved_at else ""))
    else:
        st.caption("Upload the 3 source reports once. After processing, they are reused for future date changes.")

if SS.get("daily_rcm_show_setup", False):
    with st.container(border=True):
        st.markdown("### Source Reports")
        u1, u2, u3 = st.columns(3)
        with u1:
            reg_up = st.file_uploader(
                "1) Registration Report (.xls / .xlsx)",
                type=["xls", "xlsx"],
                key="daily_rcm_registration_upload",
                help="Used only for Total Patients / unique Visit No.",
            )
        with u2:
            rev_up = st.file_uploader(
                "2) Daily Revenue — Daily Collection Details (.xls / .xlsx)",
                type=["xls", "xlsx"],
                key="daily_rcm_revenue_upload",
                help="Used for doctor visits, lab/procedure counts and insurance amount.",
            )
        with u3:
            sub_up = st.file_uploader(
                "3) Submission Report (.xls / .xlsx)",
                type=["xls", "xlsx"],
                key="daily_rcm_submission_upload",
                help="Used for Submitted / Ready / Pending / Coding TAT analysis.",
            )

        if st.button("▶ Save & Process These Reports", type="primary", use_container_width=True, key="daily_rcm_process_uploads"):
            if reg_up is None or rev_up is None or sub_up is None:
                st.error("Please upload all 3 reports before processing.")
            else:
                try:
                    new_bundle = {
                        "registration": {"name": reg_up.name, "bytes": reg_up.getvalue()},
                        "revenue": {"name": rev_up.name, "bytes": rev_up.getvalue()},
                        "submission": {"name": sub_up.name, "bytes": sub_up.getvalue()},
                        "saved_at": datetime.now().strftime("%d %b %Y %H:%M"),
                    }
                    # Validate immediately before saving.
                    _reg_df, _rev_df, _sub_df = read_bundle(new_bundle)
                    SS["daily_rcm_source_bundle"] = new_bundle
                    bundle = new_bundle
                    if s3_ok:
                        save_source_bundle_to_s3(new_bundle)

                    # Determine latest date common to all three newly uploaded files.
                    _reg_dates = {x.date() for x in _available_dates_from_report(_reg_df, ["Reg:Date", "Reg Date", "Registration Date", "Date"])}
                    _rev_dates = {x.date() for x in _available_dates_from_report(_rev_df, ["Visit Date", "VisitDate"])}
                    _sub_dates = {x.date() for x in _available_dates_from_report(_sub_df, ["Visit Date", "VisitDate", "Enc Date", "Encounter Date"])}
                    _common = _reg_dates & _rev_dates & _sub_dates
                    if _common:
                        _latest_common = max(_common)
                        SS["daily_rcm_start_date"] = _latest_common
                        SS["daily_rcm_end_date"] = _latest_common
                    SS["daily_rcm_quick_period"] = "Custom"
                    SS["daily_rcm_show_setup"] = False
                    st.rerun()
                except Exception as exc:
                    st.error(f"Could not read/process the source reports: {exc}")

# ---------------------------------------------------------
# Reporting period: ALWAYS VISIBLE
# ---------------------------------------------------------
# Prefer the last processed range, then latest common date in saved sources.
_default_start = date.today()
_default_end = date.today()
if current is not None:
    try:
        _default_start = pd.to_datetime(current.get("report_start", current.get("report_day"))).date()
        _default_end = pd.to_datetime(current.get("report_end", current.get("report_day"))).date()
    except Exception:
        pass
elif bundle is not None:
    try:
        _reg_df, _rev_df, _sub_df = read_bundle(bundle)
        _reg_dates = {x.date() for x in _available_dates_from_report(_reg_df, ["Reg:Date", "Reg Date", "Registration Date", "Date"])}
        _rev_dates = {x.date() for x in _available_dates_from_report(_rev_df, ["Visit Date", "VisitDate"])}
        _sub_dates = {x.date() for x in _available_dates_from_report(_sub_df, ["Visit Date", "VisitDate", "Enc Date", "Encounter Date"])}
        _common = _reg_dates & _rev_dates & _sub_dates
        if _common:
            _default_start = _default_end = max(_common)
    except Exception:
        pass

if "daily_rcm_start_date" not in SS:
    SS["daily_rcm_start_date"] = _default_start
if "daily_rcm_end_date" not in SS:
    SS["daily_rcm_end_date"] = _default_end
if "daily_rcm_quick_period" not in SS:
    SS["daily_rcm_quick_period"] = "Custom"

st.markdown("### Select Reporting Period")
p1, p2, p3 = st.columns([1, 1, 1.25])
with p1:
    start_day = st.date_input(
        "From",
        key="daily_rcm_start_date",
        help="First date included in all three reports.",
    )
with p2:
    end_day = st.date_input(
        "To",
        key="daily_rcm_end_date",
        help="Last date included in all three reports.",
    )
with p3:
    quick_period = st.selectbox(
        "Quick Period",
        ["Custom", "Single Day", "Last 7 Days", "This Month", "Previous Month"],
        key="daily_rcm_quick_period",
        help="Optional shortcut. Custom uses the From/To dates.",
    )

_anchor = pd.Timestamp(end_day)
if quick_period == "Single Day":
    _qs = _qe = _anchor.date()
elif quick_period == "Last 7 Days":
    _qe = _anchor.date()
    _qs = (_anchor - pd.Timedelta(days=6)).date()
elif quick_period == "This Month":
    _qs = _anchor.replace(day=1).date()
    _qe = (_anchor + pd.offsets.MonthEnd(0)).date()
elif quick_period == "Previous Month":
    _prev = _anchor.replace(day=1) - pd.Timedelta(days=1)
    _qs = _prev.replace(day=1).date()
    _qe = _prev.date()
else:
    _qs, _qe = start_day, end_day

if _qs > _qe:
    _qs, _qe = _qe, _qs
selected_period = (_qs, _qe)

st.caption(
    f"Applied identically to Registration + Revenue + Submission: "
    f"**{pd.Timestamp(_qs).strftime('%d %b %Y')} → {pd.Timestamp(_qe).strftime('%d %b %Y')}**"
)

# ---------------------------------------------------------
# AUTO-RECALCULATE whenever From/To/Quick Period changes.
# No upload and no Process button are required after sources are saved.
# ---------------------------------------------------------
if bundle is not None:
    _current_period = None
    if current is not None:
        try:
            _current_period = (
                pd.to_datetime(current.get("report_start", current.get("report_day"))).date(),
                pd.to_datetime(current.get("report_end", current.get("report_day"))).date(),
            )
        except Exception:
            _current_period = None

    if _current_period != selected_period:
        try:
            with st.spinner("Updating all 3 reports for the selected period..."):
                current = build_result_from_bundle(bundle, selected_period)
                SS["daily_rcm_result"] = current

                # Save the latest selected result as well, so it survives restart.
                if s3_ok:
                    try:
                        sub_meta = bundle["submission"]
                        save_analysis_to_s3(current, sub_meta["bytes"], sub_meta["name"])
                    except Exception:
                        pass
            st.rerun()
        except Exception as exc:
            st.error(f"Could not update the selected period: {exc}")

# If reports were just uploaded and there is still no result, calculate once now.
if bundle is not None and SS.get("daily_rcm_result") is None:
    try:
        current = build_result_from_bundle(bundle, selected_period)
        SS["daily_rcm_result"] = current
        if s3_ok:
            try:
                sub_meta = bundle["submission"]
                save_analysis_to_s3(current, sub_meta["bytes"], sub_meta["name"])
            except Exception:
                pass
        st.rerun()
    except Exception as exc:
        st.error(f"Could not calculate the report: {exc}")

# ---------------------------------------------------------
# Saved history remains optional; sources themselves are retained separately.
# ---------------------------------------------------------
hist = load_history()
if not hist.empty:
    with st.expander("Saved Reports", expanded=False):
        saved_days = list(hist["day"].dt.normalize().drop_duplicates().sort_values())
        pick_day = st.selectbox(
            "Select saved report ending on",
            options=saved_days,
            index=len(saved_days) - 1,
            format_func=lambda d: pd.to_datetime(d).strftime("%A, %d %b %Y"),
            key="daily_rcm_saved_day",
        )
        if st.button("Load Saved Report", use_container_width=True, key="daily_rcm_load_saved"):
            loaded = load_saved_day(pick_day)
            if loaded is None:
                st.error("Saved report could not be loaded.")
            else:
                SS["daily_rcm_result"] = loaded
                try:
                    SS["daily_rcm_start_date"] = pd.to_datetime(loaded.get("report_start", loaded["report_day"])).date()
                    SS["daily_rcm_end_date"] = pd.to_datetime(loaded.get("report_end", loaded["report_day"])).date()
                    SS["daily_rcm_quick_period"] = "Custom"
                except Exception:
                    pass
                st.rerun()

# ---------------------------------------------------------
# Display current/latest
# ---------------------------------------------------------
current = SS.get("daily_rcm_result")
if current is not None:
    st.markdown("---")
    render_result(current)
else:
    if bundle is None:
        st.info("Click **Upload / Replace Reports** once. After processing, the reports are saved and date changes will work without uploading again.")
    else:
        st.info("Saved source reports are available. Select a reporting period above.")

