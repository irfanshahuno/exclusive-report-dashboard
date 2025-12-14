#!/usr/bin/env python3
import sys, time, glob
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =================== CONFIG ===================
OUTPUT_NAME     = "Pharmacy_Exclusive_Report_with_Aging.xlsx"
TINY_THRESHOLD  = 4        # ≤ 4 goes to Accepted (not Balance)
DECIMALS        = 2
BINS            = [-1, 30, 45, 60, 90, float("inf")]
LABELS          = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]

# =================== HELPERS ===================
def find_input_file():
    # If dashboard provided a path, use it
    if len(sys.argv) >= 2:
        # support --out too
        args = sys.argv[1:]
        positional = [a for a in args if not a.startswith("-")]
        if positional:
            p = Path(positional[0])
            if p.exists() and p.suffix.lower() == ".xlsx":
                return str(p), parse_out(args, default=(p.parent / OUTPUT_NAME))
    # Fallback to first *.xlsx
    files = [f for f in glob.glob("*.xlsx") if "Rejection_Report" not in f]
    if not files:
        raise FileNotFoundError("❌ No Excel file (.xlsx) found in this folder.")
    p = Path(files[0])
    return str(p), p.parent / OUTPUT_NAME

def parse_out(args, default: Path):
    if "--out" in args:
        i = args.index("--out")
        if i + 1 >= len(args):
            raise SystemExit("error: --out requires a filename")
        return Path(args[i + 1])
    return default

def ci_get(df, names):
    lower_map = {c.lower(): c for c in df.columns}
    for n in names:
        if n.lower() in lower_map:
            return lower_map[n.lower()]
    return None

def style_headers(wb):
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # blue
    total_fill  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # light orange
    for ws in wb.worksheets:
        # headers
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # grand total highlights
        if ws.title in ("Balance_Aging_Summary", "Insurance_Totals", "Balance_Aging_Plan"):
            for r in range(2, ws.max_row + 1):
                first = ws.cell(row=r, column=1).value
                if str(first).strip().lower() in ("grand total", "grand_total", "totals", "total"):
                    for c in range(1, ws.max_column + 1):
                        gt = ws.cell(row=r, column=c)
                        gt.fill = total_fill
                        gt.font = Font(bold=True)
            # color the last column (Grand Total)
            last_col = ws.max_column
            for r in range(1, ws.max_row + 1):
                gt = ws.cell(row=r, column=last_col)
                gt.fill = total_fill
                gt.font = Font(bold=True)

# =================== TIMER START ===================
t0 = time.time()
print("▶️ Starting Pharmacy report…")

# =================== LOAD ===================
print("📂 Locating & loading Excel…")
input_file, out_path = find_input_file()
print(f"   Using: {input_file}")
print(f"   Output: {out_path}")
df = pd.read_excel(input_file, engine="openpyxl")
df.columns = df.columns.str.strip()
t1 = time.time()
print(f"⏳ Load time: {t1 - t0:.2f}s")

# ===== Detect columns (case-insensitive, pharmacy-friendly) =====
col_net   = ci_get(df, ["Claim Amount","Claim Amount (Net)","NetAmount","Net Amount","TotalAmount","Total Amount","Net"])
col_paid  = ci_get(df, ["Remitted Amount","Remitted Amount (Paid)","Paid","Remit Amount","RemitAmount"])
col_stat  = ci_get(df, ["ClaimStatus","Status","ResponseType"])
col_payer = ci_get(df, ["Insurance","PayerName","Insurer","Plan","InsurancePlan"])
col_date  = ci_get(df, ["ClaimDate","RxDate","DispenseDate","SubmissionDate","VisitDate","DOS","DateOfService"])
# NEW: Plan (Pharmacy)
col_plan  = ci_get(df, ["Plan","InsurancePlan","PlanName","PolicyPlan","Product","Policy","Plan Code","Plan_Name"])

missing = []
if not col_net:  missing.append("Claim Amount (net)")
if not col_paid: missing.append("Remitted Amount (paid)")
if not col_stat: missing.append("ClaimStatus/Status")
if missing:
    raise ValueError("❌ Required columns missing: " + ", ".join(missing))
if not col_payer:
    col_payer = "Insurance"; df[col_payer] = "Not Available"
if not col_date:
    col_date = "ClaimDate"; df[col_date] = pd.NaT
if not col_plan:
    col_plan = "Plan"; df[col_plan] = "-"  # ensure we can build the Plan sheet even if missing

# =================== PROCESS ===================
print("⚙️ Processing metrics…")
# Numerics & date
for c in [col_net, col_paid]:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
df[col_date] = pd.to_datetime(df[col_date], errors="coerce", dayfirst=True)

lower_status = df[col_stat].astype(str).str.lower()
net  = df[col_net].clip(lower=0)
paid = df[col_paid].clip(lower=0)
diff = (net - paid).clip(lower=0)

df["Rejected"] = 0.0
df["Accepted"] = 0.0
df["Balance"]  = 0.0

# Rejected = full net when denied/rejected
mask_denied = lower_status.isin(["denied", "rejected"])
df.loc[mask_denied, "Rejected"] = net

# Accepted = tiny leftover (≤4) when paid > 0 and not denied
mask_paid  = paid > 0
mask_tiny  = diff <= TINY_THRESHOLD
mask_acc   = (~mask_denied) & mask_paid & mask_tiny
df.loc[mask_acc, "Accepted"] = diff
df.loc[mask_acc, "Balance"]  = 0.0

# Balance = residual > 4 for non-denied
mask_bal = (~mask_denied) & (diff > TINY_THRESHOLD)
df.loc[mask_bal, "Balance"] = diff

# zero out accepted/balance for denied
df.loc[mask_denied, ["Accepted","Balance"]] = 0.0

# Standardized names for output
df.rename(columns={
    col_net: "NetAmount",
    col_paid: "Paid",
    col_payer: "Insurance",
    col_date: "RefDate",
    col_plan: "Plan",
}, inplace=True)

# Round money cols
money_cols = ["NetAmount","Paid","Balance","Rejected","Accepted"]
for c in money_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").round(DECIMALS)

# Per-row identity: Net = Paid + Balance + Rejected + Accepted
sum_cols = (df["Paid"] + df["Balance"] + df["Rejected"] + df["Accepted"]).round(DECIMALS)
drift = (df["NetAmount"].round(DECIMALS) - sum_cols).round(DECIMALS)
df["Accepted"] = (df["Accepted"] + drift).round(DECIMALS)

# Aging buckets
today = pd.Timestamp(datetime.today().date())
df["DaysDiff"] = (today - df["RefDate"]).dt.days
df["AgingBucket"] = pd.cut(df["DaysDiff"], bins=BINS, labels=LABELS)

# Balance-only views
balance_df = df.loc[df["Balance"] > 0].copy()

# Aging summary (robust if empty)
if balance_df.empty:
    pivot_summary = pd.DataFrame({"Insurance": []})
    for lab in LABELS:
        pivot_summary[lab] = []
    pivot_summary["Grand Total"] = []
else:
    pivot_summary = pd.pivot_table(
        balance_df,
        index="Insurance",
        columns="AgingBucket",
        values="Balance",
        aggfunc="sum",
        fill_value=0,
        observed=False
    )
    pivot_summary = pivot_summary.reindex(columns=LABELS).fillna(0)
    pivot_summary["Grand Total"] = pivot_summary.sum(axis=1)
    pivot_summary.loc["Grand Total"] = pivot_summary.sum(axis=0)
    pivot_summary.reset_index(inplace=True)

# Insurance totals
insurance_totals = (
    df.groupby("Insurance", dropna=False)[["NetAmount","Paid","Balance","Rejected","Accepted"]]
      .sum().reset_index()
)
# Grand Total row
gt = {
    "Insurance": "Grand Total",
    "NetAmount": insurance_totals["NetAmount"].sum(),
    "Paid":      insurance_totals["Paid"].sum(),
    "Balance":   insurance_totals["Balance"].sum(),
    "Rejected":  insurance_totals["Rejected"].sum(),
    "Accepted":  insurance_totals["Accepted"].sum(),
}
insurance_totals = pd.concat([insurance_totals, pd.DataFrame([gt])], ignore_index=True)

# Pretty column names for dashboard
insurance_totals = insurance_totals.rename(columns={"NetAmount": "Net Amount"})

# ===== NEW: Balance_Aging_Plan (Insurance + Plan) =====
if balance_df.empty:
    pivot_plan = pd.DataFrame({"Insurance": [], "Plan": []})
    for lab in LABELS:
        pivot_plan[lab] = []
    pivot_plan["Grand Total"] = []
else:
    plan_tmp = balance_df.copy()
    plan_tmp["Plan"] = plan_tmp["Plan"].fillna("-").astype(str).str.strip()
    pivot_plan = pd.pivot_table(
        plan_tmp,
        index=["Insurance", "Plan"],
        columns="AgingBucket",
        values="Balance",
        aggfunc="sum",
        fill_value=0,
        observed=False
    )
    pivot_plan = pivot_plan.reindex(columns=LABELS).fillna(0)
    pivot_plan["Grand Total"] = pivot_plan.sum(axis=1)

    # Add an overall Grand Total row (across all insurance+plan)
    overall = {"Insurance": "Grand Total", "Plan": ""}
    for col in LABELS + ["Grand Total"]:
        overall[col] = pivot_plan[col].sum() if col in pivot_plan.columns else 0
    pivot_plan = pivot_plan.reset_index()
    pivot_plan = pd.concat([pivot_plan, pd.DataFrame([overall])], ignore_index=True)

t2 = time.time()
print(f"⚙️ Processing time: {t2 - t1:.2f}s")

# =================== VALIDATION ===================
print("🧪 Running self-checks…")
checks = {}

row_sum = (df["Paid"] + df["Balance"] + df["Rejected"] + df["Accepted"]).round(DECIMALS)
row_diff = (df["NetAmount"].round(DECIMALS) - row_sum).round(DECIMALS)
checks["A_rows_balanced"] = bool((row_diff == 0).all())
checks["A_rows_unbalanced_count"] = int((row_diff != 0).sum())
checks["A_max_abs_row_drift"] = float(row_diff.abs().max()) if len(row_diff) else 0.0

tot_left  = float(df["NetAmount"].sum().round(DECIMALS))
tot_right = float((df["Paid"] + df["Balance"] + df["Rejected"] + df["Accepted"]).sum().round(DECIMALS))
checks["B_totals_match"] = bool(tot_left == tot_right)
checks["B_totals_left_net"] = tot_left
checks["B_totals_right_sum"] = tot_right

if balance_df.empty:
    checks["C_aging_detail_equals_summary"] = True
    checks["C_balance_detail_sum"] = 0.0
    checks["C_balance_summary_sum"] = 0.0
else:
    balance_total_detail = float(balance_df["Balance"].sum().round(DECIMALS))
    summary_no_gt = pivot_summary[pivot_summary["Insurance"] != "Grand Total"]
    balance_total_summary = float(summary_no_gt.drop(columns=["Insurance"]).sum(axis=1).sum().round(DECIMALS))
    checks["C_aging_detail_equals_summary"] = bool(abs(balance_total_detail - balance_total_summary) < 0.01)
    checks["C_balance_detail_sum"] = balance_total_detail
    checks["C_balance_summary_sum"] = balance_total_summary

# =================== SAVE ===================
out_path = Path(out_path)
out_path.parent.mkdir(parents=True, exist_ok=True)

print("💾 Saving Excel…")
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Exclusive_Report", index=False)
    insurance_totals.to_excel(writer, sheet_name="Insurance_Totals", index=False)
    pivot_summary.to_excel(writer, sheet_name="Balance_Aging_Summary", index=False)
    balance_df.to_excel(writer, sheet_name="Balance_Aging_Detail", index=False)
    # NEW sheet:
    pivot_plan.to_excel(writer, sheet_name="Balance_Aging_Plan", index=False)
    pd.DataFrame([checks]).to_excel(writer, sheet_name="Validation", index=False)

wb = load_workbook(out_path)
style_headers(wb)
wb.save(out_path)
t3 = time.time()
print(f"💾 Save time: {t3 - t2:.2f}s")

# =================== SUMMARY ===================
print(f"✅ Total time: {t3 - t0:.2f}s")
print("✅ Self-checks summary:")
for k, v in checks.items():
    print(f"   - {k}: {v}")
print(f"✅ Saved: {out_path.name}")


