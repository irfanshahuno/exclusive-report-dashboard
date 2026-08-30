#!/usr/bin/env python3

import sys, os, hashlib, argparse
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================
# Toggle: write the raw "Exclusive_Report" sheet?
# =========================================
WRITE_EXCLUSIVE_SHEET = True

# -------------------- helpers --------------------
def sha1_short(path: str) -> str:
    h = hashlib.sha1()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:12]

def parse_args():
    p = argparse.ArgumentParser(
        description="Build Exclusive_Report_with_Aging from an input .xlsx"
    )
    p.add_argument("input_xlsx", help="Path to source Excel (.xlsx)")
    p.add_argument("--out", dest="out_xlsx", required=True,
                   help="Path to write the output workbook (.xlsx)")
    args = p.parse_args()

    if not os.path.exists(args.input_xlsx):
        raise FileNotFoundError(f"❌ File not found: {args.input_xlsx}")
    if not args.input_xlsx.lower().endswith(".xlsx"):
        raise ValueError("❌ Input must be .xlsx")
    out_dir = os.path.dirname(os.path.abspath(args.out_xlsx)) or "."
    os.makedirs(out_dir, exist_ok=True)
    return args

# -------------------- ETL parts --------------------
def load_data(input_file: str) -> pd.DataFrame:
    df = pd.read_excel(input_file, engine="openpyxl")
    # Source exports contain leading/trailing and sometimes non-breaking spaces.
    df.columns = [str(c).replace("\xa0", " ").strip() for c in df.columns]
    return df

def ensure_numeric(df: pd.DataFrame) -> pd.DataFrame:
    num_cols = [
        "ActivityIns",
        "actRemitInsShare", "actResub1RemitInsShare",
        "actResub2RemitInsShare", "actResub3RemitInsShare", "actResub4RemitInsShare",
        "TKBKAmountAct",
    ]
    for c in num_cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df

def compute_measures(df: pd.DataFrame) -> pd.DataFrame:
    # Status is normalized only for matching; the original Status column is unchanged.
    status = df.get("Status", pd.Series("", index=df.index)).fillna("").astype(str)
    status = status.str.replace(r"\s+", "", regex=True).str.lower()

    # Approved resubmission amounts are moved to their own columns.
    # They remain included in Paid exactly once.
    for n in range(1, 5):
        source = f"actResub{n}RemitInsShare"
        target = f"Resub{n} Approved"
        df[target] = 0.0
        approved_mask = status.eq(f"approved(resub-{n})")
        df.loc[approved_mask, target] = df.loc[approved_mask, source]
        df.loc[approved_mask, source] = 0.0

    # Paid = all available remit columns + TKBKAmountAct.
    # This includes Resub1–Resub4 Approved once through their dedicated columns.
    paid_sources = ["actRemitInsShare", "TKBKAmountAct"] + [f"actResub{n}RemitInsShare" for n in range(1, 5)]
    paid_approved = [f"Resub{n} Approved" for n in range(1, 5)]
    df["Paid"] = df[paid_sources + paid_approved].sum(axis=1)

    # ActivityIns is the agreed Net Amount. ABS keeps Under Process positive.
    df["UnderProcess"] = (df["ActivityIns"] - df["Paid"]).abs()

    for c in ["Rejection", "Resub1 Rejection", "Resub2 Rejection", "Resub3 Rejection",
              "Accepted", "Resub1 Accepted", "Resub2 Accepted"]:
        df[c] = 0.0

    # Initial rejection: Status must be Rejected AND DenialCode must contain a value.
    denial = df.get("DenialCode", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
    initial_rejected = status.eq("rejected") & denial.ne("")
    df.loc[initial_rejected, "Rejection"] = df.loc[initial_rejected, "UnderProcess"]
    df.loc[initial_rejected, "UnderProcess"] = 0.0

    # Resubmission rejections: move the value out of Under Process into the matching stage.
    for n in range(1, 4):
        rejected_mask = status.eq(f"rejected(resub-{n})")
        target = f"Resub{n} Rejection"
        df.loc[rejected_mask, target] = df.loc[rejected_mask, "UnderProcess"]
        df.loc[rejected_mask, "UnderProcess"] = 0.0

    # Rejection Accepted statuses are reported separately and removed from Under Process.
    for status_value, target in {
        "rejectionaccepted": "Accepted",
        "rejectionaccepted(resub-1)": "Resub1 Accepted",
        "rejectionaccepted(resub-2)": "Resub2 Accepted",
    }.items():
        accepted_mask = status.eq(status_value)
        df.loc[accepted_mask, target] = df.loc[accepted_mask, "UnderProcess"]
        df.loc[accepted_mask, "UnderProcess"] = 0.0

    df["Rejection"] = df[["Rejection", "Resub1 Rejection", "Resub2 Rejection", "Resub3 Rejection"]].sum(axis=1)
    df["Accepted"] = df[["Accepted", "Resub1 Accepted", "Resub2 Accepted"]].sum(axis=1)

    return df

def add_aging(df: pd.DataFrame) -> pd.DataFrame:
    date_candidates = [c for c in ["SubmissionDate", "ClaimDate", "VisitDate"] if c in df.columns]
    if date_candidates:
        for c in date_candidates:
            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
        df["RefDate"] = df[date_candidates].bfill(axis=1).iloc[:, 0]
    else:
        df["RefDate"] = pd.NaT

    today = pd.Timestamp(datetime.today().date())
    df["DaysDiff"] = (today - df["RefDate"]).dt.days

    bins   = [-1, 30, 45, 60, 90, float("inf")]
    labels = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]
    df["AgingBucket"] = pd.cut(df["DaysDiff"], bins=bins, labels=labels)
    return df

def ensure_insurance_column(df: pd.DataFrame) -> pd.DataFrame:
    insurance_col = next((c for c in ["Insurance", "PayerName", "Insurer", "Plan"] if c in df.columns), None)
    if insurance_col is None:
        df["Insurance"] = "Not Available"
    elif insurance_col != "Insurance":
        df["Insurance"] = df[insurance_col]
    return df

def build_balance_aging_summary(balance_df: pd.DataFrame) -> pd.DataFrame:
    labels = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]
    pivot_summary = pd.pivot_table(
        balance_df,
        index="Insurance",
        columns="AgingBucket",
        values="UnderProcess",
        aggfunc="sum",
        fill_value=0,
        observed=False,
    ).reindex(columns=labels)
    pivot_summary["Grand Total"] = pivot_summary.sum(axis=1)
    pivot_summary.loc["Grand Total"] = pivot_summary.sum(axis=0)
    pivot_summary.reset_index(inplace=True)
    return pivot_summary

def build_insurance_totals(df: pd.DataFrame) -> pd.DataFrame:
    insurance_totals = (
        df.groupby("Insurance", dropna=False)[["ActivityIns", "Paid", "Rejection", "Accepted", "UnderProcess"]]
          .sum()
          .reset_index()
    )
    insurance_totals = insurance_totals.rename(columns={
        "ActivityIns": "Net Amount",
        "Rejection":   "Rejected",
        "UnderProcess": "Under Process",
    })
    insurance_totals = insurance_totals[["Insurance", "Net Amount", "Paid", "Under Process", "Rejected", "Accepted"]]
    total_row = {
        "Insurance":    "Grand Total",
        "Net Amount":   insurance_totals["Net Amount"].sum(),
        "Paid":         insurance_totals["Paid"].sum(),
        "Under Process":insurance_totals["Under Process"].sum(),
        "Rejected":     insurance_totals["Rejected"].sum(),
        "Accepted":     insurance_totals["Accepted"].sum(),
    }
    insurance_totals = pd.concat([insurance_totals, pd.DataFrame([total_row])], ignore_index=True)
    return insurance_totals

def build_denial_code_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Rejection count and amount by the applicable denial-code stage."""
    rows = []
    mapping = [("DenialCode", "Rejection")] + [(f"DenialCode{n}", f"Resub{n} Rejection") for n in range(1, 4)]
    for code_col, amount_col in mapping:
        if code_col not in df.columns:
            continue
        part = df.loc[df[amount_col] > 0, [code_col, amount_col]].copy()
        part[code_col] = part[code_col].fillna("").astype(str).str.strip()
        part = part.loc[part[code_col].ne("")]
        part = part.rename(columns={code_col: "Denial Code", amount_col: "Rejected Amount"})
        rows.append(part)
    if not rows:
        return pd.DataFrame(columns=["Denial Code", "Rejection Count", "Rejected Amount"])
    detail = pd.concat(rows, ignore_index=True)
    return (detail.groupby("Denial Code", as_index=False)
            .agg(**{"Rejection Count": ("Rejected Amount", "size"), "Rejected Amount": ("Rejected Amount", "sum")})
            .sort_values(["Rejected Amount", "Rejection Count"], ascending=False))

# -------------------- styling --------------------
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TOTAL_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

def style_headers(ws):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_styling(output_file: str):
    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        style_headers(ws)
        if ws.title in ("Balance_Aging_Summary", "Insurance_Totals"):
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == "Grand Total":
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = TOTAL_FILL
                        cell.font = Font(bold=True)
        if ws.title == "Balance_Aging_Summary":
            last_col = ws.max_column
            for r in range(1, ws.max_row + 1):
                cell = ws.cell(row=r, column=last_col)
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)
    wb.save(output_file)

# -------------------- main --------------------
def main():
    args = parse_args()
    input_file = os.path.abspath(args.input_xlsx)
    out_file   = os.path.abspath(args.out_xlsx)

    print(f"📂 Using input : {input_file}")
    print(f"📄 Output file : {out_file}")
    print(f"🔑 Input SHA1  : {sha1_short(input_file)}")

    df = load_data(input_file)
    df = ensure_numeric(df)
    df = compute_measures(df)
    df = add_aging(df)
    df = ensure_insurance_column(df)

    # Use UnderProcess > 0 for aging sheets
    balance_df    = df.loc[df["UnderProcess"] > 0].copy()
    pivot_summary = build_balance_aging_summary(balance_df)
    insurance_totals = build_insurance_totals(df)
    denial_summary = build_denial_code_summary(df)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        if WRITE_EXCLUSIVE_SHEET:
            df.to_excel(writer, sheet_name="Exclusive_Report", index=False)
        insurance_totals.to_excel(writer, sheet_name="Insurance_Totals", index=False)
        denial_summary.to_excel(writer, sheet_name="Denial_Code_Summary", index=False)
        pivot_summary.to_excel(writer, sheet_name="Balance_Aging_Summary", index=False)
        balance_df.to_excel(writer, sheet_name="Balance_Aging_Detail", index=False)
        meta = pd.DataFrame([{
            "InputFile":               os.path.basename(input_file),
            "InputSHA1":               sha1_short(input_file),
            "GeneratedAt":             datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Exclusive_Report_Written": WRITE_EXCLUSIVE_SHEET,
        }])
        meta.to_excel(writer, sheet_name="Meta", index=False)

    apply_styling(out_file)
    print("✅ Done.")

if __name__ == "__main__":
    main()
