import io
import hashlib
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================
# Rejection Analysis (Streamlit Module)
# Call run_rejection_app() from app.py
# =========================================

def run_rejection_app():
    # -------------------- helpers --------------------
    def sha1_short_bytes(b: bytes) -> str:
        return hashlib.sha1(b).hexdigest()[:12]

    def load_data_from_bytes(xlsx_bytes: bytes) -> pd.DataFrame:
        df = pd.read_excel(io.BytesIO(xlsx_bytes), engine="openpyxl")
        df.columns = df.columns.str.strip()
        return df

    # -------------------- ETL parts --------------------
    def ensure_numeric(df: pd.DataFrame) -> pd.DataFrame:
        num_cols = [
            "ActivityIns",
            "actRemitInsShare", "actResub1RemitInsShare",
            "actResub2RemitInsShare", "actResub3RemitInsShare",
            "TKBKAmountAct",
        ]
        for c in num_cols:
            if c not in df.columns:
                df[c] = 0
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
        return df

    def compute_paid(df: pd.DataFrame) -> pd.DataFrame:
        df["Paid"] = df[
            ["actRemitInsShare", "actResub1RemitInsShare",
             "actResub2RemitInsShare", "actResub3RemitInsShare",
             "TKBKAmountAct"]
        ].sum(axis=1)
        return df

    def ensure_insurance_column(df: pd.DataFrame) -> pd.DataFrame:
        insurance_col = next(
            (c for c in ["Insurance", "PayerName", "Insurer", "Plan"] if c in df.columns),
            "Insurance"
        )
        if insurance_col not in df.columns:
            df["Insurance"] = "Not Available"
        elif insurance_col != "Insurance":
            df["Insurance"] = df[insurance_col]
        return df

    def add_refdate_and_aging(df: pd.DataFrame) -> pd.DataFrame:
        from datetime import datetime as dt
        date_candidates = [c for c in ["SubmissionDate", "ClaimDate", "VisitDate"] if c in df.columns]
        if date_candidates:
            for c in date_candidates:
                df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
            df["RefDate"] = df[date_candidates].bfill(axis=1).iloc[:, 0]
        else:
            df["RefDate"] = pd.NaT

        today = pd.Timestamp(dt.today().date())
        df["DaysDiff"] = (today - df["RefDate"]).dt.days

        bins = [-1, 30, 45, 60, 90, float("inf")]
        labels = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]
        df["AgingBucket"] = pd.cut(df["DaysDiff"], bins=bins, labels=labels)
        return df

    def normalize_denial_code(df: pd.DataFrame) -> pd.DataFrame:
        if "DenialCode" not in df.columns:
            df["DenialCode"] = ""
        df["DenialCode"] = df["DenialCode"].astype(str).fillna("").str.strip()
        df.loc[df["DenialCode"].str.lower().isin(["nan", "none", "null"]), "DenialCode"] = ""
        return df

    def build_rejected_df(df: pd.DataFrame) -> pd.DataFrame:
        if "ActivityStatus" not in df.columns:
            return df.iloc[0:0].copy()

        status = df["ActivityStatus"].astype(str).fillna("").str.strip().str.lower()
        mask_rejected = (df["Paid"] == 0) & (status == "rejected") & (df["DenialCode"] != "")
        rejected_df = df.loc[mask_rejected].copy()

        rejected_df["RejectedAmount"] = rejected_df["ActivityIns"]
        rejected_df["RejectedCount"] = 1
        return rejected_df

    # -------------------- pivots --------------------
    def pivot_by_insurance(rej: pd.DataFrame) -> pd.DataFrame:
        out = (
            rej.groupby("Insurance", dropna=False)[["RejectedAmount", "RejectedCount"]]
              .sum()
              .reset_index()
        )
        total_row = {
            "Insurance": "Grand Total",
            "RejectedAmount": out["RejectedAmount"].sum(),
            "RejectedCount": int(out["RejectedCount"].sum()),
        }
        return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

    def pivot_by_denialcode(rej: pd.DataFrame) -> pd.DataFrame:
        out = (
            rej.groupby("DenialCode", dropna=False)[["RejectedAmount", "RejectedCount"]]
              .sum()
              .reset_index()
              .sort_values("RejectedAmount", ascending=False)
        )
        total_row = {
            "DenialCode": "Grand Total",
            "RejectedAmount": out["RejectedAmount"].sum(),
            "RejectedCount": int(out["RejectedCount"].sum()),
        }
        return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

    def pivot_insurance_x_denialcode(rej: pd.DataFrame) -> pd.DataFrame:
        pv = pd.pivot_table(
            rej,
            index="Insurance",
            columns="DenialCode",
            values="RejectedAmount",
            aggfunc="sum",
            fill_value=0,
            observed=False,
        )
        pv["Grand Total"] = pv.sum(axis=1)
        pv.loc["Grand Total"] = pv.sum(axis=0)
        pv.reset_index(inplace=True)
        return pv

    def pivot_rejection_aging(rej: pd.DataFrame) -> pd.DataFrame:
        labels = ["0–30 Days", "31–45 Days", "46–60 Days", "61–90 Days", ">90 Days"]
        pv = pd.pivot_table(
            rej,
            index="Insurance",
            columns="AgingBucket",
            values="RejectedAmount",
            aggfunc="sum",
            fill_value=0,
            observed=False,
        ).reindex(columns=labels)
        pv["Grand Total"] = pv.sum(axis=1)
        pv.loc["Grand Total"] = pv.sum(axis=0)
        pv.reset_index(inplace=True)
        return pv

    # -------------------- styling --------------------
    HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    TOTAL_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    def style_headers(ws):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def highlight_grand_total_rows(ws, label_col=1, label_value="Grand Total"):
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=label_col).value == label_value:
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = TOTAL_FILL
                    cell.font = Font(bold=True)

    def highlight_last_col(ws):
        last_col = ws.max_column
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=last_col)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)

    def apply_styling_in_memory(xlsx_bytes: bytes) -> bytes:
        in_buf = io.BytesIO(xlsx_bytes)
        wb = load_workbook(in_buf)

        for ws in wb.worksheets:
            style_headers(ws)
            if ws.title in [
                "Rejected_By_Insurance",
                "Rejected_By_DenialCode",
                "Rejected_Ins_x_DenialCode",
                "Rejected_Aging_Summary",
            ]:
                highlight_grand_total_rows(ws, label_col=1, label_value="Grand Total")
                if ws.title in ["Rejected_Ins_x_DenialCode", "Rejected_Aging_Summary"]:
                    highlight_last_col(ws)

        out_buf = io.BytesIO()
        wb.save(out_buf)
        return out_buf.getvalue()

    def build_rejection_workbook_bytes(input_xlsx_bytes: bytes, input_filename: str):
        df = load_data_from_bytes(input_xlsx_bytes)
        df = ensure_numeric(df)
        df = compute_paid(df)
        df = normalize_denial_code(df)
        df = ensure_insurance_column(df)
        df = add_refdate_and_aging(df)

        rejected_df = build_rejected_df(df)

        by_ins = pivot_by_insurance(rejected_df) if len(rejected_df) else pd.DataFrame(
            [{"Insurance": "Grand Total", "RejectedAmount": 0.0, "RejectedCount": 0}]
        )
        by_code = pivot_by_denialcode(rejected_df) if len(rejected_df) else pd.DataFrame(
            [{"DenialCode": "Grand Total", "RejectedAmount": 0.0, "RejectedCount": 0}]
        )
        ins_x_code = pivot_insurance_x_denialcode(rejected_df) if len(rejected_df) else pd.DataFrame(
            [{"Insurance": "Grand Total", "Grand Total": 0.0}]
        )
        aging_sum = pivot_rejection_aging(rejected_df) if len(rejected_df) else pd.DataFrame(
            [{"Insurance": "Grand Total", "Grand Total": 0.0}]
        )

        rejected_detail = rejected_df.copy()

        meta = pd.DataFrame([{
            "InputFile": input_filename,
            "InputSHA1": sha1_short_bytes(input_xlsx_bytes),
            "GeneratedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "RejectedRule": "Paid==0 AND lower(ActivityStatus)=='rejected' AND DenialCode not empty",
            "RejectedRows": int(len(rejected_df)),
        }])

        out_buf = io.BytesIO()
        with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
            by_ins.to_excel(writer, sheet_name="Rejected_By_Insurance", index=False)
            by_code.to_excel(writer, sheet_name="Rejected_By_DenialCode", index=False)
            ins_x_code.to_excel(writer, sheet_name="Rejected_Ins_x_DenialCode", index=False)
            aging_sum.to_excel(writer, sheet_name="Rejected_Aging_Summary", index=False)
            rejected_detail.to_excel(writer, sheet_name="Rejected_Detail", index=False)
            meta.to_excel(writer, sheet_name="Meta", index=False)

        styled_bytes = apply_styling_in_memory(out_buf.getvalue())

        preview = {
            "RejectedRows": int(len(rejected_detail)),
            "ByInsurance": by_ins,
            "ByDenialCode": by_code,
            "AgingSummary": aging_sum,
            "InsXCode": ins_x_code,
            "Detail": rejected_detail,
            "Meta": meta,
        }
        return styled_bytes, preview

    # -------------------- UI --------------------
    st.subheader("Rejection Analysis (Excel Upload → Excel Output)")
    st.caption("Rule: Paid==0 AND ActivityStatus=='rejected' AND DenialCode not empty")

    uploaded = st.file_uploader("Upload your source .xlsx file", type=["xlsx"], key="rej_uploader")

    if uploaded is None:
        st.info("Upload an Excel file to generate the Rejection Analysis workbook.")
        return

    input_bytes = uploaded.getvalue()
    input_name = uploaded.name

    colA, colB = st.columns([1, 1])
    with colA:
        st.write("**File:**", input_name)
    with colB:
        st.write("**SHA1 (short):**", sha1_short_bytes(input_bytes))

    run = st.button("Generate Rejection Analysis", type="primary", key="rej_run")

    if run:
        with st.spinner("Building rejection analysis…"):
            out_bytes, preview = build_rejection_workbook_bytes(input_bytes, input_name)

        st.success("Done ✅")

        out_name = f"Rejection_Analysis_{sha1_short_bytes(input_bytes)}.xlsx"
        st.download_button(
            label="Download Rejection Analysis Excel",
            data=out_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="rej_download",
        )

        st.divider()
        st.subheader("Preview")

        k1, k2 = st.columns(2)
        with k1:
            st.metric("Rejected rows", preview["RejectedRows"])
        with k2:
            st.metric("Generated at", preview["Meta"].loc[0, "GeneratedAt"])

        tabs = st.tabs([
            "Rejected_By_Insurance",
            "Rejected_By_DenialCode",
            "Rejected_Aging_Summary",
            "Rejected_Ins_x_DenialCode",
            "Rejected_Detail",
            "Meta",
        ])

        with tabs[0]:
            st.dataframe(preview["ByInsurance"], use_container_width=True)
        with tabs[1]:
            st.dataframe(preview["ByDenialCode"], use_container_width=True)
        with tabs[2]:
            st.dataframe(preview["AgingSummary"], use_container_width=True)
        with tabs[3]:
            st.dataframe(preview["InsXCode"], use_container_width=True)
        with tabs[4]:
            st.dataframe(preview["Detail"], use_container_width=True)
        with tabs[5]:
            st.dataframe(preview["Meta"], use_container_width=True)
