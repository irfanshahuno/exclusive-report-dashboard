# exclusive_dashboard.py — Main dashboard KPIs at TOP (Doc Performance unchanged)
# NOTE: This is your original dashboard with ONLY the minimal additions:
#   • Optional Balance_Aging_InsGroup tab (already supported)
#   • Optional Balance_Aging_Plan tab (new) with Insurance filter
#   • S.No hidden and display index starts at 1
#   • Grand Total row (any of 'Grand Total' / 'Total') is shown LAST in tables
#   • NEW: View password gate (Emc@2026)
# ✅ NEEDFUL CHANGES:
#   • After password: landing page with ONLY 2 buttons (2024/2025) + Change Year
#   • ✅ NEW: Hide/remove EasyHealth in 2024 (ONLY) — no other changes
# ✅ PREMIUM SOOTHING UI (ONLY VISUAL):
#   • Soft background + premium light-blue buttons
#   • KPI section becomes premium cards (soothing)
#   • Center names dark navy
#   • Balance card clickable and same bold with slight different color
# ✅ FIX:
#   • KPI numbers auto-fit inside KPI box (no overflow)
# Nothing else is changed.

import sys
import subprocess
import re
from pathlib import Path
from datetime import datetime, date

import pandas as pd
import streamlit as st

import streamlit.components.v1 as components  # used only for the home-card link

import base64
import time
import json
import hmac
import hashlib

# ====================== ✅ NEEDFUL S3 IMPORTS (NEW) ======================
import boto3
from botocore.exceptions import ClientError


# ====================== VIEW PASSWORD (NEW) ======================
# Keep default, but allow Streamlit Secrets to override (needful)
VIEW_PASSWORD = st.secrets.get("VIEW_PASSWORD", "Emc@2026")


# ====================== ✅ URL TOKEN (NEW) ======================
# Opening the dashboard in a NEW TAB creates a NEW Streamlit session, so it will ask password again.
# To avoid that, we generate a short-lived signed token in the URL and auto-auth the new session.
TOKEN_SECRET = st.secrets.get("TOKEN_SECRET", None)  # set in Streamlit Secrets for security
TOKEN_TTL_SECONDS = int(st.secrets.get("TOKEN_TTL_SECONDS", 600))  # 10 minutes default

def _b64url_encode(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).decode("utf-8").rstrip("=")

def _b64url_decode(s: str) -> bytes:
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)

def make_url_token(payload: dict) -> str:
    if not TOKEN_SECRET:
        return ""  # token disabled if no secret configured
    data = dict(payload)
    data["iat"] = int(time.time())
    body = json.dumps(data, separators=(",", ":"), sort_keys=True).encode("utf-8")
    sig = hmac.new(TOKEN_SECRET.encode("utf-8"), body, hashlib.sha256).digest()
    return _b64url_encode(body) + "." + _b64url_encode(sig)

def verify_url_token(token: str) -> dict | None:
    if not TOKEN_SECRET:
        return None
    try:
        body_b64, sig_b64 = token.split(".", 1)
        body = _b64url_decode(body_b64)
        sig = _b64url_decode(sig_b64)
        expected = hmac.new(TOKEN_SECRET.encode("utf-8"), body, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, expected):
            return None
        data = json.loads(body.decode("utf-8"))
        iat = int(data.get("iat", 0))
        if int(time.time()) - iat > TOKEN_TTL_SECONDS:
            return None
        return data
    except Exception:
        return None

def auto_auth_from_token():
    # If token is valid, auto-enable view access for this new session
    tok = st.query_params.get("token")
    if not tok:
        return
    data = verify_url_token(tok)
    if not data:
        return
    st.session_state.is_view_auth = True
    # optional defaults
    if data.get("year") and not st.session_state.get("year"):
        try:
            st.session_state.year = int(data["year"])
        except Exception:
            pass
    if data.get("center") and not st.session_state.get("center_key"):
        st.session_state.center_key = data["center"]

# ✅ attempt auto-auth BEFORE showing password screen
auto_auth_from_token()

# ====================== ✅ FIX: Restore session from query params (nav=balance) ======================
# When the Balance card link is clicked in the same session (same tab), query params carry
# center/year.  We need to restore rcm_year/center_key BEFORE require_year_selection() runs,
# otherwise that gate fires again and looks like a fresh session.
def _restore_session_from_query_params():
    y_param = st.query_params.get("year")
    c_param = st.query_params.get("center")
    # ✅ FIX: carry auth across navigation (Balance link opens in same tab but new rerun loses session)
    # We include a simple HMAC of the password in the URL so the new session can re-auth silently.
    auth_param = st.query_params.get("auth")

    if auth_param:
        # Verify: auth = HMAC-SHA256(secret, "view_auth")
        _secret = st.secrets.get("TOKEN_SECRET", VIEW_PASSWORD)
        expected = hmac.new(_secret.encode("utf-8"), b"view_auth", hashlib.sha256).hexdigest()[:16]
        if auth_param == expected:
            st.session_state.is_view_auth = True

    if y_param and not st.session_state.get("rcm_year"):
        try:
            y_int = int(y_param)
            st.session_state.rcm_year = y_int
            st.session_state.year = y_int
        except Exception:
            pass

    if c_param and not st.session_state.get("center_key"):
        st.session_state.center_key = c_param

_restore_session_from_query_params()
# ======================================================================================================


def require_view_access() -> None:
    """
    View-only password gate.
    - Blocks everything unless correct view password is entered.
    - Does NOT affect Admin password logic (admin is separate).
    """
    if st.session_state.get("is_view_auth", False):
        return

    st.set_page_config(page_title="Excellent Medical Group", layout="wide")
    st.set_option("client.showErrorDetails", False)

    st.title("🔒 Dashboard Access")
    st.info("Enter the view password to open the dashboard.")

    pwd = st.text_input("View Password", type="password", key="view_pwd")
    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("Enter Dashboard", use_container_width=True):
            if pwd == VIEW_PASSWORD:
                st.session_state.is_view_auth = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    with c2:
        st.caption("")

    st.stop()


# ====================== Page & base folders ======================
st.set_page_config(page_title="Excellent Medical Group", layout="wide")
st.set_option("client.showErrorDetails", False)

# NEW: enforce view login first
require_view_access()

# ====================== ✅ NEEDFUL: BALANCE PAGE INSIDE SAME APP ======================
BALANCE_PAGE_PATH = "pages/3_Balance_Attempt_Aging.py"


DAILY_REPORT_PAGE_PATH = "pages/4_Registration_Summary.py"
# External Daily Report (separate Streamlit app)
DAILY_REPORT_EXTERNAL_BASE = "https://exclusive-report-dashboard-ctan8jpussjzffxz2arkgh.streamlit.app"
DAILY_REPORT_EXTERNAL_PAGE = "/Registration_View"
# ✅ Daily Report button target (requested)
DAILY_REPORT_BUTTON_BASE = "https://exclusive-report-dashboard-ad74eqssyakelmpbeztqob.streamlit.app/"
def _make_auth_param() -> str:
    """Generate a short HMAC token to carry view-auth across URL navigation."""
    _secret = st.secrets.get("TOKEN_SECRET", VIEW_PASSWORD)
    return hmac.new(_secret.encode("utf-8"), b"view_auth", hashlib.sha256).hexdigest()[:16]

def build_balance_url(center: str, year: int) -> str:
    auth = _make_auth_param()
    return f"?nav=balance&center={center}&year={year}&auth={auth}"
# ================================================================================


# =========================================================
# ✅ PREMIUM + SOOTHING UI (ONLY STYLES) + ✅ AUTO-FIT KPI
# =========================================================
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

/* ---------- Page background (premium deep gradient) ---------- */
.stApp{
  background: linear-gradient(145deg, #EDF2FB 0%, #F8FAFF 40%, #FAFCFF 100%) !important;
  font-family: 'Inter', sans-serif !important;
}

/* Reduce harsh separators */
hr{ border: none !important; height:1px !important; background: linear-gradient(90deg, transparent, #C8D9F0, transparent) !important; }

/* ---------- Premium Buttons ---------- */
a.navlink{
  display: inline-block;
  width: 100%;
  text-align: center;
  min-height: 58px;
  padding: 14px 22px;
  font-size: 17px;
  font-weight: 700;
  font-family: 'Inter', sans-serif;
  background: linear-gradient(160deg, #FFFFFF 0%, #EEF4FF 100%);
  color: #0A2647;
  border: 1.5px solid #C5D8F5;
  border-radius: 14px;
  box-shadow: 0 2px 8px rgba(10, 38, 71, 0.08), inset 0 1px 0 rgba(255,255,255,0.9);
  text-decoration: none !important;
  line-height: 28px;
  transition: all 0.2s ease;
}
a.navlink:hover{
  background: linear-gradient(160deg, #E8F1FF 0%, #D6E8FF 100%);
  border-color: #7DAAEE;
  box-shadow: 0 6px 20px rgba(10, 38, 71, 0.15), inset 0 1px 0 rgba(255,255,255,0.9);
  transform: translateY(-1px);
  color: #0A2647;
}
a.navlink:active, a.navlink:focus, a.navlink:focus-visible{
  background: linear-gradient(160deg, #0A2647 0%, #154B8A 100%);
  color: #ffffff;
  border-color: #0A2647;
  outline: none;
  box-shadow: 0 4px 12px rgba(10, 38, 71, 0.3);
  transform: translateY(0px);
}

div.stButton > button{
  width: 100% !important;
  min-height: 58px !important;
  padding: 14px 22px !important;
  font-size: 17px !important;
  font-weight: 700 !important;
  font-family: 'Inter', sans-serif !important;
  background: linear-gradient(160deg, #FFFFFF 0%, #EEF4FF 100%) !important;
  color: #0A2647 !important;
  border: 1.5px solid #C5D8F5 !important;
  border-radius: 14px !important;
  box-shadow: 0 2px 8px rgba(10, 38, 71, 0.08), inset 0 1px 0 rgba(255,255,255,0.9) !important;
  transition: all 0.2s ease !important;
  letter-spacing: 0.1px !important;
}
div.stButton > button:hover{
  background: linear-gradient(160deg, #E8F1FF 0%, #D6E8FF 100%) !important;
  border-color: #7DAAEE !important;
  box-shadow: 0 6px 20px rgba(10, 38, 71, 0.15) !important;
  transform: translateY(-1px) !important;
}
div.stButton > button:active,
div.stButton > button:focus,
div.stButton > button:focus-visible{
  background: linear-gradient(160deg, #0A2647 0%, #154B8A 100%) !important;
  color: #ffffff !important;
  border-color: #0A2647 !important;
  outline: none !important;
  box-shadow: 0 4px 12px rgba(10, 38, 71, 0.3) !important;
}

/* Center titles */
.center-title{
  color: #0A2647 !important;
  font-weight: 900 !important;
  font-family: 'Inter', sans-serif !important;
  letter-spacing: -0.5px !important;
  margin-bottom: 0 !important;
}

/* ---------- KPI Cards (glassmorphism premium) ---------- */
.kpi-grid{
  display:grid;
  grid-template-columns: repeat(5, minmax(0, 1fr));
  gap: 14px;
  margin-top: 10px;
  margin-bottom: 10px;
}
.kpi-card{
  background: rgba(255, 255, 255, 0.85);
  backdrop-filter: blur(12px);
  -webkit-backdrop-filter: blur(12px);
  border: 1.5px solid rgba(197, 216, 245, 0.7);
  border-radius: 18px;
  padding: 16px 18px;
  box-shadow: 0 4px 16px rgba(10, 38, 71, 0.07), 0 1px 3px rgba(10, 38, 71, 0.05), inset 0 1px 0 rgba(255,255,255,0.95);
  min-width: 0;
  transition: all 0.2s ease;
}
.kpi-label{
  font-size: 12px;
  color: #8A9BB5;
  font-weight: 600;
  font-family: 'Inter', sans-serif;
  letter-spacing: 0.6px;
  text-transform: uppercase;
  margin-bottom: 8px;
}

/* ✅ AUTO-FIT number inside card */
.kpi-value{
  font-size: clamp(17px, 2.1vw, 28px);
  font-weight: 800;
  color: #0D1B2E;
  letter-spacing: -0.5px;
  font-family: 'Inter', sans-serif;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}

/* Balance card — gold accent premium */
.kpi-card.balance{
  background: linear-gradient(145deg, rgba(10,38,71,0.96) 0%, rgba(15,56,110,0.96) 100%);
  border-color: rgba(180, 210, 255, 0.25);
  box-shadow: 0 6px 24px rgba(10, 38, 71, 0.25), 0 1px 3px rgba(0,0,0,0.1), inset 0 1px 0 rgba(255,255,255,0.08);
}
.kpi-card.balance .kpi-label{
  color: rgba(180, 205, 255, 0.75);
}
.kpi-card.balance .kpi-value{
  color: #FFFFFF;
}

/* Links inside cards look clean */
.kpi-link{
  text-decoration:none !important;
  color: inherit !important;
  display:block !important;
}
.kpi-link:hover .kpi-card{
  border-color: #7DAAEE;
  box-shadow: 0 8px 28px rgba(10, 38, 71, 0.14), inset 0 1px 0 rgba(255,255,255,0.95);
  transform: translateY(-2px);
}
.kpi-link:hover .kpi-card.balance{
  box-shadow: 0 10px 32px rgba(10, 38, 71, 0.35), inset 0 1px 0 rgba(255,255,255,0.08);
  transform: translateY(-2px);
}

/* Mobile */
@media (max-width: 1100px){
  .kpi-grid{ grid-template-columns: repeat(2, minmax(0, 1fr)); }
}
</style>
""",
    unsafe_allow_html=True,
)

def render_kpi_cards(net, paid, bal, rej, acc, balance_url: str, rejection_url: str = ""):
    """Premium KPI cards (Balance clickable, Rejected clickable)."""
    def fmt(x):
        try:
            return f"{float(x):,.2f}"
        except Exception:
            return "—"

    rejected_block = f"""
      <div class="kpi-card" title="{fmt(rej)}">
        <div class="kpi-label">Rejected</div>
        <div class="kpi-value">{fmt(rej)}</div>
      </div>
    """

    if rejection_url:
        rejected_block = f"""
      <a class="kpi-link" href="{rejection_url}" title="{fmt(rej)}">
        <div class="kpi-card" title="{fmt(rej)}">
          <div class="kpi-label">Rejected</div>
          <div class="kpi-value">{fmt(rej)}</div>
        </div>
      </a>
        """

    html = f"""
    <div class="kpi-grid">
      <div class="kpi-card" title="{fmt(net)}">
        <div class="kpi-label">Net Amount</div>
        <div class="kpi-value">{fmt(net)}</div>
      </div>

      <div class="kpi-card" title="{fmt(paid)}">
        <div class="kpi-label">Paid</div>
        <div class="kpi-value">{fmt(paid)}</div>
      </div>

      <!-- ✅ NEEDFUL: remove target=_blank so it opens in same app -->
      <a class="kpi-link" href="{balance_url}" title="{fmt(bal)}">
        <div class="kpi-card balance">
          <div class="kpi-label">Balance</div>
          <div class="kpi-value">{fmt(bal)}</div>
        </div>
      </a>

      {rejected_block}

      <div class="kpi-card" title="{fmt(acc)}">
        <div class="kpi-label">Accepted</div>
        <div class="kpi-value">{fmt(acc)}</div>
      </div>
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


# ====================== NEW: YEAR LANDING PAGE (NEEDFUL) ======================
def reset_year_selection():
    # back to landing page
    st.session_state.rcm_year = None
    st.session_state.center_key = None
    st.session_state.year = None
    try:
        if "center" in st.query_params:
            del st.query_params["center"]
        if "year" in st.query_params:
            del st.query_params["year"]
    except Exception:
        pass
    st.rerun()


def require_year_selection():
    """
    After view password:
    Show ONLY buttons:
      - Revenue Management Cycle 2024
      - Revenue Management Cycle 2025
      - Revenue Management Cycle 2026
    """
    if st.session_state.get("rcm_year") in (2024, 2025, 2026):
        return

    st.title("📊 Excellent Medical Group")
    st.caption("Revenue Cycle Management")
    st.markdown("### Select Report Year")

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("Revenue Management Cycle 2024", use_container_width=True, key="rcm_btn_2024"):
            st.session_state.rcm_year = 2024
            st.session_state.center_key = None
            st.session_state.year = 2024
            st.rerun()
    with c2:
        if st.button("Revenue Management Cycle 2025", use_container_width=True, key="rcm_btn_2025"):
            st.session_state.rcm_year = 2025
            st.session_state.center_key = None
            st.session_state.year = 2025
            st.rerun()
    with c3:
        if st.button("Revenue Management Cycle 2026", use_container_width=True, key="rcm_btn_2026"):
            st.session_state.rcm_year = 2026
            st.session_state.center_key = None
            st.session_state.year = 2026
            st.rerun()

    st.stop()


require_year_selection()
# =================================================================

BASE = Path(__file__).parent
DATA_DIR = BASE / "data"
(DATA_DIR / "easyhealth").mkdir(parents=True, exist_ok=True)
(DATA_DIR / "excellent").mkdir(parents=True, exist_ok=True)
(DATA_DIR / "excellent_pharmacy").mkdir(parents=True, exist_ok=True)

YEARS = [2024, 2025, 2026]

# Canonical sheet names for main Aging report
SHEET_INS_TOT = "Insurance_Totals"
SHEET_SUMMARY = "Balance_Aging_Summary"
SHEET_DETAIL = "Balance_Aging_Detail"
SHEET_INGROUP = "Balance_Aging_InsGroup"  # optional tab if present
SHEET_IPLAN = "Balance_Aging_Plan"        # optional tab if present (PHARMACY uses Plan)

# Robust Grand Total match (handles 'Grand Total', 'total', spacing, case)
GT_PAT = re.compile(r'^\s*(grand\s*total|total)\s*$', re.I)


# ====================== Centers config ======================
CENTERS = {
    "easyhealth": {
        "key": "easyhealth",
        "name": "Easy Health Medical Clinic (MF8031)",
        "folder_root": DATA_DIR / "easyhealth",
        "src_name": "source.xlsx",
        "out_name": "report.xlsx",
        "generator": BASE / "exclusive_report_with_aging_final.py",
    },
    "excellent": {
        "key": "excellent",
        "name": "Excellent Medical Center (MF4777)",
        "folder_root": DATA_DIR / "excellent",
        "src_name": "source.xlsx",
        "out_name": "report.xlsx",
        "generator": BASE / "exclusive_report_with_aging_final.py",
    },
    "pharmacy": {
        "key": "pharmacy",
        "name": "Excellent Pharmacy (PF3205)",
        "folder_root": DATA_DIR / "excellent_pharmacy",
        "src_name": "source.xlsx",
        "out_name": "Pharmacy_Exclusive_Report_with_Aging.xlsx",
        "generator": BASE / "pharmacy_exclusive_report_with_aging.py",
    },
}

# ====================== ✅ NEEDFUL: NAV HANDLER for Balance ======================
nav = st.query_params.get("nav")
if nav == "daily":
    # ✅ Direct-open the Daily Report page (skip center selection on home)
    # New tab => new session, so token is used to auto-auth above.
    y = st.query_params.get("year")
    if y:
        try:
            st.session_state.year = int(y)
            st.session_state.rcm_year = int(y)
        except Exception:
            pass
    st.switch_page(DAILY_REPORT_PAGE_PATH)

if nav == "balance":
    c = st.query_params.get("center")
    y = st.query_params.get("year")

    if c in CENTERS:
        st.session_state.center_key = c
    try:
        if y:
            y_int = int(y)
            st.session_state.year = y_int
            st.session_state.rcm_year = y_int
    except Exception:
        pass

    try:
        del st.query_params["nav"]
    except Exception:
        pass

    st.switch_page(BALANCE_PAGE_PATH)
# ================================================================================



# ====================== ✅ NEEDFUL S3 HELPERS (NEW) ======================
def _get_s3_cfg():
    access_key = st.secrets.get("AWS_ACCESS_KEY_ID", "")
    secret_key = st.secrets.get("AWS_SECRET_ACCESS_KEY", "")

    region = (
        st.secrets.get("AWS_REGION")
        or st.secrets.get("AWS_DEFAULT_REGION")
        or "eu-north-1"
    )

    bucket = (
        st.secrets.get("S3_BUCKET_NAME")
        or st.secrets.get("S3_BUCKET")
        or ""
    )

    prefix = st.secrets.get("S3_PREFIX", "").strip().strip("/")

    if not (access_key and secret_key and bucket):
        return None

    return {
        "access_key": access_key,
        "secret_key": secret_key,
        "region": region,
        "bucket": bucket,
        "prefix": prefix,
    }


@st.cache_resource(show_spinner=False)
def _s3_client_cached(access_key: str, secret_key: str, region: str):
    # Cache the boto3 client per-credential set to avoid recreating it on every rerun
    return boto3.client(
        "s3",
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        region_name=region,
    )

def _s3_client(cfg):
    return _s3_client_cached(cfg["access_key"], cfg["secret_key"], cfg["region"])


def s3_key_for(center_key: str, year: int, filename: str) -> str:
    """Build the S3 object key.

    We keep ONE consistent structure in S3:
      s3://<bucket>/streamlit/<center>/<year>/<filename>

    If you set secrets S3_PREFIX="streamlit", it will become:
      <prefix>/<center>/<year>/<filename>
    """
    cfg = _get_s3_cfg()
    if not cfg:
        return ""

    # If prefix is empty, default to "streamlit" to match your existing uploads
    prefix = (cfg.get("prefix") or "streamlit").strip().strip("/")
    return f"{prefix}/{center_key}/{year}/{filename}"


def upload_to_s3(local_path: Path, center_key: str, year: int) -> str:
    cfg = _get_s3_cfg()
    if cfg is None:
        return ""

    if not local_path.exists():
        raise FileNotFoundError(f"Local file not found: {local_path}")

    key = s3_key_for(center_key, year, local_path.name)
    client = _s3_client(cfg)

    try:
        client.upload_file(str(local_path), cfg["bucket"], key)
        return f"s3://{cfg['bucket']}/{key}"
    except ClientError as e:
        raise RuntimeError(f"S3 upload failed: {e}")


def download_from_s3(dest_path: Path, center_key: str, year: int, filename: str) -> bool:
    """Download a file from S3 to local disk (returns True if downloaded)."""
    cfg = _get_s3_cfg()
    if cfg is None:
        return False

    key = s3_key_for(center_key, year, filename)
    if not key:
        return False

    client = _s3_client(cfg)
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        client.download_file(cfg["bucket"], key, str(dest_path))
        return dest_path.exists() and dest_path.stat().st_size > 0
    except ClientError:
        return False



def resolve_existing_report(folder: Path, preferred_name: str) -> Path:
    """Return the best local report path (preferred -> report.xlsx -> newest excel)."""
    candidates = [preferred_name]
    if preferred_name.lower() != "report.xlsx":
        candidates.append("report.xlsx")
    # common variants
    candidates += ["Report.xlsx", "report.xlsb", "report.xlsm"]
    for name in candidates:
        p = folder / name
        if p.exists():
            return p
    # fallback: newest excel-like file
    for pattern in ("*.xlsx", "*.xlsb", "*.xlsm"):
        files = sorted(folder.glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True)
        if files:
            return files[0]
    return folder / preferred_name


def ensure_report_available(folder: Path, center_key: str, year: int, preferred_name: str) -> Path:
    """Ensure report exists locally; if missing, try to download from S3."""
    folder.mkdir(parents=True, exist_ok=True)

    # 1) local
    local_best = resolve_existing_report(folder, preferred_name)
    if local_best.exists():
        return local_best

    # 2) try S3 download (preferred, then report.xlsx)
    candidates = [preferred_name]
    if preferred_name.lower() != "report.xlsx":
        candidates.append("report.xlsx")

    for fn in candidates:
        dest = folder / fn
        if download_from_s3(dest, center_key, year, fn):
            return dest

    # nothing found
    return folder / preferred_name


# ====================== Small helpers ======================
def mtime_token(p: Path) -> float:
    try:
        return p.stat().st_mtime
    except FileNotFoundError:
        return 0.0

# --------- (rest of your script unchanged below) ---------

def _run(cmd):
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        raise RuntimeError(
            "Command failed:\n"
            + " ".join(cmd)
            + "\n\nSTDOUT:\n"
            + (res.stdout or "(empty)")
            + "\n\nSTDERR:\n"
            + (res.stderr or "(empty)")
        )
    return res


def rebuild_report(gen_path: Path, src_path: Path, out_path: Path) -> str:
    py = sys.executable
    out_path.parent.mkdir(parents=True, exist_ok=True)
    cmd = [py, str(gen_path), str(src_path)]
    try:
        res = _run(cmd + ["--out", str(out_path)])
        return res.stdout or "OK"
    except Exception:
        res = _run([py, str(gen_path), "--out", str(out_path), str(src_path)])
        return res.stdout or "OK"


def resolve_source_path(folder: Path, preferred: str = "source.xlsx") -> Path:
    for p in [folder / "source.xlsb", folder / "source.xlsx", folder / "source.xlsm"]:
        if p.exists():
            return p
    return folder / preferred


def save_uploaded_source(folder: Path, upload) -> Path:
    ext = Path(upload.name).suffix.lower()
    if ext not in {".xlsb", ".xlsx", ".xlsm"}:
        raise ValueError("Please upload an .xlsb, .xlsx, or .xlsm file.")
    dst = folder / f"source{ext}"
    folder.mkdir(parents=True, exist_ok=True)
    dst.write_bytes(upload.read())
    return dst


@st.cache_data(max_entries=6, show_spinner=False)
def get_report_bytes(path: str) -> bytes:
    return Path(path).read_bytes()


@st.cache_data(show_spinner=True)
def load_core_sheets(path: str, _token: float):
    """Load the two core sheets efficiently.

    Notes:
    - Uses ExcelFile once to avoid reopening the workbook multiple times.
    - _token (mtime) is included to invalidate cache when the file changes.
    """
    ext = Path(path).suffix.lower()
    engine = "pyxlsb" if ext == ".xlsb" else "openpyxl"
    try:
        xls = pd.ExcelFile(path, engine=engine)
        df_ins = pd.read_excel(xls, sheet_name=SHEET_INS_TOT)
        df_sum = pd.read_excel(xls, sheet_name=SHEET_SUMMARY)
        return df_ins, df_sum, [SHEET_INS_TOT, SHEET_SUMMARY]
    except Exception as e:
        try:
            names = pd.ExcelFile(path, engine=engine).sheet_names
        except Exception:
            names = []
        raise RuntimeError(
            f"Required sheets not found or failed to load. "
            f"Available: {', '.join(names) if names else '(none)'}\nOriginal error: {e}"
        )



def trim_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df2 = df.dropna(how="all")
    if df2.empty:
        return df2
    blank_rows = df2.fillna("").astype(str).apply(lambda r: "".join(r).strip() == "", axis=1)
    return df2.loc[~blank_rows]


def drop_empty_insurance(df: pd.DataFrame, name_col: str = "Insurance") -> pd.DataFrame:
    if df is None or df.empty or name_col not in df.columns:
        return df
    series = df[name_col].astype(str).fillna("").str.strip()
    bad = series.str.lower().isin(["", "none", "nan", "null", "na", "-", "--"])
    keep_grand = series.str.contains("grand total", case=False, na=False)
    return df.loc[~bad | keep_grand].copy()


def ensure_grand_total(df: pd.DataFrame, name_col: str = "Insurance") -> pd.DataFrame:
    """Ensure a Grand Total/Total row exists; if not, append one computed from numeric cols."""
    if df is None or df.empty or name_col not in df.columns:
        return df
    if df[name_col].astype(str).str.match(GT_PAT).any():
        return df
    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    gt_vals = {c: pd.to_numeric(df[c], errors="coerce").sum() for c in num_cols}
    row = {c: "" for c in df.columns}
    row.update(gt_vals)
    row[name_col] = "Grand Total"
    return pd.concat([df, pd.DataFrame([row])], ignore_index=True)


def move_grand_total_last(df: pd.DataFrame) -> pd.DataFrame:
    """Put the (Grand) Total row at the bottom; if missing, create it first."""
    if df is None or df.empty:
        return df
    first = df.columns[0]
    if not df[first].astype(str).str.match(GT_PAT).any():
        df = ensure_grand_total(df, first)
    mask = df[first].astype(str).str.match(GT_PAT)
    body = df.loc[~mask]
    gt = df.loc[mask]
    return pd.concat([body, gt], ignore_index=True)


def drop_gt(df: pd.DataFrame) -> pd.DataFrame:
    """Drop GT/Total rows (for KPI sums only)."""
    if df is None or df.empty:
        return df
    first = df.columns[0]
    return df.loc[~df[first].astype(str).str.match(GT_PAT)]


def full_height(df, row_px: int = 45, header_px: int = 70, padding_px: int = 150) -> int:
    n = 0 if df is None else len(df)
    return header_px + (n * row_px) + padding_px


def ksum(df: pd.DataFrame, *cands):
    for col in cands:
        if col in df.columns:
            return float(pd.to_numeric(df[col], errors="coerce").sum())
    return 0.0
    
def build_rejection_url(center, year):
    return f"/Rejection_Analysis?center={center}&year={year}"

# IMPORTANT: keep build_balance_url (nav=balance) defined earlier.
# This helper is only for display/debug if you need a plain page URL.
def build_balance_page_url(center, year):
    return f"/Balance_Attempt_Aging?center={center}&year={year}"
   
def is_admin_mode() -> bool:
    secret_pwd = st.secrets.get("ADMIN_PASSWORD", "")
    if secret_pwd:
        if st.session_state.get("is_admin", False):
            return True
        with st.popover("🔒 Admin login"):
            pwd = st.text_input("Password", type="password", key="admin_pwd")
            if st.button("Login", key="admin_login_btn"):
                if pwd == secret_pwd:
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    st.error("Wrong password")
        return False
    else:
        return st.toggle("Admin mode", value=st.session_state.get("is_admin", False))


# ====================== (Home KPIs helper) ======================
def load_center_kpis(center_key: str, year: int):
    """
    IMPORTANT:
    - Uses ONLY the selected year.
    - If report for that year is missing, returns zeros (NO fallback to other years).
    """
    cfg0 = CENTERS[center_key]

    if year not in YEARS:
        return year, 0.0, 0.0, 0.0, 0.0, 0.0

    folder_y = cfg0["folder_root"] / str(year)
    outp = ensure_report_available(folder_y, center_key, year, cfg0["out_name"])
    if not outp.exists():
        return year, 0.0, 0.0, 0.0, 0.0, 0.0

    tok = mtime_token(outp)
    if tok == 0.0:
        return year, 0.0, 0.0, 0.0, 0.0, 0.0

    try:
        totals, _, _ = load_core_sheets(str(outp), tok)
        totals = totals.copy()

        if "Insurance" not in totals.columns and len(totals.columns) > 0:
            totals = totals.rename(columns={totals.columns[0]: "Insurance"})

        for a in ["NetAmount", "Net amount", "Net"]:
            if a in totals.columns and "Net Amount" not in totals.columns:
                totals = totals.rename(columns={a: "Net Amount"})

        totals = trim_empty_rows(totals)
        totals = drop_empty_insurance(totals, "Insurance")
        totals = ensure_grand_total(totals, "Insurance")

        totals_no_gt = drop_gt(totals)

        net = ksum(totals_no_gt, "Net Amount", "NetAmount", "Net")
        paid = ksum(totals_no_gt, "Paid")
        bal = ksum(totals_no_gt, "Balance")
        rej = ksum(totals_no_gt, "Rejected", "Rejection")
        acc = ksum(totals_no_gt, "Accepted")

        return year, net, paid, bal, rej, acc
    except Exception:
        return year, 0.0, 0.0, 0.0, 0.0, 0.0


# ====================== Header & routing ======================
t1, t2, t3 = st.columns([6, 2, 2])
with t1:
    st.title("📊 Excellent Medical Group")
with t2:
    # ✅ Open External Daily Report (view page) in a NEW browser tab
    yr = int(st.session_state.get("rcm_year") or st.session_state.get("year") or 2026)
    ck = st.session_state.get("center_key") or st.query_params.get("center")
    # Build URL with auto-passed center/year when available
    params = []
    if ck:
        params.append(f"center={ck}")
    params.append(f"year={yr}")
    qp = ("?" + "&".join(params)) if params else ""
    # ✅ Daily Report button opens this app link (with center/year)
    base = DAILY_REPORT_BUTTON_BASE.rstrip("/")
    # keep same params for deep-linking
    daily_url = DAILY_REPORT_BUTTON_BASE

    st.markdown(f'<a class="navlink" href="{daily_url}" target="_blank">📅 Daily Report</a>', unsafe_allow_html=True)
with t3:
    if st.button("⬅ Change Year", use_container_width=True, key="btn_change_year"):
        reset_year_selection()

st.session_state.is_admin = is_admin_mode()

qs = st.query_params
if st.session_state.get("center_key") is None and qs.get("center"):
    ck_qs = qs.get("center")
    if ck_qs in CENTERS:
        st.session_state.center_key = ck_qs

if st.session_state.get("year") is None and qs.get("year"):
    try:
        st.session_state.year = int(qs.get("year"))
    except Exception:
        pass

if st.session_state.get("year") is None and st.session_state.get("rcm_year") in YEARS:
    st.session_state.year = st.session_state.get("rcm_year")

if (st.session_state.get("center_key") != st.session_state.get("last_center_key")) or \
   (st.session_state.get("year") != st.session_state.get("last_year")):
    load_core_sheets.clear()
    get_report_bytes.clear()
    st.session_state.last_center_key = st.session_state.get("center_key")
    st.session_state.last_year = st.session_state.get("year")

st.caption(
    f"Mode: **{'admin' if st.session_state.get('is_admin') else 'view'}** · "
    f"Center: **{st.session_state.get('center_key') or 'none'}** · "
    f"Year: **{st.session_state.get('year') or 'none'}**"
)
# ====================== ✅ HIDE EASYHEALTH IN 2024 (ONLY) ======================
# Block direct access via URL or session state when year selection is 2024
if st.session_state.get("rcm_year") == 2024:
    if st.session_state.get("center_key") == "easyhealth" or st.query_params.get("center") == "easyhealth":
        st.warning("Easy Health is available only in 2025.")
        st.session_state.center_key = None
        st.session_state.year = 2024
        try:
            if "center" in st.query_params:
                del st.query_params["center"]
            if "year" in st.query_params:
                del st.query_params["year"]
        except Exception:
            pass
        st.rerun()
# ============================================================================

# ====================== Home cards ======================
ck = st.session_state.get("center_key")
if ck not in CENTERS:
    st.subheader("Choose a center")

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.container(border=True).button(CENTERS["excellent"]["name"], use_container_width=True, key="home_exc"):
            st.session_state.center_key = "excellent"
            st.session_state.year = st.session_state.get("rcm_year")
            st.rerun()

    with c2:
        if st.container(border=True).button(CENTERS["pharmacy"]["name"], use_container_width=True, key="home_pharm"):
            st.session_state.center_key = "pharmacy"
            st.session_state.year = st.session_state.get("rcm_year")
            st.rerun()

    with c3:
        # ✅ EasyHealth hidden in 2024 only
        if st.session_state.get("rcm_year") != 2024:
            if st.container(border=True).button(CENTERS["easyhealth"]["name"], use_container_width=True, key="home_easy"):
                st.session_state.center_key = "easyhealth"
                st.session_state.year = st.session_state.get("rcm_year")
                st.rerun()

    st.markdown("---")
    st.subheader("Key metrics (All centers)")

    sel_year = st.session_state.get("rcm_year")

    y_exc, net_exc, paid_exc, bal_exc, rej_exc, acc_exc = load_center_kpis("excellent", sel_year)
    y_ph,  net_ph,  paid_ph,  bal_ph,  rej_ph,  acc_ph  = load_center_kpis("pharmacy", sel_year)

    # ✅ EasyHealth KPI section hidden in 2024 only
    if st.session_state.get("rcm_year") != 2024:
        y_eh, net_eh, paid_eh, bal_eh, rej_eh, acc_eh = load_center_kpis("easyhealth", sel_year)
    else:
        y_eh = sel_year
        net_eh = paid_eh = bal_eh = rej_eh = acc_eh = 0.0

    # ══════════════════════════════════════════════════════════════
    # ✅ PREMIUM CIRCLE DASHBOARD — donut charts per center
    # ══════════════════════════════════════════════════════════════
    def _build_circle_dashboard(
        year,
        net_exc, paid_exc, bal_exc, rej_exc, acc_exc,
        net_ph,  paid_ph,  bal_ph,  rej_ph,  acc_ph,
        net_eh,  paid_eh,  bal_eh,  rej_eh,  acc_eh,
        show_easyhealth: bool,
        bal_url_exc, bal_url_ph, bal_url_eh,
        rej_url_exc, rej_url_ph, rej_url_eh,
    ) -> str:
        import json

        centers_data = [
            {
                "cls": "c-exc", "accentColor": "#4F9EFF",
                "name": "Excellent Medical Center", "id": "MF4777", "tag": "Medical Center",
                "net": net_exc, "paid": paid_exc, "bal": bal_exc,
                "rej": rej_exc, "acc": acc_exc,
                "bal_url": bal_url_exc, "rej_url": rej_url_exc,
            },
            {
                "cls": "c-pharm", "accentColor": "#3DD9A0",
                "name": "Excellent Pharmacy", "id": "PF3205", "tag": "Pharmacy",
                "net": net_ph, "paid": paid_ph, "bal": bal_ph,
                "rej": rej_ph, "acc": acc_ph,
                "bal_url": bal_url_ph, "rej_url": rej_url_ph,
            },
        ]
        if show_easyhealth:
            centers_data.append({
                "cls": "c-easy", "accentColor": "#FFD166",
                "name": "Easy Health Medical Clinic", "id": "MF8031", "tag": "Medical Clinic",
                "net": net_eh, "paid": paid_eh, "bal": bal_eh,
                "rej": rej_eh, "acc": acc_eh,
                "bal_url": bal_url_eh, "rej_url": rej_url_eh,
            })

        centers_json = json.dumps(centers_data)

        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;900&family=DM+Sans:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
<style>
:root{{
  --bg:#06091A; --surface:rgba(255,255,255,0.04); --border:rgba(255,255,255,0.08);
  --text:#E8EEF8; --muted:rgba(180,195,230,0.55);
  --c-paid:#3DD9A0; --c-bal:#4F9EFF; --c-rej:#FF6B8A; --c-acc:#FFD166;
}}
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0;}}
body{{
  background:var(--bg);font-family:'DM Sans',sans-serif;color:var(--text);
  padding:24px 16px 40px;
}}
body::before{{
  content:'';position:fixed;inset:0;z-index:0;
  background:
    radial-gradient(ellipse 60% 50% at 15% 20%,rgba(79,158,255,.12) 0%,transparent 70%),
    radial-gradient(ellipse 50% 45% at 85% 75%,rgba(61,217,160,.10) 0%,transparent 70%),
    radial-gradient(ellipse 40% 35% at 50% 50%,rgba(255,107,138,.07) 0%,transparent 70%);
  pointer-events:none;
}}
body::after{{
  content:'';position:fixed;inset:0;z-index:0;
  background-image:linear-gradient(rgba(255,255,255,.02) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,.02) 1px,transparent 1px);
  background-size:60px 60px;pointer-events:none;
}}
.wrap{{position:relative;z-index:1;max-width:1280px;margin:0 auto;}}
.section-label{{
  font-size:.7rem;letter-spacing:.15em;text-transform:uppercase;color:var(--muted);
  margin-bottom:20px;display:flex;align-items:center;gap:10px;
}}
.section-label::after{{content:'';flex:1;height:1px;background:var(--border);}}
.cards-grid{{
  display:grid;grid-template-columns:repeat(auto-fit,minmax(340px,1fr));
  gap:24px;margin-bottom:40px;
}}
.card{{
  background:var(--surface);border:1px solid var(--border);border-radius:22px;
  padding:28px 24px 22px;backdrop-filter:blur(20px);position:relative;overflow:hidden;
  transition:transform .3s ease,box-shadow .3s ease;animation:fadeUp .5s ease both;
}}
.card:hover{{transform:translateY(-5px);box-shadow:0 20px 50px rgba(0,0,0,.4);}}
.card::before{{content:'';position:absolute;top:-80px;right:-80px;width:200px;height:200px;border-radius:50%;pointer-events:none;}}
.card.c-exc::before{{background:radial-gradient(circle,rgba(79,158,255,.18),transparent 70%);}}
.card.c-pharm::before{{background:radial-gradient(circle,rgba(61,217,160,.18),transparent 70%);}}
.card.c-easy::before{{background:radial-gradient(circle,rgba(255,209,102,.18),transparent 70%);}}
.card-header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:20px;}}
.card-name{{font-size:.92rem;font-weight:700;}}
.card-id{{font-size:.7rem;color:var(--muted);margin-top:3px;letter-spacing:.08em;}}
.card-tag{{font-size:.68rem;font-weight:700;padding:4px 11px;border-radius:50px;letter-spacing:.05em;}}
.card.c-exc  .card-tag{{background:rgba(79,158,255,.15);color:#7DC4FF;border:1px solid rgba(79,158,255,.25);}}
.card.c-pharm .card-tag{{background:rgba(61,217,160,.15);color:#5DE8B2;border:1px solid rgba(61,217,160,.25);}}
.card.c-easy .card-tag{{background:rgba(255,209,102,.15);color:#FFD166;border:1px solid rgba(255,209,102,.25);}}
.donut-row{{display:flex;align-items:center;gap:24px;margin-bottom:20px;}}
.donut-wrap{{flex-shrink:0;position:relative;width:150px;height:150px;}}
.donut-wrap svg{{transform:rotate(-90deg);}}
.donut-center{{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;pointer-events:none;}}
.dc-label{{font-size:.6rem;color:var(--muted);letter-spacing:.1em;text-transform:uppercase;margin-bottom:2px;}}
.dc-value{{font-size:1.1rem;font-weight:700;}}
.dc-pct{{font-size:.72rem;color:var(--muted);margin-top:1px;}}
.legend{{flex:1;display:flex;flex-direction:column;gap:10px;}}
.legend-item{{display:flex;align-items:center;gap:9px;}}
.legend-dot{{width:9px;height:9px;border-radius:3px;flex-shrink:0;}}
.legend-name{{font-size:.75rem;color:var(--muted);flex:1;}}
.legend-val{{font-size:.82rem;font-weight:700;}}
.legend-pct{{font-size:.72rem;color:var(--muted);min-width:36px;text-align:right;}}
.legend-link{{text-decoration:none!important;color:inherit!important;display:contents;}}
.legend-link:hover .legend-val{{text-decoration:underline;opacity:.85;}}
.metric-pills{{display:flex;gap:8px;flex-wrap:wrap;}}
.pill{{flex:1;min-width:90px;background:rgba(255,255,255,.03);border:1px solid var(--border);border-radius:10px;padding:9px 12px;transition:background .2s;}}
.pill:hover{{background:rgba(255,255,255,.06);}}
.p-label{{font-size:.62rem;color:var(--muted);letter-spacing:.08em;text-transform:uppercase;margin-bottom:3px;}}
.p-val{{font-size:.9rem;font-weight:700;}}
.summary-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:36px;}}
.sum-card{{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:18px 16px;display:flex;flex-direction:column;align-items:center;gap:5px;position:relative;overflow:hidden;transition:transform .25s;}}
.sum-card:hover{{transform:translateY(-3px);}}
.sum-card::after{{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;border-radius:0 0 16px 16px;}}
.sum-card.sc-paid::after{{background:#3DD9A0;box-shadow:0 0 10px rgba(61,217,160,.5);}}
.sum-card.sc-bal::after {{background:#4F9EFF;box-shadow:0 0 10px rgba(79,158,255,.5);}}
.sum-card.sc-rej::after {{background:#FF6B8A;box-shadow:0 0 10px rgba(255,107,138,.5);}}
.sum-card.sc-net::after {{background:#FFD166;box-shadow:0 0 10px rgba(255,209,102,.5);}}
.sc-icon{{font-size:1.3rem;}}
.sc-label{{font-size:.65rem;color:var(--muted);text-transform:uppercase;letter-spacing:.1em;}}
.sc-val{{font-size:1.05rem;font-weight:700;}}
.sc-sub{{font-size:.68rem;color:var(--muted);}}
.comp-row{{display:flex;flex-direction:column;gap:12px;}}
.comp-item{{display:flex;align-items:center;gap:12px;}}
.comp-name{{width:200px;flex-shrink:0;font-size:.76rem;color:var(--muted);}}
.comp-bar-track{{flex:1;height:9px;border-radius:50px;background:rgba(255,255,255,.06);overflow:hidden;}}
.comp-bar-fill{{height:100%;border-radius:50px;transition:width 1.4s cubic-bezier(.16,1,.3,1);width:0%;}}
.comp-val{{width:80px;text-align:right;font-size:.8rem;font-weight:700;}}
@keyframes fadeUp{{from{{opacity:0;transform:translateY(16px)}}to{{opacity:1;transform:translateY(0)}}}}
.card:nth-child(1){{animation-delay:.05s}}
.card:nth-child(2){{animation-delay:.12s}}
.card:nth-child(3){{animation-delay:.20s}}
@media(max-width:680px){{
  .summary-grid{{grid-template-columns:repeat(2,1fr);}}
  .cards-grid{{grid-template-columns:1fr;}}
  .comp-name{{width:110px;font-size:.68rem;}}
}}
</style>
</head>
<body>
<div class="wrap">
  <div class="section-label">Financial Performance by Center — Year {year}</div>
  <div class="cards-grid" id="cardsGrid"></div>
  <div class="section-label">Group Totals — All Centers Combined</div>
  <div class="summary-grid" id="summaryGrid"></div>
  <div class="section-label">Net Amount Comparison</div>
  <div class="comp-row" id="compRow"></div>
</div>
<script>
const centers={centers_json};
const fmt=n=>{{
  const v=Math.abs(n);
  if(v>=1e6)return (n/1e6).toFixed(2)+'M';
  if(v>=1e3)return (n/1e3).toFixed(1)+'K';
  return n.toLocaleString('en-US',{{minimumFractionDigits:2,maximumFractionDigits:2}});
}};
function buildDonut(c,r=64,sw=13){{
  const vals=[c.paid,c.bal,Math.max(c.rej,0),Math.max(c.acc,0)];
  const total=vals.reduce((s,v)=>s+v,0);
  const colors=['#3DD9A0','#4F9EFF','#FF6B8A','#FFD166'];
  const glows=['rgba(61,217,160,.6)','rgba(79,158,255,.6)','rgba(255,107,138,.6)','rgba(255,209,102,.6)'];
  const circ=2*Math.PI*r;
  let paths=`<circle cx="75" cy="75" r="${{r}}" fill="none" stroke="rgba(255,255,255,.05)" stroke-width="${{sw}}"/>`;
  let off=0;
  vals.forEach((v,i)=>{{
    const pct=total>0?v/total:0;
    const dash=pct*circ,gap=circ-dash;
    paths+=`<circle cx="75" cy="75" r="${{r}}" fill="none" stroke="${{colors[i]}}" stroke-width="${{sw}}"
      stroke-dasharray="${{dash}} ${{gap}}" stroke-dashoffset="${{{-off*circ}}}"
      style="filter:drop-shadow(0 0 5px ${{glows[i]}});transition:stroke-dasharray 1.2s cubic-bezier(.16,1,.3,1) ${{i*0.08}}s;"/>`;
    off+=pct;
  }});
  return `<svg width="150" height="150" viewBox="0 0 150 150">${{paths}}</svg>`;
}}
function renderCards(){{
  const grid=document.getElementById('cardsGrid');
  centers.forEach(c=>{{
    const vals=[c.paid,c.bal,Math.max(c.rej,0),Math.max(c.acc,0)];
    const total=vals.reduce((s,v)=>s+v,0);
    const pcts=vals.map(v=>total>0?(v/total*100).toFixed(1):0);
    const labels=['Paid','Balance','Rejected','Accepted'];
    const colors=['#3DD9A0','#4F9EFF','#FF6B8A','#FFD166'];
    const urls=[c.bal_url,c.bal_url,c.rej_url,c.rej_url];
    const legendHTML=labels.map((l,i)=>{{
      const isLink=(i===1||i===2)&&urls[i];
      const inner=`<div class="legend-dot" style="background:${{colors[i]}};box-shadow:0 0 5px ${{colors[i]}}77;"></div>
        <span class="legend-name">${{l}}</span>
        <span class="legend-val">${{fmt(vals[i])}}</span>
        <span class="legend-pct">${{pcts[i]}}%</span>`;
      return isLink
        ?`<div class="legend-item"><a class="legend-link" href="${{urls[i]}}" title="Open ${{l}} detail">${{inner}}</a></div>`
        :`<div class="legend-item">${{inner}}</div>`;
    }}).join('');
    const div=document.createElement('div');
    div.className=`card ${{c.cls}}`;
    div.innerHTML=`
      <div class="card-header">
        <div><div class="card-name">${{c.name}}</div><div class="card-id">${{c.id}}</div></div>
        <span class="card-tag">${{c.tag}}</span>
      </div>
      <div class="donut-row">
        <div class="donut-wrap">
          ${{buildDonut(c)}}
          <div class="donut-center">
            <span class="dc-label">Net</span>
            <span class="dc-value" style="color:${{c.accentColor}}">${{fmt(c.net)}}</span>
            <span class="dc-pct">AED</span>
          </div>
        </div>
        <div class="legend">${{legendHTML}}</div>
      </div>
      <div class="metric-pills">
        <div class="pill"><div class="p-label">Collection</div><div class="p-val" style="color:#3DD9A0">${{c.net>0?(c.paid/c.net*100).toFixed(1):0}}%</div></div>
        <div class="pill"><div class="p-label">Balance</div><div class="p-val" style="color:#4F9EFF">${{c.net>0?(c.bal/c.net*100).toFixed(1):0}}%</div></div>
        <div class="pill"><div class="p-label">Rejection</div><div class="p-val" style="color:#FF6B8A">${{c.net>0?(Math.max(c.rej,0)/c.net*100).toFixed(1):0}}%</div></div>
      </div>`;
    grid.appendChild(div);
  }});
}}
function renderSummary(){{
  const totNet=centers.reduce((s,c)=>s+c.net,0);
  const totPaid=centers.reduce((s,c)=>s+c.paid,0);
  const totBal=centers.reduce((s,c)=>s+c.bal,0);
  const totRej=centers.reduce((s,c)=>s+Math.max(c.rej,0),0);
  const items=[
    {{cls:'sc-paid',icon:'💚',label:'Total Paid',val:totPaid,sub:`${{(totPaid/totNet*100).toFixed(1)}}% of net`}},
    {{cls:'sc-bal', icon:'🔵',label:'Total Balance',val:totBal,sub:`${{(totBal/totNet*100).toFixed(1)}}% of net`}},
    {{cls:'sc-rej', icon:'🔴',label:'Total Rejected',val:totRej,sub:`${{(totRej/totNet*100).toFixed(1)}}% of net`}},
    {{cls:'sc-net', icon:'🟡',label:'Net Amount',val:totNet,sub:'All centers combined'}},
  ];
  const grid=document.getElementById('summaryGrid');
  items.forEach(it=>{{
    const div=document.createElement('div');
    div.className=`sum-card ${{it.cls}}`;
    div.innerHTML=`<div class="sc-icon">${{it.icon}}</div><div class="sc-label">${{it.label}}</div><div class="sc-val">${{fmt(it.val)}}</div><div class="sc-sub">${{it.sub}}</div>`;
    grid.appendChild(div);
  }});
}}
function renderComparison(){{
  const maxNet=Math.max(...centers.map(c=>c.net));
  const row=document.getElementById('compRow');
  centers.forEach(c=>{{
    const w=(c.net/maxNet*100).toFixed(1);
    const item=document.createElement('div');
    item.className='comp-item';
    item.innerHTML=`
      <div class="comp-name" style="color:${{c.accentColor}}44">${{c.name.replace('Excellent ','').replace('Easy Health ','')}} <span style="color:var(--muted);font-size:.68rem">(${{c.id}})</span></div>
      <div class="comp-bar-track"><div class="comp-bar-fill" style="background:linear-gradient(90deg,${{c.accentColor}}88,${{c.accentColor}});box-shadow:0 0 8px ${{c.accentColor}}44;" data-w="${{w}}"></div></div>
      <div class="comp-val" style="color:${{c.accentColor}}">${{fmt(c.net)}}</div>`;
    row.appendChild(item);
  }});
  requestAnimationFrame(()=>{{
    document.querySelectorAll('.comp-bar-fill').forEach(el=>{{
      setTimeout(()=>{{el.style.width=el.dataset.w+'%';}},120);
    }});
  }});
}}
renderCards();renderSummary();renderComparison();
</script>
</body>
</html>
"""

    # Build URLs for navigation
    _bal_url_exc = build_balance_url("excellent", sel_year)
    _bal_url_ph  = build_balance_url("pharmacy",  sel_year)
    _bal_url_eh  = build_balance_url("easyhealth", sel_year)
    _rej_url_exc = build_rejection_url("excellent", sel_year)
    _rej_url_ph  = build_rejection_url("pharmacy",  sel_year)
    _rej_url_eh  = build_rejection_url("easyhealth", sel_year)

    _show_easy = (st.session_state.get("rcm_year") != 2024)

    _circle_html = _build_circle_dashboard(
        year=sel_year,
        net_exc=net_exc, paid_exc=paid_exc, bal_exc=bal_exc, rej_exc=rej_exc, acc_exc=acc_exc,
        net_ph=net_ph,   paid_ph=paid_ph,   bal_ph=bal_ph,   rej_ph=rej_ph,   acc_ph=acc_ph,
        net_eh=net_eh,   paid_eh=paid_eh,   bal_eh=bal_eh,   rej_eh=rej_eh,   acc_eh=acc_eh,
        show_easyhealth=_show_easy,
        bal_url_exc=_bal_url_exc, bal_url_ph=_bal_url_ph, bal_url_eh=_bal_url_eh,
        rej_url_exc=_rej_url_exc, rej_url_ph=_rej_url_ph, rej_url_eh=_rej_url_eh,
    )

    # Render circle dashboard (height adjusts for 2 or 3 centers)
    _card_height = 720 if _show_easy else 580
    components.html(_circle_html, height=_card_height, scrolling=False)

    st.stop()

# ====================== MAIN aging dashboard ======================
if st.session_state.get("rcm_year") is None:
    st.subheader("Select Year")
    ycols = st.columns(len(YEARS))
    for i, y in enumerate(YEARS):
        with ycols[i]:
            if st.session_state.get("year") == y:
                st.markdown(
                    f"""
                    <div style="
                      background-color:#0B2D5C;color:white;text-align:center;
                      padding:0.85em;border-radius:14px;font-weight:900;font-size:1.1em;
                      border:2px solid #0B2D5C;
                      box-shadow: 0 6px 16px rgba(11,45,92,0.18);
                    ">
                      {y}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            else:
                if st.button(str(y), use_container_width=True, key=f"year_btn_{y}"):
                    st.session_state.year = y
                    st.rerun()

if st.session_state.get("year") is None:
    if st.session_state.get("rcm_year") in YEARS:
        st.session_state.year = st.session_state.get("rcm_year")
        st.rerun()

    cfg_tmp = CENTERS[st.session_state.get("center_key")]
    found = None
    for y in reversed(YEARS):
        out_try = (cfg_tmp["folder_root"] / str(y) / cfg_tmp["out_name"])
        if out_try.exists():
            found = y
            break
    st.session_state.year = found or YEARS[-1]
    st.rerun()

cfg = CENTERS[st.session_state.get("center_key")]
folder = cfg["folder_root"] / str(st.session_state.get("year"))
folder.mkdir(parents=True, exist_ok=True)

src_path = resolve_source_path(folder, preferred=cfg["src_name"])
out_path = ensure_report_available(folder, st.session_state.get("center_key"), st.session_state.get("year"), cfg["out_name"])
gen_path = cfg["generator"]

if (st.query_params.get("center") != st.session_state.get("center_key")) or \
   (st.query_params.get("year") != str(st.session_state.get("year"))):
    st.query_params["center"] = st.session_state.get("center_key")
    st.query_params["year"] = str(st.session_state.get("year"))

mt = mtime_token(out_path)
built = "—" if not mt else datetime.fromtimestamp(mt).strftime("%Y-%m-%d %H:%M")

st.markdown(
    f"""
    <div style="
        background: #F5FAFF;
        border: 1.5px solid #CFE3FF;
        padding: 14px 18px;
        border-radius: 16px;
        margin-bottom: 8px;
        box-shadow: 0 6px 18px rgba(11, 45, 92, 0.08);
    ">
        <div style="font-size:26px;font-weight:900;color:#0B2D5C;">
            {cfg["name"]}
        </div>
        <div style="font-size:14px;color:#334155;margin-top:2px;font-weight:750;">
            Year: {st.session_state.get("year")}
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.caption(f"Built: **{built}** · Source: {src_path} · Report: {out_path.name}")

if st.button("◀ Choose another center", key="btn_back_center"):
    st.session_state.center_key = None
    st.session_state.year = st.session_state.get("rcm_year")
    try:
        if "center" in st.query_params:
            del st.query_params["center"]
        if "year" in st.query_params:
            del st.query_params["year"]
    except Exception:
        st.experimental_set_query_params()
    st.rerun()

if st.session_state.get("is_admin"):
    st.success("You are in **ADMIN** mode — upload/rebuild is enabled.")

    with st.expander("⬆️ Upload/replace source Excel for this year", expanded=False):
        up = st.file_uploader(
            f"Upload source Excel for {st.session_state.get('year')} (.xlsb/.xlsx/.xlsm)",
            type=["xlsb", "xlsx", "xlsm"],
            key=f"uploader_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
        )
        if up:
            try:
                saved = save_uploaded_source(folder, up)
                st.success(f"Saved to {saved.name}")

                # ✅ NEEDFUL: upload source to S3 (optional but recommended)
                try:
                    s3_uri_src = upload_to_s3(saved, st.session_state.get("center_key"), st.session_state.get("year"))
                    if s3_uri_src:
                        st.info(f"Source uploaded to S3: {s3_uri_src}")
                except Exception as e:
                    st.warning(f"S3 source upload skipped/failed: {e}")

            except Exception as e:
                st.error(str(e))

    if st.button("↻ Rebuild report", use_container_width=True, key=f"rebuild_{st.session_state.get('center_key')}_{st.session_state.get('year')}"):
        try:
            if not gen_path.exists():
                st.error(f"Generator not found: {gen_path}")
            elif not src_path.exists():
                st.error("No source file found. Please upload source.xlsb/.xlsx first.")
            else:
                t0 = datetime.now()
                msg = rebuild_report(gen_path, src_path, out_path)
                t1 = datetime.now()
                st.success(f"Report rebuilt successfully in {(t1 - t0).total_seconds():.1f}s.")
                if msg.strip():
                    st.code(msg, language="bash")
                load_core_sheets.clear()
                get_report_bytes.clear()

                # ✅ NEEDFUL: upload rebuilt report to S3
                try:
                    s3_uri = upload_to_s3(out_path, st.session_state.get("center_key"), st.session_state.get("year"))
                    if s3_uri:
                        st.info(f"Report uploaded to S3: {s3_uri}")
                except Exception as e:
                    st.warning(f"S3 upload skipped/failed: {e}")

        except Exception as e:
            st.error(str(e))

token = mtime_token(out_path)
if token == 0.0:
    # ✅ NEEDFUL: If local report is missing (Streamlit Cloud redeploy/restart),
    # try to pull it back from S3 (because you already uploaded it).
    downloaded = download_from_s3(
        out_path,
        st.session_state.get("center_key"),
        st.session_state.get("year"),
        out_path.name,
    )
    if downloaded:
        load_core_sheets.clear()
        get_report_bytes.clear()
        token = mtime_token(out_path)

if token == 0.0:
    msg = f"Report not found for {cfg['name']} ({st.session_state.get('year')})."
    if st.session_state.get("is_admin"):
        msg += " (Upload source and click Rebuild.)"
    msg += " (If you uploaded to S3, check your S3 secrets: bucket/prefix/region.)"
    st.warning(msg)
    st.stop()

try:
    totals, summary, _ = load_core_sheets(str(out_path), token)
    totals = totals.copy()

    if "Insurance" not in totals.columns and len(totals.columns) > 0:
        totals = totals.rename(columns={totals.columns[0]: "Insurance"})

    for a in ["NetAmount", "Net amount", "Net"]:
        if a in totals.columns and "Net Amount" not in totals.columns:
            totals = totals.rename(columns={a: "Net Amount"})

    totals = trim_empty_rows(totals)
    totals = drop_empty_insurance(totals, "Insurance")
    totals = ensure_grand_total(totals, "Insurance")

    summary = trim_empty_rows(summary)
    if not summary.empty:
        summary = ensure_grand_total(summary, summary.columns[0])

    ext = Path(str(out_path)).suffix.lower()
    engine = "pyxlsb" if ext == ".xlsb" else "openpyxl"

    try:
        insgroup_df = pd.read_excel(str(out_path), sheet_name=SHEET_INGROUP, engine=engine)
        insgroup_df = trim_empty_rows(insgroup_df)
    except Exception:
        insgroup_df = None

    try:
        plan_df = pd.read_excel(str(out_path), sheet_name=SHEET_IPLAN, engine=engine)
        plan_df = trim_empty_rows(plan_df)
    except Exception:
        plan_df = None

    totals_no_gt = drop_gt(totals)

    net = ksum(totals_no_gt, "Net Amount", "NetAmount", "Net")
    paid = ksum(totals_no_gt, "Paid")
    bal = ksum(totals_no_gt, "Balance")
    rej = ksum(totals_no_gt, "Rejected", "Rejection")
    acc = ksum(totals_no_gt, "Accepted")

    st.markdown(f"### Key metrics — {st.session_state.get('year')}")
    render_kpi_cards(net, paid, bal, rej, acc,
                 build_balance_url(st.session_state.get("center_key"), st.session_state.get("year")),
                 rejection_url=build_rejection_url(st.session_state.get("center_key"), st.session_state.get("year")))
    st.markdown("---")

    tab_labels = [SHEET_INS_TOT, SHEET_SUMMARY]
    if insgroup_df is not None:
        tab_labels.append(SHEET_INGROUP)
    if plan_df is not None:
        tab_labels.append(SHEET_IPLAN)
    tab_labels.append("Downloads")

    if insgroup_df is not None and st.session_state.pop("_stay_on_ig", False):
        tab_labels = [SHEET_INGROUP] + [x for x in tab_labels if x != SHEET_INGROUP]

    t_tabs = st.tabs(tab_labels)
    tab_map = {name: t for name, t in zip(tab_labels, t_tabs)}

    t1 = tab_map[SHEET_INS_TOT]
    t2 = tab_map[SHEET_SUMMARY]
    t3 = tab_map["Downloads"]
    tIG = tab_map.get(SHEET_INGROUP)
    tPL = tab_map.get(SHEET_IPLAN)

    def _display_df(df: pd.DataFrame) -> pd.DataFrame:
        d = df.drop(columns=["S.No"], errors="ignore").reset_index(drop=True)
        d.index = range(1, len(d) + 1)
        d.index.name = None
        return d

    with t1:
        st.dataframe(
            _display_df(move_grand_total_last(totals)),
            use_container_width=True,
            height=full_height(totals),
        )
        st.download_button(
            "⬇️ Export Insurance Totals (CSV)",
            totals.to_csv(index=False).encode("utf-8"),
            file_name=f"{cfg['key']}_{st.session_state.get('year')}_insurance_totals.csv",
            use_container_width=True,
            key=f"dl_csv_totals_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
        )

    with t2:
        st.dataframe(
            _display_df(move_grand_total_last(summary)),
            use_container_width=True,
            height=full_height(summary),
        )
        st.download_button(
            "⬇️ Export Summary (CSV)",
            summary.to_csv(index=False).encode("utf-8"),
            file_name=f"{cfg['key']}_{st.session_state.get('year')}_summary.csv",
            use_container_width=True,
            key=f"dl_csv_summary_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
        )

    if tIG is not None and insgroup_df is not None:
        with tIG:
            insurers = (
                insgroup_df["Insurance"]
                .dropna()
                .astype(str)
                .loc[lambda s: ~s.str.match(GT_PAT)]
                .sort_values()
                .unique()
                .tolist()
            )
            ig_key = f"insgroup_select_{st.session_state.get('center_key')}_{st.session_state.get('year')}"

            with st.form(key=f"ig_form_{st.session_state.get('center_key')}_{st.session_state.get('year')}"):
                choice = st.selectbox(
                    "Filter by Insurance",
                    ["All"] + insurers,
                    index=(["All"] + insurers).index(st.session_state.get(ig_key, "All")),
                )
                apply_btn = st.form_submit_button("Apply")

            if apply_btn:
                st.session_state[ig_key] = choice
                st.session_state["_stay_on_ig"] = True
                st.rerun()

            choice = st.session_state.get(ig_key, "All")
            view_df = insgroup_df.copy()
            if choice != "All":
                view_df = (
                    view_df.loc[view_df["Insurance"].astype(str) == choice]
                    .drop(columns=["Insurance"], errors="ignore")
                )

            st.caption(f"Showing InsGroup aging for **{choice}**")
            st.dataframe(
                _display_df(move_grand_total_last(view_df)),
                use_container_width=True,
                height=full_height(view_df),
            )
            st.download_button(
                "⬇️ Export InsGroup (CSV) — current view",
                view_df.to_csv(index=False).encode("utf-8"),
                file_name=f"{cfg['key']}_{st.session_state.get('year')}_insgroup{'_' + choice if choice != 'All' else ''}.csv",
                use_container_width=True,
                key=f"dl_csv_insgroup_view_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
            )

    if tPL is not None and plan_df is not None:
        with tPL:
            insurers_pl = (
                plan_df["Insurance"]
                .dropna()
                .astype(str)
                .loc[lambda s: ~s.str.match(GT_PAT)]
                .sort_values()
                .unique()
                .tolist()
            )
            pl_key = f"plan_select_{st.session_state.get('center_key')}_{st.session_state.get('year')}"

            with st.form(key=f"pl_form_{st.session_state.get('center_key')}_{st.session_state.get('year')}"):
                choice_pl = st.selectbox(
                    "Filter by Insurance",
                    ["All"] + insurers_pl,
                    index=(["All"] + insurers_pl).index(st.session_state.get(pl_key, "All")),
                )
                apply_btn_pl = st.form_submit_button("Apply")

            if apply_btn_pl:
                st.session_state[pl_key] = choice_pl
                st.rerun()

            choice_pl = st.session_state.get(pl_key, "All")
            view_pl = plan_df.copy()
            if choice_pl != "All":
                view_pl = (
                    view_pl.loc[view_pl["Insurance"].astype(str) == choice_pl]
                    .drop(columns=["Insurance"], errors="ignore")
                )

            st.caption(f"Showing Plan aging for **{choice_pl}**")
            st.dataframe(
                _display_df(move_grand_total_last(view_pl)),
                use_container_width=True,
                height=full_height(view_pl),
            )
            st.download_button(
                "⬇️ Export Plan (CSV) — current view",
                view_pl.to_csv(index=False).encode("utf-8"),
                file_name=f"{cfg['key']}_{st.session_state.get('year')}_plan{'_' + choice_pl if choice_pl != 'All' else ''}.csv",
                use_container_width=True,
                key=f"dl_csv_plan_view_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
            )

    with t3:
        st.markdown("### Report Download")
        st.write("Open the XLSX locally to inspect **Balance_Aging_Detail** if needed.")
        st.download_button(
            "⬇️ Download full report (.xlsx)",
            get_report_bytes(str(out_path)),
            file_name=out_path.name,
            use_container_width=True,
            key=f"dl_xlsx_full_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
        )

        # ====================== MONTHLY BREAKDOWN (DOWNLOAD ONLY) ======================
        st.markdown("---")
        st.markdown("### 📅 Monthly Breakdown")

        try:
            _ext2 = Path(str(out_path)).suffix.lower()
            _eng2 = "pyxlsb" if _ext2 == ".xlsb" else "openpyxl"

            try:
                _monthly = pd.read_excel(str(out_path), sheet_name="Monthly_Totals", engine=_eng2)
                _monthly = trim_empty_rows(_monthly)
            except Exception:
                _monthly = None

            try:
                _mid = pd.read_excel(str(out_path), sheet_name="Monthly_Insurance_Detail", engine=_eng2)
                _mid = trim_empty_rows(_mid)
            except Exception:
                _mid = None

            if _monthly is None or _monthly.empty:
                st.info("Monthly_Totals sheet not found. Please rebuild the report.")
            else:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                # Month color palette (one per month, soft colors)
                _MONTH_COLORS = [
                    "DDEEFF", "D5F5E3", "FFF3CD", "F9EBEA", "E8DAEF",
                    "D6EAF8", "FDEBD0", "D0ECE7", "FDEDEC", "EBF5FB",
                    "F4ECF7", "E8F8F5",
                ]
                _HEADER_COLOR  = "2E4057"   # dark navy
                _GT_COLOR      = "FCE4D6"   # soft orange for Grand Total
                _MONTH_HDR_CLR = "1A5276"   # deep blue for month header row

                def _thin_border():
                    s = Side(style="thin", color="BBBBBB")
                    return Border(left=s, right=s, top=s, bottom=s)

                def _col_widths(ws):
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                max_len = max(max_len, len(str(cell.value or "")))
                            except Exception:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                wb = Workbook()

                # ── Sheet 1: Monthly_Totals ──────────────────────────────
                ws1 = wb.active
                ws1.title = "Monthly_Totals"

                # Sort months chronologically
                _month_order = ["January","February","March","April","May","June",
                                "July","August","September","October","November","December"]
                def _month_sort_key(val):
                    for i, m in enumerate(_month_order):
                        if str(val).startswith(m):
                            # extract year
                            parts = str(val).split()
                            yr = int(parts[1]) if len(parts) > 1 else 0
                            return (yr, i)
                    return (9999, 99)

                _monthly_sorted = _monthly[_monthly.iloc[:,0] != "Grand Total"].copy()
                _monthly_sorted = _monthly_sorted.sort_values(
                    by=_monthly_sorted.columns[0],
                    key=lambda s: s.map(_month_sort_key)
                ).reset_index(drop=True)
                _gt_row = _monthly[_monthly.iloc[:,0] == "Grand Total"]
                _monthly_sorted = pd.concat([_monthly_sorted, _gt_row], ignore_index=True)

                # Write header
                headers1 = list(_monthly_sorted.columns)
                for ci, h in enumerate(headers1, 1):
                    cell = ws1.cell(row=1, column=ci, value=h)
                    cell.fill = PatternFill("solid", fgColor=_HEADER_COLOR)
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _thin_border()
                ws1.row_dimensions[1].height = 22

                # Write data rows
                for ri, row in _monthly_sorted.iterrows():
                    is_gt = str(row.iloc[0]) == "Grand Total"
                    color = _GT_COLOR if is_gt else _MONTH_COLORS[ri % len(_MONTH_COLORS)]
                    for ci, val in enumerate(row, 1):
                        cell = ws1.cell(row=ri+2, column=ci, value=val)
                        cell.fill = PatternFill("solid", fgColor=color)
                        cell.font = Font(bold=is_gt, size=10)
                        cell.alignment = Alignment(horizontal="right" if ci > 1 else "left", vertical="center")
                        cell.border = _thin_border()
                        if ci > 1 and not is_gt:
                            try:
                                cell.number_format = "#,##0.00"
                            except Exception:
                                pass
                        if is_gt:
                            cell.font = Font(bold=True, size=11, color="8B0000")

                ws1.freeze_panes = "A2"
                _col_widths(ws1)

                # ── Sheet 2: Monthly_Insurance_Detail ───────────────────
                if _mid is not None and not _mid.empty:
                    ws2 = wb.create_sheet("Monthly_Insurance_Detail")

                    # Sort months
                    _all_months = sorted(
                        _mid["Month"].dropna().unique().tolist(),
                        key=_month_sort_key
                    )

                    # Write header row
                    _detail_cols = list(_mid.columns)  # Month, Insurance, Net Amount, ...
                    for ci, h in enumerate(_detail_cols, 1):
                        cell = ws2.cell(row=1, column=ci, value=h)
                        cell.fill = PatternFill("solid", fgColor=_HEADER_COLOR)
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = _thin_border()
                    ws2.row_dimensions[1].height = 22
                    ws2.freeze_panes = "A2"

                    cur_row = 2
                    for m_idx, _mon in enumerate(_all_months):
                        _df_mon = _mid[_mid["Month"] == _mon].drop(columns=["Month"]).copy()
                        _color = _MONTH_COLORS[m_idx % len(_MONTH_COLORS)]
                        _num_c = [c for c in _df_mon.columns if c != "Insurance"]

                        # Grand total row
                        _gt = {"Insurance": "Grand Total"}
                        for _nc in _num_c:
                            _gt[_nc] = pd.to_numeric(_df_mon[_nc], errors="coerce").sum()
                        _df_mon = pd.concat([_df_mon, pd.DataFrame([_gt])], ignore_index=True)

                        first_row_of_month = cur_row

                        for r_idx, row in _df_mon.iterrows():
                            is_gt = str(row.get("Insurance","")) == "Grand Total"
                            row_color = _GT_COLOR if is_gt else _color

                            # Col A: Month — only write on first data row, blank for rest
                            month_val = _mon if r_idx == 0 else ""
                            cell_a = ws2.cell(row=cur_row, column=1, value=month_val)
                            cell_a.fill = PatternFill("solid", fgColor=row_color)
                            cell_a.font = Font(bold=(r_idx == 0 or is_gt), size=10,
                                               color=("1A5276" if r_idx == 0 else ("8B0000" if is_gt else "000000")))
                            cell_a.alignment = Alignment(horizontal="left", vertical="center")
                            cell_a.border = _thin_border()

                            # Remaining cols
                            row_vals = [row.get(c, "") for c in _df_mon.columns]
                            for ci, val in enumerate(row_vals, 2):
                                cell = ws2.cell(row=cur_row, column=ci, value=val)
                                cell.fill = PatternFill("solid", fgColor=row_color)
                                cell.font = Font(bold=is_gt, size=10,
                                                 color=("8B0000" if is_gt else "000000"))
                                cell.alignment = Alignment(horizontal="right" if ci > 2 else "left",
                                                           vertical="center")
                                cell.border = _thin_border()
                                try:
                                    cell.number_format = "#,##0.00"
                                except Exception:
                                    pass

                            cur_row += 1

                    _col_widths(ws2)

                _buf = io.BytesIO()
                wb.save(_buf)
                _buf.seek(0)

                st.download_button(
                    "⬇️ Download Monthly Breakdown (Excel)",
                    _buf.read(),
                    file_name=f"{cfg['key']}_{st.session_state.get('year')}_monthly_breakdown.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_monthly_{st.session_state.get('center_key')}_{st.session_state.get('year')}",
                )

        except Exception as _me:
            st.warning(f"Monthly breakdown could not be generated: {_me}")

except Exception as e:
    try:
        ext = Path(str(out_path)).suffix.lower()
        eng = "pyxlsb" if ext == ".xlsb" else "openpyxl"
        names = pd.ExcelFile(str(out_path), engine=eng).sheet_names
    except Exception:
        names = []
    st.error(f"{e}\n\nAvailable sheets: {', '.join(names) if names else '(none)'}")









